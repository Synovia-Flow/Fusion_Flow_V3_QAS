#!/usr/bin/env python3
"""Fusion Flow V3 QAS - Database review / data-model analysis exporter.

Lists the schemas in the database, lets you select some (e.g. "1,3,5") or All,
and writes a STYLED Excel workbook to the Documentation_Layer\\Database_Analysis
output folder:

  - Tab 1  "DataModel Analysis"  : schema inventory, every table (clickable ->
                                    its data tab), and the foreign-key joins.
  - Tab 2+ one per table         : all rows of that table (capped by --max-rows),
                                    with a link back to the analysis tab.

Connection comes from Configuration/Fusion_Flow_QAS.ini ([database]).
Output root comes from CFG.Application_Parameters.DOCUMENTATION_OUTPUT_ROOT
(falling back to the known UNC path), under a Database_Analysis subfolder.

Usage:
  python export_datamodel.py                 # interactive schema picker
  python export_datamodel.py --all
  python export_datamodel.py --schemas CFG,ING,EXC
  python export_datamodel.py --schemas-only  # skip the per-table data tabs
"""

from __future__ import annotations

import argparse
import configparser
import os
import sys
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

os.environ.setdefault("NO_COLOR", "1")

REVIEW_DIR = Path(__file__).resolve().parent
REPO_ROOT = REVIEW_DIR.parents[1]
DEFAULT_INI = REPO_ROOT / "Configuration" / "Fusion_Flow_QAS.ini"
FALLBACK_DOC_ROOT = r"\\PL-AZ-SDF-PLINT\Fusion_Production\Synovia_Flow_Quality\Documentation_Layer"

HEADER_FILL = "1F4E78"   # navy
ACCENT_FILL = "2E75B6"   # teal/blue
TITLE_FILL = "0F243E"


# --------------------------------------------------------------------------- #
# Connection
# --------------------------------------------------------------------------- #
def load_db_config(ini_path: Path) -> dict[str, str]:
    if not ini_path.exists():
        raise FileNotFoundError(
            f"Connection file not found: {ini_path}. "
            f"Copy Fusion_Flow_QAS.example.ini to Fusion_Flow_QAS.ini and set the password.")
    parser = configparser.ConfigParser()
    parser.read(ini_path, encoding="utf-8")
    if "database" not in parser:
        raise ValueError(f"No [database] section in {ini_path}")
    return {k.lower(): v for k, v in parser["database"].items()}


def build_connection_string(db: dict[str, str]) -> str:
    parts = [
        f"Driver={db.get('driver', '{ODBC Driver 17 for SQL Server}')}",
        f"Server={db['server']}",
        f"Database={db['database']}",
    ]
    if db.get("user"):
        parts += [f"Uid={db['user']}", f"Pwd={db.get('password', '')}"]
    else:
        parts.append("Trusted_Connection=yes")
    yes = lambda v: str(v).lower() in ("yes", "true", "1")
    parts.append(f"Encrypt={'yes' if yes(db.get('encrypt', 'yes')) else 'no'}")
    parts.append(f"TrustServerCertificate={'yes' if yes(db.get('trust_server_certificate', 'no')) else 'no'}")
    return ";".join(parts) + ";"


def query(conn: Any, sql: str, *params: Any) -> list[dict[str, Any]]:
    cur = conn.cursor()
    cur.execute(sql, *params)
    cols = [c[0] for c in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]


# --------------------------------------------------------------------------- #
# Metadata reads
# --------------------------------------------------------------------------- #
def list_schemas(conn: Any) -> list[dict[str, Any]]:
    return query(conn, """
        SELECT s.name AS SchemaName, COUNT(t.object_id) AS TableCount
        FROM sys.schemas s
        LEFT JOIN sys.tables t ON t.schema_id = s.schema_id
        WHERE s.name NOT IN ('sys','guest','INFORMATION_SCHEMA')
          AND s.name NOT LIKE 'db[_]%'
        GROUP BY s.name
        HAVING COUNT(t.object_id) > 0
        ORDER BY s.name
    """)


def list_tables(conn: Any, schemas: list[str]) -> list[dict[str, Any]]:
    placeholders = ",".join("?" for _ in schemas)
    return query(conn, f"""
        SELECT s.name AS SchemaName, t.name AS TableName,
               (SELECT COUNT(*) FROM sys.columns c WHERE c.object_id = t.object_id) AS ColumnCount
        FROM sys.tables t
        JOIN sys.schemas s ON s.schema_id = t.schema_id
        WHERE s.name IN ({placeholders})
        ORDER BY s.name, t.name
    """, *schemas)


def list_foreign_keys(conn: Any, schemas: list[str]) -> list[dict[str, Any]]:
    placeholders = ",".join("?" for _ in schemas)
    return query(conn, f"""
        SELECT fk.name AS FKName,
               sp.name AS ParentSchema, tp.name AS ParentTable, cp.name AS ParentColumn,
               sr.name AS RefSchema,    tr.name AS RefTable,     cr.name AS RefColumn
        FROM sys.foreign_keys fk
        JOIN sys.foreign_key_columns fkc ON fkc.constraint_object_id = fk.object_id
        JOIN sys.tables  tp ON tp.object_id = fk.parent_object_id
        JOIN sys.schemas sp ON sp.schema_id = tp.schema_id
        JOIN sys.columns cp ON cp.object_id = tp.object_id AND cp.column_id = fkc.parent_column_id
        JOIN sys.tables  tr ON tr.object_id = fk.referenced_object_id
        JOIN sys.schemas sr ON sr.schema_id = tr.schema_id
        JOIN sys.columns cr ON cr.object_id = tr.object_id AND cr.column_id = fkc.referenced_column_id
        WHERE sp.name IN ({placeholders}) OR sr.name IN ({placeholders})
        ORDER BY sp.name, tp.name, fk.name
    """, *(schemas + schemas))


def row_count(conn: Any, schema: str, table: str) -> int:
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) FROM [{schema}].[{table}]")
    return int(cur.fetchone()[0])


def fetch_table(conn: Any, schema: str, table: str, max_rows: int) -> tuple[list[str], list[tuple]]:
    cur = conn.cursor()
    cur.execute(f"SELECT TOP ({max_rows}) * FROM [{schema}].[{table}]")
    cols = [c[0] for c in cur.description]
    return cols, cur.fetchall()


# --------------------------------------------------------------------------- #
# Selection
# --------------------------------------------------------------------------- #
def choose_schemas(schemas: list[dict[str, Any]], args: argparse.Namespace) -> list[str]:
    names = [s["SchemaName"] for s in schemas]
    if args.all:
        return names
    if args.schemas:
        wanted = {x.strip().upper() for x in args.schemas.split(",") if x.strip()}
        chosen = [n for n in names if n.upper() in wanted]
        if not chosen:
            raise SystemExit(f"None of {sorted(wanted)} match available schemas {names}")
        return chosen

    print("\nSchemas in this database:\n")
    for i, s in enumerate(schemas, 1):
        print(f"  {i:>2}. {s['SchemaName']:<6}  ({s['TableCount']} tables)")
    print("   A. All\n")
    raw = input("Select schemas (e.g. 1,3,5  or  A for All): ").strip()
    if raw.lower() in ("a", "all", "*"):
        return names
    chosen = []
    for tok in raw.replace(" ", "").split(","):
        if tok.isdigit() and 1 <= int(tok) <= len(names):
            chosen.append(names[int(tok) - 1])
    if not chosen:
        raise SystemExit("No valid selection made.")
    return chosen


# --------------------------------------------------------------------------- #
# Workbook
# --------------------------------------------------------------------------- #
def safe_sheet_name(base: str, used: set[str]) -> str:
    import re
    name = re.sub(r"[\\/*?:\[\]]", "_", base)[:31]
    candidate, n = name, 1
    while candidate.lower() in used:
        suffix = f"~{n}"
        candidate = name[:31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def sanitize(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool, datetime)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (bytes, bytearray)):
        return f"0x{value.hex()[:60]}" + ("..." if len(value) > 30 else "")
    return str(value)


def style_header(ws, ncols: int, fill_hex: str = HEADER_FILL) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    if ncols:
        from openpyxl.utils import get_column_letter
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def autosize(ws, ncols: int, sample_rows: int = 200) -> None:
    from openpyxl.utils import get_column_letter
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        width = 10
        for row in range(1, min(ws.max_row, sample_rows) + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                width = max(width, min(len(str(v)) + 2, 60))
        ws.column_dimensions[letter].width = width


def internal_link(sheet_name: str) -> str:
    return f"#'{sheet_name}'!A1"


def build_workbook(conn: Any, chosen: list[str], schemas_meta: list[dict[str, Any]],
                   include_data: bool, max_rows: int) -> Any:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    db_cfg_name = query(conn, "SELECT DB_NAME() AS n")[0]["n"]
    tables = list_tables(conn, chosen)
    fks = list_foreign_keys(conn, chosen)

    # Pre-compute sheet names + row counts for every table.
    used_names: set[str] = {"datamodel analysis"}
    sheet_for: dict[tuple[str, str], str] = {}
    counts: dict[tuple[str, str], int] = {}
    for t in tables:
        key = (t["SchemaName"], t["TableName"])
        counts[key] = row_count(conn, *key)
        sheet_for[key] = safe_sheet_name(f"{t['SchemaName']}.{t['TableName']}", used_names)

    wb = Workbook()
    front = wb.active
    front.title = "DataModel Analysis"

    title_font = Font(color="FFFFFF", bold=True, size=16)
    sub_font = Font(color="FFFFFF", bold=True)
    link_font = Font(color="0563C1", underline="single")

    front["A1"] = "Fusion Flow V3 QAS - DataModel Analysis"
    front["A1"].font = title_font
    front["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    front.merge_cells("A1:F1")
    front.row_dimensions[1].height = 26

    meta = [
        ("Database", db_cfg_name),
        ("Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
        ("Schemas", ", ".join(chosen)),
        ("Tables", str(len(tables))),
        ("Foreign keys", str(len(fks))),
    ]
    r = 3
    for label, val in meta:
        front.cell(row=r, column=1, value=label).font = Font(bold=True)
        front.cell(row=r, column=2, value=val)
        r += 1

    # --- Schema summary ---
    r += 1
    front.cell(row=r, column=1, value="SCHEMAS").font = sub_font
    front.cell(row=r, column=1).fill = PatternFill("solid", fgColor=ACCENT_FILL)
    front.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    head_row = r
    for i, h in enumerate(["Schema", "Tables", "Rows"], 1):
        front.cell(row=r, column=i, value=h)
    for col in range(1, 4):
        c = front.cell(row=head_row, column=col); c.font = Font(color="FFFFFF", bold=True)
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
    r += 1
    chosen_set = set(chosen)
    for sname in chosen:
        s_tables = [t for t in tables if t["SchemaName"] == sname]
        s_rows = sum(counts[(t["SchemaName"], t["TableName"])] for t in s_tables)
        front.cell(row=r, column=1, value=sname)
        front.cell(row=r, column=2, value=len(s_tables))
        front.cell(row=r, column=3, value=s_rows)
        r += 1

    # --- Tables index (clickable) ---
    r += 1
    front.cell(row=r, column=1, value="TABLES (click a table name to open its data tab)").font = sub_font
    front.cell(row=r, column=1).fill = PatternFill("solid", fgColor=ACCENT_FILL)
    front.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    for i, h in enumerate(["Schema", "Table", "Columns", "Rows"], 1):
        c = front.cell(row=r, column=i, value=h)
        c.font = Font(color="FFFFFF", bold=True); c.fill = PatternFill("solid", fgColor=HEADER_FILL)
    r += 1
    for t in tables:
        key = (t["SchemaName"], t["TableName"])
        front.cell(row=r, column=1, value=t["SchemaName"])
        cell = front.cell(row=r, column=2, value=t["TableName"])
        if include_data:
            cell.hyperlink = internal_link(sheet_for[key])
            cell.font = link_font
        front.cell(row=r, column=3, value=t["ColumnCount"])
        front.cell(row=r, column=4, value=counts[key])
        r += 1

    # --- Relationships / joins ---
    r += 1
    front.cell(row=r, column=1, value="RELATIONSHIPS (FOREIGN KEY JOINS)").font = sub_font
    front.cell(row=r, column=1).fill = PatternFill("solid", fgColor=ACCENT_FILL)
    front.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    for i, h in enumerate(["FK Name", "From (table.column)", "->", "To (table.column)"], 1):
        c = front.cell(row=r, column=i, value=h)
        c.font = Font(color="FFFFFF", bold=True); c.fill = PatternFill("solid", fgColor=HEADER_FILL)
    r += 1
    if fks:
        for fk in fks:
            front.cell(row=r, column=1, value=fk["FKName"])
            front.cell(row=r, column=2, value=f"{fk['ParentSchema']}.{fk['ParentTable']}.{fk['ParentColumn']}")
            front.cell(row=r, column=3, value="->")
            front.cell(row=r, column=4, value=f"{fk['RefSchema']}.{fk['RefTable']}.{fk['RefColumn']}")
            r += 1
    else:
        front.cell(row=r, column=1, value="(no foreign keys in the selected schemas)")
        r += 1

    for letter, width in (("A", 22), ("B", 40), ("C", 10), ("D", 40), ("E", 14), ("F", 14)):
        front.column_dimensions[letter].width = width

    # --- Per-table data tabs ---
    if include_data:
        for t in tables:
            key = (t["SchemaName"], t["TableName"])
            ws = wb.create_sheet(sheet_for[key])
            back = ws.cell(row=1, column=1, value="<- DataModel Analysis")
            back.hyperlink = internal_link("DataModel Analysis")
            back.font = link_font
            ws.cell(row=2, column=1, value=f"{t['SchemaName']}.{t['TableName']}").font = Font(bold=True, size=12)
            cols, rows = fetch_table(conn, key[0], key[1], max_rows)
            header_row = 4
            for ci, cname in enumerate(cols, 1):
                c = ws.cell(row=header_row, column=ci, value=cname)
                c.font = Font(color="FFFFFF", bold=True)
                c.fill = PatternFill("solid", fgColor=HEADER_FILL)
            for ri, row in enumerate(rows, header_row + 1):
                for ci, val in enumerate(row, 1):
                    ws.cell(row=ri, column=ci, value=sanitize(val))
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
            autosize(ws, len(cols))
            if counts[key] > max_rows:
                ws.cell(row=3, column=1,
                        value=f"NOTE: showing first {max_rows} of {counts[key]} rows (raise with --max-rows).").font = Font(color="C00000", italic=True)

    return wb


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #
def resolve_output_dir(conn: Any, override: str | None) -> Path:
    if override:
        return Path(override)
    rows = query(conn, "SELECT ParameterValue FROM CFG.Application_Parameters "
                       "WHERE ParameterKey = 'DOCUMENTATION_OUTPUT_ROOT' AND IsActive = 1")
    root = rows[0]["ParameterValue"] if rows else FALLBACK_DOC_ROOT
    return Path(root) / "Database_Analysis"


def main() -> int:
    p = argparse.ArgumentParser(description="Fusion Flow V3 QAS - data-model analysis exporter.")
    p.add_argument("--ini", type=Path, default=DEFAULT_INI)
    p.add_argument("--server", help="Override server from the .ini")
    p.add_argument("--schemas", help="Comma list, e.g. CFG,ING (non-interactive)")
    p.add_argument("--all", action="store_true", help="Select all schemas")
    p.add_argument("--schemas-only", action="store_true", help="Skip per-table data tabs")
    p.add_argument("--max-rows", type=int, default=100000, help="Max rows per data tab (default 100000)")
    p.add_argument("--out", help="Output folder override")
    args = p.parse_args()

    db_cfg = load_db_config(args.ini)
    if args.server:
        db_cfg["server"] = args.server

    import pyodbc  # lazy
    conn = pyodbc.connect(build_connection_string(db_cfg), autocommit=True)
    try:
        schemas = list_schemas(conn)
        if not schemas:
            print("No user schemas with tables found."); return 1
        chosen = choose_schemas(schemas, args)
        print(f"\nSelected: {', '.join(chosen)}")
        wb = build_workbook(conn, chosen, schemas, include_data=not args.schemas_only, max_rows=args.max_rows)

        out_dir = resolve_output_dir(conn, args.out)
        out_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        out_path = out_dir / f"Fusion_DataModel_Analysis_{stamp}.xlsx"
        wb.save(out_path)
        print(f"\nWorkbook written: {out_path}")
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
