#!/usr/bin/env python3
"""Fusion Flow V3 QAS - full database snapshot to Excel.

Downloads EVERY base table in the database to a single .xlsx workbook:

    * one worksheet PER populated table (header row + all rows)
    * tables with 0 rows are NOT given their own (empty) tab - they are listed
      together on a single "Zero Records" tab
    * a "Summary" tab listing every table with its row count, column count,
      whether it has a primary key, and which worksheet holds its data
    * a "Column Analysis" tab - the detailed analysis: every column of every
      table with data type, length/precision, nullability, PK / identity flags,
      default, and (for populated tables) the non-null count and % populated.

No CLI by design (matches the other runners). The connection comes from
Configuration/Fusion_Flow_QAS.ini [database]; the output folder comes from
CFG.Application_Parameters 'DB_SNAPSHOT_OUTPUT_DIR' (falling back to
'DOCUMENTATION_OUTPUT_ROOT'), i.e. by default:

    \\\\PL-AZ-SDF-PLINT\\Fusion_Production\\Synovia_Flow_Quality\\Documentation_Layer

An optional single positional argument overrides the output folder for ad-hoc runs:

    python export_db_snapshot.py "D:\\some\\other\\folder"
"""

from __future__ import annotations

import configparser
import os
import sys
from datetime import date, datetime, time as dtime
from decimal import Decimal
from pathlib import Path
from typing import Any

HERE = Path(__file__).resolve()
REPO_ROOT = HERE.parents[2]                       # .../Modules/Global/<file> -> repo root
DEFAULT_INI = REPO_ROOT / "Configuration" / "Fusion_Flow_QAS.ini"

# Fallback output root (also held in CFG; CFG wins when present).
FALLBACK_OUTPUT_DIR = r"\\PL-AZ-SDF-PLINT\Fusion_Production\Synovia_Flow_Quality\Documentation_Layer"

# Guards. Excel allows 1,048,576 rows per sheet (1 used by the header) and
# 32,767 chars per cell; we stay safely inside both and note any truncation.
MAX_DATA_ROWS = 1_048_575
MAX_CELL_CHARS = 32_767

# Header styling (Synovia blue).
HEADER_FILL = "FF1F3A5F"
HEADER_FONT = "FFFFFFFF"


# --------------------------------------------------------------------------- #
# Config / connection (self-contained, mirrors the other Global scripts).
# --------------------------------------------------------------------------- #
def load_db_config(ini_path: Path) -> dict[str, str]:
    if not ini_path.exists():
        raise FileNotFoundError(
            f"Connection file not found: {ini_path}. Copy Fusion_Flow_QAS.example.ini "
            f"to Fusion_Flow_QAS.ini and set the password.")
    cp = configparser.ConfigParser()
    cp.read(ini_path, encoding="utf-8")
    if "database" not in cp:
        raise ValueError(f"No [database] section in {ini_path}")
    return {k.lower(): v for k, v in cp["database"].items()}


def conn_str(db: dict[str, str]) -> str:
    parts = [
        f"Driver={db.get('driver', '{ODBC Driver 17 for SQL Server}')}",
        f"Server={db['server']}",
        f"Database={db['database']}",
    ]
    if db.get("user"):
        parts += [f"Uid={db['user']}", f"Pwd={db.get('password', '')}"]
    else:
        parts.append("Trusted_Connection=yes")
    parts.append(f"Encrypt={'yes' if db.get('encrypt', 'yes').lower() in ('yes', 'true', '1') else 'no'}")
    parts.append(f"TrustServerCertificate="
                 f"{'yes' if db.get('trust_server_certificate', 'no').lower() in ('yes', 'true', '1') else 'no'}")
    return ";".join(parts) + ";"


def q(cur, sql: str, *params: Any) -> list[dict[str, Any]]:
    cur.execute(sql, *params)
    cols = [c[0] for c in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]


def fetch_param(cur, key: str, default: str = "") -> str:
    try:
        rows = q(cur, "SELECT ParameterValue FROM CFG.Application_Parameters "
                      "WHERE ParameterKey = ? AND IsActive = 1", key)
    except Exception:
        return default
    return rows[0]["ParameterValue"] if rows and rows[0]["ParameterValue"] is not None else default


# --------------------------------------------------------------------------- #
# Catalog introspection
# --------------------------------------------------------------------------- #
def list_tables(cur) -> list[dict[str, Any]]:
    """Every base table with an accurate row count, ordered schema, table."""
    rows = q(cur, """
        SELECT s.name AS SchemaName, t.name AS TableName,
               SUM(CASE WHEN p.index_id IN (0,1) THEN p.rows ELSE 0 END) AS ApproxRows
        FROM sys.tables t
        JOIN sys.schemas s ON s.schema_id = t.schema_id
        JOIN sys.partitions p ON p.object_id = t.object_id
        GROUP BY s.name, t.name
        ORDER BY s.name, t.name""")
    out: list[dict[str, Any]] = []
    for r in rows:
        sch, tab = r["SchemaName"], r["TableName"]
        # Accurate count (partition rows can be stale/approximate).
        try:
            cnt = q(cur, f"SELECT COUNT_BIG(*) AS n FROM [{sch}].[{tab}]")[0]["n"]
        except Exception:
            cnt = int(r["ApproxRows"] or 0)
        out.append({"schema": sch, "table": tab, "rows": int(cnt)})
    return out


def table_columns(cur, schema: str, table: str) -> list[dict[str, Any]]:
    """Structural metadata for every column, including PK / identity / default."""
    return q(cur, """
        SELECT c.ORDINAL_POSITION              AS Ordinal,
               c.COLUMN_NAME                    AS ColumnName,
               c.DATA_TYPE                      AS DataType,
               c.CHARACTER_MAXIMUM_LENGTH       AS CharLen,
               c.NUMERIC_PRECISION              AS NumPrecision,
               c.NUMERIC_SCALE                  AS NumScale,
               c.IS_NULLABLE                    AS IsNullable,
               c.COLUMN_DEFAULT                 AS ColumnDefault,
               COLUMNPROPERTY(OBJECT_ID(QUOTENAME(c.TABLE_SCHEMA)+'.'+QUOTENAME(c.TABLE_NAME)),
                              c.COLUMN_NAME, 'IsIdentity') AS IsIdentity,
               CASE WHEN pk.COLUMN_NAME IS NOT NULL THEN 1 ELSE 0 END AS IsPrimaryKey
        FROM INFORMATION_SCHEMA.COLUMNS c
        LEFT JOIN (
            SELECT ku.TABLE_SCHEMA, ku.TABLE_NAME, ku.COLUMN_NAME
            FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS tc
            JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE ku
              ON tc.CONSTRAINT_NAME = ku.CONSTRAINT_NAME
             AND tc.TABLE_SCHEMA = ku.TABLE_SCHEMA
            WHERE tc.CONSTRAINT_TYPE = 'PRIMARY KEY'
        ) pk ON pk.TABLE_SCHEMA = c.TABLE_SCHEMA
            AND pk.TABLE_NAME = c.TABLE_NAME
            AND pk.COLUMN_NAME = c.COLUMN_NAME
        WHERE c.TABLE_SCHEMA = ? AND c.TABLE_NAME = ?
        ORDER BY c.ORDINAL_POSITION""", schema, table)


def nonnull_counts(cur, schema: str, table: str, cols: list[str]) -> dict[str, int]:
    """One-scan non-null count per column (COUNT(col) ignores NULLs)."""
    if not cols:
        return {}
    select = ", ".join(f"COUNT([{c}]) AS [c{i}]" for i, c in enumerate(cols))
    try:
        row = q(cur, f"SELECT {select} FROM [{schema}].[{table}]")[0]
    except Exception:
        return {}
    return {c: int(row[f"c{i}"] or 0) for i, c in enumerate(cols)}


# --------------------------------------------------------------------------- #
# Value coercion for Excel cells
# --------------------------------------------------------------------------- #
def cell_value(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        return v.replace(tzinfo=None)            # openpyxl rejects tz-aware datetimes
    if isinstance(v, (date, dtime)):
        return v
    if isinstance(v, (bytes, bytearray, memoryview)):
        s = bytes(v).hex()
    else:
        s = str(v)
    return s if len(s) <= MAX_CELL_CHARS else s[:MAX_CELL_CHARS - 14] + "...[TRUNCATED]"


def safe_sheet_name(raw: str, used: set[str]) -> str:
    """Excel sheet names: <=31 chars, none of []:*?/\\, unique."""
    name = raw
    for ch in "[]:*?/\\":
        name = name.replace(ch, "_")
    name = name.strip() or "Sheet"
    name = name[:31]
    base, i = name, 1
    while name.lower() in used:
        suffix = f"~{i}"
        name = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(name.lower())
    return name


# --------------------------------------------------------------------------- #
# Workbook build
# --------------------------------------------------------------------------- #
def build_workbook(cur, database: str, generated_at: datetime):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    head_fill = PatternFill("solid", fgColor=HEADER_FILL)
    head_font = Font(bold=True, color=HEADER_FONT)
    link_font = Font(color="FF0563C1", underline="single")
    wrap_top = Alignment(vertical="top", wrap_text=False)

    def sheet_link(label: str) -> str:
        """Internal hyperlink to cell A1 of another worksheet (quotes doubled)."""
        return f"#'{label.replace(chr(39), chr(39) * 2)}'!A1"

    def style_header(ws, ncols: int) -> None:
        for ci in range(1, ncols + 1):
            c = ws.cell(row=1, column=ci)
            c.fill = head_fill
            c.font = head_font
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"

    def autosize(ws, ncols: int, sample_rows: int = 200) -> None:
        for ci in range(1, ncols + 1):
            width = 8
            for ri in range(1, min(ws.max_row, sample_rows + 1) + 1):
                val = ws.cell(row=ri, column=ci).value
                if val is not None:
                    width = max(width, min(60, len(str(val)) + 2))
            ws.column_dimensions[get_column_letter(ci)].width = width

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    zero = wb.create_sheet("Zero Records")
    analysis = wb.create_sheet("Column Analysis")

    used_names = {"summary", "zero records", "column analysis"}
    tables = list_tables(cur)

    # --- per-table data tabs + collect summary/analysis rows ---------------- #
    summary_rows: list[list[Any]] = []
    zero_rows: list[list[Any]] = []
    analysis_rows: list[list[Any]] = []

    for idx, t in enumerate(tables, 1):
        sch, tab, rows = t["schema"], t["table"], t["rows"]
        cols_meta = table_columns(cur, sch, tab)
        col_names = [c["ColumnName"] for c in cols_meta]
        has_pk = any(c["IsPrimaryKey"] for c in cols_meta)
        nn = nonnull_counts(cur, sch, tab, col_names) if rows else {}

        if rows == 0:
            sheet_label = "Zero Records"
            zero_rows.append([sch, tab, len(col_names), "Yes" if has_pk else "No"])
        else:
            sheet_label = safe_sheet_name(f"{sch}.{tab}", used_names)
            ws = wb.create_sheet(sheet_label)
            ws.append(col_names)
            style_header(ws, len(col_names))
            written = 0
            cur.execute(f"SELECT * FROM [{sch}].[{tab}]")
            colnames = [d[0] for d in cur.description]
            for r in cur.fetchall():
                if written >= MAX_DATA_ROWS:
                    break
                ws.append([cell_value(v) for v in r])
                written += 1
            autosize(ws, len(colnames))
            if written < rows:
                t["note"] = f"TRUNCATED to {written:,} of {rows:,} rows (Excel sheet limit)"

        summary_rows.append([
            idx, sch, tab, rows, len(col_names),
            "Yes" if has_pk else "No",
            "EMPTY" if rows == 0 else "POPULATED",
            sheet_label, t.get("note", ""),
        ])

        for c in cols_meta:
            length = c["CharLen"]
            if length is None and c["NumPrecision"] is not None:
                length = f"{c['NumPrecision']},{c['NumScale']}" if c["NumScale"] is not None else c["NumPrecision"]
            nn_c = nn.get(c["ColumnName"])
            pct = round(100.0 * nn_c / rows, 1) if rows and nn_c is not None else ("" if rows else 0)
            analysis_rows.append([
                sch, tab, c["Ordinal"], c["ColumnName"], c["DataType"],
                length if length is not None else "",
                "Yes" if c["IsNullable"] == "YES" else "No",
                "Yes" if c["IsPrimaryKey"] else "No",
                "Yes" if c["IsIdentity"] else "No",
                (c["ColumnDefault"] or "")[:200] if c["ColumnDefault"] else "",
                "" if not rows else (nn_c if nn_c is not None else ""),
                pct,
            ])

    # --- Summary tab -------------------------------------------------------- #
    total_tables = len(tables)
    populated = sum(1 for t in tables if t["rows"])
    empties = total_tables - populated
    total_rows = sum(t["rows"] for t in tables)

    summary.append([f"Fusion Flow V3 QAS - Database Snapshot: {database}"])
    summary.append([f"Generated (UTC): {generated_at:%Y-%m-%d %H:%M:%S}"])
    summary.append([f"Tables: {total_tables}   Populated: {populated}   "
                    f"Empty: {empties}   Total rows: {total_rows:,}"])
    summary.append([])
    header = ["#", "Schema", "Table", "Rows", "Columns", "Has PK",
              "Status", "Worksheet", "Notes"]
    summary.append(header)
    head_row = summary.max_row
    for ci in range(1, len(header) + 1):
        c = summary.cell(row=head_row, column=ci)
        c.fill = head_fill
        c.font = head_font
    for row in summary_rows:
        summary.append(row)
        r = summary.max_row
        label = row[7]                                  # target worksheet ("<schema>.<table>" or "Zero Records")
        link = sheet_link(label)
        for col in (3, 8):                              # Table name + Worksheet cells -> clickable
            cell = summary.cell(row=r, column=col)
            cell.hyperlink = link
            cell.font = link_font
    summary.freeze_panes = f"A{head_row + 1}"
    summary.auto_filter.ref = f"A{head_row}:{get_column_letter(len(header))}{head_row}"
    summary.cell(row=1, column=1).font = Font(bold=True, size=14, color=HEADER_FILL)
    widths = [5, 12, 34, 12, 9, 8, 11, 34, 50]
    for ci, w in enumerate(widths, 1):
        summary.column_dimensions[get_column_letter(ci)].width = w

    # --- Zero Records tab --------------------------------------------------- #
    zhdr = ["Schema", "Table", "Columns", "Has PK"]
    zero.append([f"Tables with 0 records ({len(zero_rows)})"])
    zero.append([])
    zero.append(zhdr)
    zh = zero.max_row
    for ci in range(1, len(zhdr) + 1):
        zero.cell(row=zh, column=ci).fill = head_fill
        zero.cell(row=zh, column=ci).font = head_font
    for row in sorted(zero_rows):
        zero.append(row)
    zero.cell(row=1, column=1).font = Font(bold=True, size=12, color=HEADER_FILL)
    for ci, w in enumerate([12, 38, 9, 8], 1):
        zero.column_dimensions[get_column_letter(ci)].width = w
    if zero_rows:
        zero.freeze_panes = f"A{zh + 1}"

    # --- Column Analysis tab ------------------------------------------------ #
    ahdr = ["Schema", "Table", "Ordinal", "Column", "Data Type", "Length/Prec",
            "Nullable", "PK", "Identity", "Default", "Non-Null", "% Populated"]
    analysis.append(ahdr)
    style_header(analysis, len(ahdr))
    for row in analysis_rows:
        analysis.append(row)
    for ci, w in enumerate([12, 30, 8, 28, 14, 12, 9, 5, 9, 28, 10, 11], 1):
        analysis.column_dimensions[get_column_letter(ci)].width = w

    return wb, {"tables": total_tables, "populated": populated,
                "empty": empties, "rows": total_rows}


# --------------------------------------------------------------------------- #
# Entry point
# --------------------------------------------------------------------------- #
def run(ini_path: Path = DEFAULT_INI, output_override: str | None = None) -> int:
    import pyodbc

    db_cfg = load_db_config(ini_path)
    database = db_cfg.get("database", "Fusion_Flow_V3_QAS")
    conn = pyodbc.connect(conn_str(db_cfg), autocommit=True)
    try:
        cur = conn.cursor()
        out_dir = (output_override
                   or fetch_param(cur, "DB_SNAPSHOT_OUTPUT_DIR")
                   or fetch_param(cur, "DOCUMENTATION_OUTPUT_ROOT")
                   or FALLBACK_OUTPUT_DIR).strip()

        generated_at = datetime.utcnow()
        print(f"Snapshotting {database} -> {out_dir}")
        wb, stats = build_workbook(cur, database, generated_at)

        try:
            os.makedirs(out_dir, exist_ok=True)
        except OSError as e:
            print(f"[WARN] Could not create '{out_dir}' ({e}); writing to repo Documentation/ instead.")
            out_dir = str(REPO_ROOT / "Documentation")
            os.makedirs(out_dir, exist_ok=True)

        fname = f"DB_Snapshot_{database}_{generated_at:%Y%m%d_%H%M%S}.xlsx"
        out_path = os.path.join(out_dir, fname)
        wb.save(out_path)
        print(f"OK  tables={stats['tables']} populated={stats['populated']} "
              f"empty={stats['empty']} rows={stats['rows']:,}")
        print(f"Saved: {out_path}")
        return 0
    finally:
        conn.close()


def main() -> int:
    ini_path = Path(os.environ.get("FUSION_FLOW_INI", str(DEFAULT_INI)))
    override = sys.argv[1] if len(sys.argv) > 1 else None
    return run(ini_path, override)


if __name__ == "__main__":
    raise SystemExit(main())
