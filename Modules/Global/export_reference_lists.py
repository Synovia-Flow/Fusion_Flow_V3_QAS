#!/usr/bin/env python3
"""Fusion Flow V3 QAS - reference / option lists to Excel (for analysis).

Exports the CURATED configuration "option" lists - the enumerations and lookup
tables that define the platform's allowed values - into a single, separate
workbook (distinct from the full-DB snapshot). One tab per list:

    Vocabulary          <- CFG.Status_Vocabulary   (the process/status model)
    Clients, Jobs, Choice Fields, Choice Field Map, Translations,
    Processing Profiles, Field Map, Carriers, Ingestion Sources,
    API Versions, API Process Map, TSS Environments, Parameters

An "Index" tab lists every sheet with its row count and a clickable link.

Secret-ish values are masked (any column whose name looks like a password/secret/
token/key, plus Application_Parameters values for secret-named keys) -> '***'.

No CLI (house pattern). Connection from the .ini; output folder from CFG
DB_SNAPSHOT_OUTPUT_DIR (falls back to DOCUMENTATION_OUTPUT_ROOT, default the
Documentation_Layer share). Optional positional arg overrides the folder.

    python export_reference_lists.py ["D:\\some\\folder"]
"""

from __future__ import annotations

import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

# Reuse the snapshot's connection + workbook helpers (single source of truth).
try:
    from .export_db_snapshot import (  # type: ignore
        load_db_config, conn_str, q, fetch_param, cell_value, safe_sheet_name,
        DEFAULT_INI, REPO_ROOT, FALLBACK_OUTPUT_DIR, HEADER_FILL, HEADER_FONT)
except Exception:  # pragma: no cover - script context
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from export_db_snapshot import (  # type: ignore
        load_db_config, conn_str, q, fetch_param, cell_value, safe_sheet_name,
        DEFAULT_INI, REPO_ROOT, FALLBACK_OUTPUT_DIR, HEADER_FILL, HEADER_FONT)

# Curated reference/option lists: (schema, table, tab name, preferred ORDER BY).
# Missing tables are skipped gracefully; a bad ORDER BY falls back to unordered.
REFERENCE_TABLES: list[tuple[str, str, str, str]] = [
    ("CFG", "Status_Vocabulary",     "Vocabulary",          "SortOrder"),
    ("CFG", "Clients",               "Clients",             "ClientCode"),
    ("CFG", "Job",                   "Jobs",                "ModuleName, StepNo, JobCode"),
    ("CFG", "Choice_Field_Registry", "Choice Fields",       "ChoiceField"),
    ("CFG", "Choice_Field_Map",      "Choice Field Map",    "ChoiceField"),
    ("CFG", "Value_Translation",     "Translations",        "ClientCode, TargetField, IncomingValue"),
    ("CFG", "Processing_Profile",    "Processing Profiles", "ClientCode, EntityKind"),
    ("CFG", "Processing_Field_Map",  "Field Map",           "ClientCode, EntityKind, StepNo"),
    ("CFG", "Carrier_Master",        "Carriers",            "Eori"),
    ("CFG", "Ingestion_Source",      "Ingestion Sources",   "ClientCode"),
    ("CFG", "API_Version",           "API Versions",        "ClientCode"),
    ("CFG", "API_Process_Map",       "API Process Map",     "ClientCode, StepOrder"),
    ("CFG", "TSS_Environment",       "TSS Environments",    "EnvCode"),
    ("CFG", "Application_Parameters","Parameters",          "ParameterKey"),
]

SECRET_COL = re.compile(r"(password|secret|pwd|token|apikey|api_key|\bkey\b|connectionstring)", re.I)
SECRET_KEY = re.compile(r"(password|secret|pwd|token|apikey|api_key|clientsecret)", re.I)
MASK = "***"


def _rows_for(cur, schema: str, table: str, order_by: str):
    """(columns, rows) for a table, or (None, None) if it doesn't exist."""
    if not q(cur, "SELECT OBJECT_ID(?, 'U') AS oid", f"{schema}.{table}")[0]["oid"]:
        return None, None
    for clause in (f" ORDER BY {order_by}" if order_by else "", ""):
        try:
            cur.execute(f"SELECT * FROM [{schema}].[{table}]{clause}")
            cols = [d[0] for d in cur.description]
            return cols, cur.fetchall()
        except Exception:
            continue
    return [], []


def build_workbook(cur, database: str, generated_at: datetime):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    head_fill = PatternFill("solid", fgColor=HEADER_FILL)
    head_font = Font(bold=True, color=HEADER_FONT)
    link_font = Font(color="FF0563C1", underline="single")

    def sheet_link(label: str) -> str:
        return f"#'{label.replace(chr(39), chr(39) * 2)}'!A1"

    def style_header(ws, ncols: int) -> None:
        for ci in range(1, ncols + 1):
            c = ws.cell(row=1, column=ci)
            c.fill = head_fill
            c.font = head_font
        ws.freeze_panes = "A2"
        if ncols:
            ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"

    def autosize(ws, ncols: int, sample: int = 300) -> None:
        for ci in range(1, ncols + 1):
            w = 8
            for ri in range(1, min(ws.max_row, sample + 1) + 1):
                v = ws.cell(row=ri, column=ci).value
                if v is not None:
                    w = max(w, min(70, len(str(v)) + 2))
            ws.column_dimensions[get_column_letter(ci)].width = w

    wb = Workbook()
    index = wb.active
    index.title = "Index"
    used = {"index"}
    index_rows: list[tuple[str, str, int]] = []   # (tab label, "schema.table", rowcount)

    for schema, table, tab, order_by in REFERENCE_TABLES:
        cols, data = _rows_for(cur, schema, table, order_by)
        if cols is None:
            continue                                     # table not deployed - skip
        label = safe_sheet_name(tab, used)
        ws = wb.create_sheet(label)
        ws.append(cols)
        style_header(ws, len(cols))
        mask_idx = {i for i, c in enumerate(cols) if SECRET_COL.search(c or "")}
        key_i = cols.index("ParameterKey") if "ParameterKey" in cols else None
        val_i = cols.index("ParameterValue") if "ParameterValue" in cols else None
        for r in data:
            out = []
            for i, v in enumerate(r):
                if i in mask_idx and v is not None:
                    out.append(MASK)
                elif key_i is not None and val_i is not None and i == val_i \
                        and SECRET_KEY.search(str(r[key_i] or "")):
                    out.append(MASK)
                else:
                    out.append(cell_value(v))
            ws.append(out)
        autosize(ws, len(cols))
        index_rows.append((label, f"{schema}.{table}", len(data)))

    # --- Index tab ---------------------------------------------------------- #
    index.append([f"Fusion Flow V3 QAS - Reference / Option Lists: {database}"])
    index.append([f"Generated (UTC): {generated_at:%Y-%m-%d %H:%M:%S}"])
    index.append([f"Lists: {len(index_rows)}"])
    index.append([])
    header = ["#", "List", "Source table", "Rows"]
    index.append(header)
    hr = index.max_row
    for ci in range(1, len(header) + 1):
        index.cell(row=hr, column=ci).fill = head_fill
        index.cell(row=hr, column=ci).font = head_font
    for i, (label, src, n) in enumerate(index_rows, 1):
        index.append([i, label, src, n])
        r = index.max_row
        c = index.cell(row=r, column=2)
        c.hyperlink = sheet_link(label)
        c.font = link_font
    index.cell(row=1, column=1).font = Font(bold=True, size=14, color=HEADER_FILL)
    index.freeze_panes = f"A{hr + 1}"
    index.auto_filter.ref = f"A{hr}:{get_column_letter(len(header))}{hr}"
    for ci, w in enumerate([5, 22, 30, 8], 1):
        index.column_dimensions[get_column_letter(ci)].width = w

    return wb, len(index_rows)


def run(ini_path: Path = DEFAULT_INI, output_override: str | None = None) -> int:
    import pyodbc

    db_cfg = load_db_config(ini_path)
    database = db_cfg.get("database", "Fusion_Flow_V3_QAS")
    conn = pyodbc.connect(conn_str(db_cfg), autocommit=True)
    try:
        cur = conn.cursor()
        out_dir = (output_override
                   or fetch_param(cur, "DB_SNAPSHOT_OUTPUT_DIR")
                   or fetch_param(cur, "DOCUMENTATION_OUTPUT_ROOT")
                   or FALLBACK_OUTPUT_DIR).strip()

        generated_at = datetime.utcnow()
        print(f"Exporting reference lists from {database} -> {out_dir}")
        wb, n = build_workbook(cur, database, generated_at)

        try:
            os.makedirs(out_dir, exist_ok=True)
        except OSError as e:
            print(f"[WARN] Could not create '{out_dir}' ({e}); writing to repo Documentation/ instead.")
            out_dir = str(REPO_ROOT / "Documentation")
            os.makedirs(out_dir, exist_ok=True)

        fname = f"Reference_Lists_{database}_{generated_at:%Y%m%d_%H%M%S}.xlsx"
        out_path = os.path.join(out_dir, fname)
        wb.save(out_path)
        print(f"OK  lists={n}")
        print(f"Saved: {out_path}")
        return 0
    finally:
        conn.close()


def main() -> int:
    ini_path = Path(os.environ.get("FUSION_FLOW_INI", str(DEFAULT_INI)))
    override = sys.argv[1] if len(sys.argv) > 1 else None
    return run(ini_path, override)


if __name__ == "__main__":
    raise SystemExit(main())
