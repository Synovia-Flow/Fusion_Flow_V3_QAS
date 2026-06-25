#!/usr/bin/env python3
"""Fusion Flow V3 QAS - Microsoft Graph email ingestion route.

Implements the EMAIL channel for Module 1, adapting the proven approach in
Inbound/Graph_Inbox_Analyzer.py to the CFG/ING/EXC architecture:

  app-only MSAL token -> scan the mailbox (Inbox + sub-folders, skipping system
  folders and the processed target) -> download every NON-IMAGE attachment ->
  land each verbatim into ING (Inbound_File + Raw_Record) and record the email
  in ING.Source_Email -> move the scanned message into Inbox/<processed folder>.

Configuration is read from CFG.Application_Parameters (seeded by 006):
  GRAPH_CLIENT_ID, GRAPH_TENANT_ID, GRAPH_CLIENT_SECRET_REF, GRAPH_MAILBOX,
  GRAPH_PROCESSED_FOLDER, GRAPH_AUTHORITY, GRAPH_SCOPE.

The client secret is resolved at runtime (never stored in the DB):
  1. env var GRAPH_CLIENT_SECRET, else
  2. Azure Key Vault via GRAPH_CLIENT_SECRET_REF ("<vault>/<secret>") if
     azure-identity + azure-keyvault-secrets are installed.

Every step is logged to EXC.Execution / LOG via the passed IngestionDb.
"""

from __future__ import annotations

import csv
import io
import os
import time
from datetime import datetime, timezone
from typing import Any

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
SYSTEM_FOLDERS = {
    "sent items", "drafts", "deleted items", "junk email", "outbox",
    "archive", "conversation history", "clutter", "rss feeds", "sync issues",
}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".tif", ".webp", ".svg", ".heic"}


# --------------------------------------------------------------------------- #
# Secret resolution (never stored in the DB)
# --------------------------------------------------------------------------- #
def resolve_client_secret(params: dict[str, str]) -> str:
    """Resolve the Graph client secret. Order (design: secret lives in the table):
      1. CFG.Application_Parameters.GRAPH_CLIENT_SECRET (passed in `params`)
      2. env var GRAPH_CLIENT_SECRET
      3. Key Vault via GRAPH_CLIENT_SECRET_REF ("<vault>/<secret>")
    """
    direct = (params.get("GRAPH_CLIENT_SECRET") or "").strip()
    if direct and not direct.startswith("<"):
        return direct
    env = os.environ.get("GRAPH_CLIENT_SECRET", "").strip()
    if env:
        return env
    secret_ref = (params.get("GRAPH_CLIENT_SECRET_REF") or "").strip()
    if secret_ref and "/" in secret_ref and not secret_ref.startswith("<"):
        try:
            from azure.identity import DefaultAzureCredential
            from azure.keyvault.secrets import SecretClient
            vault, name = secret_ref.split("/", 1)
            url = f"https://{vault}.vault.azure.net" if not vault.startswith("http") else vault
            client = SecretClient(vault_url=url, credential=DefaultAzureCredential())
            return client.get_secret(name).value
        except Exception as error:  # noqa: BLE001
            raise RuntimeError(
                f"Could not resolve Graph secret from Key Vault ref '{secret_ref}': {error}."
            ) from error
    raise RuntimeError(
        "No Graph client secret. Set CFG.Application_Parameters.GRAPH_CLIENT_SECRET, "
        "env GRAPH_CLIENT_SECRET, or a Key Vault ref in GRAPH_CLIENT_SECRET_REF.")


def acquire_token(authority_base: str, tenant_id: str, client_id: str, client_secret: str, scope: str) -> str:
    import msal
    authority = authority_base.rstrip("/") + "/" + tenant_id
    app = msal.ConfidentialClientApplication(
        client_id=client_id, client_credential=client_secret, authority=authority)
    result = app.acquire_token_for_client(scopes=[scope])
    if "access_token" not in result:
        raise RuntimeError(f"Token acquisition failed: {result.get('error')} - {result.get('error_description')}")
    return result["access_token"]


class GraphClient:
    """Thin Graph REST wrapper with paging + 429/503 backoff (from the analyzer)."""

    def __init__(self, token: str, timeout: int = 60) -> None:
        import requests
        self._s = requests.Session()
        self._s.headers.update({"Authorization": f"Bearer {token}", "Accept": "application/json"})
        self._timeout = timeout

    def _retry(self, fn):
        for attempt in range(6):
            resp = fn()
            if resp.status_code in (429, 503):
                time.sleep(int(resp.headers.get("Retry-After", 2 ** attempt)))
                continue
            if resp.status_code >= 400:
                raise RuntimeError(f"Graph {resp.status_code}: {resp.text[:500]}")
            return resp
        raise RuntimeError("Graph kept throttling")

    def get(self, url: str, params: dict | None = None) -> dict:
        full = url if url.startswith("http") else GRAPH_BASE + url
        return self._retry(lambda: self._s.get(full, params=params, timeout=self._timeout)).json()

    def get_all(self, url: str, params: dict | None = None) -> list[dict]:
        items, page = [], self.get(url, params)
        items.extend(page.get("value", []))
        nxt = page.get("@odata.nextLink")
        while nxt:
            page = self.get(nxt)
            items.extend(page.get("value", []))
            nxt = page.get("@odata.nextLink")
        return items

    def get_bytes(self, url: str) -> bytes:
        full = url if url.startswith("http") else GRAPH_BASE + url
        return self._retry(lambda: self._s.get(full, timeout=self._timeout)).content

    def post(self, url: str, body: dict | None = None) -> dict:
        full = url if url.startswith("http") else GRAPH_BASE + url
        resp = self._retry(lambda: self._s.post(full, json=body, timeout=self._timeout))
        return resp.json() if resp.content else {}


# --------------------------------------------------------------------------- #
# Mailbox helpers
# --------------------------------------------------------------------------- #
def resolve_inbox_id(client: GraphClient, mailbox: str) -> str:
    return client.get(f"/users/{mailbox}/mailFolders/inbox", {"$select": "id"})["id"]


def ensure_processed_folder(client: GraphClient, mailbox: str, inbox_id: str,
                            name: str, subfolder: str | None = None) -> str:
    """Ensure Inbox/<name> (and optionally Inbox/<name>/<subfolder>) exist; return the
    deepest folder id. Processed mail is moved into a per-client subfolder, e.g.
    Fusion_Processed/BKD."""
    def ensure_child(parent_id: str, child_name: str) -> str:
        for c in client.get_all(f"/users/{mailbox}/mailFolders/{parent_id}/childFolders",
                                {"$select": "id,displayName"}):
            if c.get("displayName", "").lower() == child_name.lower():
                return c["id"]
        return client.post(f"/users/{mailbox}/mailFolders/{parent_id}/childFolders",
                           {"displayName": child_name})["id"]

    processed_id = ensure_child(inbox_id, name)
    return ensure_child(processed_id, subfolder) if subfolder else processed_id


def scan_folders(client: GraphClient, mailbox: str, inbox_id: str,
                 skip_names: set[str] | None = None) -> list[dict]:
    """Inbox sub-folders (recursive), skipping system folders and the processed tree."""
    skip = {s.lower() for s in (skip_names or set())}
    out: list[dict] = []

    def walk(folder_id: str):
        for f in client.get_all(f"/users/{mailbox}/mailFolders/{folder_id}/childFolders",
                                 {"$select": "id,displayName,childFolderCount", "$top": 100}):
            name = f.get("displayName", "").lower()
            if name in SYSTEM_FOLDERS or name in skip:
                continue
            out.append(f)
            if f.get("childFolderCount", 0):
                walk(f["id"])

    walk(inbox_id)
    return out


def _ext(name: str) -> str:
    return ("." + name.rsplit(".", 1)[-1].lower()) if "." in name else ""


def parse_rows(name: str, content: bytes) -> list[dict[str, Any]]:
    """Parse CSV/XLSX bytes into verbatim row dicts (best-effort; other types skipped)."""
    ext = _ext(name)
    if ext == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        return [dict(r) for r in csv.DictReader(io.StringIO(text))]
    if ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or f"col{i}") for i, h in enumerate(rows[0], 1)]
        return [dict(zip(headers, r)) for r in rows[1:] if any(v is not None for v in r)]
    return []  # non-tabular attachment: landed verbatim as a file, not parsed into rows


# --------------------------------------------------------------------------- #
# Entry point used by ingest.py EmailGraphChannel
# --------------------------------------------------------------------------- #
def run_email_ingest(db: Any, client_code: str, params: dict[str, str], dry_run: bool) -> dict[str, int]:
    """Scan the mailbox, land non-image attachments into ING, move processed mail.

    `db` is ingest.IngestionDb (already has an open EXC.Execution). `params` is a
    dict of CFG.Application_Parameters values. Returns counts.
    """
    stats = {"messages": 0, "files": 0, "rows": 0, "moved": 0, "errors": 0}

    mailbox = params.get("GRAPH_MAILBOX") or ""
    tenant = params.get("GRAPH_TENANT_ID") or ""
    client_id = params.get("GRAPH_CLIENT_ID") or ""
    authority = params.get("GRAPH_AUTHORITY") or "https://login.microsoftonline.com/"
    scope = params.get("GRAPH_SCOPE") or "https://graph.microsoft.com/.default"
    processed_name = params.get("GRAPH_PROCESSED_FOLDER") or "Fusion_Processed"

    if not (mailbox and tenant and client_id) or tenant.startswith("<"):
        db.log("EMAIL", "Graph config incomplete (GRAPH_TENANT_ID/CLIENT_ID/MAILBOX). Skipping.", "WARN")
        return stats

    if dry_run:
        db.log("EMAIL", f"[dry-run] would scan {mailbox} and land non-image attachments.")
        return stats

    secret = resolve_client_secret(params)
    token = acquire_token(authority, tenant, client_id, secret, scope)
    client = GraphClient(token)
    db.log("EMAIL", f"Authenticated to Graph for {mailbox}.")

    inbox_id = resolve_inbox_id(client, mailbox)
    # Move processed mail into a per-client subfolder: Fusion_Processed/<CLIENT>.
    processed_id = ensure_processed_folder(client, mailbox, inbox_id, processed_name, subfolder=client_code)
    folders = scan_folders(client, mailbox, inbox_id, skip_names={processed_name})
    db.log("EMAIL", f"Scanning {len(folders)} folder(s) under Inbox (-> processed: {processed_name}/{client_code}).")

    for folder in folders:
        msgs = client.get_all(
            f"/users/{mailbox}/mailFolders/{folder['id']}/messages",
            {"$select": "id,subject,from,receivedDateTime,hasAttachments,internetMessageId,bodyPreview",
             "$filter": "hasAttachments eq true", "$top": 50})
        for msg in msgs:
            stats["messages"] += 1
            try:
                _record_email(db, client_code, mailbox, msg)
                atts = client.get_all(f"/users/{mailbox}/messages/{msg['id']}/attachments",
                                      {"$select": "id,name,contentType,size,isInline,@odata.type"})
                saved_any = False
                for att in atts:
                    if att.get("@odata.type") != "#microsoft.graph.fileAttachment" or att.get("isInline"):
                        continue
                    name = att.get("name") or "attachment"
                    if _ext(name) in IMAGE_EXTS or str(att.get("contentType", "")).startswith("image/"):
                        continue
                    content = client.get_bytes(f"/users/{mailbox}/messages/{msg['id']}/attachments/{att['id']}/$value")
                    sender = (msg.get("from", {}).get("emailAddress", {}) or {}).get("address", "")
                    file_id = db.land_inbound_file(client_code, "EMAIL", name, content,
                                                   sender=sender, mailbox=mailbox,
                                                   content_type=att.get("contentType", ""))
                    if file_id:
                        stats["files"] += 1
                        stats["rows"] += db.land_raw_rows(file_id, client_code, parse_rows(name, content))
                        saved_any = True
                # Move the scanned message into the processed folder.
                client.post(f"/users/{mailbox}/messages/{msg['id']}/move", {"destinationId": processed_id})
                stats["moved"] += 1
            except Exception as error:  # noqa: BLE001
                stats["errors"] += 1
                db.log_error("EMAIL", f"Message {str(msg.get('id'))[:12]}: {error}", "GraphIngestError")

    db.log("EMAIL", f"Email route done: {stats}", "OK" if stats["errors"] == 0 else "WARN")
    return stats


def _record_email(db: Any, client_code: str, mailbox: str, msg: dict) -> None:
    if db.dry_run or not db.execution_id:
        return
    sender = (msg.get("from", {}).get("emailAddress", {}) or {}).get("address", "")
    received = msg.get("receivedDateTime")
    cur = db.conn.cursor()
    cur.execute(
        "INSERT INTO ING.Source_Email (ExecutionID, TransactionID, ClientCode, Mailbox, GraphMessageID, "
        "InternetMessageID, Sender, SenderDomain, Subject, ReceivedUtc, HasAttachments, BodyText, Status) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'INGESTED')",
        db.execution_id, db.transaction_id, client_code, mailbox, str(msg.get("id"))[:450],
        (msg.get("internetMessageId") or "")[:1000], sender[:320],
        (sender.split("@")[-1] if "@" in sender else "")[:320], (msg.get("subject") or "")[:998],
        _parse_dt(received), 1, (msg.get("bodyPreview") or "")[:4000])
    db.conn.commit()


def _parse_dt(value: str | None):
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).astimezone(timezone.utc).replace(tzinfo=None)
    except ValueError:
        return None
