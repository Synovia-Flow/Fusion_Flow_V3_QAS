#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Graph_Inbox_Analyzer.py
=======================

Connects to a Microsoft 365 mailbox via the Microsoft Graph API (app-only /
client-credentials flow), enumerates every mail folder together with the number
of emails it holds, performs an analysis of the Inbox, and writes the results to
an Excel workbook.

Inputs
------
* Parameters workbook : F:\\Synovia_Flow_Quality\\Documentation_Layer\\Project_Build_Files\\Ingestion_Setup.xlsx
    The "Parameters" tab must contain the following key/value rows
    (parameter name in the first column, value in the second):
        GRAPH_CLIENT_SECRET
        GRAPH_CLIENT_ID
        GRAPH_TENANT_ID
* App manifest        : E:\\Git\\FLow_3_1\\Flow_3_1_Development\\Development_Tools\\Config\\Manifest_V2.json
    Used to validate that the client id in the workbook matches the registered
    application ("Fusion Flow Mail Reader") and to confirm the requested Graph
    permissions before any call is made.

Output
------
* Excel report        : F:\\Synovia_Flow_Quality\\Documentation_Layer\\Graph\\
    A timestamped workbook (Mailbox_Report_<mailbox>_<YYYYMMDD_HHMMSS>.xlsx)
    containing:
        - Folder Summary  : every folder (recursively), full path, total &
                            unread email counts.
        - Inbox Analysis  : headline inbox metrics (totals, read/unread, top
                            senders, oldest/newest, volume by day).
        - Run Info        : metadata about the run (mailbox, app, timestamp).

The target mailbox defaults to ``nexus@synoviaflow.cloud`` and can be overridden
with ``--mailbox``.

This tool is read-only: it requires the Mail.Read / Mail.ReadBasic.All
*application* permissions that are already declared in Manifest_V2.json.

Usage
-----
    python Graph_Inbox_Analyzer.py
    python Graph_Inbox_Analyzer.py --mailbox someone@synoviaflow.cloud
    python Graph_Inbox_Analyzer.py --setup-xlsx "C:\\path\\Ingestion_Setup.xlsx" \\
                                   --output-dir "C:\\path\\out" --inbox-sample 1000

Dependencies (see requirements.txt):
    msal, requests, openpyxl
"""

from __future__ import annotations

import argparse
import collections
import datetime as _dt
import json
import logging
import os
import sys
import time
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    import requests
except ImportError:  # pragma: no cover - dependency guidance
    sys.exit("Missing dependency 'requests'. Install with: pip install -r requirements.txt")

try:
    import msal
except ImportError:  # pragma: no cover - dependency guidance
    sys.exit("Missing dependency 'msal'. Install with: pip install -r requirements.txt")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - dependency guidance
    sys.exit("Missing dependency 'openpyxl'. Install with: pip install -r requirements.txt")


# --------------------------------------------------------------------------- #
# Defaults - these mirror the deployment environment described in the request. #
# Every value can be overridden on the command line.                          #
# --------------------------------------------------------------------------- #
DEFAULT_SETUP_XLSX = r"F:\Synovia_Flow_Quality\Documentation_Layer\Project_Build_Files\Ingestion_Setup.xlsx"
DEFAULT_MANIFEST = r"E:\Git\FLow_3_1\Flow_3_1_Development\Development_Tools\Config\Manifest_V2.json"
DEFAULT_OUTPUT_DIR = r"F:\Synovia_Flow_Quality\Documentation_Layer\Graph"
DEFAULT_MAILBOX = "nexus@synoviaflow.cloud"
DEFAULT_PARAMETERS_SHEET = "Parameters"
DEFAULT_INBOX_SAMPLE = 2000  # max inbox messages to pull for the analysis sheet

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
GRAPH_SCOPE = ["https://graph.microsoft.com/.default"]
AUTHORITY_TEMPLATE = "https://login.microsoftonline.com/{tenant}"

# Microsoft Graph well-known resource app id, used to read the manifest's
# requiredResourceAccess so we can report the granted permissions.
GRAPH_RESOURCE_APP_ID = "00000003-0000-0000-c000-000000000000"

REQUIRED_PARAMETERS = ("GRAPH_CLIENT_SECRET", "GRAPH_CLIENT_ID", "GRAPH_TENANT_ID")

LOG = logging.getLogger("graph_inbox_analyzer")


# --------------------------------------------------------------------------- #
# Configuration loading                                                       #
# --------------------------------------------------------------------------- #
def read_parameters_from_xlsx(path: str, sheet_name: str) -> Dict[str, str]:
    """Read the key/value pairs from the Parameters tab of the setup workbook.

    The tab is expected to hold a parameter name in the first column and its
    value in the second. Matching of the sheet name is case-insensitive and
    tolerant of the common 'Paramters' misspelling.
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Parameters workbook not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        target = _resolve_sheet(wb.sheetnames, sheet_name)
        if target is None:
            raise ValueError(
                f"Could not find a '{sheet_name}' tab in {path}. "
                f"Available tabs: {', '.join(wb.sheetnames)}"
            )
        ws = wb[target]

        params: Dict[str, str] = {}
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            if not key:
                continue
            value = row[1] if len(row) > 1 else None
            params[key] = "" if value is None else str(value).strip()
        return params
    finally:
        wb.close()


def _resolve_sheet(sheet_names: Iterable[str], wanted: str) -> Optional[str]:
    """Find a worksheet name, ignoring case and the 'Paramters' typo."""
    wanted_norm = wanted.strip().lower()
    candidates = {wanted_norm, "parameters", "paramters", "params"}
    for name in sheet_names:
        if name.strip().lower() in candidates:
            return name
    # Fall back to any tab that starts with 'param'.
    for name in sheet_names:
        if name.strip().lower().startswith("param"):
            return name
    return None


def load_manifest(path: str) -> Optional[Dict[str, Any]]:
    """Load the app manifest if present; missing/invalid manifests are non-fatal."""
    if not os.path.isfile(path):
        LOG.warning("Manifest not found at %s - continuing without manifest validation.", path)
        return None
    try:
        with open(path, "r", encoding="utf-8-sig") as fh:
            return json.load(fh)
    except (OSError, json.JSONDecodeError) as exc:
        LOG.warning("Could not parse manifest %s (%s) - continuing.", path, exc)
        return None


def validate_against_manifest(manifest: Optional[Dict[str, Any]], client_id: str) -> List[str]:
    """Cross-check the workbook client id with the manifest and list permissions.

    Returns a human-readable list of the Graph permissions declared in the
    manifest. A client-id mismatch is logged as a warning (not fatal) so the
    operator is alerted but the run can still proceed if intended.
    """
    permissions: List[str] = []
    if not manifest:
        return permissions

    manifest_app_id = manifest.get("appId")
    if manifest_app_id and client_id and manifest_app_id.lower() != client_id.lower():
        LOG.warning(
            "Client id in workbook (%s) does not match manifest appId (%s).",
            client_id,
            manifest_app_id,
        )
    else:
        LOG.info("Client id matches manifest app '%s'.", manifest.get("displayName", "<unknown>"))

    for resource in manifest.get("requiredResourceAccess", []) or []:
        if resource.get("resourceAppId") != GRAPH_RESOURCE_APP_ID:
            continue
        for access in resource.get("resourceAccess", []) or []:
            permissions.append(f"{access.get('id')} ({access.get('type')})")
    return permissions


# --------------------------------------------------------------------------- #
# Authentication                                                              #
# --------------------------------------------------------------------------- #
def acquire_token(tenant_id: str, client_id: str, client_secret: str) -> str:
    """Acquire an app-only access token via MSAL client-credentials flow."""
    authority = AUTHORITY_TEMPLATE.format(tenant=tenant_id)
    app = msal.ConfidentialClientApplication(
        client_id=client_id,
        client_credential=client_secret,
        authority=authority,
    )
    LOG.info("Requesting application token from %s", authority)
    result = app.acquire_token_for_client(scopes=GRAPH_SCOPE)
    if "access_token" not in result:
        raise RuntimeError(
            "Failed to acquire token: "
            f"{result.get('error')} - {result.get('error_description')}"
        )
    LOG.info("Token acquired successfully.")
    return result["access_token"]


# --------------------------------------------------------------------------- #
# Graph client                                                                #
# --------------------------------------------------------------------------- #
class GraphClient:
    """Thin Microsoft Graph REST wrapper with paging and throttling handling."""

    def __init__(self, token: str, timeout: int = 60) -> None:
        self._session = requests.Session()
        self._session.headers.update(
            {"Authorization": f"Bearer {token}", "Accept": "application/json"}
        )
        self._timeout = timeout

    def get(self, url: str, params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """GET a single Graph resource, honouring 429/503 Retry-After backoff."""
        full_url = url if url.startswith("http") else f"{GRAPH_BASE}{url}"
        for attempt in range(6):
            resp = self._session.get(full_url, params=params, timeout=self._timeout)
            if resp.status_code in (429, 503):
                wait = int(resp.headers.get("Retry-After", 2 ** attempt))
                LOG.warning("Throttled (%s). Waiting %ss before retry.", resp.status_code, wait)
                time.sleep(wait)
                continue
            if resp.status_code >= 400:
                raise RuntimeError(
                    f"Graph request failed ({resp.status_code}) for {full_url}: {resp.text}"
                )
            return resp.json()
        raise RuntimeError(f"Graph request kept being throttled: {full_url}")

    def get_all(self, url: str, params: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """GET a collection, following @odata.nextLink until exhausted."""
        items: List[Dict[str, Any]] = []
        page = self.get(url, params=params)
        items.extend(page.get("value", []))
        next_link = page.get("@odata.nextLink")
        while next_link:
            page = self.get(next_link)
            items.extend(page.get("value", []))
            next_link = page.get("@odata.nextLink")
        return items


# --------------------------------------------------------------------------- #
# Mailbox enumeration                                                         #
# --------------------------------------------------------------------------- #
FOLDER_SELECT = "id,displayName,parentFolderId,childFolderCount,unreadItemCount,totalItemCount,isHidden"


def enumerate_folders(client: GraphClient, mailbox: str) -> List[Dict[str, Any]]:
    """Recursively walk the mailbox folder tree, returning one record per folder.

    Each record carries the full folder path and email counts. Hidden folders
    are included so the report reflects the true mailbox layout.
    """
    folders: List[Dict[str, Any]] = []

    def walk(folder_id: str, parent_path: str, depth: int) -> None:
        url = f"/users/{mailbox}/mailFolders/{folder_id}/childFolders"
        children = client.get_all(
            url, params={"includeHiddenFolders": "true", "$select": FOLDER_SELECT, "$top": 100}
        )
        for child in children:
            name = child.get("displayName", "<unnamed>")
            path = f"{parent_path}/{name}" if parent_path else name
            folders.append(_folder_record(child, path, depth))
            if (child.get("childFolderCount") or 0) > 0:
                walk(child["id"], path, depth + 1)

    top = client.get_all(
        f"/users/{mailbox}/mailFolders",
        params={"includeHiddenFolders": "true", "$select": FOLDER_SELECT, "$top": 100},
    )
    for folder in top:
        name = folder.get("displayName", "<unnamed>")
        folders.append(_folder_record(folder, name, 0))
        if (folder.get("childFolderCount") or 0) > 0:
            walk(folder["id"], name, 1)

    return folders


def _folder_record(folder: Dict[str, Any], path: str, depth: int) -> Dict[str, Any]:
    return {
        "id": folder.get("id"),
        "name": folder.get("displayName", "<unnamed>"),
        "path": path,
        "depth": depth,
        "total": int(folder.get("totalItemCount") or 0),
        "unread": int(folder.get("unreadItemCount") or 0),
        "child_count": int(folder.get("childFolderCount") or 0),
        "hidden": bool(folder.get("isHidden")),
    }


INBOX_MESSAGE_SELECT = "id,subject,from,receivedDateTime,isRead,hasAttachments,importance"


def fetch_inbox_messages(client: GraphClient, mailbox: str, sample: int) -> List[Dict[str, Any]]:
    """Pull up to ``sample`` most-recent Inbox messages for analysis."""
    messages: List[Dict[str, Any]] = []
    params = {
        "$select": INBOX_MESSAGE_SELECT,
        "$orderby": "receivedDateTime desc",
        "$top": 100,
    }
    url = f"/users/{mailbox}/mailFolders/inbox/messages"
    page = client.get(url, params=params)
    messages.extend(page.get("value", []))
    next_link = page.get("@odata.nextLink")
    while next_link and len(messages) < sample:
        page = client.get(next_link)
        messages.extend(page.get("value", []))
        next_link = page.get("@odata.nextLink")
    return messages[:sample]


def analyse_inbox(messages: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Compute headline metrics over a list of Inbox messages."""
    total = len(messages)
    unread = sum(1 for m in messages if not m.get("isRead", True))
    with_attachments = sum(1 for m in messages if m.get("hasAttachments"))
    high_importance = sum(1 for m in messages if m.get("importance") == "high")

    senders: "collections.Counter[str]" = collections.Counter()
    by_day: "collections.Counter[str]" = collections.Counter()
    received_times: List[_dt.datetime] = []

    for m in messages:
        sender = _sender_address(m)
        if sender:
            senders[sender] += 1
        received = _parse_dt(m.get("receivedDateTime"))
        if received:
            received_times.append(received)
            by_day[received.date().isoformat()] += 1

    oldest = min(received_times).isoformat() if received_times else None
    newest = max(received_times).isoformat() if received_times else None

    return {
        "total": total,
        "read": total - unread,
        "unread": unread,
        "with_attachments": with_attachments,
        "high_importance": high_importance,
        "unique_senders": len(senders),
        "top_senders": senders.most_common(25),
        "by_day": sorted(by_day.items()),
        "oldest": oldest,
        "newest": newest,
    }


def _sender_address(message: Dict[str, Any]) -> Optional[str]:
    frm = (message.get("from") or {}).get("emailAddress") or {}
    return frm.get("address") or frm.get("name")


def _parse_dt(value: Optional[str]) -> Optional[_dt.datetime]:
    if not value:
        return None
    try:
        return _dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


# --------------------------------------------------------------------------- #
# Excel report                                                                #
# --------------------------------------------------------------------------- #
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=14, bold=True, color="1F4E78")


def _style_header(ws, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, max_width: int = 70) -> None:
    for col_cells in ws.columns:
        length = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def write_report(
    output_path: str,
    mailbox: str,
    folders: List[Dict[str, Any]],
    inbox_stats: Dict[str, Any],
    run_meta: Dict[str, Any],
) -> None:
    """Build the multi-sheet Excel workbook and save it to ``output_path``."""
    wb = Workbook()

    _write_folder_sheet(wb.active, mailbox, folders)
    _write_inbox_sheet(wb.create_sheet("Inbox Analysis"), mailbox, inbox_stats)
    _write_run_info_sheet(wb.create_sheet("Run Info"), run_meta)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


def _write_folder_sheet(ws, mailbox: str, folders: List[Dict[str, Any]]) -> None:
    ws.title = "Folder Summary"
    ws["A1"] = f"Mailbox folder summary - {mailbox}"
    ws["A1"].font = _TITLE_FONT

    headers = ["Folder", "Full Path", "Depth", "Total Emails", "Unread", "Sub-folders", "Hidden"]
    ws.append([])  # spacer row (row 2)
    ws.append(headers)
    _style_header(ws, ws.max_row, len(headers))

    # Indent the display name by depth so the tree is readable at a glance.
    for f in sorted(folders, key=lambda x: x["path"].lower()):
        indent = "    " * f["depth"]
        ws.append(
            [
                f"{indent}{f['name']}",
                f["path"],
                f["depth"],
                f["total"],
                f["unread"],
                f["child_count"],
                "Yes" if f["hidden"] else "No",
            ]
        )

    # Totals row.
    total_emails = sum(f["total"] for f in folders)
    total_unread = sum(f["unread"] for f in folders)
    ws.append([])
    ws.append(["TOTAL", f"{len(folders)} folders", "", total_emails, total_unread, "", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.freeze_panes = "A4"
    _autosize(ws)


def _write_inbox_sheet(ws, mailbox: str, stats: Dict[str, Any]) -> None:
    ws["A1"] = f"Inbox analysis - {mailbox}"
    ws["A1"].font = _TITLE_FONT

    row = 3
    metrics: List[Tuple[str, Any]] = [
        ("Messages analysed (sample)", stats["total"]),
        ("Read", stats["read"]),
        ("Unread", stats["unread"]),
        ("With attachments", stats["with_attachments"]),
        ("High importance", stats["high_importance"]),
        ("Unique senders", stats["unique_senders"]),
        ("Oldest in sample", stats["oldest"] or "n/a"),
        ("Newest in sample", stats["newest"] or "n/a"),
    ]
    ws.cell(row=row, column=1, value="Metric")
    ws.cell(row=row, column=2, value="Value")
    _style_header(ws, row, 2)
    row += 1
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Top senders block.
    row += 1
    ws.cell(row=row, column=1, value="Top senders").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Sender")
    ws.cell(row=row, column=2, value="Email count")
    _style_header(ws, row, 2)
    row += 1
    for sender, count in stats["top_senders"]:
        ws.cell(row=row, column=1, value=sender)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Volume by day block (alongside, starting at column D).
    day_row = 3
    ws.cell(row=day_row, column=4, value="Date")
    ws.cell(row=day_row, column=5, value="Emails received")
    _style_header(ws, day_row, 0)  # no-op fill guard
    for col in (4, 5):
        c = ws.cell(row=day_row, column=col)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
    day_row += 1
    for day, count in stats["by_day"]:
        ws.cell(row=day_row, column=4, value=day)
        ws.cell(row=day_row, column=5, value=count)
        day_row += 1

    _autosize(ws)


def _write_run_info_sheet(ws, meta: Dict[str, Any]) -> None:
    ws["A1"] = "Run information"
    ws["A1"].font = _TITLE_FONT
    row = 3
    for key, value in meta.items():
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(value))
        row += 1
    _autosize(ws)


# --------------------------------------------------------------------------- #
# Orchestration                                                               #
# --------------------------------------------------------------------------- #
def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--setup-xlsx", default=DEFAULT_SETUP_XLSX, help="Path to Ingestion_Setup.xlsx")
    parser.add_argument("--parameters-sheet", default=DEFAULT_PARAMETERS_SHEET, help="Name of the parameters tab")
    parser.add_argument("--manifest", default=DEFAULT_MANIFEST, help="Path to the app manifest JSON")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="Directory for the Excel report")
    parser.add_argument("--mailbox", default=DEFAULT_MAILBOX, help="Target mailbox (UPN) to analyse")
    parser.add_argument("--inbox-sample", type=int, default=DEFAULT_INBOX_SAMPLE,
                        help="Maximum number of Inbox messages to pull for analysis")
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging")
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)-7s %(message)s",
    )

    LOG.info("Reading parameters from %s", args.setup_xlsx)
    params = read_parameters_from_xlsx(args.setup_xlsx, args.parameters_sheet)

    missing = [p for p in REQUIRED_PARAMETERS if not params.get(p)]
    if missing:
        LOG.error("Missing required parameter(s) on the '%s' tab: %s",
                  args.parameters_sheet, ", ".join(missing))
        return 2

    client_id = params["GRAPH_CLIENT_ID"]
    tenant_id = params["GRAPH_TENANT_ID"]
    client_secret = params["GRAPH_CLIENT_SECRET"]

    manifest = load_manifest(args.manifest)
    permissions = validate_against_manifest(manifest, client_id)
    if permissions:
        LOG.info("Manifest declares Graph permissions: %s", ", ".join(permissions))

    token = acquire_token(tenant_id, client_id, client_secret)
    client = GraphClient(token)

    LOG.info("Enumerating mail folders for %s", args.mailbox)
    folders = enumerate_folders(client, args.mailbox)
    LOG.info("Found %d folders.", len(folders))

    LOG.info("Fetching up to %d Inbox messages for analysis", args.inbox_sample)
    messages = fetch_inbox_messages(client, args.mailbox, args.inbox_sample)
    inbox_stats = analyse_inbox(messages)
    LOG.info("Inbox sample: %d messages (%d unread).", inbox_stats["total"], inbox_stats["unread"])

    timestamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_mailbox = args.mailbox.replace("@", "_at_").replace(".", "_")
    output_path = os.path.join(args.output_dir, f"Mailbox_Report_{safe_mailbox}_{timestamp}.xlsx")

    run_meta = {
        "Mailbox": args.mailbox,
        "Application (clientId)": client_id,
        "Tenant": tenant_id,
        "App display name": (manifest or {}).get("displayName", "n/a"),
        "Graph permissions": ", ".join(permissions) if permissions else "n/a",
        "Folders enumerated": len(folders),
        "Inbox messages sampled": inbox_stats["total"],
        "Generated (local time)": _dt.datetime.now().isoformat(timespec="seconds"),
        "Parameters source": args.setup_xlsx,
    }

    LOG.info("Writing report to %s", output_path)
    write_report(output_path, args.mailbox, folders, inbox_stats, run_meta)
    LOG.info("Done. Report saved: %s", output_path)
    print(output_path)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:  # pragma: no cover
        sys.exit(130)
    except Exception as exc:  # noqa: BLE001 - top-level guard for operator clarity
        LOG.error("Fatal error: %s", exc)
        sys.exit(1)
