"""
Consignment CRUD Blueprint — Fusion Flow V2 BKD Portal
Uses existing app.db module for database access.
"""
from collections import namedtuple
import io
import json
import logging
import os
import re
from datetime import date, datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, Response

logger = logging.getLogger(__name__)
from app.db import query_all, query_one, execute, db_cursor, insert_api_call_log
from app.ingestion.defaults import resolve_ingest_defaults
from app.pipeline_validation import auto_validate_consignment_record, normalise_package_type
from app.search_utils import search_matches_values
from app.sdi_links import attach_sdi_links_to_consignments, load_prd_sdi_links_for_context, merge_sdi_links
from app.status_utils import (
    TSS_FILTER_STATUS_TABS,
    badge_class_for_status,
    canonical_filter_status,
    consignment_should_discover_sdi,
    effective_tss_filter_status,
    fixed_consignment_local_status,
    local_goods_status_after_parent_sync,
    normalize_status_key,
    status_filter_tabs,
    tss_allows_data_changes,
)
from app.tenant import get_tenant
from app.tss_guidance import explain_tss_error

CVOption = namedtuple('CVOption', ['value', 'name'])

consignments_bp = Blueprint('consignments', __name__,
    template_folder='../../templates/consignments',
    url_prefix='/consignments')

S = 'BKD'

PRD_LOCAL_CONSIGNMENT_EDIT_STATUSES = {
    'CREATED',
    'DRAFT',
    'PENDING',
    'PENDING REVIEW',
    'FAILED',
    'INVALID',
    'VALIDATED',
}
PRD_TSS_CONSIGNMENT_EDIT_STATUSES = {'DRAFT', 'TRADER INPUT REQUIRED'}
PRD_LOCAL_GOODS_DELETE_STATUSES = PRD_LOCAL_CONSIGNMENT_EDIT_STATUSES | {'DRAFT'}


def _safe_local_next_url(value, fallback):
    text = str(value or '').strip()
    if text.startswith('/') and not text.startswith('//') and '\\' not in text:
        return text
    return fallback


def _prd_tss_allows_consignment_edit(status):
    return normalize_status_key(status) in PRD_TSS_CONSIGNMENT_EDIT_STATUSES


def _prd_consignment_allows_edit(cons):
    if not cons:
        return False
    has_tss_ref = bool(str(cons.get('tss_consignment_ref') or '').strip())
    if has_tss_ref:
        return _prd_tss_allows_consignment_edit(cons.get('TssStatus') or cons.get('cons_tss_status'))
    return normalize_status_key(cons.get('sub_status')) in PRD_LOCAL_CONSIGNMENT_EDIT_STATUSES


def _table_column(cursor, table_name, *candidates):
    schema = get_tenant()["schema"]
    cursor.execute(
        """
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
        """,
        [schema, table_name],
    )
    columns = {str(row[0]).lower(): str(row[0]) for row in cursor.fetchall()}
    for candidate in candidates:
        if candidate and candidate.lower() in columns:
            return columns[candidate.lower()]
    return None


def _delete_supdecs_for_consignment_ids(cursor, consignment_ids):
    if not consignment_ids:
        return 0, 0

    if not _table_column(cursor, 'StagingSupDecHeaders', 'staging_cons_id'):
        return 0, 0
    supdec_header_id_col = _table_column(cursor, 'StagingSupDecHeaders', 'staging_id', 'id')
    supdec_goods_parent_col = _table_column(cursor, 'StagingSupDecGoods', 'staging_supdec_id', 'supdec_header_id')
    placeholders = ','.join('?' for _ in consignment_ids)
    deleted_goods = 0

    if supdec_header_id_col and supdec_goods_parent_col:
        cursor.execute(
            f"""
            DELETE g
            FROM {S}.StagingSupDecGoods g
            JOIN {S}.StagingSupDecHeaders sd
              ON sd.[{supdec_header_id_col}] = g.[{supdec_goods_parent_col}]
            WHERE sd.staging_cons_id IN ({placeholders})
            """,
            consignment_ids,
        )
        deleted_goods = cursor.rowcount

    cursor.execute(
        f"DELETE FROM {S}.StagingSupDecHeaders WHERE staging_cons_id IN ({placeholders})",
        consignment_ids,
    )
    return deleted_goods, cursor.rowcount


def _ingest_defaults_context():
    try:
        defaults = resolve_ingest_defaults(tenant_code=get_tenant()["code"])
    except Exception:
        defaults = None
    mode = getattr(defaults, 'mode', 'review_required')
    mode = 'auto_create_if_clean' if mode == 'auto_create_if_clean' else 'review_required'
    return {'ingest_defaults': defaults, 'creation_mode': mode}

CONS_EDIT_FIELD_META = {
    'ens_parent': {
        'label': 'ENS Header',
        'aliases': ['ens header', 'staging_ens_id', 'carrier_eori', 'carrier eori'],
        'suggestion': 'Select or re-link the correct ENS Header for this consignment. If TSS flagged carrier EORI, open the linked ENS Header and fill Carrier EORI there.',
    },
    'goods_description': {
        'label': 'Goods Description',
        'aliases': ['goods description', 'goods_description'],
        'suggestion': 'Enter a clear goods description, for example "Golf clubs" or "Textile garments".',
    },
    'trader_reference': {
        'label': 'Trader Reference',
        'aliases': ['trader reference', 'trader_reference'],
        'suggestion': 'Use the commercial reference for this consignment, if one is required by the movement.',
    },
    'transport_document_number': {
        'label': 'Transport Document Number',
        'aliases': ['transport document number', 'transport_document_number', 'transport doc'],
        'suggestion': 'Use the transport or booking reference that should identify this movement.',
    },
    'controlled_goods': {
        'label': 'Controlled Goods',
        'aliases': ['controlled goods', 'controlled_goods', 'yes/no'],
        'suggestion': 'Choose Yes or No so TSS knows whether controlled goods rules apply.',
    },
    'importer_eori': {
        'label': 'Importer EORI',
        'aliases': ['importer eori', 'importer_eori'],
        'suggestion': 'Use an EORI/TIN that TSS accepts in the selected environment/account. If TSS asks for it, provide the importer name/address as well.',
    },
    'importer_address_required': {
        'label': 'Importer Address Required',
        'aliases': ['importer address required', 'importer_address_required'],
        'suggestion': 'Set this to true when TSS warns that importer name/address must be provided for the importer EORI.',
    },
    'importer_name': {
        'label': 'Importer Name',
        'aliases': ['importer name', 'importer_name'],
        'suggestion': 'If the importer EORI is blank, enter the importer name exactly as it should be sent to TSS.',
    },
    'exporter_eori': {
        'label': 'Exporter EORI',
        'aliases': ['exporter eori', 'exporter_eori'],
        'suggestion': 'Use an EORI/TIN that TSS accepts in the selected environment/account. If TSS asks for the address, set Exporter Address Required / EORI Unknown to true and provide the full address.',
    },
    'exporter_address_required': {
        'label': 'Exporter Address Required / EORI Unknown',
        'aliases': ['exporter address required', 'exporter eori unknown', 'exporter_address_required'],
        'suggestion': 'Set this to true when TSS needs the exporter name/address, or when the exporter EORI is unknown.',
    },
    'exporter_name': {
        'label': 'Exporter Name',
        'aliases': ['exporter name', 'exporter_name'],
        'suggestion': 'If the exporter EORI is blank, enter the exporter name exactly as it should be sent to TSS.',
    },
    'consignor_eori': {
        'label': 'Consignor EORI',
        'aliases': ['consignor eori', 'consignor_eori'],
        'suggestion': 'Use a consignor EORI/TIN accepted by TSS. If TSS asks for the address, set Consignor Address Required / EORI Unknown to true and provide the full address.',
    },
    'consignor_address_required': {
        'label': 'Consignor Address Required / EORI Unknown',
        'aliases': ['consignor address required', 'consignor eori unknown', 'consignor_address_required'],
        'suggestion': 'Set this to true when TSS needs the consignor name/address, or when the consignor EORI is unknown.',
    },
    'consignor_name': {
        'label': 'Consignor Name',
        'aliases': ['consignor name', 'consignor_name'],
        'suggestion': 'If the consignor EORI is blank, enter the consignor name exactly as it should be sent to TSS.',
    },
    'consignee_eori': {
        'label': 'Consignee EORI',
        'aliases': ['consignee eori', 'consignee_eori'],
        'suggestion': 'Use a consignee EORI/TIN accepted by TSS. If the EORI is unknown or TSS needs the address, set Consignee Address Required / EORI Unknown to true and provide the full consignee name/address.',
    },
    'consignee_address_required': {
        'label': 'Consignee Address Required / EORI Unknown',
        'aliases': ['consignee address required', 'eori unknown', 'consignee_address_required'],
        'suggestion': 'Set this to true when TSS needs the consignee name/address, or when the consignee EORI is unknown.',
    },
    'consignee_name': {
        'label': 'Consignee Name',
        'aliases': ['consignee name', 'consignee_name'],
        'suggestion': 'If the consignee EORI is blank, enter the consignee name exactly as it should be sent to TSS.',
    },
    'buyer_same_as_importer': {
        'label': 'Buyer Same as Importer',
        'aliases': ['buyer same as importer', 'buyer_same_as_importer'],
        'suggestion': 'Choose Yes when the buyer is the importer, otherwise choose No and provide buyer details where required.',
    },
    'buyer_eori': {
        'label': 'Buyer EORI',
        'aliases': ['buyer eori', 'buyer_eori'],
        'suggestion': 'Use a buyer EORI/TIN accepted by TSS when the buyer is not the importer, or provide the required buyer address fields.',
    },
    'seller_same_as_exporter': {
        'label': 'Seller Same as Exporter',
        'aliases': ['seller same as exporter', 'seller_same_as_exporter'],
        'suggestion': 'Choose Yes when the seller is the exporter, otherwise choose No and provide seller details where required.',
    },
    'seller_eori': {
        'label': 'Seller EORI',
        'aliases': ['seller eori', 'seller_eori'],
        'suggestion': 'Use a seller EORI/TIN accepted by TSS when the seller is not the exporter, or provide the required seller address fields.',
    },
    'container_indicator': {
        'label': 'Container Indicator',
        'aliases': ['container indicator', 'container_indicator'],
        'suggestion': 'Choose 0 for Uncontainerised or 1 for Containerised before validating or sending cargo to TSS.',
    },
    'goods_domestic_status': {
        'label': 'Goods Domestic Status',
        'aliases': ['goods domestic status', 'goods_domestic_status', 'domestic status'],
        'suggestion': 'If Controlled Goods is "Yes", choose the matching domestic status.',
    },
    'destination_country': {
        'label': 'Destination Country',
        'aliases': ['destination country', 'destination_country'],
        'suggestion': 'Select the destination country from the official country list.',
    },
    'ducr': {
        'label': 'DUCR',
        'aliases': ['ducr'],
        'suggestion': 'Enter the DUCR only if this movement requires one, in the trader-issued format.',
    },
    'supervising_customs_office': {
        'label': 'Supervising Customs Office',
        'aliases': ['supervising customs office', 'supervising_customs_office'],
        'suggestion': 'Enter the supervising customs office when the declaration choice requires it.',
    },
    'align_ukims': {
        'label': 'Align UKIMS',
        'aliases': ['align ukims', 'align_ukims'],
        'suggestion': 'Choose Yes or No only when UKIMS alignment applies to this consignment.',
    },
    'customs_warehouse_identifier': {
        'label': 'Customs Warehouse Identifier',
        'aliases': ['customs warehouse identifier', 'customs_warehouse_identifier'],
        'suggestion': 'Enter the customs warehouse identifier when the declaration choice requires it.',
    },
    'use_importer_sde': {
        'label': 'Use Importer SDE',
        'aliases': ['use importer sde', 'use_importer_sde'],
        'suggestion': 'Choose Yes or No when this consignment needs importer SDE handling.',
    },
    'declaration_choice': {
        'label': 'Declaration Choice',
        'aliases': ['declaration choice', 'declaration_choice'],
        'suggestion': 'Choose the declaration route expected for this consignment.',
    },
    'generate_SD': {
        'label': 'Generate SDI/SupDec',
        'aliases': ['generate sd', 'generate_sd'],
        'suggestion': 'Choose Yes when the downstream SDI/SupDec requirement is known in advance. Fusion will still discover SDI from TSS once an SFD exists.',
    },
    'no_sfd_reason': {
        'label': 'No SFD Reason',
        'aliases': ['no sfd reason', 'no_sfd_reason'],
        'suggestion': 'Select a TSS No SFD reason when this should be ENS-only or when TSS requires it because the importer EORI is not registered for SFD creation.',
    },
}

PARTY_DETAIL_ROLES = {
    'consignor': 'Consignor',
    'consignee': 'Consignee',
    'importer': 'Importer',
    'exporter': 'Exporter',
    'buyer': 'Buyer',
    'seller': 'Seller',
}
PARTY_ADDRESS_FIELD_META = {
    'street_number': ('Street / name of street', 'street or name of street'),
    'city': ('City', 'city'),
    'postcode': ('Postcode', 'postcode'),
    'country': ('Country', 'country'),
}
for _role, _role_label in PARTY_DETAIL_ROLES.items():
    for _suffix, (_label_suffix, _alias_suffix) in PARTY_ADDRESS_FIELD_META.items():
        _field = f"{_role}_street_and_number" if _role in {'buyer', 'seller'} and _suffix == 'street_number' else f"{_role}_{_suffix}"
        CONS_EDIT_FIELD_META.setdefault(_field, {
            'label': f"{_role_label} {_label_suffix}",
            'aliases': [
                _field,
                _field.replace('_', ' '),
                f"{_role} {_alias_suffix}",
                f"{_role_label.lower()} {_alias_suffix}",
            ],
            'suggestion': f"Fill the {_role_label.lower()} {_label_suffix.lower()} exactly as it should be sent to TSS.",
        })

NULLISH_FORM_VALUES = {'', 'none', 'null', 'undefined'}
UNSAFE_SAMPLE_EORIS = {
    'GB000000000000',
}
EORI_FORM_FIELDS = {
    'consignor_eori',
    'consignee_eori',
    'importer_eori',
    'exporter_eori',
    'buyer_eori',
    'seller_eori',
}
YES_NO_FORM_FIELDS = {
    'buyer_same_as_importer',
    'seller_same_as_exporter',
    'controlled_goods',
    'align_ukims',
    'use_importer_sde',
    'generate_SD',
}
TRUE_FALSE_FORM_FIELDS = {
    'consignor_address_required',
    'consignee_address_required',
    'importer_address_required',
    'exporter_address_required',
}
ADDRESS_REQUIRED_VALIDATION_PREFIXES = ('consignee',)
ADDRESS_REQUIRED_FIELD_SUFFIXES = ('name', 'street_number', 'city', 'postcode', 'country')
COUNTRY_FORM_FIELDS = {
    'destination_country',
    'consignor_country',
    'consignee_country',
    'importer_country',
    'exporter_country',
    'buyer_country',
    'seller_country',
}
BUYER_DETAIL_FIELDS = {
    'buyer_eori',
    'buyer_name',
    'buyer_street_and_number',
    'buyer_city',
    'buyer_postcode',
    'buyer_country',
}
SELLER_DETAIL_FIELDS = {
    'seller_eori',
    'seller_name',
    'seller_street_and_number',
    'seller_city',
    'seller_postcode',
    'seller_country',
}
CONSIGNMENT_CREATE_PARAM_COLUMNS = [
    'staging_ens_id',
    'label',
    'goods_description',
    'trader_reference',
    'transport_document_number',
    'controlled_goods',
    'goods_domestic_status',
    'destination_country',
    'supervising_customs_office',
    'customs_warehouse_identifier',
    'ducr',
    'no_sfd_reason',
    'consignor_eori',
    'consignor_name',
    'consignor_street_number',
    'consignor_city',
    'consignor_postcode',
    'consignor_country',
    'consignee_eori',
    'consignee_name',
    'consignee_street_number',
    'consignee_city',
    'consignee_postcode',
    'consignee_country',
    'importer_eori',
    'importer_name',
    'importer_street_number',
    'importer_city',
    'importer_postcode',
    'importer_country',
    'exporter_eori',
    'exporter_name',
    'exporter_street_number',
    'exporter_city',
    'exporter_postcode',
    'exporter_country',
    'buyer_same_as_importer',
    'seller_same_as_exporter',
    'buyer_eori',
    'buyer_name',
    'buyer_street_and_number',
    'buyer_city',
    'buyer_postcode',
    'buyer_country',
    'seller_eori',
    'seller_name',
    'seller_street_and_number',
    'seller_city',
    'seller_postcode',
    'seller_country',
    'container_indicator',
    'align_ukims',
    'use_importer_sde',
    'declaration_choice',
    'generate_SD',
]
CONSIGNMENT_CREATE_FIXED_COLUMNS = [
    'status',
    'retry_count',
    'max_retries',
    'created_at',
]


def _clean_consignment_form_value(value):
    if value is None:
        return None
    text = str(value).strip()
    if text.lower() in NULLISH_FORM_VALUES:
        return None
    return text


def _consignment_field_label(field):
    meta = CONS_EDIT_FIELD_META.get(field, {})
    return meta.get('label') or field.replace('_', ' ').title()


def _add_consignment_form_warning(warnings, field, message):
    if warnings is None:
        return
    warnings.setdefault(field, []).append(message)


def _is_truthy_form_flag(value):
    return str(value or '').strip().lower() in {'1', 'true', 'yes', 'on'}


def _party_address_field_names(prefix):
    return [f'{prefix}_{suffix}' for suffix in ADDRESS_REQUIRED_FIELD_SUFFIXES]


def _missing_party_address_fields(record, prefix):
    return [
        field for field in _party_address_field_names(prefix)
        if not str((record or {}).get(field) or '').strip()
    ]


def _address_required_missing_message(prefix):
    label = PARTY_DETAIL_ROLES.get(prefix, prefix.title())
    return (
        f"{label} Address Required / EORI Unknown is true, so complete the full "
        f"{label.lower()} name/address before saving or rerunning TSS."
    )


def _add_address_required_form_errors(errors, cleaned):
    for prefix in ADDRESS_REQUIRED_VALIDATION_PREFIXES:
        if not _is_truthy_form_flag((cleaned or {}).get(f'{prefix}_address_required')):
            continue
        for field in _missing_party_address_fields(cleaned, prefix):
            _add_consignment_form_warning(errors, field, _address_required_missing_message(prefix))


def _clean_consignment_form(form, *, field_warnings=None):
    cleaned = {key: _clean_consignment_form_value(value) for key, value in form.items()}

    for field in YES_NO_FORM_FIELDS:
        if field in cleaned and cleaned[field] is not None:
            value = str(cleaned[field]).strip().lower()
            cleaned[field] = value if value in {'yes', 'no'} else cleaned[field]

    for field in TRUE_FALSE_FORM_FIELDS:
        if field in cleaned:
            value = str(cleaned.get(field) or '').strip().lower()
            cleaned[field] = 'true' if value in {'1', 'true', 'yes', 'on'} else 'false'

    for field in COUNTRY_FORM_FIELDS:
        value = cleaned.get(field)
        if value is not None and len(str(value).strip()) > 2:
            cleaned[field] = None

    for field in EORI_FORM_FIELDS:
        value = cleaned.get(field)
        if value is not None and str(value).strip().upper() in UNSAFE_SAMPLE_EORIS:
            _add_consignment_form_warning(
                field_warnings,
                field,
                (
                    f"{_consignment_field_label(field)} '{value}' was not saved because "
                    "it looks like a placeholder/test EORI. Use a real EORI/TIN accepted "
                    "by TSS, or leave it blank only when the matching Address Required / "
                    "EORI Unknown option applies."
                ),
            )
            cleaned[field] = None

    _add_address_required_form_errors(field_warnings, cleaned)

    # Hidden buyer/seller inputs can still be posted by the browser. When the
    # same-as flags are active, TSS derives these parties from importer/exporter.
    if cleaned.get('buyer_same_as_importer', 'yes') != 'no':
        for field in BUYER_DETAIL_FIELDS:
            cleaned[field] = None
        cleaned['buyer_same_as_importer'] = 'yes'
    if cleaned.get('seller_same_as_exporter', 'yes') != 'no':
        for field in SELLER_DETAIL_FIELDS:
            cleaned[field] = None
        cleaned['seller_same_as_exporter'] = 'yes'

    return cleaned


def _party_address_required_display(row, prefix, *, force_when_address=False):
    record = dict(row or {})
    raw = record.get(f'{prefix}_address_required')
    if raw is not None:
        return 'true' if str(raw).strip().lower() in {'1', 'true', 'yes', 'on'} else 'false'

    address_fields = (
        f'{prefix}_name',
        f'{prefix}_street_number',
        f'{prefix}_city',
        f'{prefix}_postcode',
        f'{prefix}_country',
    )
    has_any_address = any(str(record.get(field) or '').strip() for field in address_fields)
    if force_when_address and has_any_address:
        return 'true'

    if str(record.get(f'{prefix}_eori') or '').strip():
        return 'false'

    return 'true' if has_any_address else 'false'


def _clean_consignment_display_row(row):
    display = {
        key: '' if _clean_consignment_form_value(value) is None else value
        for key, value in dict(row or {}).items()
    }
    for prefix in ('consignor', 'consignee', 'importer', 'exporter'):
        display[f'{prefix}_address_required'] = _party_address_required_display(
            row,
            prefix,
            force_when_address=prefix == 'importer',
        )
    return display


def _safe_master_eori(row):
    for field in ('eori_xi', 'eori_gb'):
        value = str((row or {}).get(field) or '').strip().upper()
        if value and value not in UNSAFE_SAMPLE_EORIS:
            return value
    return ''


def _load_company_master_defaults():
    try:
        row = query_one(
            f"""
            SELECT TOP 1 company_name, trading_name, eori_xi, eori_gb,
                   address_line1, city, postcode, country
            FROM {S}.CompanyMaster
            ORDER BY
                CASE
                    WHEN eori_xi IS NOT NULL AND LTRIM(RTRIM(eori_xi)) <> '' THEN 0
                    WHEN eori_gb IS NOT NULL AND LTRIM(RTRIM(eori_gb)) <> '' THEN 1
                    ELSE 2
                END,
                id
            """
        )
    except Exception:
        return {}

    eori = _safe_master_eori(row)
    name = str((row or {}).get('trading_name') or (row or {}).get('company_name') or '').strip()
    address = {
        'eori': eori,
        'name': name[:35],
        'street_number': str((row or {}).get('address_line1') or '').strip()[:35],
        'city': str((row or {}).get('city') or '').strip()[:35],
        'postcode': str((row or {}).get('postcode') or '').strip()[:35],
        'country': str((row or {}).get('country') or '').strip().upper()[:2],
    }
    return address if any(address.values()) else {}


def _default_consignment_form():
    form = {
        'controlled_goods': 'no',
        'goods_domestic_status': 'D',
        'destination_country': 'GB',
        'buyer_same_as_importer': 'yes',
        'seller_same_as_exporter': 'yes',
    }
    company = _load_company_master_defaults()
    if not company:
        return form

    for role in ('consignor', 'importer', 'exporter'):
        form[f'{role}_eori'] = company.get('eori', '')
        form[f'{role}_name'] = company.get('name', '')
        form[f'{role}_street_number'] = company.get('street_number', '')
        form[f'{role}_city'] = company.get('city', '')
        form[f'{role}_postcode'] = company.get('postcode', '')
        form[f'{role}_country'] = company.get('country', '')
    if company.get('country'):
        form['destination_country'] = company['country']
    return form


def _consignment_create_params(staging_ens_id, form):
    return [
        staging_ens_id,
        form.get('label', ''),
        form.get('goods_description', ''),
        form.get('trader_reference', '') or None,
        form.get('transport_document_number', ''),
        form.get('controlled_goods', 'no'),
        form.get('goods_domestic_status', '') or None,
        form.get('destination_country', '') or None,
        form.get('supervising_customs_office', '') or None,
        form.get('customs_warehouse_identifier', '') or None,
        form.get('ducr', '') or None,
        form.get('no_sfd_reason', '') or None,
        form.get('consignor_eori', '') or None,
        form.get('consignor_name', '') or None,
        form.get('consignor_street_number', '') or None,
        form.get('consignor_city', '') or None,
        form.get('consignor_postcode', '') or None,
        form.get('consignor_country', '') or None,
        form.get('consignee_eori', '') or None,
        form.get('consignee_name', '') or None,
        form.get('consignee_street_number', '') or None,
        form.get('consignee_city', '') or None,
        form.get('consignee_postcode', '') or None,
        form.get('consignee_country', '') or None,
        form.get('importer_eori', '') or None,
        form.get('importer_name', '') or None,
        form.get('importer_street_number', '') or None,
        form.get('importer_city', '') or None,
        form.get('importer_postcode', '') or None,
        form.get('importer_country', '') or None,
        form.get('exporter_eori', '') or None,
        form.get('exporter_name', '') or None,
        form.get('exporter_street_number', '') or None,
        form.get('exporter_city', '') or None,
        form.get('exporter_postcode', '') or None,
        form.get('exporter_country', '') or None,
        form.get('buyer_same_as_importer', 'yes'),
        form.get('seller_same_as_exporter', 'yes'),
        form.get('buyer_eori', '') or None,
        form.get('buyer_name', '') or None,
        form.get('buyer_street_and_number', '') or None,
        form.get('buyer_city', '') or None,
        form.get('buyer_postcode', '') or None,
        form.get('buyer_country', '') or None,
        form.get('seller_eori', '') or None,
        form.get('seller_name', '') or None,
        form.get('seller_street_and_number', '') or None,
        form.get('seller_city', '') or None,
        form.get('seller_postcode', '') or None,
        form.get('seller_country', '') or None,
        form.get('container_indicator', '') or None,
        form.get('align_ukims', '') or None,
        form.get('use_importer_sde', '') or None,
        form.get('declaration_choice', '') or None,
        form.get('generate_SD', '') or None,
    ]


def _consignment_create_sql(schema):
    columns = CONSIGNMENT_CREATE_PARAM_COLUMNS + CONSIGNMENT_CREATE_FIXED_COLUMNS
    column_sql = ', '.join(columns)
    placeholder_sql = ', '.join('?' for _ in CONSIGNMENT_CREATE_PARAM_COLUMNS)
    return f"""
        INSERT INTO {schema}.StagingConsignments (
            {column_sql}
        )
        OUTPUT INSERTED.staging_id
        VALUES (
            {placeholder_sql},
            'PENDING', 0, 3, SYSUTCDATETIME()
        )
    """

def badge_class(status):
    return badge_class_for_status(status)


def _consignment_display_status(status='', tss_status=''):
    return effective_tss_filter_status(status, tss_status)


def _apply_consignment_goods_status(goods, cons):
    normalised = []
    for item in goods or []:
        row = dict(item)
        row['status'] = local_goods_status_after_parent_sync(
            row.get('status'),
            goods_tss_status=row.get('tss_status'),
            parent_local_status=(cons or {}).get('status'),
            parent_tss_status=(cons or {}).get('tss_status'),
        )
        normalised.append(row)
    return normalised


def _consignment_needs_tss_sync(cons):
    has_reference = bool((cons or {}).get('dec_reference') or (cons or {}).get('sfd_reference'))
    tss_status = normalize_status_key((cons or {}).get('tss_status'))
    return has_reference and (not tss_status or tss_status in {'PENDING SYNC', 'IMPORTED', 'SYNC PENDING'})


def _consignment_status_tabs(counts, selected='ALL'):
    base = TSS_FILTER_STATUS_TABS
    return status_filter_tabs(counts, base, selected)


def _consignment_arrival_sort_key(row):
    value = (row or {}).get('arrival_date_time')
    if not value:
        return datetime.min
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    normalised = text.replace('Z', '+00:00')
    if '.' in normalised:
        head, tail = normalised.split('.', 1)
        for marker in ('+', '-'):
            if marker in tail:
                fraction, timezone = tail.split(marker, 1)
                tail = f"{fraction[:6]}{marker}{timezone}"
                break
        else:
            tail = tail[:6]
        normalised = f"{head}.{tail}"
    try:
        return datetime.fromisoformat(normalised)
    except ValueError:
        pass
    for fmt in (
        '%d/%m/%Y %H:%M:%S',
        '%d/%m/%Y %H:%M',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%dT%H:%M:%S',
        '%Y-%m-%d',
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return datetime.min


def _consignment_created_date_matches(row, expected_date):
    if not expected_date:
        return True
    value = (row or {}).get('created_at')
    if not value:
        return False
    if isinstance(value, datetime):
        return value.date() == expected_date
    parsed = _consignment_arrival_sort_key({'arrival_date_time': value})
    return parsed != datetime.min and parsed.date() == expected_date


def _parse_iso_filter_date(value):
    text = str(value or '').strip()
    if not text:
        return None, ''
    try:
        parsed = datetime.strptime(text, '%Y-%m-%d').date()
    except ValueError:
        return None, ''
    return parsed, text


def _parse_arrival_month(value):
    text = str(value or '').strip()
    if not text:
        return None, None, ''
    try:
        first = datetime.strptime(text, '%Y-%m').date().replace(day=1)
    except ValueError:
        return None, None, ''
    if first.month == 12:
        last = date(first.year, 12, 31)
    else:
        last = date(first.year, first.month + 1, 1)
        last = date.fromordinal(last.toordinal() - 1)
    return first, last, text


def _arrival_filter_bounds(values):
    getter = values.get if hasattr(values, 'get') else lambda key, default=None: default
    arrival_month_raw = getter('arrival_month', '')
    arrival_from_raw = getter('arrival_from', '')
    arrival_to_raw = getter('arrival_to', '')
    legacy_created_on = getter('created_on', '')

    month_from, month_to, arrival_month = _parse_arrival_month(arrival_month_raw)
    if arrival_month:
        return month_from, month_to, '', '', arrival_month

    arrival_from_date, arrival_from = _parse_iso_filter_date(arrival_from_raw)
    arrival_to_date, arrival_to = _parse_iso_filter_date(arrival_to_raw)

    if not arrival_from_date and not arrival_to_date and legacy_created_on:
        legacy_date, legacy_text = _parse_iso_filter_date(legacy_created_on)
        if legacy_date:
            return legacy_date, legacy_date, legacy_text, legacy_text, ''

    if arrival_from_date and arrival_to_date and arrival_from_date > arrival_to_date:
        arrival_from_date, arrival_to_date = arrival_to_date, arrival_from_date
        arrival_from, arrival_to = arrival_to, arrival_from
    return arrival_from_date, arrival_to_date, arrival_from, arrival_to, ''


def _consignment_arrival_date(row):
    parsed = _consignment_arrival_sort_key(row)
    if parsed == datetime.min:
        return None
    return parsed.date()


def _consignment_arrival_date_in_range(row, start_date, end_date):
    if not start_date and not end_date:
        return True
    arrival = _consignment_arrival_date(row)
    if not arrival:
        return False
    if start_date and arrival < start_date:
        return False
    if end_date and arrival > end_date:
        return False
    return True


def _apply_consignment_sort(rows, sort):
    if sort == 'arrival_asc':
        rows.sort(key=_consignment_arrival_sort_key)
    elif sort == 'arrival_desc':
        rows.sort(key=_consignment_arrival_sort_key, reverse=True)
    return rows


def _safe_positive_int(value, default=1):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        parsed = int(default)
    return max(1, parsed)


def _consignment_page_size(value, *, show_all=False):
    if show_all or str(value or '').strip().lower() in {'all', 'show_all'}:
        return None, True, 'all'
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        parsed = 20
    if parsed not in {10, 20, 50}:
        parsed = 20
    return parsed, False, str(parsed)


def _can_change_consignment_data(cons):
    return tss_allows_data_changes((cons or {}).get('tss_status'), (cons or {}).get('status'))


def _can_edit_consignment_data(cons):
    if not _can_change_consignment_data(cons):
        return False
    local_status = normalize_status_key((cons or {}).get('status'))
    tss_status = normalize_status_key((cons or {}).get('tss_status'))
    repairable_tss_statuses = {
        'TRADER INPUT REQUIRED',
        'AMENDMENT REQUIRED',
        'ERROR',
        'DO NOT LOAD',
    }
    return local_status in {'PENDING', 'PENDING REVIEW', 'FAILED', 'INVALID'} or tss_status in repairable_tss_statuses


def _can_validate_consignment_pipeline(cons, goods=None):
    if not _can_change_consignment_data(cons):
        return False
    if (cons or {}).get('dec_reference') or (cons or {}).get('sfd_reference'):
        return False
    local_status = normalize_status_key((cons or {}).get('status'))
    if local_status not in {'PENDING', 'PENDING REVIEW'}:
        return False
    return goods is None or len(goods) > 0


def _can_add_goods_to_consignment(cons):
    return _can_change_consignment_data(cons)


def _can_cancel_consignment_in_tss(cons):
    cons = cons or {}
    dec_ref = (
        cons.get('dec_reference')
        or cons.get('tss_consignment_ref')
        or cons.get('ConsignmentReference')
    )
    if not dec_ref:
        return False

    local_status = normalize_status_key(cons.get('sub_status') or cons.get('status'))
    tss_status = normalize_status_key(
        cons.get('TssStatus')
        or cons.get('cons_tss_status')
        or cons.get('tss_status')
    )
    blocked_local_statuses = {
        'CANCELLED',
        'CANCELED',
        'DELETED',
        'COMPLETED',
    }
    blocked_tss_statuses = {
        'ARRIVED',
        'CANCELLED',
        'CANCELED',
        'CLOSED',
        'COMPLETED',
        'DELETED',
    }
    if tss_status.startswith('ARRIVED'):
        return False
    return local_status not in blocked_local_statuses and tss_status not in blocked_tss_statuses


def _can_recreate_cancelled_consignment(cons):
    cons = cons or {}
    local_status = normalize_status_key(cons.get('sub_status') or cons.get('status'))
    tss_status = normalize_status_key(
        cons.get('TssStatus')
        or cons.get('cons_tss_status')
        or cons.get('tss_status')
    )
    return bool(
        cons.get('stg_header_id')
        and (cons.get('tss_ens_header_ref') or cons.get('ens_reference'))
        and (local_status in {'CANCELLED', 'CANCELED'} or tss_status in {'CANCELLED', 'CANCELED'})
    )


class _SqlNow:
    pass


_SQL_NOW = _SqlNow()


def _sql_now():
    return _SQL_NOW


def _stg_table_columns(cursor, table_name):
    cursor.execute(
        """
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = 'STG' AND TABLE_NAME = ?
        ORDER BY ORDINAL_POSITION
        """,
        [table_name],
    )
    return [str(row[0]) for row in cursor.fetchall()]


def _cursor_row_as_dict(cursor):
    row = cursor.fetchone()
    if not row:
        return {}
    columns = [item[0] for item in cursor.description or []]
    return {column: row[idx] for idx, column in enumerate(columns)}


def _insert_stg_clone(cursor, table_name, identity_column, source_record, overrides):
    columns = _stg_table_columns(cursor, table_name)
    source_by_lower = {str(key).lower(): value for key, value in (source_record or {}).items()}
    overrides_by_lower = {str(key).lower(): value for key, value in (overrides or {}).items()}
    identity_key = str(identity_column or '').lower()

    insert_columns = []
    values = []
    for column in columns:
        key = column.lower()
        if key == identity_key:
            continue
        if key in overrides_by_lower:
            value = overrides_by_lower[key]
        elif key in source_by_lower:
            value = source_by_lower[key]
        else:
            continue
        insert_columns.append(column)
        values.append(value)

    if not insert_columns:
        raise RuntimeError(f'No cloneable columns found for STG.{table_name}.')

    placeholders = []
    params = []
    for value in values:
        if value is _SQL_NOW:
            placeholders.append('SYSUTCDATETIME()')
        else:
            placeholders.append('?')
            params.append(value)

    cursor.execute(
        f"""
        INSERT INTO [STG].[{table_name}] ({', '.join(f'[{column}]' for column in insert_columns)})
        OUTPUT INSERTED.[{identity_column}]
        VALUES ({', '.join(placeholders)})
        """,
        params,
    )
    row = cursor.fetchone()
    return int(row[0])


def _clone_cancelled_consignment_for_resend(cursor, cons):
    cons = dict(cons or {})
    new_cons_id = _insert_stg_clone(
        cursor,
        'BKD_ENS_Consignments',
        'stg_consignment_id',
        cons,
        {
            'sub_status': 'PENDING',
            'source': 'RECREATED_CANCELLED',
            'ing_consignment_id': None,
            'tss_consignment_ref': None,
            'tss_ens_header_ref': cons.get('tss_ens_header_ref') or cons.get('ens_reference'),
            'tss_api_http_status': None,
            'metadata_json': json.dumps({
                'recreated_from_cancelled_stg_consignment_id': cons.get('stg_consignment_id'),
                'recreated_from_cancelled_dec': cons.get('tss_consignment_ref'),
            }, default=str),
            'validated_at': None,
            'submitted_at': None,
            'completed_at': None,
            'last_sub_status_change': _sql_now(),
            'stg_created_at': _sql_now(),
            'updated_at': _sql_now(),
        },
    )

    cursor.execute(
        """
        SELECT *
        FROM [STG].[BKD_GoodsItems]
        WHERE ClientCode = ?
          AND stg_consignment_id = ?
          AND UPPER(COALESCE(goods_stage, 'ENS')) = 'ENS'
        ORDER BY COALESCE(item_seq, stg_item_id), stg_item_id
        """,
        [cons.get('ClientCode'), cons.get('stg_consignment_id')],
    )
    goods_rows = []
    columns = [item[0] for item in cursor.description or []]
    for row in cursor.fetchall():
        goods_rows.append({column: row[idx] for idx, column in enumerate(columns)})

    cloned_goods = 0
    for goods in goods_rows:
        _insert_stg_clone(
            cursor,
            'BKD_GoodsItems',
            'stg_item_id',
            goods,
            {
                'stg_consignment_id': new_cons_id,
                'sub_status': 'PENDING',
                'source': 'RECREATED_CANCELLED',
                'goods_stage': 'ENS',
                'ing_item_id': None,
                'tss_hex_id': None,
                'tss_consignment_ref': None,
                'tss_sup_dec_number': None,
                'tss_sfd_number': None,
                'stg_sdi_id': None,
                'source_stg_item_id': goods.get('stg_item_id'),
                'error_message': None,
                'validated_at': None,
                'submitted_at': None,
                'completed_at': None,
                'last_sub_status_change': _sql_now(),
                'stg_created_at': _sql_now(),
                'updated_at': _sql_now(),
            },
        )
        cloned_goods += 1

    return new_cons_id, cloned_goods


def _local_failed_cleanup_filter(search='', ens_ref=''):
    where = """
        WHERE c.status = 'FAILED'
          AND c.dec_reference IS NULL
          AND c.sfd_reference IS NULL
    """
    params = []
    if search:
        where += """
            AND (
                c.goods_description LIKE ?
                OR c.dec_reference LIKE ?
                OR c.transport_document_number LIKE ?
                OR c.importer_eori LIKE ?
                OR c.tss_status LIKE ?
                OR c.consignor_eori LIKE ?
                OR c.consignee_eori LIKE ?
                OR c.exporter_eori LIKE ?
            )
        """
        params.extend([f"%{search}%"] * 8)
    if ens_ref:
        where += " AND e.ens_reference = ?"
        params.append(ens_ref)
    return where, params


def _detail_action(label, href=None, kind='primary', phase=None, post_to=None):
    action = {'label': label, 'kind': kind}
    if href:
        action['href'] = href
    if phase:
        if phase in {'sync', 'sync_pipeline', 'sync_gmr', 'sync_tss'}:
            phase = 'sync_all'
            if str(label).lower().startswith('sync'):
                action['label'] = 'Sync All TSS Data'
        action['phase'] = phase
        action['confirm_text'] = f"Run {action['label']} now?"
    if post_to:
        action['post_to'] = post_to
        action['confirm_text'] = action.get('confirm_text') or f'Run {label} now?'
    return action


def _requires_sdi(record):
    return consignment_should_discover_sdi(record)


def _with_linked_sfd_reference(cons, linked_sfd=None):
    enriched = dict(cons or {})
    linked_sfd = linked_sfd or {}
    linked_ref = linked_sfd.get('sfd_reference') or linked_sfd.get('sfd_number') or linked_sfd.get('reference')
    if linked_ref:
        enriched.setdefault('synced_sfd_reference', linked_ref)
    return enriched


def _load_linked_sfd_for_consignment(cons):
    refs = [
        (cons or {}).get('dec_reference'),
        (cons or {}).get('sfd_reference'),
    ]
    refs = [str(ref).strip() for ref in refs if ref]
    if not refs:
        return None

    placeholders = ', '.join('?' for _ in refs)
    params = refs * 4
    try:
        return query_one(f"""
            SELECT TOP 1 id, sfd_number, sfd_reference, ens_consignment_reference,
                   declaration_number, tss_status, mrn, movement_reference_number,
                   eori_for_eidr, updated_at, created_at
            FROM {S}.Sfds
            WHERE ens_consignment_reference IN ({placeholders})
               OR declaration_number IN ({placeholders})
               OR sfd_reference IN ({placeholders})
               OR sfd_number IN ({placeholders})
            ORDER BY id DESC
        """, params)
    except Exception:
        logger.exception("Failed to load linked SFD for consignment %s", (cons or {}).get('staging_id'))
        return None


def _build_sfd_route_a_state(cons, linked_sfd=None):
    cons = cons or {}
    linked_sfd = linked_sfd or {}
    tss_status = normalize_status_key(cons.get('tss_status') or cons.get('status'))
    if tss_status.startswith('TSS:'):
        tss_status = tss_status.split(':', 1)[1].strip()
    reference = (
        linked_sfd.get('sfd_reference')
        or linked_sfd.get('sfd_number')
        or cons.get('sfd_reference')
        or ''
    )
    mrn = (
        linked_sfd.get('movement_reference_number')
        or linked_sfd.get('mrn')
        or cons.get('sfd_mrn')
        or ''
    )
    eidr = linked_sfd.get('eori_for_eidr') or ''
    status = linked_sfd.get('tss_status') or cons.get('sfd_status') or ''
    parent_reference = (
        linked_sfd.get('declaration_number')
        or linked_sfd.get('ens_consignment_reference')
        or cons.get('dec_reference')
        or ''
    )
    updated_at = linked_sfd.get('updated_at') or linked_sfd.get('created_at')
    if hasattr(updated_at, 'strftime'):
        updated_label = updated_at.strftime('%d/%m/%Y %H:%M')
    else:
        updated_label = str(updated_at or '')
    authorised = tss_status in ('AUTHORISED FOR MOVEMENT', 'AUTHORIZED FOR MOVEMENT', 'ARRIVED')
    visible = bool(reference or mrn or eidr or status or authorised)

    if reference:
        title = 'SFD created by TSS'
        detail = 'TSS has generated or synced the SFD for this consignment.'
        tone = 'success'
    elif authorised:
        title = 'Authorised for Movement - SFD pending sync'
        detail = (
            'TSS has authorised the consignment. If TSS has already generated the SFD, '
            'run Sync Cargo Statuses or Sync TSS Tables to pull the SFD reference into Fusion.'
        )
        tone = 'info'
    else:
        title = 'SFD not created yet'
        detail = 'The SFD appears after TSS progresses the consignment far enough.'
        tone = 'muted'

    customs_label = 'EIDR' if eidr and not mrn else 'MRN'
    customs_reference = eidr if eidr and not mrn else mrn
    return {
        'visible': visible,
        'reference': reference,
        'parent_reference': parent_reference,
        'customs_reference': customs_reference,
        'customs_label': customs_label,
        'mrn': mrn,
        'eidr': eidr,
        'status': status,
        'updated_at': updated_at,
        'updated_label': updated_label,
        'title': title,
        'detail': detail,
        'tone': tone,
        'has_synced_row': bool(linked_sfd),
    }


def _pipeline_job_markers():
    rows = query_all(f"""
        SELECT call_type, MAX(called_at) AS called_at
        FROM {S}.ApiCallLog
        WHERE call_type IN ('JOB_VALIDATE_PIPELINE', 'JOB_SUBMIT_PIPELINE', 'JOB_SYNC_PIPELINE')
        GROUP BY call_type
    """)
    return {row['call_type']: row['called_at'] for row in rows}


def _clean_pipeline_error(message):
    text = (message or '').strip()
    if not text:
        return ''
    if text.startswith('{'):
        try:
            payload = json.loads(text)
            result = payload.get('result') or {}
            return (result.get('process_message') or result.get('message') or text).strip()
        except Exception:
            return text
    return text


def _is_invalid_consignment_submit_error(message):
    cleaned = _clean_pipeline_error(message).lower()
    return 'invalid op_type' in cleaned and 'submit' in cleaned


def _build_consignment_error_explanation(cons):
    cleaned = _clean_pipeline_error(cons.get('error_message'))
    if not _is_invalid_consignment_submit_error(cleaned):
        return explain_tss_error(
            cleaned,
            local_status=cons.get('status'),
            tss_status=cons.get('tss_status'),
            entity_label='this consignment',
        )

    dec_match = re.search(r"'(DEC\d+)'", cleaned, flags=re.IGNORECASE)
    dec_ref = cons.get('dec_reference') or (dec_match.group(1) if dec_match else 'this DEC')
    local_status = (cons.get('status') or 'Unknown').strip()
    tss_status = (cons.get('tss_status') or 'Pending sync').strip()

    return {
        'title': 'TSS rejected the submit action',
        'summary': (
            f'This does not mean {dec_ref} is Submitted. '
            f'Local {local_status} means the DEC was created and linked; '
            f'TSS {tss_status} is the current status reported by TSS.'
        ),
        'detail': (
            'The word submit appears because the cargo pipeline tried to run op_type=submit '
            'for this DEC, and TSS rejected that operation for the DEC in its current state.'
        ),
        'next_step': (
            'Run Sync TSS Status to confirm the latest TSS state. If it remains Draft, '
            'do not read this error as a successful submission; treat it as a rejected action.'
        ),
        'technical': cleaned,
        'raw': cons.get('error_message') or cleaned,
    }


def _is_local_validation_failure(message):
    upper = _clean_pipeline_error(message).upper()
    if not upper:
        return False
    local_markers = (
        'REQUIRED:',
        'FORMAT:',
        'INVALID:',
        'LENGTH:',
        'MUST LINK TO AN ENS HEADER',
        'DOMESTIC STATUS REQUIRED',
    )
    return any(marker in upper for marker in local_markers)


def _looks_like_tss_rejection(message, api_calls=None):
    raw = (message or '').strip()
    upper = _clean_pipeline_error(message).upper()
    create_attempt = any((call.get('call_type') or '').upper() == 'CREATE_CONSIGNMENT' for call in (api_calls or []))
    return bool(
        create_attempt
        or raw.startswith('{')
        or 'PROCESS_MESSAGE' in raw.upper()
        or 'MANDATORY FIELD' in upper
        or upper.startswith('ERROR:')
    )


def _match_consignment_error_field(message):
    lower = (message or '').lower()
    for candidate in re.findall(r"'([a-z_]+)'", lower):
        if candidate in CONS_EDIT_FIELD_META:
            return candidate
    if 'must link to an ens header' in lower or 'staging_ens_id' in lower:
        return 'ens_parent'
    for field, meta in CONS_EDIT_FIELD_META.items():
        if any(alias in lower for alias in meta['aliases']):
            return field
    return None


def _extract_tss_pointer_names(message):
    pointers = []
    for match in re.finditer(r"PointerNames:([^;]+)", message or '', flags=re.IGNORECASE):
        pointer = match.group(1).strip()
        if pointer:
            pointers.append(pointer)
    return pointers


def _consignment_field_from_pointer(pointer):
    compact = re.sub(r'\s+', '', (pointer or '').replace('\\', '/').lower())
    role = None
    for candidate in ('consignee', 'consignor', 'importer', 'exporter', 'buyer', 'seller', 'carrier'):
        if f'/{candidate}/' in compact or compact.endswith(f'/{candidate}') or f'goodsshipment/{candidate}/' in compact:
            role = candidate
            break
    if role == 'carrier':
        return 'ens_parent'
    if not role:
        return None

    if 'identificationnumber' in compact or compact.endswith('/eori') or 'eorinumber' in compact:
        return f'{role}_eori'
    if 'nameofstreet' in compact or 'streetandnumber' in compact or 'streetnumber' in compact or 'addressline' in compact:
        return f'{role}_street_and_number' if role in {'buyer', 'seller'} else f'{role}_street_number'
    if '/name' in compact or compact.endswith('name'):
        return f'{role}_name'
    if 'postcode' in compact or 'postalcode' in compact:
        return f'{role}_postcode'
    if 'city' in compact or 'town' in compact:
        return f'{role}_city'
    if 'country' in compact:
        return f'{role}_country'
    return None


def _consignment_pointer_fields(message):
    fields = []
    seen = set()
    for pointer in _extract_tss_pointer_names(message):
        field = _consignment_field_from_pointer(pointer)
        if not field or field in seen:
            continue
        seen.add(field)
        meta = CONS_EDIT_FIELD_META.get(field, {})
        fields.append({
            'label': meta.get('label') or _consignment_field_label(field),
            'field': field,
            'pointer': pointer,
            'message': f"TSS flagged {meta.get('label') or field}.",
            'suggestion': meta.get('suggestion'),
        })
    return fields


def _yes_no_value(value, default='yes'):
    text = (value or '').strip().lower()
    if text in {'yes', 'no'}:
        return text
    return default


def _resolve_consignment_tss_fields(cons, fields):
    """Map TSS buyer/seller pointers to the visible source fields when same-as flags are active."""
    resolved = []
    buyer_same = _yes_no_value((cons or {}).get('buyer_same_as_importer'), default='yes')
    seller_same = _yes_no_value((cons or {}).get('seller_same_as_exporter'), default='yes')
    buyer_has_direct_value = bool(((cons or {}).get('buyer_eori') or '').strip())
    seller_has_direct_value = bool(((cons or {}).get('seller_eori') or '').strip())

    for item in fields or []:
        current = dict(item)
        field = current.get('field')
        if field == 'buyer_eori' and buyer_same != 'no' and not buyer_has_direct_value:
            current.update({
                'label': 'Buyer EORI/TIN via Importer EORI',
                'field': 'importer_eori',
                'source_field': 'buyer_eori',
                'message': 'TSS flagged Buyer EORI/TIN. Because Buyer same as Importer is Yes, fix the Importer EORI.',
                'suggestion': 'Correct importer_eori, or change Buyer same as Importer to No and provide separate buyer details.',
            })
        elif field == 'seller_eori' and seller_same != 'no' and not seller_has_direct_value:
            current.update({
                'label': 'Seller EORI/TIN via Exporter EORI',
                'field': 'exporter_eori',
                'source_field': 'seller_eori',
                'message': 'TSS flagged Seller EORI/TIN. Because Seller same as Exporter is Yes, fix the Exporter EORI.',
                'suggestion': 'Correct exporter_eori, or change Seller same as Exporter to No and provide separate seller details.',
            })
        elif field == 'carrier_eori':
            current.update({
                'label': 'Carrier EORI/TIN on ENS Header',
                'field': 'ens_parent',
                'source_field': 'carrier_eori',
                'message': 'TSS flagged Carrier EORI/TIN. This value lives on the linked ENS Header, not this consignment.',
                'suggestion': 'Open the ENS Header selected above, fill Carrier EORI and carrier address, save, then rerun the cargo flow.',
            })
        elif field == 'no_sfd_reason':
            current.update({
                'label': 'No SFD Reason',
                'field': 'no_sfd_reason',
                'message': 'TSS requires No SFD Reason because the importer EORI is not registered for SFD creation.',
                'suggestion': CONS_EDIT_FIELD_META['no_sfd_reason']['suggestion'],
            })
        resolved.append(current)
    return resolved


def _address_required_field_items(cons, item):
    field = (item or {}).get('field') or ''
    if not field.endswith('_eori'):
        return None
    prefix = field[:-5]
    if prefix not in ADDRESS_REQUIRED_VALIDATION_PREFIXES:
        return None
    if not _is_truthy_form_flag((cons or {}).get(f'{prefix}_address_required')):
        return None

    missing_fields = _missing_party_address_fields(cons or {}, prefix)
    if not missing_fields:
        return None

    role_label = PARTY_DETAIL_ROLES.get(prefix, prefix.title())
    items = []
    for missing_field in missing_fields:
        meta = CONS_EDIT_FIELD_META.get(missing_field, {})
        label = meta.get('label') or _consignment_field_label(missing_field)
        items.append({
            'label': label,
            'field': missing_field,
            'source_field': field,
            'pointer': (item or {}).get('pointer'),
            'message': (
                f"TSS flagged {role_label} EORI/TIN, but {role_label} Address Required / "
                f"EORI Unknown is true. Complete {label}."
            ),
            'suggestion': meta.get('suggestion') or _address_required_missing_message(prefix),
        })
    return items


def _apply_address_required_tss_field_mapping(cons, fields):
    mapped = []
    seen = set()
    for item in fields or []:
        replacements = _address_required_field_items(cons or {}, item)
        for candidate in replacements or [item]:
            key = candidate.get('field') or candidate.get('source_field') or candidate.get('label')
            if key in seen:
                continue
            seen.add(key)
            mapped.append(candidate)
    return mapped


def _consignment_tss_field_detail(fields, fallback=None):
    if not fields:
        return fallback
    if any((item.get('field') or '') == 'no_sfd_reason' for item in fields):
        return fallback
    field_names = ', '.join(
        f"{item.get('label')} ({item.get('field')})" if item.get('field') else (item.get('label') or 'Party EORI/TIN')
        for item in fields
    )
    return (
        f"Review these fields: {field_names}. Values must be real EORI/TIN numbers, not company names, "
        "blanks, placeholders, sample values, or values with spaces/separators. A value can pass local "
        "format validation and still be rejected if TSS does not recognise it in the selected environment "
        "or for the submitting account. If TSS cannot auto-populate that EORI, provide the matching party "
        "name and full address from Master Data. If buyer/seller are marked as the same party as importer/exporter, "
        "TSS validates those linked importer/exporter EORI values."
    )


def _consignment_tss_suggestions(suggestions):
    merged = list(suggestions or [])
    hint = (
        'Format checks only confirm the shape of the value. TSS also checks that the EORI/TIN is recognised '
        'for the selected environment/account. When Buyer same as Importer is Yes, fix buyer errors on '
        'Importer EORI; when Seller same as Exporter is Yes, fix seller errors on Exporter EORI.'
    )
    address_hint = (
        'If TSS shows an address-required warning, update Master Data with the exact legal name, street, '
        'city, postcode and country for that EORI before sending cargo again.'
    )
    if address_hint not in merged:
        merged.append(address_hint)
    if hint not in merged:
        merged.append(hint)
    return merged


def _build_consignment_edit_guidance(cons):
    cleaned = _clean_pipeline_error(cons.get('error_message'))
    if not cleaned:
        return None

    tss_explanation = explain_tss_error(
        cleaned,
        local_status=cons.get('status'),
        tss_status=cons.get('tss_status'),
        entity_label='this consignment',
    )
    if tss_explanation and tss_explanation.get('fields'):
        resolved_tss_fields = _resolve_consignment_tss_fields(cons, tss_explanation.get('fields') or [])
        pointer_fields = _consignment_pointer_fields(cleaned)
        if pointer_fields:
            existing_fields = {field.get('field') for field in resolved_tss_fields}
            existing_source_fields = {field.get('source_field') for field in resolved_tss_fields}
            resolved_tss_fields.extend(
                field for field in pointer_fields
                if field.get('field') not in existing_fields
                and field.get('field') not in existing_source_fields
            )
        resolved_tss_fields = _apply_address_required_tss_field_mapping(cons, resolved_tss_fields)
        field_errors = {}
        field_suggestions = {}
        issues = []
        for item in resolved_tss_fields:
            field = item.get('field')
            label = item.get('label') or 'Party EORI/TIN'
            meta = CONS_EDIT_FIELD_META.get(field, {})
            message = (
                item.get('message') or f"TSS flagged {label} as invalid."
                if field else
                f"TSS flagged {label}; confirm the matching buyer/seller/importer/exporter EORI fields."
            )
            suggestion = item.get('suggestion') or meta.get('suggestion') or (
                'Use a real EORI/TIN value from the party record; avoid company names, blanks, placeholders and values with separators.'
            )
            if field:
                field_errors.setdefault(field, []).append(message)
                field_suggestions[field] = suggestion
            issues.append({
                'field': field,
                'label': meta.get('label', label),
                'message': message,
                'suggestion': suggestion,
            })
        return {
            'title': tss_explanation.get('title') or 'TSS needs corrected consignment data',
            'detail': _consignment_tss_field_detail(resolved_tss_fields, fallback=tss_explanation.get('detail') or tss_explanation.get('summary')),
            'issues': issues,
            'field_errors': field_errors,
            'field_suggestions': field_suggestions,
        }

    pointer_fields = _consignment_pointer_fields(cleaned)
    if pointer_fields:
        pointer_fields = _apply_address_required_tss_field_mapping(cons, pointer_fields)
        field_errors = {}
        field_suggestions = {}
        issues = []
        for item in pointer_fields:
            field = item.get('field')
            message = item.get('message') or f"TSS flagged {item.get('label') or field}."
            if field:
                field_errors.setdefault(field, []).append(message)
                if item.get('suggestion'):
                    field_suggestions[field] = item['suggestion']
            issues.append({
                'field': field,
                'label': item.get('label') or _consignment_field_label(field),
                'message': message,
                'suggestion': item.get('suggestion'),
            })
        return {
            'title': 'TSS needs corrected consignment data',
            'detail': _consignment_tss_field_detail(pointer_fields, fallback='TSS flagged one or more party address fields.'),
            'issues': issues,
            'field_errors': field_errors,
            'field_suggestions': field_suggestions,
        }

    parts = [part.strip() for part in cleaned.split(' | ') if part.strip()] or [cleaned]
    field_errors = {}
    field_suggestions = {}
    issues = []

    for part in parts:
        field = _match_consignment_error_field(part)
        meta = CONS_EDIT_FIELD_META.get(field, {})
        if field:
            field_errors.setdefault(field, []).append(part)
            if meta.get('suggestion'):
                field_suggestions[field] = meta['suggestion']
        issues.append({
            'field': field,
            'label': meta.get('label', 'General issue'),
            'message': part,
            'suggestion': meta.get('suggestion'),
        })

    return {
        'title': 'This consignment needs fixes before it can move forward',
        'detail': 'The highlighted fields below are the most likely blockers. Update them, save, and the record will be revalidated automatically.',
        'issues': issues,
        'field_errors': field_errors,
        'field_suggestions': field_suggestions,
    }


def _latest_consignment_tss_issue(cons, api_calls=None):
    """Surface the most useful TSS action-required message on the detail page."""
    tss_status = normalize_status_key((cons or {}).get('tss_status'))
    if tss_status not in {'TRADER INPUT REQUIRED', 'AMENDMENT REQUIRED', 'ERROR', 'DO NOT LOAD'}:
        return None

    for call in api_calls or []:
        http_status = call.get('http_status') or 0
        fallback_message = _clean_pipeline_error(
            call.get('error_detail')
            or call.get('response_message')
            or call.get('response_json')
            or ''
        )
        if http_status >= 400 or fallback_message:
            explanation = explain_tss_error(
                call.get('error_detail'),
                call.get('response_message'),
                call.get('response_json'),
                local_status=(cons or {}).get('status'),
                tss_status=(cons or {}).get('tss_status'),
                http_status=http_status,
                entity_label='this consignment',
            )
            if explanation:
                fields = _resolve_consignment_tss_fields(cons or {}, explanation.get('fields') or [])
                pointer_fields = _consignment_pointer_fields(
                    ' '.join(str(value or '') for value in (
                        call.get('error_detail'),
                        call.get('response_message'),
                        call.get('response_json'),
                    ))
                )
                existing_fields = {field.get('field') for field in fields}
                existing_source_fields = {field.get('source_field') for field in fields}
                fields.extend(
                    field for field in pointer_fields
                    if field.get('field') not in existing_fields
                    and field.get('field') not in existing_source_fields
                )
                fields = _apply_address_required_tss_field_mapping(cons or {}, fields)
                ens_fields = [
                    field for field in fields
                    if (field.get('field') == 'ens_parent' or field.get('source_field') == 'carrier_eori')
                ]
                consignment_fields = [
                    field for field in fields
                    if field not in ens_fields
                ]
                compact = tss_status == 'TRADER INPUT REQUIRED'
                return {
                    'title': 'TSS: Trader Input Required' if compact else explanation.get('title'),
                    'summary': (
                        'TSS needs corrections before this DEC can move forward.'
                        if compact else explanation.get('summary')
                    ),
                    'message': _consignment_tss_field_detail(fields, fallback=explanation.get('detail')),
                    'next_step': explanation.get('next_step'),
                    'suggestions': [] if compact else _consignment_tss_suggestions(explanation.get('suggestions') or []),
                    'fields': fields,
                    'ens_fields': ens_fields,
                    'consignment_fields': consignment_fields,
                    'compact': compact,
                    'technical': explanation.get('technical'),
                    'call_type': call.get('call_type'),
                    'called_at': call.get('called_at'),
                }
            pointer_fields = _consignment_pointer_fields(fallback_message)
            if pointer_fields:
                pointer_fields = _apply_address_required_tss_field_mapping(cons or {}, pointer_fields)
                ens_fields = [
                    field for field in pointer_fields
                    if field.get('field') == 'ens_parent' or field.get('source_field') == 'carrier_eori'
                ]
                consignment_fields = [field for field in pointer_fields if field not in ens_fields]
                compact = tss_status == 'TRADER INPUT REQUIRED'
                return {
                    'title': 'TSS: Trader Input Required' if compact else f'TSS action required: {tss_status}',
                    'summary': 'TSS needs corrections before this DEC can move forward.',
                    'message': _consignment_tss_field_detail(pointer_fields, fallback=fallback_message),
                    'next_step': 'Open Edit, correct the highlighted values, save, then rerun the cargo flow.',
                    'suggestions': [] if compact else _consignment_tss_suggestions([]),
                    'fields': pointer_fields,
                    'ens_fields': ens_fields,
                    'consignment_fields': consignment_fields,
                    'compact': compact,
                    'technical': fallback_message,
                    'call_type': call.get('call_type'),
                    'called_at': call.get('called_at'),
                }
            return {
                'title': f'TSS action required: {tss_status}',
                'summary': 'TSS is waiting for corrected trader data before this consignment can move forward.',
                'message': fallback_message or 'Open the API Log tab to inspect the latest TSS response.',
                'next_step': 'Open Edit, correct the blocking values, save, then rerun the pipeline.',
                'suggestions': [],
                'fields': [],
                'technical': fallback_message,
                'call_type': call.get('call_type'),
                'called_at': call.get('called_at'),
            }

    return {
        'title': f'TSS action required: {tss_status}',
        'summary': 'TSS is waiting for corrected trader data before this consignment can move forward.',
        'message': (cons or {}).get('error_message') or 'No linked API error was found. Check the TSS portal or rerun Sync TSS Status to refresh the latest response.',
        'next_step': 'Open Edit, correct the blocking values, save, then rerun the pipeline.',
        'suggestions': [],
        'fields': [],
        'technical': (cons or {}).get('error_message'),
        'call_type': None,
        'called_at': None,
    }


def _consignment_detail_issue_fields(cons, tss_action_required=None):
    fields = set()
    edit_guidance = _build_consignment_edit_guidance(cons or {}) or {}
    fields.update((edit_guidance.get('field_errors') or {}).keys())
    if tss_action_required:
        for item in tss_action_required.get('fields') or []:
            if item.get('field') and item.get('field') != 'ens_parent':
                fields.add(item['field'])
        for item in tss_action_required.get('consignment_fields') or []:
            if item.get('field') and item.get('field') != 'ens_parent':
                fields.add(item['field'])
    return sorted(fields)


def _build_consignment_form_guidance(errors):
    if not errors:
        return None

    field_suggestions = {}
    issues = []
    for field, messages in errors.items():
        meta = CONS_EDIT_FIELD_META.get(field, {})
        if meta.get('suggestion'):
            field_suggestions[field] = meta['suggestion']
        if not isinstance(messages, (list, tuple)):
            messages = [messages]
        for message in messages:
            issues.append({
                'field': field if field in CONS_EDIT_FIELD_META else None,
                'label': meta.get('label', 'General issue'),
                'message': message,
                'suggestion': meta.get('suggestion'),
            })

    return {
        'title': 'Fix highlighted fields before saving this consignment',
        'detail': 'The form can stay draft-friendly, but these fields are already known blockers for validation or TSS submission.',
        'issues': issues,
        'field_errors': errors,
        'field_suggestions': field_suggestions,
    }


def _build_dec_reference_state(cons, api_calls=None):
    local_status = (cons.get('status') or '').upper()
    tss_status = (cons.get('tss_status') or '').upper()
    cleaned_error = _clean_pipeline_error(cons.get('error_message'))

    if cons.get('dec_reference'):
        if tss_status in {'TRADER INPUT REQUIRED', 'AMENDMENT REQUIRED', 'ERROR', 'DO NOT LOAD'}:
            return {
                'tone': 'warning',
                'label': 'DEC linked, correction required',
                'detail': (
                    'TSS has this DEC reference, but it has not accepted the current consignment data. '
                    'Correct the highlighted fields, then run Send Cargo Pipeline and Sync TSS Status.'
                ),
            }
        detail = 'TSS accepted the consignment create step and returned this DEC reference.'
        if tss_status:
            detail += f' Current TSS status: {tss_status}.'
            return {'tone': 'success', 'label': 'DEC linked successfully', 'detail': detail}
        detail += ' Run Sync TSS Status if you need the latest downstream TSS status.'
        return {'tone': 'info', 'label': 'DEC linked, waiting for status sync', 'detail': detail}

    if local_status == 'FAILED':
        if _is_local_validation_failure(cleaned_error):
            return {
                'tone': 'danger',
                'label': 'DEC blocked by local validation failure',
                'detail': 'No DEC was created because local validation failed before a successful TSS create. '
                          'Fix the field errors, reset to pending, then rerun Validate Pipeline and create the DEC in TSS.',
            }
        if _looks_like_tss_rejection(cleaned_error, api_calls=api_calls):
            detail = 'No DEC was created because TSS rejected the consignment create request.'
            if cleaned_error:
                detail += f' Latest TSS response: {cleaned_error}'
            detail += ' Correct the blocking data, then rerun the pipeline.'
            return {
                'tone': 'danger',
                'label': 'DEC not created because TSS rejected submission',
                'detail': detail,
            }
        return {
            'tone': 'danger',
            'label': 'DEC creation blocked by pipeline failure',
            'detail': 'This consignment failed before a DEC could be linked. Review the failure details, then retry the pipeline.',
        }

    if local_status == 'VALIDATED':
        return {
            'tone': 'info',
            'label': 'Ready for DEC creation',
            'detail': 'Local validation is complete. The DEC reference will be created when the cargo send job creates this consignment in TSS successfully.',
        }

    if local_status in ('PENDING', 'INVALID'):
        return {
            'tone': 'warning',
            'label': 'Waiting for first successful pipeline pass',
            'detail': 'No DEC exists yet because this consignment has not completed local validation and TSS submission.',
        }

    return {
        'tone': 'info',
        'label': 'Waiting for first successful TSS consignment create',
        'detail': 'A DEC reference will appear here after the first successful consignment create in TSS.',
    }


def _ens_header_route_a_state(staging_ens_id):
    if not staging_ens_id:
        return None
    counts = query_one(f"""
        SELECT COUNT(*) AS cons_count,
               SUM(CASE WHEN tss_status IN ('AUTHORISED_FOR_MOVEMENT', 'Authorised for Movement') THEN 1 ELSE 0 END) AS auth_count
        FROM {S}.StagingConsignments
        WHERE staging_ens_id = ?
    """, [staging_ens_id]) or {}
    cons_count = counts.get('cons_count') or 0
    auth_count = counts.get('auth_count') or 0
    return {
        'cons_count': cons_count,
        'auth_count': auth_count,
        'remaining_count': max(cons_count - auth_count, 0),
        'ready': cons_count > 0 and cons_count == auth_count,
    }


def _build_consignment_guidance(cons, goods, linked_gmr=None, linked_supdecs=None, linked_sfd=None, api_calls=None, header_route_a_state=None):
    local_status = (cons.get('status') or '').upper()
    tss_status = (cons.get('tss_status') or '').upper()
    goods_total = len(goods or [])
    job_markers = _pipeline_job_markers()
    updated_at = cons.get('updated_at') or cons.get('created_at')
    last_validate = job_markers.get('JOB_VALIDATE_PIPELINE')
    last_submit = job_markers.get('JOB_SUBMIT_PIPELINE')
    last_sync = job_markers.get('JOB_SYNC_PIPELINE')
    requires_sdi = _requires_sdi(_with_linked_sfd_reference(cons, linked_sfd))

    if local_status == 'FAILED':
        cleaned_error = _clean_pipeline_error(cons.get('error_message'))
        if _is_local_validation_failure(cleaned_error):
            return {
                'tone': 'danger',
                'title': 'Local validation failed before TSS submission',
                'detail': f'No DEC reference could be created because local validation failed. {cleaned_error}',
                'actions': [
                    _detail_action('Edit Consignment', href=url_for('consignments.edit', sid=cons['staging_id']), kind='warning'),
                    _detail_action('Reset to Pending', post_to=url_for('consignments.retry', sid=cons['staging_id']), kind='warning'),
                    _detail_action('Validate Pipeline', kind='warning', phase='validate_pipeline'),
                ],
            }
        if _looks_like_tss_rejection(cleaned_error, api_calls=api_calls):
            detail = 'TSS rejected the consignment create request before issuing a DEC reference.'
            if cleaned_error:
                detail += f' Response: {cleaned_error}'
            detail += ' Fix the blocking field values, reset this record to PENDING, then rerun Validate Pipeline and create the DEC in TSS.'
            return {
                'tone': 'danger',
                'title': 'TSS rejected the consignment create request',
                'detail': detail,
                'actions': [
                    _detail_action('Edit Consignment', href=url_for('consignments.edit', sid=cons['staging_id']), kind='warning'),
                    _detail_action('Reset to Pending', post_to=url_for('consignments.retry', sid=cons['staging_id']), kind='warning'),
                    _detail_action('Validate Pipeline', kind='warning', phase='validate_pipeline'),
                ],
            }
        return {
            'tone': 'danger',
            'title': 'Blocked by a local error',
            'detail': cleaned_error or 'Fix the failed fields, then retry validation or submission.',
            'actions': [
                _detail_action('Edit Consignment', href=url_for('consignments.edit', sid=cons['staging_id']), kind='warning'),
                _detail_action('Reset to Pending', post_to=url_for('consignments.retry', sid=cons['staging_id']), kind='warning'),
            ],
        }

    if goods_total == 0:
        return {
            'tone': 'warning',
            'title': 'Goods items are still missing',
            'detail': 'Add at least one goods item before this consignment can be validated and submitted to TSS.',
            'actions': [
                _detail_action('Add Goods Item', href=url_for('goods.create', cons_id=cons['staging_id']), kind='primary'),
            ],
        }

    if not cons.get('dec_reference'):
        if local_status in ('PENDING', 'PENDING REVIEW'):
            detail = 'Goods are present, but this consignment has not been through the local pipeline yet. Run Validate Pipeline, then create the DEC in TSS.'
            if last_validate and updated_at and last_validate < updated_at:
                detail = 'This consignment was created or changed after the last Validate Pipeline run, so no TSS call has been attempted yet.'
            return {
                'tone': 'warning',
                'title': 'Waiting for Validate Pipeline',
                'detail': detail,
                'actions': [
                    _detail_action('Validate Pipeline', kind='warning', phase='validate_pipeline'),
                ],
            }
        if local_status == 'VALIDATED':
            detail = 'Validation is complete. The next step is to create the DEC reference in TSS.'
            if last_submit and updated_at and last_submit < updated_at:
                detail = 'This consignment became VALIDATED after the last cargo send run, so it is still waiting to be sent to TSS.'
            return {
                'tone': 'info',
                'title': 'Ready for TSS submission',
                'detail': detail,
                'actions': [
                    _detail_action('Create DEC in TSS', kind='primary', phase='submit_pipeline'),
                ],
            }

    if cons.get('dec_reference') and not tss_status:
        detail = 'A DEC reference exists, but no TSS status has been synced back yet. Run Sync TSS Status.'
        if last_sync and updated_at and last_sync < updated_at:
            detail = 'This consignment changed after the last sync run, so the current TSS status has not been refreshed yet.'
        return {
            'tone': 'info',
            'title': 'Waiting for TSS status sync',
            'detail': detail,
            'actions': [
                _detail_action('Sync TSS Status', kind='primary', phase='sync_pipeline'),
            ],
        }

    if cons.get('dec_reference') and tss_status in {'TRADER INPUT REQUIRED', 'AMENDMENT REQUIRED', 'ERROR', 'DO NOT LOAD'}:
        return {
            'tone': 'warning',
            'title': 'TSS needs corrected consignment data',
            'detail': (
                f'TSS is holding {cons.get("dec_reference")} at {cons.get("tss_status")}. '
                'Edit the highlighted consignment fields, save, run Send Cargo Pipeline to update the existing DEC, '
                'then run Sync TSS Status to confirm the new TSS response.'
            ),
            'actions': [
                _detail_action('Edit Consignment', href=url_for('consignments.edit', sid=cons['staging_id']), kind='warning'),
                _detail_action('Send Cargo Pipeline to TSS', kind='primary', phase='submit_pipeline'),
                _detail_action('Sync TSS Status', kind='secondary', phase='sync_pipeline'),
            ],
        }

    if cons.get('dec_reference') and local_status == 'CREATED' and tss_status in {'DRAFT', 'CREATED', 'UPDATED'}:
        error_explanation = _build_consignment_error_explanation(cons)
        if error_explanation:
            return {
                'tone': 'warning',
                'title': 'DEC exists, but TSS rejected the submit action',
                'detail': (
                    f'{cons.get("dec_reference")} is still local CREATED and TSS reports {cons.get("tss_status") or "Pending sync"}. '
                    'The saved error is from an attempted submit operation, not proof that the DEC became Submitted.'
                ),
                'actions': [
                    _detail_action('Sync TSS Status', kind='primary', phase='sync_pipeline'),
                ],
            }

        ready_statuses = {'CREATED', 'IMPORTED', 'SYNCED', 'SUBMITTED'}
        blocked_goods = [
            g for g in (goods or [])
            if (g.get('status') or '').upper() not in ready_statuses
        ]
        if blocked_goods:
            return {
                'tone': 'warning',
                'title': 'Created in TSS, awaiting goods',
                'detail': f'DEC {cons.get("dec_reference")} exists in TSS, but {len(blocked_goods)} goods item(s) still need to be created before the consignment can be submitted.',
                'actions': [
                    _detail_action('Send Cargo Pipeline to TSS', kind='primary', phase='submit_pipeline'),
                ],
            }
        return {
            'tone': 'info',
            'title': 'Created in TSS, ready to submit',
            'detail': f'DEC {cons.get("dec_reference")} and all goods items exist in TSS. Submit the consignment to move it out of Draft, then sync jobs will read the live TSS/SFD progression.',
            'actions': [
                _detail_action('Send Cargo Pipeline to TSS', kind='primary', phase='submit_pipeline'),
                _detail_action('Sync TSS Status', kind='secondary', phase='sync_pipeline'),
            ],
        }

    if tss_status == 'ARRIVED':
        if not requires_sdi:
            return {
                'tone': 'success',
                'title': 'ARRIVED - no supplementary declaration required',
                'detail': 'This consignment is configured not to generate SFD/SDI, so no supplementary declaration action is needed here.',
                'actions': [],
            }
        if not linked_supdecs:
            if not linked_sfd:
                return {
                    'tone': 'warning',
                    'title': 'ARRIVED - waiting for SFD sync before SDI',
                    'detail': 'Goods have arrived, but no synced SFD DEC is linked yet. Sync TSS tables first so the SDI is created from the real SFD, not directly from the consignment.',
                    'actions': [
                        _detail_action('Sync TSS Tables', kind='warning', phase='sync_tss'),
                    ],
                }
            return {
                'tone': 'warning',
                'title': 'ARRIVED - SDI is next',
                'detail': 'Goods have arrived. The next business step is to create or link the supplementary declaration.',
                'actions': [
                    _detail_action(
                        'Start SDI',
                        href=url_for('supdec.create', cons_id=cons['staging_id']),
                        kind='primary',
                    ),
                ],
            }
        return {
            'tone': 'success',
            'title': 'ARRIVED - downstream declaration linked',
            'detail': 'This consignment has reached ARRIVED status and already has its SDI chain attached.',
            'actions': [],
        }

    if tss_status == 'AUTHORISED_FOR_MOVEMENT':
        if not linked_gmr:
            if header_route_a_state and not header_route_a_state['ready']:
                remaining = header_route_a_state['remaining_count']
                return {
                    'tone': 'warning',
                    'title': 'Waiting for the rest of the ENS Header before GMR',
                    'detail': f'This consignment is AUTHORISED_FOR_MOVEMENT, but Route A requires every consignment on the ENS Header to be ready before GMR creation. {remaining} consignments are still not authorised.',
                    'actions': [
                        _detail_action(
                            'Open ENS Header',
                            href=url_for('declarations.detail_by_ref', ens_ref=cons.get('ens_reference')) if cons.get('ens_reference') else url_for('consignments.list_view'),
                            kind='warning',
                        ),
                        _detail_action('Sync TSS Status', kind='primary', phase='sync_pipeline'),
                    ],
                }
            return {
                'tone': 'primary',
                'title': 'Authorised for Movement',
                'detail': 'The next business step is to create the GMR on the linked ENS header.',
                'actions': [
                    _detail_action(
                        'Create GMR',
                        href=url_for('gmr.create') + f"?ens_id={cons['staging_ens_id']}",
                        kind='primary',
                    ),
                ],
            }
        if linked_gmr and not linked_supdecs:
            if not requires_sdi:
                return {
                    'tone': 'success',
                    'title': 'GMR exists - no SDI required',
                    'detail': 'This consignment is configured not to generate SFD/SDI, so the downstream supplementary declaration step does not apply.',
                    'actions': [],
                }
            return {
                'tone': 'info',
                'title': 'GMR exists - waiting for arrival before SDI',
                'detail': 'The movement is authorised and the GMR is already in place. The supplementary declaration stage starts once the goods reach ARRIVED in NI.',
                'actions': [
                    _detail_action('Sync TSS Status', kind='primary', phase='sync_pipeline'),
                ],
            }
        return {
            'tone': 'success',
            'title': 'Downstream flow linked',
            'detail': 'This consignment already has its follow-on movement records attached.',
            'actions': [],
        }

    if cons.get('dec_reference'):
        current_tss = cons.get('tss_status') or cons.get('status')
        detail = 'Run Sync TSS Status if this looks stale or if you expect a newer TSS status.'
        if current_tss:
            if requires_sdi:
                detail = (
                    f'TSS has accepted the consignment, but it is still at {current_tss}. '
                    'GMR and SDI stay blocked until this consignment reaches AUTHORISED_FOR_MOVEMENT. '
                    'Run Sync TSS Status if you expect TSS to have moved it on.'
                )
            else:
                detail = (
                    f'TSS has accepted the consignment, but it is still at {current_tss}. '
                    'Downstream movement stays blocked until this consignment reaches AUTHORISED_FOR_MOVEMENT. '
                    'Run Sync TSS Status if you expect TSS to have moved it on.'
                )
        return {
            'tone': 'info',
            'title': f'Waiting for TSS to progress beyond {current_tss}',
            'detail': detail,
            'actions': [
                _detail_action('Sync TSS Status', kind='primary', phase='sync_pipeline'),
            ],
        }

    if not api_calls:
        return {
            'tone': 'warning',
            'title': 'No consignment API calls have been recorded yet',
            'detail': 'That is expected until this consignment reaches TSS submission or TSS status sync.',
            'actions': [],
        }

    return {
        'tone': 'info',
        'title': 'Review the current workflow state',
        'detail': 'Use the linked ENS, goods, GMR and SDI records to continue the Route A journey.',
        'actions': [],
    }


def _consignment_detail_target(cons):
    ref = (cons or {}).get('dec_reference') or (cons or {}).get('sfd_reference')
    if ref:
        return url_for('consignments.detail_by_ref', cons_ref=ref)
    return url_for('consignments.detail', sid=(cons or {}).get('staging_id', 0))


def _flash_auto_validation_result(staging_id, result):
    if not result:
        return
    if result['ok']:
        flash(f'Consignment #{staging_id} saved and auto-validated.', 'success')
        return
    first_error = (result.get('errors') or ['Validation failed.'])[0]
    flash(f'Consignment #{staging_id} saved, but local validation failed: {first_error}', 'warning')

def get_cv(table, val_col='value', name_col='name'):
    """Load choice values from a TSS.CV_* table."""
    try:
        rows = query_all(f"SELECT [{val_col}], [{name_col}] FROM TSS.[{table}] ORDER BY [{name_col}]")
        return [CVOption(r[val_col], r[name_col]) for r in rows]
    except:
        return []

def get_ens_parents():
    """Staging ENS headers (any active status) + validated/draft from declarations + synced TSS mirror."""
    results = []
    seen_refs = set()
    try:
        rows = query_all(f"""
            SELECT staging_id, ens_reference, label, status, tss_status
            FROM {S}.StagingEnsHeaders
            WHERE status NOT IN ('CANCELLED')
            ORDER BY staging_id DESC""")
        for r in rows:
            if not tss_allows_data_changes(r.get('tss_status'), r.get('status')):
                continue
            tss = r.get('tss_status') or r.get('status') or ''
            lbl = r.get('label') or ''
            ref = r.get('ens_reference') or ''
            if ref:
                seen_refs.add(ref)
            results.append({'id': str(r['staging_id']),
                'display': f"{ref or '(pending ref)'} — {lbl} [{tss}]"})
    except: pass
    try:
        # ENS declarations: local validated or submitted records not yet mirrored into staging.
        rows = query_all(f"""
            SELECT id, external_ref, external_status, status,
                   movement_type, arrival_port, carrier_name
            FROM {S}.StagingDeclarations
            WHERE status IN ('Validated', 'Submitted')
            ORDER BY id DESC""")
        for r in rows:
            if not tss_allows_data_changes(r.get('external_status'), r.get('status')):
                continue
            ref = r.get('external_ref') or ''
            if ref and ref in seen_refs:
                continue  # already covered by StagingEnsHeaders stub
            seen_refs.add(ref)
            carrier = r.get('carrier_name') or r.get('arrival_port') or ''
            state   = r.get('external_status') or r.get('status') or ''
            lbl     = f"{ref or 'Not submitted'} — {carrier} [{state}]"
            results.append({'id': f"dec:{r['id']}", 'display': lbl})
    except: pass
    try:
        rows = query_all(f"""
            SELECT id, declaration_number, tss_status
            FROM {S}.EnsHeaders ORDER BY id DESC""")
        for r in rows:
            if not tss_allows_data_changes(r.get('tss_status')):
                continue
            ref = r.get('declaration_number') or ''
            if ref and ref in seen_refs:
                continue
            results.append({'id': f"synced:{r['id']}",
                'display': f"[Synced] {ref} ({r.get('tss_status','')})"})
    except: pass
    return results


def _ensure_staging_ens_option(options, staging_id):
    """Make sure a resolved staging ENS appears in the dropdown."""
    if not staging_id:
        return options
    sid = str(staging_id)
    if any(o['id'] == sid for o in options):
        return options
    row = query_one(
        f"SELECT staging_id, ens_reference, label, status, tss_status FROM {S}.StagingEnsHeaders WHERE staging_id = ?",
        [staging_id]
    )
    if row:
        ref = row.get('ens_reference') or '(pending ref)'
        lbl = row.get('label') or ''
        state = row.get('tss_status') or row.get('status') or ''
        options.insert(0, {'id': sid, 'display': f"{ref} - {lbl} [{state}]"})
    return options


def _resolve_synced_ens(synced_id):
    """Find or create a StagingEnsHeaders stub from a BKD.EnsHeaders record.
    Returns staging_id on success, None on failure."""
    try:
        synced_id = int(synced_id)
        synced = query_one(
            f"SELECT declaration_number, movement_type, arrival_port FROM {S}.EnsHeaders WHERE id = ?",
            [synced_id])
        if not synced:
            return None
        dec_num = synced['declaration_number']
        # Return existing staging record if already linked
        existing = query_one(
            f"SELECT staging_id FROM {S}.StagingEnsHeaders WHERE ens_reference = ?",
            [dec_num])
        if existing:
            return existing['staging_id']
        # Auto-create a minimal stub so the FK can be set
        with db_cursor() as cur:
            cur.execute(f"""
                INSERT INTO {S}.StagingEnsHeaders
                    (label, ens_reference, status, movement_type, arrival_port, created_at)
                VALUES (?, ?, 'CREATED', ?, ?, SYSUTCDATETIME())
            """, [f"Synced ENS {dec_num}", dec_num,
                  synced.get('movement_type'), synced.get('arrival_port')])
            cur.execute("SELECT SCOPE_IDENTITY() AS id")
            row = cur.fetchone()
            return int(row[0]) if row and row[0] else None
    except Exception:
        return None


def _resolve_dec_ens(dec_id):
    """Find or create a StagingEnsHeaders stub from a BKD.StagingDeclarations record.
    Returns staging_id on success, None on failure."""
    try:
        dec_id = int(dec_id)
        dec = query_one(
            f"SELECT id, external_ref, external_status, status,"
            f" movement_type, arrival_port, carrier_name, carrier_eori"
            f" FROM {S}.StagingDeclarations WHERE id = ?",
            [dec_id])
        if not dec:
            return None
        existing = query_one(
            f"SELECT staging_id FROM {S}.StagingEnsHeaders WHERE staging_declaration_id = ?",
            [dec_id]
        )
        if existing:
            return existing['staging_id']
        ext_ref = dec.get('external_ref') or ''
        # If already has a TSS ref, check for existing stub first
        if ext_ref:
            existing = query_one(
                f"SELECT staging_id FROM {S}.StagingEnsHeaders WHERE ens_reference = ?",
                [ext_ref])
            if existing:
                return existing['staging_id']
        carrier = dec.get('carrier_name') or dec.get('carrier_eori') or ''
        lbl = (f"{ext_ref} — {carrier}" if carrier else ext_ref) if ext_ref else f"Dec #{dec_id}"
        tss_st = dec.get('external_status') or ('Draft' if dec.get('status') == 'Submitted' else None)
        with db_cursor() as cur:
            cur.execute(f"""
                INSERT INTO {S}.StagingEnsHeaders
                    (label, ens_reference, status, tss_status, source, created_at, staging_declaration_id)
                VALUES (?, ?, 'SUBMITTED', ?, 'declarations_portal', SYSUTCDATETIME(), ?)
            """, [lbl, ext_ref or None, tss_st, dec_id])
            cur.execute("SELECT SCOPE_IDENTITY() AS id")
            row = cur.fetchone()
            return int(row[0]) if row and row[0] else None
    except Exception:
        return None


def _resolve_ref_ens(raw_ref):
    """Find or create a StagingEnsHeaders stub from a raw ENS reference string (typed/pasted).
    Returns staging_id on success, None on failure."""
    try:
        raw_ref = (raw_ref or '').strip().upper()
        if not raw_ref:
            return None
        existing = query_one(
            f"SELECT staging_id FROM {S}.StagingEnsHeaders WHERE ens_reference = ?", [raw_ref])
        if existing:
            return existing['staging_id']
        # Fall back to synced mirror
        synced = query_one(
            f"SELECT id FROM {S}.EnsHeaders WHERE declaration_number = ?", [raw_ref])
        if synced:
            return _resolve_synced_ens(str(synced['id']))
        # Create minimal stub for this ref
        with db_cursor() as cur:
            cur.execute(f"""
                INSERT INTO {S}.StagingEnsHeaders
                    (label, ens_reference, status, source, created_at)
                VALUES (?, ?, 'PENDING', 'manual_entry', SYSUTCDATETIME())
            """, [raw_ref, raw_ref])
            cur.execute("SELECT SCOPE_IDENTITY() AS id")
            row = cur.fetchone()
            return int(row[0]) if row and row[0] else None
    except Exception:
        return None


def load_cons_choices():
    return {
        'countries': get_cv('CV_country'),
        'goods_domestic_status': get_cv('CV_goods_domestic_status'),
        'no_sfd_reason': get_cv('CV_no_sfd_reason'),
        'declaration_choice': get_cv('CV_sfd_declaration_choice'),
        'prev_doc_types': get_cv('CV_previous_document_type'),
        'auth_type_codes': get_cv('CV_auth_type_code'),
    }


def _party_master_option(
    source,
    source_id,
    partner_type,
    name,
    eori,
    street,
    city,
    postcode,
    country,
    account_ref='',
):
    name = (name or '').strip()
    eori = (eori or '').strip().upper()
    account_ref = (account_ref or '').strip()
    postcode = (postcode or '').strip().upper()
    label_bits = [bit for bit in (account_ref, name, eori, postcode) if bit]
    return {
        'option_id': f'{source}:{source_id}',
        'source': source,
        'partner_type': partner_type or '',
        'label': ' - '.join(label_bits) or f'{source} #{source_id}',
        'name': name,
        'eori': eori,
        'street': (street or '').strip(),
        'city': (city or '').strip(),
        'postcode': postcode,
        'country': (country or '').strip().upper()[:2],
        'account_ref': account_ref,
    }


def _load_party_masterdata():
    """Return saved party records grouped by consignment form role."""
    role_types = {
        'consignor': {'Consignor', 'Exporter', 'Carrier', 'Declarant', 'Agent'},
        'consignee': {'Consignee', 'Importer', 'Customer'},
        'importer': {'Importer', 'Consignee', 'Customer'},
        'exporter': {'Exporter', 'Consignor', 'Carrier', 'Declarant', 'Agent'},
        'buyer': {'Importer', 'Consignee', 'Customer'},
        'seller': {'Exporter', 'Consignor', 'Vendor'},
    }
    grouped = {role: [] for role in role_types}

    try:
        company = query_one(f"""
            SELECT TOP 1 id, company_name, trading_name, eori_xi, eori_gb,
                   address_line1, city, postcode, country
            FROM {S}.CompanyMaster
            ORDER BY
                CASE
                    WHEN eori_xi IS NOT NULL AND LTRIM(RTRIM(eori_xi)) <> '' THEN 0
                    WHEN eori_gb IS NOT NULL AND LTRIM(RTRIM(eori_gb)) <> '' THEN 1
                    ELSE 2
                END,
                id
        """)
    except Exception:
        company = None

    if company:
        company_option = _party_master_option(
            'company',
            company.get('id') or 1,
            'CompanyMaster',
            company.get('trading_name') or company.get('company_name'),
            _safe_master_eori(company),
            company.get('address_line1'),
            company.get('city'),
            company.get('postcode'),
            company.get('country'),
            'BKD',
        )
        for role in ('consignor', 'importer', 'exporter', 'seller'):
            grouped[role].append(company_option)

    try:
        partners = query_all(f"""
            SELECT TOP 750 id, partner_type, partner_name, eori, eori_gb,
                   address_line1, city, postcode, country, account_ref
            FROM {S}.Partners
            WHERE active = 1
            ORDER BY
                CASE WHEN account_ref IS NOT NULL AND LTRIM(RTRIM(account_ref)) <> '' THEN 0 ELSE 1 END,
                partner_type,
                partner_name
        """)
    except Exception:
        partners = []

    for partner in partners or []:
        option = _party_master_option(
            'partner',
            partner.get('id'),
            partner.get('partner_type'),
            partner.get('partner_name'),
            partner.get('eori') or partner.get('eori_gb'),
            partner.get('address_line1'),
            partner.get('city'),
            partner.get('postcode'),
            partner.get('country'),
            partner.get('account_ref'),
        )
        if not option['name']:
            continue
        for role, allowed_types in role_types.items():
            if option['partner_type'] in allowed_types:
                grouped[role].append(option)

    for role, options in grouped.items():
        seen = set()
        deduped = []
        for option in options:
            key = (
                option['source'],
                option['option_id'],
                option['eori'],
                option['name'].upper(),
                option['postcode'],
            )
            if key in seen:
                continue
            seen.add(key)
            deduped.append(option)
        grouped[role] = deduped[:250]
    return grouped


# ── LIST ──────────────────────────────────────────────

@consignments_bp.route('/')
def list_view():
    status_filter = canonical_filter_status(request.args.get('status') or 'ALL') or 'ALL'
    search = (request.args.get('q') or '').strip()
    ens_ref = (request.args.get('ens_ref') or '').strip()
    arrival_from_date, arrival_to_date, arrival_from, arrival_to, arrival_month = _arrival_filter_bounds(request.args)
    sort = request.args.get('sort') or 'arrival_desc'
    if sort not in {'arrival_asc', 'arrival_desc'}:
        sort = 'arrival_desc'
    page = _safe_positive_int(request.args.get('page'), 1)
    page_size, show_all, page_size_param = _consignment_page_size(
        request.args.get('page_size'),
        show_all=str(request.args.get('show_all') or '').strip().lower() in {'1', 'true', 'yes', 'all'},
    )
    client_code = (get_tenant().get('code') or 'BKD').upper()
    consignments = [dict(row) for row in query_all("""
        SELECT
            c.stg_consignment_id,
            c.stg_header_id,
            COALESCE(c.trader_reference, c.transport_document_number, c.tss_consignment_ref) AS document_no,
            c.trader_reference,
            c.transport_document_number,
            c.importer_eori,
            c.consignor_eori,
            c.consignee_eori,
            c.ducr,
            c.generate_SD,
            c.no_sfd_reason,
            c.sub_status,
            c.tss_consignment_ref,
            c.goods_description,
            c.source,
            c.updated_at,
            c.stg_created_at AS created_at,
            (SELECT COUNT(*)
             FROM STG.BKD_GoodsItems g
             WHERE g.ClientCode = c.ClientCode
               AND g.stg_consignment_id = c.stg_consignment_id
               AND UPPER(COALESCE(g.goods_stage, 'ENS')) <> 'SDI'
               AND UPPER(COALESCE(g.sub_status, '')) NOT IN ('CANCELLED', 'DELETED')) AS goods_count,
            (SELECT STRING_AGG(CAST(CONCAT(
                        ' ', g2.stg_item_id,
                        ' ', COALESCE(g2.sku, ''),
                        ' ', COALESCE(g2.tss_hex_id, ''),
                        ' ', COALESCE(g2.goods_description, ''),
                        ' ', COALESCE(g2.commodity_code, ''),
                        ' ', COALESCE(CONVERT(NVARCHAR(30), g2.item_seq), ''),
                        ' ', COALESCE(g2.sub_status, ''),
                        ' ', COALESCE(g2.error_message, '')
                    ) AS NVARCHAR(MAX)), N' ')
             FROM STG.BKD_GoodsItems g2
             WHERE g2.ClientCode = c.ClientCode
               AND g2.stg_consignment_id = c.stg_consignment_id
               AND UPPER(COALESCE(g2.goods_stage, 'ENS')) <> 'SDI'
               AND UPPER(COALESCE(g2.sub_status, '')) NOT IN ('CANCELLED', 'DELETED')) AS goods_search_text,
            (SELECT STRING_AGG(CAST(CONCAT(
                        ' ', COALESCE(sd.tss_sup_dec_number, ''),
                        ' ', COALESCE(sd.tss_sfd_consignment_ref, ''),
                        ' ', COALESCE(sd.tss_consignment_ref, ''),
                        ' ', COALESCE(sd.trader_reference, ''),
                        ' ', COALESCE(sd.transport_document_number, ''),
                        ' ', COALESCE(sd.sub_status, ''),
                        ' ', COALESCE(sd.tss_status, ''),
                        ' ', COALESCE(sd.validation_errors_json, ''),
                        ' ', COALESCE(sd.auto_submit_error, '')
                    ) AS NVARCHAR(MAX)), N' ')
             FROM STG.BKD_SDI_Headers sd
             WHERE sd.ClientCode = c.ClientCode
               AND (
                    sd.stg_consignment_id = c.stg_consignment_id
                 OR sd.tss_consignment_ref = c.tss_consignment_ref
                 OR sd.tss_sfd_consignment_ref = c.tss_consignment_ref
               )) AS sdi_search_text,
            h.conveyance_ref,
            h.arrival_date_time,
            h.tss_ens_header_ref,
            tc.TssStatus,
            COALESCE(sfd_track.tss_sfd_number, tss_sfd.SfdReference) AS sfd_reference,
            COALESCE(
                NULLIF(LTRIM(RTRIM(sfd_track.tss_movement_reference_number)), ''),
                NULLIF(LTRIM(RTRIM(tss_sfd.MovementReferenceNumber)), ''),
                CASE WHEN ISJSON(tc.RawJson) = 1 THEN NULLIF(LTRIM(RTRIM(JSON_VALUE(tc.RawJson, '$.movement_reference_number'))), '') END,
                CASE WHEN ISJSON(tc.RawJson) = 1 THEN NULLIF(LTRIM(RTRIM(JSON_VALUE(tc.RawJson, '$.movementReferenceNumber'))), '') END,
                CASE WHEN ISJSON(tc.RawJson) = 1 THEN NULLIF(LTRIM(RTRIM(JSON_VALUE(tc.RawJson, '$.mrn'))), '') END
            ) AS sfd_mrn,
            sfd_track.tss_eori_for_eidr AS sfd_eidr,
            COALESCE(
                CASE WHEN UPPER(LTRIM(RTRIM(COALESCE(sfd_track.tss_sfd_status, '')))) IN ('OK', 'SUCCESS') THEN NULL ELSE sfd_track.tss_sfd_status END,
                CASE WHEN UPPER(LTRIM(RTRIM(COALESCE(tss_sfd.TssStatus, '')))) IN ('OK', 'SUCCESS') THEN NULL ELSE tss_sfd.TssStatus END,
                tc.TssStatus
            ) AS sfd_status
        FROM STG.BKD_ENS_Consignments c
        LEFT JOIN STG.BKD_ENS_Headers h
            ON h.ClientCode = c.ClientCode
           AND h.stg_header_id = c.stg_header_id
        LEFT JOIN TSS.BKD_ENS_Consignments tc
            ON tc.ClientCode = c.ClientCode
           AND tc.ConsignmentReference = c.tss_consignment_ref
        OUTER APPLY (
            SELECT TOP 1
                t.tss_sfd_number,
                t.tss_movement_reference_number,
                t.tss_eori_for_eidr,
                t.tss_sfd_status
            FROM STG.BKD_SFD_Tracking t
            WHERE t.ClientCode = c.ClientCode
              AND t.tss_consignment_ref = c.tss_consignment_ref
            ORDER BY t.stg_polled_at DESC, t.stg_tracking_id DESC
        ) sfd_track
        OUTER APPLY (
            SELECT TOP 1
                s.SfdReference,
                s.MovementReferenceNumber,
                s.TssStatus
            FROM TSS.BKD_SFD s
            WHERE s.ClientCode = c.ClientCode
              AND (
                    s.DeclarationNumber = c.tss_consignment_ref
                 OR (
                        NULLIF(LTRIM(RTRIM(COALESCE(s.DeclarationNumber, ''))), '') IS NULL
                    AND s.EnsReference = c.tss_ens_header_ref
                    )
              )
            ORDER BY
                CASE WHEN s.DeclarationNumber = c.tss_consignment_ref THEN 0 ELSE 1 END,
                COALESCE(s.UpdatedAt, s.LastSyncedAt, s.CreatedAt) DESC
        ) tss_sfd
        WHERE c.ClientCode = ?
        ORDER BY
            CASE WHEN h.arrival_date_time IS NULL THEN 1 ELSE 0 END,
            h.arrival_date_time DESC,
            c.stg_consignment_id DESC
    """, [client_code]) or []]

    filtered = []
    counts = {}
    for raw in consignments or []:
        row = dict(raw)
        filter_status = canonical_filter_status(row.get('TssStatus') or row.get('sub_status')) or 'PENDING'
        row['filter_status'] = filter_status
        row['sdi_link_search_text'] = row.get('sdi_search_text') or ''
        counts[filter_status] = counts.get(filter_status, 0) + 1
        if status_filter != 'ALL' and filter_status != status_filter:
            continue
        if ens_ref and ens_ref.casefold() not in str(row.get('tss_ens_header_ref') or row.get('conveyance_ref') or '').casefold():
            continue
        if not _consignment_arrival_date_in_range(row, arrival_from_date, arrival_to_date):
            continue
        if search:
            if not search_matches_values(search, [
                row.get('stg_consignment_id'),
                row.get('document_no'),
                row.get('trader_reference'),
                row.get('transport_document_number'),
                row.get('goods_description'),
                row.get('tss_consignment_ref'),
                row.get('importer_eori'),
                row.get('consignor_eori'),
                row.get('consignee_eori'),
                row.get('ducr'),
                row.get('TssStatus'),
                row.get('sub_status'),
                row.get('tss_ens_header_ref'),
                row.get('conveyance_ref'),
                row.get('sfd_reference'),
                row.get('sfd_mrn'),
                row.get('sfd_eidr'),
                row.get('goods_count'),
                row.get('goods_search_text'),
                row.get('sdi_search_text'),
                row.get('sdi_link_search_text'),
                row.get('source'),
            ]):
                continue
        filtered.append(row)

    _apply_consignment_sort(filtered, sort)
    filtered_total = len(filtered)
    date_filter_active = bool(arrival_month or arrival_from_date or arrival_to_date)
    if date_filter_active:
        show_all = True
        page_size = None
        page_size_param = 'all'
    if show_all:
        total_pages = 1
        page = 1
        paged_consignments = filtered
    else:
        total_pages = max(1, (filtered_total + page_size - 1) // page_size) if filtered_total else 1
        page = min(page, total_pages)
        page_start = (page - 1) * page_size
        paged_consignments = filtered[page_start:page_start + page_size]

    prd_sdi_links = load_prd_sdi_links_for_context(
        client_code=client_code,
        consignment_ids=[row.get('stg_consignment_id') for row in paged_consignments],
        consignment_refs=[
            value
            for row in paged_consignments
            for value in (
                row.get('tss_consignment_ref'),
                row.get('document_no'),
                row.get('trader_reference'),
                row.get('transport_document_number'),
                row.get('tss_ens_header_ref'),
            )
        ],
        sfd_refs=[row.get('sfd_reference') for row in paged_consignments],
        limit=max(250, min(max(len(paged_consignments), 1) * 5, 5000)),
    )
    attach_sdi_links_to_consignments(paged_consignments, prd_sdi_links)
    for row in paged_consignments:
        row['sdi_required'] = _requires_sdi({
            **row,
            'tss_status': row.get('TssStatus'),
            'synced_sfd_reference': row.get('sfd_reference'),
        })
        row['sdi_references'] = ' '.join(
            str(link.get('sup_dec_number') or link.get('staging_id') or '').strip()
            for link in row.get('linked_supdecs') or []
            if (link.get('sup_dec_number') or link.get('staging_id'))
        )

    status_tabs = ['ALL', 'PENDING', 'VALIDATED', 'FAILED', 'SUBMITTED',
                   'PROCESSING', 'AUTHORISED FOR MOVEMENT', 'ARRIVED',
                   'TRADER INPUT REQUIRED', 'CANCELLED']
    return render_template(
        'consignments/list.html',
        consignments=paged_consignments,
        status_filter=status_filter,
        search=search,
        ens_ref=ens_ref,
        created_on=arrival_from if arrival_from and arrival_from == arrival_to else '',
        arrival_from=arrival_from,
        arrival_to=arrival_to,
        arrival_month=arrival_month,
        sort=sort,
        status_tabs=status_tabs,
        status_counts=counts,
        total=len(consignments or []),
        filtered_total=filtered_total,
        page=page,
        page_size=page_size,
        page_size_param=page_size_param,
        page_size_options=(10, 20, 50),
        total_pages=total_pages,
        show_all=show_all,
        badge_class=badge_class,
    )


@consignments_bp.route('/export.csv')
def export_csv():
    flash('This legacy portal view is not available in Automation PRD. Use Ingestion to monitor email automation and use STG/TSS-backed pages only.', 'info')
    return redirect(url_for('ingest.queue'))
# ── CREATE ────────────────────────────────────────────

@consignments_bp.route('/create', methods=['GET', 'POST'])
def create():
    flash('This legacy portal view is not available in Automation PRD. Use Ingestion to monitor email automation and use STG/TSS-backed pages only.', 'info')
    return redirect(url_for('ingest.queue'))
# ── SPEC / PRINT VIEW ────────────────────────────────

@consignments_bp.route('/<int:sid>/spec')
def spec(sid):
    flash('This legacy portal view is not available in Automation PRD. Use Ingestion to monitor email automation and use STG/TSS-backed pages only.', 'info')
    return redirect(url_for('ingest.queue'))
# ── EDIT ──────────────────────────────────────────────

@consignments_bp.route('/<int:sid>/edit', methods=['GET', 'POST'])
def edit(sid):
    cons = query_one("""
        SELECT
            c.stg_consignment_id, c.stg_header_id, c.ClientCode,
            c.sub_status, c.tss_consignment_ref,
            c.goods_description, c.trader_reference, c.transport_document_number,
            c.controlled_goods, c.goods_domestic_status, c.destination_country,
            c.ducr, c.container_indicator, c.generate_SD, c.no_sfd_reason,
            c.align_ukims, c.use_importer_sde, c.declaration_choice,
            c.consignor_eori, c.consignor_name,
            c.consignor_street_number, c.consignor_city, c.consignor_postcode, c.consignor_country,
            c.consignee_eori, c.consignee_name,
            c.consignee_street_number, c.consignee_city, c.consignee_postcode, c.consignee_country,
            c.importer_eori, c.importer_name,
            c.importer_street_number, c.importer_city, c.importer_postcode, c.importer_country,
            c.exporter_eori,
            c.buyer_same_as_importer, c.seller_same_as_exporter,
            c.updated_at,
            tc.TssStatus
        FROM STG.BKD_ENS_Consignments c
        LEFT JOIN TSS.BKD_ENS_Consignments tc
            ON tc.ClientCode = c.ClientCode
           AND tc.ConsignmentReference = c.tss_consignment_ref
        WHERE c.stg_consignment_id = ?
    """, [sid])
    if not cons:
        flash('Consignment not found.', 'warning')
        return redirect(url_for('consignments.list_view'))

    fallback_url = url_for('consignments.detail', sid=sid)
    next_url = _safe_local_next_url(request.form.get('next_url') or request.referrer, fallback_url)
    if not _prd_consignment_allows_edit(cons):
        flash('This consignment can only be edited while it is local repair, TSS Draft, or Trader Input Required.', 'warning')
        return redirect(next_url)
    if request.method == 'GET':
        flash('Use the edit modal from the ENS or consignment detail page for PRD cargo repairs.', 'info')
        return redirect(fallback_url)

    warnings = {}
    cleaned = _clean_consignment_form(request.form.to_dict(), field_warnings=warnings)
    errors = dict(warnings)

    def require(field, label):
        if not str(cleaned.get(field) or '').strip():
            errors.setdefault(field, []).append(f'{label} is required.')

    require('goods_description', 'Goods Description')
    require('transport_document_number', 'Transport Document Number')
    require('importer_eori', 'Importer EORI')

    if errors:
        first_field = next(iter(errors))
        flash(f"{_consignment_field_label(first_field)}: {errors[first_field][0]}", 'warning')
        return redirect(next_url)

    execute("""
        UPDATE STG.BKD_ENS_Consignments
           SET goods_description = ?,
               trader_reference = ?,
               transport_document_number = ?,
               controlled_goods = ?,
               goods_domestic_status = ?,
               destination_country = ?,
               ducr = ?,
               container_indicator = ?,
               generate_SD = ?,
               no_sfd_reason = ?,
               align_ukims = ?,
               use_importer_sde = ?,
               declaration_choice = ?,
               consignor_eori = ?,
               consignor_name = ?,
               consignor_street_number = ?,
               consignor_city = ?,
               consignor_postcode = ?,
               consignor_country = ?,
               consignee_eori = ?,
               consignee_name = ?,
               consignee_street_number = ?,
               consignee_city = ?,
               consignee_postcode = ?,
               consignee_country = ?,
               importer_eori = ?,
               importer_name = ?,
               importer_street_number = ?,
               importer_city = ?,
               importer_postcode = ?,
               importer_country = ?,
               exporter_eori = ?,
               buyer_same_as_importer = ?,
               seller_same_as_exporter = ?,
               metadata_json = NULL,
               sub_status = CASE
                   WHEN NULLIF(LTRIM(RTRIM(COALESCE(tss_consignment_ref, ''))), '') IS NULL
                   THEN 'PENDING'
                   ELSE sub_status
               END,
               last_sub_status_change = SYSUTCDATETIME(),
               updated_at = SYSUTCDATETIME()
         WHERE stg_consignment_id = ?
    """, [
        cleaned.get('goods_description'),
        cleaned.get('trader_reference'),
        cleaned.get('transport_document_number'),
        cleaned.get('controlled_goods') or 'no',
        cleaned.get('goods_domestic_status'),
        cleaned.get('destination_country'),
        cleaned.get('ducr'),
        cleaned.get('container_indicator'),
        cleaned.get('generate_SD') or 'no',
        cleaned.get('no_sfd_reason'),
        cleaned.get('align_ukims'),
        cleaned.get('use_importer_sde'),
        cleaned.get('declaration_choice'),
        cleaned.get('consignor_eori'),
        cleaned.get('consignor_name'),
        cleaned.get('consignor_street_number'),
        cleaned.get('consignor_city'),
        cleaned.get('consignor_postcode'),
        cleaned.get('consignor_country'),
        cleaned.get('consignee_eori'),
        cleaned.get('consignee_name'),
        cleaned.get('consignee_street_number'),
        cleaned.get('consignee_city'),
        cleaned.get('consignee_postcode'),
        cleaned.get('consignee_country'),
        cleaned.get('importer_eori'),
        cleaned.get('importer_name'),
        cleaned.get('importer_street_number'),
        cleaned.get('importer_city'),
        cleaned.get('importer_postcode'),
        cleaned.get('importer_country'),
        cleaned.get('exporter_eori'),
        cleaned.get('buyer_same_as_importer') or 'yes',
        cleaned.get('seller_same_as_exporter') or 'yes',
        sid,
    ])
    flash(f'Consignment #{sid} saved in STG. Automation will use the corrected values on the next cargo retry.', 'success')
    return redirect(next_url)
# ── NOTIFY ────────────────────────────────────────────

@consignments_bp.route('/<int:sid>/notify', methods=['POST'])
def notify(sid):
    flash('This legacy portal view is not available in Automation PRD. Use Ingestion to monitor email automation and use STG/TSS-backed pages only.', 'info')
    return redirect(url_for('ingest.queue'))
# ── DELETE ────────────────────────────────────────────

@consignments_bp.route('/<int:sid>/delete', methods=['POST'])
def delete(sid):
    flash('This legacy portal view is not available in Automation PRD. Use Ingestion to monitor email automation and use STG/TSS-backed pages only.', 'info')
    return redirect(url_for('ingest.queue'))
# ── RETRY ─────────────────────────────────────────────

@consignments_bp.route('/<int:sid>/cancel-tss', methods=['POST'])
def cancel_tss(sid):
    client_code = (get_tenant().get('code') or 'BKD').upper()
    cons = query_one(
        """
        SELECT
            c.stg_consignment_id, c.ClientCode, c.stg_header_id, c.sub_status,
            c.tss_consignment_ref, c.tss_ens_header_ref,
            COALESCE(c.trader_reference, c.transport_document_number, c.tss_consignment_ref) AS document_no,
            tc.TssStatus
        FROM STG.BKD_ENS_Consignments c
        LEFT JOIN TSS.BKD_ENS_Consignments tc
          ON tc.ClientCode = c.ClientCode
         AND tc.ConsignmentReference = c.tss_consignment_ref
        WHERE c.ClientCode = ? AND c.stg_consignment_id = ?
        """,
        [client_code, sid],
    )
    if not cons:
        flash('Consignment not found.', 'warning')
        return redirect(url_for('consignments.list_view'))

    dec_ref = str(cons.get('tss_consignment_ref') or '').strip()
    if not _can_cancel_consignment_in_tss(cons):
        flash(
            'This consignment cannot be cancelled in TSS from Fusion. TSS cancellation is only available before ARRIVED and before the record is already cancelled/completed.',
            'warning',
        )
        return redirect(url_for('consignments.detail', sid=sid))

    from app.tss_api import build_cfg_client
    api = build_cfg_client()
    payload = {'op_type': 'cancel', 'consignment_number': dec_ref}
    result = api.cancel_consignment(dec_ref)
    message = result.get('message') or result.get('error_message') or ''
    response_status = result.get('status') or ''

    insert_api_call_log(
        client_code,
        'CANCEL_PRD_ENS_CONSIGNMENT',
        staging_id=sid,
        http_method=result.get('method', 'POST'),
        url=result.get('url', ''),
        request_payload=payload,
        http_status=result.get('http_status'),
        response_status=response_status,
        response_message=message,
        response_json=result.get('raw_response') or result.get('response'),
        duration_ms=result.get('duration_ms'),
        error_detail='' if result.get('success') else message,
    )

    if result.get('success'):
        raw_json = json.dumps(result.get('response') or result, default=str)
        with db_cursor() as cursor:
            cursor.execute(
                """
                UPDATE STG.BKD_ENS_Consignments
                   SET sub_status = 'CANCELLED',
                       metadata_json = NULL,
                       completed_at = SYSUTCDATETIME(),
                       last_sub_status_change = SYSUTCDATETIME(),
                       updated_at = SYSUTCDATETIME()
                 WHERE ClientCode = ? AND stg_consignment_id = ?
                """,
                [client_code, sid],
            )
            cursor.execute(
                """
                UPDATE STG.BKD_GoodsItems
                   SET sub_status = 'CANCELLED',
                       updated_at = SYSUTCDATETIME()
                 WHERE ClientCode = ?
                   AND stg_consignment_id = ?
                   AND UPPER(COALESCE(sub_status, '')) NOT IN ('CANCELLED', 'DELETED')
                """,
                [client_code, sid],
            )
            cursor.execute(
                """
                UPDATE STG.BKD_SFD_Tracking
                   SET sub_status = 'CANCELLED',
                       tss_sfd_status = 'CANCELLED',
                       tss_control_status = COALESCE(NULLIF(tss_control_status, ''), 'Cancelled with parent consignment'),
                       stg_polled_at = SYSUTCDATETIME()
                 WHERE ClientCode = ?
                   AND tss_consignment_ref = ?
                   AND UPPER(COALESCE(sub_status, '')) NOT IN ('CANCELLED', 'DELETED')
                """,
                [client_code, dec_ref],
            )
            cursor.execute(
                """
                UPDATE TSS.BKD_SFD
                   SET TssStatus = 'CANCELLED',
                       LastSyncedAt = SYSUTCDATETIME(),
                       UpdatedAt = SYSUTCDATETIME()
                 WHERE ClientCode = ?
                   AND DeclarationNumber = ?
                   AND UPPER(COALESCE(TssStatus, '')) NOT IN ('CANCELLED', 'DELETED')
                """,
                [client_code, dec_ref],
            )
            cursor.execute(
                """
                MERGE TSS.BKD_ENS_Consignments AS target
                USING (SELECT ? AS ClientCode, ? AS ConsignmentReference) AS src
                   ON target.ClientCode = src.ClientCode
                  AND target.ConsignmentReference = src.ConsignmentReference
                WHEN MATCHED THEN UPDATE SET
                    DeclarationNumber = ?,
                    EnsReference = COALESCE(NULLIF(?, ''), target.EnsReference),
                    TssStatus = 'CANCELLED',
                    RawJson = ?,
                    LastSyncedAt = SYSUTCDATETIME(),
                    UpdatedAt = SYSUTCDATETIME()
                WHEN NOT MATCHED THEN
                    INSERT (ClientCode, DeclarationNumber, EnsReference, ConsignmentReference,
                            TssStatus, RawJson, LastSyncedAt, UpdatedAt)
                    VALUES (src.ClientCode, ?, NULLIF(?, ''), src.ConsignmentReference,
                            'CANCELLED', ?, SYSUTCDATETIME(), SYSUTCDATETIME());
                """,
                [
                    client_code, dec_ref,
                    dec_ref, cons.get('tss_ens_header_ref') or '', raw_json,
                    dec_ref, cons.get('tss_ens_header_ref') or '', raw_json,
                ],
            )
        try:
            if cons.get('stg_header_id'):
                from app.ingestion.ens_status_watcher import start_ens_status_watcher
                start_ens_status_watcher(int(cons['stg_header_id']), tenant_code=client_code)
        except Exception:
            logger.debug('Skipping watcher after consignment cancel for %s', sid, exc_info=True)
        flash(f'Cancelled {dec_ref} in TSS.', 'success')
        return redirect(url_for('consignments.detail', sid=sid))

    with db_cursor() as cursor:
        cursor.execute(
            """
            UPDATE STG.BKD_ENS_Consignments
               SET metadata_json = ?,
                   last_sub_status_change = SYSUTCDATETIME(),
                   updated_at = SYSUTCDATETIME()
             WHERE ClientCode = ? AND stg_consignment_id = ?
            """,
            [
                json.dumps({'cancel_error': message or 'TSS rejected cancellation.', 'result': result}, default=str)[:4000],
                client_code,
                sid,
            ],
        )
    flash(f'TSS rejected cancellation for {dec_ref}: {message or "unknown error"}', 'danger')
    return redirect(url_for('consignments.detail', sid=sid))


@consignments_bp.route('/<int:sid>/recreate-cancelled', methods=['POST'])
def recreate_cancelled(sid):
    client_code = (get_tenant().get('code') or 'BKD').upper()
    new_cons_id = None
    cloned_goods = 0
    header_id = None
    old_dec_ref = ''

    try:
        with db_cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    c.*,
                    tc.TssStatus AS _cons_tss_status,
                    h.tss_ens_header_ref AS _header_tss_ens_header_ref
                FROM [STG].[BKD_ENS_Consignments] c
                LEFT JOIN [STG].[BKD_ENS_Headers] h
                  ON h.ClientCode = c.ClientCode
                 AND h.stg_header_id = c.stg_header_id
                LEFT JOIN [TSS].[BKD_ENS_Consignments] tc
                  ON tc.ClientCode = c.ClientCode
                 AND tc.ConsignmentReference = c.tss_consignment_ref
                WHERE c.ClientCode = ? AND c.stg_consignment_id = ?
                """,
                [client_code, sid],
            )
            cons = _cursor_row_as_dict(cursor)
            if not cons:
                flash('Consignment not found.', 'warning')
                return redirect(url_for('consignments.list_view'))

            cons['TssStatus'] = cons.get('_cons_tss_status')
            cons['tss_ens_header_ref'] = (
                cons.get('tss_ens_header_ref')
                or cons.get('_header_tss_ens_header_ref')
                or ''
            )
            if not _can_recreate_cancelled_consignment(cons):
                flash('Only cancelled consignments with a live parent ENS can be recreated for a new DEC.', 'warning')
                return redirect(url_for('consignments.detail', sid=sid))

            old_dec_ref = str(cons.get('tss_consignment_ref') or '').strip()
            header_id = int(cons.get('stg_header_id') or 0)
            new_cons_id, cloned_goods = _clone_cancelled_consignment_for_resend(cursor, cons)
    except Exception as exc:
        logger.exception('Failed to recreate cancelled consignment %s', sid)
        flash(f'Could not recreate cancelled consignment: {exc}', 'danger')
        return redirect(url_for('consignments.detail', sid=sid))

    queued = False
    if header_id and cloned_goods:
        try:
            from app.blueprints.ingest.routes import _start_prd_cargo_auto_submit_worker
            result = _start_prd_cargo_auto_submit_worker(
                header_id,
                tenant_code=client_code,
                env_code='PRD',
                subject=f'Recreate cancelled DEC {old_dec_ref}',
                filename='Consignment detail',
            )
            queued = bool(result.get('queued'))
        except Exception:
            logger.debug('Cargo worker queue skipped after recreating consignment %s', sid, exc_info=True)

    if queued:
        flash(
            f'Recreated cancelled DEC {old_dec_ref or sid} as local consignment #{new_cons_id} with {cloned_goods} goods item(s). TSS creation has started in the background.',
            'success',
        )
    elif cloned_goods:
        flash(
            f'Recreated cancelled DEC {old_dec_ref or sid} as local consignment #{new_cons_id} with {cloned_goods} goods item(s). Use Retry Cargo to TSS to send it.',
            'warning',
        )
    else:
        flash(
            f'Recreated cancelled DEC {old_dec_ref or sid} as local consignment #{new_cons_id}, but no ENS goods were available to clone.',
            'warning',
        )
    return redirect(url_for('consignments.detail', sid=new_cons_id or sid))


@consignments_bp.route('/bulk-delete-local-failed', methods=['POST'])
def bulk_delete_local_failed():
    flash('This legacy portal view is not available in Automation PRD. Use Ingestion to monitor email automation and use STG/TSS-backed pages only.', 'info')
    return redirect(url_for('ingest.queue'))
@consignments_bp.route('/bulk-delete-selected', methods=['POST'])
def bulk_delete_selected():
    consignment_ids = []
    for raw in request.form.getlist('selected_ids'):
        try:
            consignment_ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    consignment_ids = sorted(set(consignment_ids))
    if not consignment_ids:
        flash('Select at least one consignment to delete.', 'warning')
        return redirect(url_for('consignments.list_view'))

    client_code = (get_tenant().get('code') or 'BKD').upper()
    placeholders = ','.join('?' for _ in consignment_ids)
    with db_cursor() as cursor:
        cursor.execute(
            f"""
            DELETE FROM STG.BKD_ENS_Consignments
            WHERE ClientCode = ?
              AND stg_consignment_id IN ({placeholders})
              AND COALESCE(tss_consignment_ref, '') = ''
            """,
            [client_code, *consignment_ids],
        )
        deleted = cursor.rowcount
    skipped = len(consignment_ids) - max(deleted, 0)
    if deleted:
        flash(f'Deleted {deleted} local consignment draft(s). Linked goods were removed by STG cascade.', 'success')
    if skipped:
        flash(f'{skipped} consignment(s) were kept because they already have a TSS reference or are not local drafts.', 'warning')
    return redirect(url_for('consignments.list_view'))


@consignments_bp.route('/bulk-export-selected', methods=['POST'])
def bulk_export_selected():
    consignment_ids = []
    export_source = 'selected'
    for raw in request.form.getlist('selected_ids'):
        try:
            consignment_ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    if not consignment_ids:
        export_source = 'visible'
        for raw in request.form.getlist('visible_ids'):
            try:
                consignment_ids.append(int(raw))
            except (TypeError, ValueError):
                continue
    consignment_ids = list(dict.fromkeys(consignment_ids))
    if not consignment_ids:
        flash('There are no visible consignments to export.', 'warning')
        return redirect(url_for('consignments.list_view'))

    client_code = (get_tenant().get('code') or 'BKD').upper()
    placeholders = ','.join('?' for _ in consignment_ids)
    rows = query_all(
        f"""
        SELECT
            c.stg_consignment_id, c.stg_header_id,
            h.tss_ens_header_ref,
            COALESCE(c.trader_reference, c.transport_document_number, c.tss_consignment_ref) AS document_no,
            c.trader_reference,
            c.transport_document_number,
            c.goods_description AS consignment_goods_description,
            c.sub_status,
            c.tss_consignment_ref,
            tc.TssStatus,
            COALESCE(sfd_track.tss_sfd_number, tss_sfd.SfdReference) AS sfd_reference,
            COALESCE(
                NULLIF(LTRIM(RTRIM(sfd_track.tss_movement_reference_number)), ''),
                NULLIF(LTRIM(RTRIM(tss_sfd.MovementReferenceNumber)), ''),
                CASE WHEN ISJSON(tc.RawJson) = 1 THEN NULLIF(LTRIM(RTRIM(JSON_VALUE(tc.RawJson, '$.movement_reference_number'))), '') END,
                CASE WHEN ISJSON(tc.RawJson) = 1 THEN NULLIF(LTRIM(RTRIM(JSON_VALUE(tc.RawJson, '$.movementReferenceNumber'))), '') END,
                CASE WHEN ISJSON(tc.RawJson) = 1 THEN NULLIF(LTRIM(RTRIM(JSON_VALUE(tc.RawJson, '$.mrn'))), '') END
            ) AS mrn,
            (SELECT STRING_AGG(CAST(sd.tss_sup_dec_number AS NVARCHAR(MAX)), N' | ')
             FROM STG.BKD_SDI_Headers sd
             WHERE sd.ClientCode = c.ClientCode
               AND NULLIF(LTRIM(RTRIM(COALESCE(sd.tss_sup_dec_number, ''))), '') IS NOT NULL
               AND (
                    sd.stg_consignment_id = c.stg_consignment_id
                 OR sd.tss_consignment_ref = c.tss_consignment_ref
                 OR sd.tss_sfd_consignment_ref = c.tss_consignment_ref
               )) AS sdi_references,
            (SELECT COUNT(*)
             FROM STG.BKD_GoodsItems gc
             WHERE gc.ClientCode = c.ClientCode
               AND gc.stg_consignment_id = c.stg_consignment_id
               AND UPPER(COALESCE(gc.goods_stage, 'ENS')) <> 'SDI'
               AND UPPER(COALESCE(gc.sub_status, '')) NOT IN ('CANCELLED', 'DELETED')) AS goods_count,
            h.conveyance_ref,
            h.arrival_date_time,
            c.source,
            c.updated_at,
            g.stg_item_id,
            g.item_seq,
            g.tss_hex_id,
            tg.TssStatus AS goods_tss_status,
            g.sub_status AS goods_local_status,
            g.sku,
            g.goods_description AS goods_item_description,
            g.commodity_code,
            g.country_of_origin,
            g.gross_mass_kg,
            g.net_mass_kg,
            g.number_of_packages,
            g.number_of_individual_pieces,
            g.type_of_packages,
            g.package_marks,
            g.item_invoice_amount,
            g.item_invoice_currency,
            g.customs_value,
            g.procedure_code,
            g.additional_procedure_code,
            g.controlled_goods,
            g.ni_additional_information_codes,
            g.invoice_number,
            g.document_references_json,
            g.error_message AS goods_error_message
        FROM STG.BKD_ENS_Consignments c
        LEFT JOIN STG.BKD_ENS_Headers h
          ON h.ClientCode = c.ClientCode
         AND h.stg_header_id = c.stg_header_id
        LEFT JOIN TSS.BKD_ENS_Consignments tc
          ON tc.ClientCode = c.ClientCode
         AND tc.ConsignmentReference = c.tss_consignment_ref
        LEFT JOIN STG.BKD_GoodsItems g
          ON g.ClientCode = c.ClientCode
         AND g.stg_consignment_id = c.stg_consignment_id
         AND UPPER(COALESCE(g.goods_stage, 'ENS')) <> 'SDI'
         AND UPPER(COALESCE(g.sub_status, '')) NOT IN ('CANCELLED', 'DELETED')
        LEFT JOIN TSS.BKD_GoodsItems tg
          ON tg.ClientCode = g.ClientCode
         AND tg.GoodsId = g.tss_hex_id
        OUTER APPLY (
            SELECT TOP 1
                t.tss_sfd_number,
                t.tss_movement_reference_number,
                t.tss_eori_for_eidr,
                t.tss_sfd_status
            FROM STG.BKD_SFD_Tracking t
            WHERE t.ClientCode = c.ClientCode
              AND t.tss_consignment_ref = c.tss_consignment_ref
            ORDER BY t.stg_polled_at DESC, t.stg_tracking_id DESC
        ) sfd_track
        OUTER APPLY (
            SELECT TOP 1
                s.SfdReference,
                s.MovementReferenceNumber,
                s.TssStatus
            FROM TSS.BKD_SFD s
            WHERE s.ClientCode = c.ClientCode
              AND (
                    s.DeclarationNumber = c.tss_consignment_ref
                 OR (
                        NULLIF(LTRIM(RTRIM(COALESCE(s.DeclarationNumber, ''))), '') IS NULL
                    AND s.EnsReference = c.tss_ens_header_ref
                    )
              )
            ORDER BY
                CASE WHEN s.DeclarationNumber = c.tss_consignment_ref THEN 0 ELSE 1 END,
                COALESCE(s.UpdatedAt, s.LastSyncedAt, s.CreatedAt) DESC
        ) tss_sfd
        WHERE c.ClientCode = ? AND c.stg_consignment_id IN ({placeholders})
        ORDER BY c.stg_consignment_id, COALESCE(g.item_seq, g.stg_item_id), g.stg_item_id
        """,
        [client_code, *consignment_ids],
    )
    rows = [dict(row) for row in rows or []]
    id_order = {value: index for index, value in enumerate(consignment_ids)}
    rows.sort(key=lambda row: (
        id_order.get(row.get('stg_consignment_id'), len(id_order)),
        _safe_positive_int(row.get('item_seq') or row.get('stg_item_id'), 999999),
    ))

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    columns = [
        ('stg_consignment_id', 'Consignment ID', 16),
        ('stg_header_id', 'ENS Header ID', 14),
        ('tss_ens_header_ref', 'ENS Ref', 18),
        ('tss_consignment_ref', 'DEC Ref', 20),
        ('local_status', 'Local Status', 18),
        ('tss_status', 'TSS Status', 22),
        ('sfd_reference', 'SFD Ref', 20),
        ('mrn', 'MRN / EIDR', 24),
        ('sdi_references', 'SUPDEC / SDI', 28),
        ('goods_count', 'Goods Count', 12),
        ('document_no', 'Document / Ref', 18),
        ('trader_reference', 'Trader Ref', 18),
        ('transport_document_number', 'Transport Doc No', 20),
        ('arrival_date_time', 'Arrival', 20),
        ('conveyance_ref', 'Conveyance', 18),
        ('consignment_goods_description', 'Consignment Description', 34),
        ('source', 'Source', 18),
        ('updated_at', 'Updated At', 20),
        ('stg_item_id', 'Goods STG ID', 14),
        ('item_seq', 'Item No', 10),
        ('tss_hex_id', 'TSS Goods ID', 36),
        ('goods_tss_status', 'Goods TSS Status', 18),
        ('goods_local_status', 'Goods Local Status', 18),
        ('sku', 'SKU', 18),
        ('goods_item_description', 'Goods Description', 42),
        ('commodity_code', 'Commodity', 16),
        ('country_of_origin', 'Origin', 10),
        ('gross_mass_kg', 'Gross Mass KG', 14),
        ('net_mass_kg', 'Net Mass KG', 14),
        ('number_of_packages', 'Packages', 12),
        ('number_of_individual_pieces', 'Pieces', 12),
        ('type_of_packages', 'Package Type', 14),
        ('package_marks', 'Package Marks', 18),
        ('item_invoice_amount', 'Item Value', 14),
        ('item_invoice_currency', 'Currency', 10),
        ('customs_value', 'Customs Value', 14),
        ('procedure_code', 'Procedure', 14),
        ('additional_procedure_code', 'Additional Procedure', 20),
        ('controlled_goods', 'Controlled Goods', 16),
        ('ni_additional_information_codes', 'NI Additional Info', 20),
        ('invoice_number', 'Invoice / N935', 18),
        ('document_references_json', 'Document References JSON', 44),
        ('goods_error_message', 'Goods Error / TSS Message', 44),
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Consignments Goods'
    title_fill = PatternFill('solid', fgColor='0B1F3A')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    goods_fill = PatternFill('solid', fgColor='D9EAD3')
    thin = Side(style='thin', color='D9E2EC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = 'Consignments with Goods Export'
    title_cell.fill = title_fill
    title_cell.font = Font(color='FFFFFF', bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    summary = [
        f'Tenant: {client_code}',
        f'Export: {export_source}',
        f'Consignments: {len(consignment_ids)}',
        f'Goods rows: {len(rows)}',
        f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}',
    ]
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    summary_cell = sheet.cell(row=2, column=1)
    summary_cell.value = ' | '.join(summary)
    summary_cell.fill = PatternFill('solid', fgColor='EAF2F8')
    summary_cell.font = Font(color='1F2937', bold=True)

    header_row = 4
    for col_index, (_, label, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=col_index)
        cell.value = label
        cell.fill = goods_fill if col_index >= 19 else header_fill
        cell.font = Font(color='000000' if col_index >= 19 else 'FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        sheet.column_dimensions[get_column_letter(col_index)].width = width

    for row_index, source_row in enumerate(rows, start=header_row + 1):
        row = dict(source_row)
        row['local_status'] = fixed_consignment_local_status(row.get('sub_status'), row.get('TssStatus'))
        row['tss_status'] = row.get('TssStatus')
        for col_index, (key, _, _) in enumerate(columns, start=1):
            value = row.get(key)
            cell = sheet.cell(row=row_index, column=col_index)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=key in {
                'consignment_goods_description',
                'goods_item_description',
                'document_references_json',
                'goods_error_message',
            })
            if row_index % 2 == 1:
                cell.fill = PatternFill('solid', fgColor='F8FAFC')
            if key in {'local_status', 'tss_status', 'goods_tss_status', 'goods_local_status'}:
                text = str(value or '').upper()
                if text in {'SUBMITTED', 'ARRIVED', 'AUTHORISED FOR MOVEMENT', 'CLOSED'}:
                    cell.fill = PatternFill('solid', fgColor='D9EAD3')
                elif text in {'TRADER INPUT REQUIRED', 'FAILED', 'REJECTED'}:
                    cell.fill = PatternFill('solid', fgColor='F4CCCC')
                elif text in {'DRAFT', 'PENDING', 'PROCESSING'}:
                    cell.fill = PatternFill('solid', fgColor='FFF2CC')

    sheet.freeze_panes = 'A5'
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{max(header_row, header_row + len(rows))}"
    for row in sheet.iter_rows(min_row=header_row + 1):
        sheet.row_dimensions[row[0].row].height = 34

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    suffix = 'selected' if export_source == 'selected' else 'visible'
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=consignments_{suffix}_with_goods.xlsx'},
    )
@consignments_bp.route('/<int:sid>/retry', methods=['POST'])
def retry(sid):
    flash('This legacy portal view is not available in Automation PRD. Use Ingestion to monitor email automation and use STG/TSS-backed pages only.', 'info')
    return redirect(url_for('ingest.queue'))
# ── DETAIL ────────────────────────────────────────────

@consignments_bp.route('/<string:cons_ref>/detail')
def detail_by_ref(cons_ref):
    flash('This legacy portal view is not available in Automation PRD. Use Ingestion to monitor email automation and use STG/TSS-backed pages only.', 'info')
    return redirect(url_for('ingest.queue'))
@consignments_bp.route('/<int:sid>')
def detail(sid):
    cons = query_one("""
        SELECT
            c.stg_consignment_id, c.ClientCode, c.stg_header_id,
            COALESCE(c.trader_reference, c.transport_document_number, c.tss_consignment_ref) AS document_no,
            c.sub_status, c.tss_consignment_ref, c.goods_description,
            c.trader_reference, c.transport_document_number,
            c.controlled_goods, c.goods_domestic_status, c.destination_country,
            c.ducr, c.container_indicator, c.generate_SD, c.no_sfd_reason,
            c.align_ukims, c.use_importer_sde, c.declaration_choice,
            c.consignor_eori, c.consignor_name,
            c.consignor_street_number, c.consignor_city, c.consignor_postcode, c.consignor_country,
            c.consignee_eori, c.consignee_name,
            c.consignee_street_number, c.consignee_city, c.consignee_postcode, c.consignee_country,
            c.importer_eori, c.importer_name,
            c.importer_street_number, c.importer_city, c.importer_postcode, c.importer_country,
            c.exporter_eori,
            c.buyer_same_as_importer, c.seller_same_as_exporter,
            c.source, c.updated_at,
            h.conveyance_ref, h.arrival_date_time, h.tss_ens_header_ref,
            tc.TssStatus,
            COALESCE(sfd_track.tss_sfd_number, tss_sfd.SfdReference) AS sfd_reference,
            COALESCE(
                NULLIF(LTRIM(RTRIM(sfd_track.tss_movement_reference_number)), ''),
                NULLIF(LTRIM(RTRIM(tss_sfd.MovementReferenceNumber)), ''),
                CASE WHEN ISJSON(tc.RawJson) = 1 THEN NULLIF(LTRIM(RTRIM(JSON_VALUE(tc.RawJson, '$.movement_reference_number'))), '') END,
                CASE WHEN ISJSON(tc.RawJson) = 1 THEN NULLIF(LTRIM(RTRIM(JSON_VALUE(tc.RawJson, '$.movementReferenceNumber'))), '') END,
                CASE WHEN ISJSON(tc.RawJson) = 1 THEN NULLIF(LTRIM(RTRIM(JSON_VALUE(tc.RawJson, '$.mrn'))), '') END
            ) AS sfd_mrn,
            sfd_track.tss_eori_for_eidr AS sfd_eidr,
            COALESCE(
                CASE WHEN UPPER(LTRIM(RTRIM(COALESCE(sfd_track.tss_sfd_status, '')))) IN ('OK', 'SUCCESS') THEN NULL ELSE sfd_track.tss_sfd_status END,
                CASE WHEN UPPER(LTRIM(RTRIM(COALESCE(tss_sfd.TssStatus, '')))) IN ('OK', 'SUCCESS') THEN NULL ELSE tss_sfd.TssStatus END,
                tc.TssStatus
            ) AS sfd_status
        FROM STG.BKD_ENS_Consignments c
        LEFT JOIN STG.BKD_ENS_Headers h ON h.stg_header_id = c.stg_header_id
        LEFT JOIN TSS.BKD_ENS_Consignments tc
            ON tc.ClientCode = c.ClientCode
           AND tc.ConsignmentReference = c.tss_consignment_ref
        OUTER APPLY (
            SELECT TOP 1
                t.tss_sfd_number,
                t.tss_movement_reference_number,
                t.tss_eori_for_eidr,
                t.tss_sfd_status
            FROM STG.BKD_SFD_Tracking t
            WHERE t.ClientCode = c.ClientCode
              AND t.tss_consignment_ref = c.tss_consignment_ref
            ORDER BY t.stg_polled_at DESC, t.stg_tracking_id DESC
        ) sfd_track
        OUTER APPLY (
            SELECT TOP 1
                s.SfdReference,
                s.MovementReferenceNumber,
                s.TssStatus
            FROM TSS.BKD_SFD s
            WHERE s.ClientCode = c.ClientCode
              AND s.DeclarationNumber = c.tss_consignment_ref
            ORDER BY COALESCE(s.UpdatedAt, s.LastSyncedAt, s.CreatedAt) DESC
        ) tss_sfd
        WHERE c.stg_consignment_id = ?
    """, [sid])
    if not cons:
        flash('Consignment not found.', 'warning')
        return redirect(url_for('consignments.list_view'))
    goods = query_all("""
        SELECT
            g.stg_item_id, g.item_seq, g.goods_description, g.commodity_code,
            g.gross_mass_kg, g.net_mass_kg,
            g.number_of_packages, g.type_of_packages,
            g.sub_status, g.error_message, g.tss_hex_id,
            COALESCE(
                NULLIF(LTRIM(RTRIM(g.sku)), ''),
                NULLIF(LTRIM(RTRIM(pc.sku)), ''),
                NULLIF(LTRIM(RTRIM(pc.product_code)), '')
            ) AS sku,
            tg.TssStatus AS item_tss_status
        FROM STG.BKD_GoodsItems g
        LEFT JOIN TSS.BKD_GoodsItems tg
            ON tg.ClientCode = g.ClientCode
           AND tg.GoodsId = g.tss_hex_id
        OUTER APPLY (
            SELECT TOP 1
                p.sku,
                p.product_code
            FROM BKD.DocProductCatalog p
            WHERE (
                    NULLIF(LTRIM(RTRIM(g.goods_description)), '') IS NOT NULL
                AND UPPER(LTRIM(RTRIM(p.description))) = UPPER(LTRIM(RTRIM(g.goods_description)))
            )
              AND (
                    NULLIF(LTRIM(RTRIM(g.commodity_code)), '') IS NULL
                 OR NULLIF(LTRIM(RTRIM(p.commodity_code)), '') IS NULL
                 OR REPLACE(LTRIM(RTRIM(p.commodity_code)), ' ', '') = REPLACE(LTRIM(RTRIM(g.commodity_code)), ' ', '')
              )
            ORDER BY
                CASE WHEN NULLIF(LTRIM(RTRIM(p.sku)), '') IS NULL THEN 1 ELSE 0 END,
                p.id DESC
        ) pc
        WHERE g.stg_consignment_id = ?
        ORDER BY g.item_seq, g.stg_item_id
    """, [sid])
    for row in goods or []:
        row['type_of_packages'] = normalise_package_type(row.get('type_of_packages'), 'PK') or 'PK'
        row['can_delete_failed_goods'] = (
            normalize_status_key(row.get('sub_status')) in PRD_LOCAL_GOODS_DELETE_STATUSES
            and not str(row.get('tss_hex_id') or '').strip()
        )
        row['can_delete_local_goods'] = row['can_delete_failed_goods']
    local_status = normalize_status_key(cons.get('sub_status'))
    pending_goods_count = sum(
        1 for g in (goods or [])
        if not str(g.get('tss_hex_id') or '').strip()
        and normalize_status_key(g.get('sub_status')) not in {'CANCELLED', 'DELETED'}
    )
    can_sync_header = bool(cons.get('stg_header_id') and str(cons.get('tss_ens_header_ref') or '').strip())
    has_cons_ref = bool(str(cons.get('tss_consignment_ref') or '').strip())
    tss_can_submit = normalize_status_key(cons.get('TssStatus')) in {'DRAFT', 'TRADER INPUT REQUIRED'}
    can_submit_to_tss = bool(
        can_sync_header
        and (
            not has_cons_ref
            or (
                tss_can_submit
                and (
                    pending_goods_count > 0
                    or local_status not in {'SUBMITTED', 'COMPLETED'}
                )
            )
        )
    )
    linked_supdecs = load_prd_sdi_links_for_context(
        client_code=cons.get('ClientCode') or (get_tenant().get('code') or 'BKD').upper(),
        consignment_ids=[cons.get('stg_consignment_id')],
        consignment_refs=[
            cons.get('tss_consignment_ref'),
            cons.get('trader_reference'),
            cons.get('transport_document_number'),
            cons.get('document_no'),
        ],
        sfd_refs=[cons.get('sfd_reference')],
    )
    linked_supdecs = merge_sdi_links(linked_supdecs)
    return render_template(
        'consignments/detail.html',
        cons=cons,
        goods=goods,
        badge_class=badge_class,
        linked_supdecs=linked_supdecs,
        can_submit_to_tss=can_submit_to_tss,
        can_sync_header=can_sync_header,
        can_edit_consignment=_prd_consignment_allows_edit(cons),
        can_cancel_tss=_can_cancel_consignment_in_tss(cons),
        can_recreate_cancelled=_can_recreate_cancelled_consignment(cons),
        pending_goods_count=pending_goods_count,
    )
def _render_detail(cons):
    goods = query_all(f"""
        SELECT staging_id, item_number, goods_description, commodity_code,
               type_of_packages, number_of_packages, gross_mass_kg,
               status, tss_status, goods_id, error_message
        FROM {S}.StagingGoodsItems WHERE staging_cons_id=?
        ORDER BY item_number, staging_id
    """, [cons['staging_id']])
    goods = _apply_consignment_goods_status(goods, cons)

    linked_supdecs = []
    try:
        linked_supdecs = query_all(f"""
            SELECT staging_id, sup_dec_number, status, submission_due_date,
                   sfd_reference, ens_consignment_ref
            FROM {S}.StagingSupDecHeaders
            WHERE staging_cons_id = ?
               OR ens_consignment_ref = ?
               OR sfd_reference = ?
            ORDER BY created_at DESC
        """, [cons['staging_id'], cons.get('dec_reference'), cons.get('sfd_reference')])
    except Exception:
        linked_supdecs = []

    prd_sdi_links = load_prd_sdi_links_for_context(
        client_code=(get_tenant().get('code') or 'BKD').upper(),
        consignment_refs=[
            cons.get('dec_reference'),
            cons.get('sfd_reference'),
            cons.get('trader_reference'),
            cons.get('transport_document_number'),
        ],
        sfd_refs=[cons.get('sfd_reference')],
    )
    if prd_sdi_links:
        linked_supdecs = merge_sdi_links(linked_supdecs, prd_sdi_links)

    linked_sfd = _load_linked_sfd_for_consignment(cons)
    sfd_state = _build_sfd_route_a_state(cons, linked_sfd=linked_sfd)

    linked_gmr = None
    if cons.get('staging_ens_id'):
        linked_gmr = query_one(f"""
            SELECT staging_id, gmr_id, status, gvms_status
            FROM {S}.StagingGmrs
            WHERE staging_ens_id = ?
            ORDER BY created_at DESC
        """, [cons['staging_ens_id']])

    api_calls = []
    try:
        api_calls = query_all(f"""
            SELECT TOP 50 id, call_type, http_method, url,
                   request_payload, http_status, response_status,
                   response_message, response_json, duration_ms,
                   error_detail, called_at
            FROM {S}.ApiCallLog
            WHERE staging_id = ?
            ORDER BY called_at DESC
        """, [cons['staging_id']])
    except Exception:
        pass

    header_route_a_state = _ens_header_route_a_state(cons.get('staging_ens_id'))

    guidance = _build_consignment_guidance(
        cons,
        goods,
        linked_gmr=linked_gmr,
        linked_supdecs=linked_supdecs,
        linked_sfd=linked_sfd,
        api_calls=api_calls,
        header_route_a_state=header_route_a_state,
    )
    dec_state = _build_dec_reference_state(cons, api_calls=api_calls)
    api_log_hint = None
    if not api_calls:
        if cons.get('error_message') and _looks_like_tss_rejection(cons.get('error_message'), api_calls=api_calls):
            api_log_hint = 'This record stores a TSS-style rejection message, but no ApiCallLog row is currently linked to this staging record.'
        else:
            api_log_hint = guidance['detail']

    tss_action_required = _latest_consignment_tss_issue(cons, api_calls=api_calls)
    detail_issue_fields = _consignment_detail_issue_fields(cons, tss_action_required=tss_action_required)
    error_explanation = None if tss_action_required else _build_consignment_error_explanation(cons)

    return render_template('consignments/detail.html', cons=cons, goods=goods,
                           badge_class=badge_class, api_calls=api_calls,
                           linked_supdecs=linked_supdecs, linked_sfd=linked_sfd, linked_gmr=linked_gmr,
                           sfd_state=sfd_state,
                           guidance=guidance, dec_state=dec_state, api_log_hint=api_log_hint,
                           error_explanation=error_explanation, tss_action_required=tss_action_required,
                           detail_issue_fields=detail_issue_fields,
                           header_route_a_state=header_route_a_state,
                           can_change_consignment_data=_can_change_consignment_data(cons),
                           can_edit_consignment_data=_can_edit_consignment_data(cons),
                           can_validate_consignment_pipeline=_can_validate_consignment_pipeline(cons, goods),
                           can_cancel_consignment_in_tss=_can_cancel_consignment_in_tss(cons),
                           can_add_goods_to_consignment=_can_add_goods_to_consignment(cons),
                           nested={'prev_docs': [], 'auth_holders': []})
