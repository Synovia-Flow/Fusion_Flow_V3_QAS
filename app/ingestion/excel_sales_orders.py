"""Excel sales-order ingestion for tenants like Birkdale.

Pipeline:
    Email (with `DETAILS FOR ...` block) + xlsx attachment
    -> EmailMetadata + ParsedSalesOrders
    -> stage.stage_sales_orders_batch
    -> ENS / consignments / goods staged with anti-duplicate keys

This module is parsing-only. No DB writes. No TSS calls.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

NON_GOODS_SKUS = {
    "CARRIAGE",
    "DELIVERY",
    "DELIVERYCHARGE",
    "DPDELIVERY",
    "FREIGHT",
    "SHIPPING",
}
NON_GOODS_UOM_MARKERS = ("\u00a3", "$", "\u20ac")


def _normalise_line_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def _is_non_goods_sales_line(sku: Any, uom_code: Any = "") -> bool:
    sku_key = _normalise_line_key(sku)
    if sku_key in NON_GOODS_SKUS:
        return True

    uom_text = str(uom_code or "").strip().lower()
    if not uom_text:
        return False
    starts_with_currency = any(uom_text.startswith(marker) for marker in NON_GOODS_UOM_MARKERS)
    return starts_with_currency and any(
        token in uom_text for token in ("del", "delivery", "freight", "carriage")
    )


def _missing_document_no_needs_warning(goods_fields: dict[str, Any]) -> bool:
    """Return true only for rows that look like real goods lines."""
    sku = goods_fields.get("sku")
    uom_code = goods_fields.get("uom_code")
    if _is_non_goods_sales_line(sku, uom_code):
        return False

    sku_key = _normalise_line_key(sku)
    if sku_key in {"TOTAL", "SUBTOTAL", "GRANDTOTAL"}:
        return False

    return any(goods_fields.get(key) not in (None, "") for key in ("sku", "line_no"))


def _can_fill_down_document_no(goods_fields: dict[str, Any], previous_doc_no: str | None) -> bool:
    if not previous_doc_no:
        return False
    if not _missing_document_no_needs_warning(goods_fields):
        return False
    return bool(_normalise_line_key(goods_fields.get("sku")))

# ── Email body parsing ─────────────────────────────────────────────────

#: Marker that introduces the carrier transport block. Appears in the
#: innermost forwarded message from carrier ops (e.g. Primeline).
DETAILS_FOR_MARKER = re.compile(
    r"(?<!TRANSPORT\s)(?:DETAILS|TSS)\s+FOR\s+"
    r"(?P<day>\d{1,2})[./-](?P<month>\d{1,2})(?:[./-](?P<year>\d{2,4}))?",
    re.IGNORECASE,
)

#: Block ends at the next `From:` header, the Primeline response mailbox, or
#: end of body. The mailbox marker prevents signatures/disclaimers/older
#: replies after the carrier instruction block from overwriting parsed fields.
NEXT_FORWARD_MARKER = re.compile(r"^\s*From\s*:", re.IGNORECASE | re.MULTILINE)
PRIMELINE_RESPONSE_STOP_MARKER = re.compile(
    r"customsadmin@primelineexpress\.co\.uk",
    re.IGNORECASE,
)

#: Field labels we extract from the carrier block. Map label → canonical key.
EMAIL_FIELD_LABELS = {
    "type of movement": "movement_type_text",
    "type of passive transport": "type_of_passive_transport",
    "identity number of transport": "identity_no_of_transport",
    "nationality of means of transport": "nationality_of_transport_text",
    "carrier eori": "carrier_eori",
    "carrier name": "carrier_name",
    "carrier street and number": "carrier_street_number",
    "carrier street / number": "carrier_street_number",
    "carrier street number": "carrier_street_number",
    "carrier city": "carrier_city",
    "carrier postcode": "carrier_postcode",
    "carrier country": "carrier_country_text",
    "transport document number": "conveyance_ref",  # ICR — goes to ENS
    "arrival date/time": "arrival_date_time_text",
    "arrival date / time": "arrival_date_time_text",
    "port of arrival": "arrival_port_text",
    "place(s) of loading": "place_of_loading",
    "is/are the place(s) of acceptance same as place(s) of loading?": "place_of_acceptance_same_as_loading_text",
    "place of acceptance": "place_of_acceptance",
    "place(s) of acceptance": "place_of_acceptance",
    "place(s) of unloading": "place_of_unloading",
    "is/are the place(s) of delivery same as place(s) of unloading?": "place_of_delivery_same_as_unloading_text",
    "place of delivery": "place_of_delivery",
    "place(s) of delivery": "place_of_delivery",
    "transport charges": "transport_charges_text",
    "haulier eori": "haulier_eori",
}

#: Movement-type human → TSS code. Best-effort, fall back to default.
MOVEMENT_TYPE_MAP = {
    "roro accompanied ics2": "3a",
    "roro accompanied": "3a",
    "roro unaccompanied ics2": "3b",
    "roro unaccompanied": "3b",
    "container": "1",
}

#: Country name → ISO2.
NATIONALITY_MAP = {
    "united kingdom": "GB",
    "uk": "GB",
    "great britain": "GB",
    "ireland": "IE",
    "republic of ireland": "IE",
    "france": "FR",
    "germany": "DE",
    "netherlands": "NL",
    "belgium": "BE",
}

#: Common port names → TSS port code (extend with CV_port lookup later).
ARRIVAL_PORT_HINTS = {
    "belfast": "GBAUBELBELBEL",
    "belfast port": "GBAUBELBELBEL",
    "larne": "GBAULARLARLAR",
    "larne port": "GBAULARLARLAR",
    "warrenpoint": "GBAUWPTWPTWPT",
}

#: Transport charges human → TSS code.
TRANSPORT_CHARGES_MAP = {
    "account holder with carrier": "Y",
    "cash": "A",
    "credit card": "B",
    "cheque": "D",
    "electronic credit transfer": "H",
}

#: Carrier descriptions -> TSS passive transport choice values.
#: Keep this deliberately small: unknown text must stay in review instead of
#: being silently accepted as a clean upload.
PASSIVE_TRANSPORT_MAP = {
    "truck": "TRAILER",
    "trailer": "TRAILER",
    "tautliner": "3103",
    "truck tautliner": "3103",
    "truck tautliner 25 tonne": "3103",
    "truck, tautliner, 25 tonne": "3103",
}


@dataclass
class EmailMetadata:
    """Transport metadata extracted from the carrier `DETAILS FOR` block."""

    raw_block: str = ""
    movement_type: str = ""
    type_of_passive_transport: str = ""
    identity_no_of_transport: str = ""
    nationality_of_transport: str = ""
    carrier_eori: str = ""
    carrier_name: str = ""
    carrier_street_number: str = ""
    carrier_city: str = ""
    carrier_postcode: str = ""
    carrier_country: str = ""
    haulier_eori: str = ""
    conveyance_ref: str = ""  # ICR
    arrival_date_time: datetime | None = None
    arrival_port: str = ""
    place_of_loading: str = ""
    place_of_acceptance_same_as_loading: str = ""
    place_of_acceptance: str = ""
    place_of_unloading: str = ""
    place_of_delivery_same_as_unloading: str = ""
    place_of_delivery: str = ""
    transport_charges: str = ""
    diff_flags: list[str] = field(default_factory=list)
    parse_warnings: list[str] = field(default_factory=list)


def _details_match_datetime(match: re.Match, received_at: datetime | None = None) -> datetime | None:
    try:
        day = int(match.group("day"))
        month = int(match.group("month"))
        raw_year = match.groupdict().get("year")
        if raw_year in (None, ""):
            if received_at is None:
                return None
            year = received_at.year
        else:
            year = int(raw_year)
        if year < 100:
            year += 2000
        return datetime(year, month, day, tzinfo=timezone.utc)
    except (IndexError, TypeError, ValueError):
        return None


def _select_details_for_match(matches: list[re.Match], received_at: datetime | None = None) -> re.Match:
    dated = []
    for match in matches:
        marker_date = _details_match_datetime(match, received_at)
        if marker_date is not None:
            candidate_tail = match.string[match.start():match.start() + 3000].lower()
            field_score = sum(
                marker in candidate_tail
                for marker in (
                    "type of movement",
                    "identity number of transport",
                    "transport document number",
                    "arrival date",
                    "port of arrival",
                )
            )
            dated.append((marker_date, field_score, match.start(), match))
    if dated:
        return max(dated, key=lambda item: (item[0], item[1], -item[2]))[3]
    return matches[0]


def parse_email_carrier_block(body_text: str, received_at: datetime | None = None) -> EmailMetadata:
    """Extract the innermost `DETAILS FOR <date>` carrier block and parse fields.

    received_at is used to resolve relative arrival expressions like "Tomorrow".
    Falls back to now() in UTC.
    """
    received_at = received_at or datetime.now(timezone.utc)
    meta = EmailMetadata()

    if not body_text:
        meta.parse_warnings.append("Empty email body")
        return meta

    # Find the newest dated DETAILS FOR occurrence. Real Outlook/Graph threads
    # keep older carrier blocks below the current one, so "last match" can be
    # months old and produce the wrong truck/ICR/arrival data.
    matches = list(DETAILS_FOR_MARKER.finditer(body_text))
    if not matches:
        meta.parse_warnings.append("No 'DETAILS FOR' marker found in body")
        return meta
    details_match = _select_details_for_match(matches, received_at)
    block_start = details_match.start()
    marker_received_at = _details_match_datetime(details_match, received_at) or received_at
    tail = body_text[block_start:]
    next_forward = NEXT_FORWARD_MARKER.search(tail, pos=20)
    response_stop = PRIMELINE_RESPONSE_STOP_MARKER.search(tail)
    block_end_candidates = []
    if next_forward is not None:
        block_end_candidates.append(next_forward.start())
    if response_stop is not None:
        block_end_candidates.append(response_stop.end())
    block = tail[: min(block_end_candidates)] if block_end_candidates else tail
    meta.raw_block = block

    # Tokenize: each label may be on its own line followed by the value on next non-empty line.
    # Match labels by startswith so suffixes like "(ICR Number)" still resolve.
    lines = [ln.strip() for ln in block.splitlines()]
    label_to_value: dict[str, str] = {}
    label_keys_sorted = sorted(EMAIL_FIELD_LABELS.keys(), key=len, reverse=True)
    i = 0
    while i < len(lines):
        line = lines[i]
        normalized = line.lower().rstrip(":").strip()
        matched_key = None
        for label in label_keys_sorted:
            if normalized.startswith(label):
                matched_key = EMAIL_FIELD_LABELS[label]
                break
        if matched_key:
            j = i + 1
            while j < len(lines) and not lines[j]:
                j += 1
            if j < len(lines):
                next_norm = lines[j].lower().rstrip(":").strip()
                next_is_label = any(next_norm.startswith(lbl) for lbl in label_keys_sorted)
                if not next_is_label:
                    label_to_value[matched_key] = lines[j]
                    i = j + 1
                    continue
        i += 1

    # Movement type
    raw_mov = label_to_value.get("movement_type_text", "").lower().strip()
    meta.movement_type = MOVEMENT_TYPE_MAP.get(raw_mov, "")
    if raw_mov and not meta.movement_type:
        meta.parse_warnings.append(f"Unmapped movement type: {label_to_value.get('movement_type_text')!r}")

    raw_passive = label_to_value.get("type_of_passive_transport", "")
    meta.type_of_passive_transport = map_passive_transport(raw_passive)
    if raw_passive and not meta.type_of_passive_transport:
        meta.parse_warnings.append(f"Unmapped passive transport: {raw_passive!r}")
    meta.identity_no_of_transport = label_to_value.get("identity_no_of_transport", "")

    raw_nat = label_to_value.get("nationality_of_transport_text", "").lower().strip()
    meta.nationality_of_transport = NATIONALITY_MAP.get(raw_nat, "")
    if raw_nat and not meta.nationality_of_transport:
        meta.parse_warnings.append(f"Unmapped nationality: {label_to_value.get('nationality_of_transport_text')!r}")

    meta.carrier_eori = label_to_value.get("carrier_eori", "").upper()
    meta.carrier_name = label_to_value.get("carrier_name", "")
    meta.carrier_street_number = label_to_value.get("carrier_street_number", "")
    meta.carrier_city = label_to_value.get("carrier_city", "")
    meta.carrier_postcode = label_to_value.get("carrier_postcode", "")
    raw_carrier_country = label_to_value.get("carrier_country_text", "")
    meta.carrier_country = NATIONALITY_MAP.get(raw_carrier_country.lower().strip()) or (
        raw_carrier_country.strip().upper() if len(raw_carrier_country.strip()) == 2 else ""
    )
    meta.haulier_eori = label_to_value.get("haulier_eori", "").upper()
    parsed_conveyance = label_to_value.get("conveyance_ref", "").strip()
    if parsed_conveyance and not re.match(r"^ICR[A-Z0-9-]{3,}$", parsed_conveyance, re.IGNORECASE):
        parsed_conveyance = ""
    meta.conveyance_ref = (
        parsed_conveyance
        or _fallback_icr_from_block(block)
    )

    raw_arrival = label_to_value.get("arrival_date_time_text", "")
    meta.arrival_date_time = parse_arrival_datetime(raw_arrival, marker_received_at)
    if raw_arrival and meta.arrival_date_time is None:
        meta.parse_warnings.append(f"Could not parse arrival: {raw_arrival!r}")

    raw_port = label_to_value.get("arrival_port_text", "").lower().strip()
    meta.arrival_port = ARRIVAL_PORT_HINTS.get(raw_port, "")
    if raw_port and not meta.arrival_port:
        meta.parse_warnings.append(f"Unmapped port: {label_to_value.get('arrival_port_text')!r}")

    meta.place_of_loading = label_to_value.get("place_of_loading", "")
    meta.place_of_acceptance_same_as_loading = _yes_no(label_to_value.get("place_of_acceptance_same_as_loading_text", ""))
    meta.place_of_acceptance = label_to_value.get("place_of_acceptance", "")
    meta.place_of_unloading = label_to_value.get("place_of_unloading", "")
    meta.place_of_delivery_same_as_unloading = _yes_no(label_to_value.get("place_of_delivery_same_as_unloading_text", ""))
    meta.place_of_delivery = label_to_value.get("place_of_delivery", "")

    raw_charges = label_to_value.get("transport_charges_text", "").lower().strip()
    meta.transport_charges = TRANSPORT_CHARGES_MAP.get(raw_charges, "")

    return meta


def _fallback_icr_from_block(block: str) -> str:
    """Find the Primeline ICR when forwarding/html flattens label/value lines."""
    if not block or not re.search(r"transport\s+document\s+number", block, re.IGNORECASE):
        return ""
    labelled = re.search(
        r"transport\s+document\s+number(?:\s*\([^)]*\))?"
        r"[\s:;\-–—]*\b(ICR[A-Z0-9-]{3,})\b",
        block,
        re.IGNORECASE,
    )
    if labelled:
        return labelled.group(1).upper()
    any_icr = re.search(r"\b(ICR[A-Z0-9-]{3,})\b", block, re.IGNORECASE)
    return any_icr.group(1).upper() if any_icr else ""


def map_passive_transport(text: Any) -> str:
    """Map carrier passive-transport text to the TSS choice value.

    Returns an empty string for unknown text so callers keep the batch in
    review until a deliberate mapping is added.
    """
    raw = str(text or "").strip()
    if not raw:
        return ""
    normalized = re.sub(r"[^a-z0-9]+", " ", raw.lower()).strip()
    if normalized in PASSIVE_TRANSPORT_MAP:
        return PASSIVE_TRANSPORT_MAP[normalized]
    return ""


def _yes_no(text: Any) -> str:
    raw = str(text or "").strip().lower()
    if raw in {"yes", "y", "true", "1"}:
        return "yes"
    if raw in {"no", "n", "false", "0"}:
        return "no"
    return ""


def parse_arrival_datetime(text: str, received_at: datetime) -> datetime | None:
    """Parse `Tomorrow's Date / 06:30` or explicit date forms.

    Supported forms (case-insensitive):
      - "Tomorrow / HH:MM" or "Tomorrow's Date / HH:MM"
      - "Today / HH:MM"
      - "DD/MM/YYYY HH:MM" or "DD-MM-YYYY HH:MM"
      - "DD/MM/YYYY" (uses 06:00 default)
    Returns naive UTC datetime or None.
    """
    if not text:
        return None
    s = text.strip().lower()

    time_match = re.search(r"(\d{1,2}):(\d{2})", s)
    hour, minute = (6, 0)
    if time_match:
        hour, minute = int(time_match.group(1)), int(time_match.group(2))

    base_date = None
    if "tomorrow" in s:
        base_date = (received_at + timedelta(days=1)).date()
    elif "today" in s:
        base_date = received_at.date()
    else:
        date_match = re.search(r"(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})", s)
        if date_match:
            d, m, y = (int(x) for x in date_match.groups())
            if y < 100:
                y += 2000
            try:
                base_date = datetime(y, m, d).date()
            except ValueError:
                return None

    if base_date is None:
        return None
    try:
        return datetime(base_date.year, base_date.month, base_date.day, hour, minute, 0)
    except ValueError:
        return None


# ── Excel parsing ──────────────────────────────────────────────────────


@dataclass
class ParsedGoodsLine:
    document_no: str
    line_no: int | None
    sku: str
    quantity: float | None = None
    quantity_base: float | None = None
    amount: float | None = None
    line_amount_excl_vat: float | None = None
    unit_price_excl_vat: float | None = None
    qty_per_uom: float | None = None
    uom_code: str = ""
    raw: dict = field(default_factory=dict)


@dataclass
class ParsedConsignment:
    document_no: str
    sell_to_customer_no: str = ""
    ship_to_name: str = ""
    ship_to_address: str = ""
    ship_to_address_2: str = ""
    ship_to_city: str = ""
    ship_to_county: str = ""
    ship_to_postcode: str = ""
    ship_to_country: str = ""
    ship_to_phone: str = ""
    ship_to_email: str = ""
    goods: list[ParsedGoodsLine] = field(default_factory=list)


@dataclass
class ParsedSalesOrders:
    consignments: list[ParsedConsignment] = field(default_factory=list)
    parse_warnings: list[str] = field(default_factory=list)
    source_filename: str = ""


# Column header → ParsedConsignment / ParsedGoodsLine attribute
EXCEL_COLUMN_MAP = {
    "Sales Header - Sell-to Customer No.": ("consignment", "sell_to_customer_no"),
    "Sales Header - Ship-to Name":        ("consignment", "ship_to_name"),
    "Sales Header - Ship-to Address":     ("consignment", "ship_to_address"),
    "Sales Header - Ship-to City":        ("consignment", "ship_to_city"),
    "Sales Header - Ship-to County":      ("consignment", "ship_to_county"),
    "Sales Header - Ship-to Post Code":   ("consignment", "ship_to_postcode"),
    "Sales Header - Ship-to Country/Region Code": ("consignment", "ship_to_country"),
    "Sales Header - Ship-to Phone No.":   ("consignment", "ship_to_phone"),
    "Sales Header - Email":               ("consignment", "ship_to_email"),
    "Document No.":                       ("consignment", "document_no"),
    "No.":                                ("goods",       "sku"),
    "Line No.":                           ("goods",       "line_no"),
    "Quantity":                           ("goods",       "quantity"),
    "Quantity (Base)":                    ("goods",       "quantity_base"),
    "Amount":                             ("goods",       "amount"),
    "Line Amount Excl. VAT":              ("goods",       "line_amount_excl_vat"),
    "Unit Price Excl. VAT":               ("goods",       "unit_price_excl_vat"),
    "SDI Nature of Transaction":          ("goods",       "nature_of_transaction"),
    "SDI NI Additional Information Codes": ("goods",      "ni_additional_information_codes"),
    "Nature of Transaction":              ("goods",       "nature_of_transaction"),
    "NI Additional Information Codes":    ("goods",       "ni_additional_information_codes"),
    "Qty. per Unit of Measure":           ("goods",       "qty_per_uom"),
    "Unit of Measure Code":               ("goods",       "uom_code"),
}


SYNOVIAFLOW_ENS_SHEET = "ENS Header"
SYNOVIAFLOW_DATA_SHEET = "Consignments+Goods"

#: Columns the SynoviaFlow data sheet exposes. Used to map row dict → ParsedConsignment / ParsedGoodsLine.
SYNOVIAFLOW_CONSIGNMENT_KEYS = {
    "trader_reference", "transport_document_number", "destination_country",
    "goods_domestic_status", "controlled_goods", "container_indicator",
    "no_sfd_reason", "ducr", "align_ukims", "use_importer_sde",
    "supervising_customs_office", "customs_warehouse_identifier",
    "generate_SD",
    "consignor_eori", "consignor_name", "consignor_street_number",
    "consignor_city", "consignor_postcode", "consignor_country",
    "consignee_eori", "consignee_name", "consignee_street_number",
    "consignee_city", "consignee_postcode", "consignee_country",
    "importer_eori", "importer_name", "importer_street_number",
    "importer_city", "importer_postcode", "importer_country",
    "buyer_same_as_importer", "buyer_eori", "buyer_name",
    "buyer_street_and_number", "buyer_city", "buyer_postcode",
    "buyer_country",
    "exporter_eori", "exporter_name", "exporter_street_number",
    "exporter_city", "exporter_postcode", "exporter_country",
    "seller_same_as_exporter", "seller_eori", "seller_name",
    "seller_street_and_number", "seller_city", "seller_postcode",
    "seller_country",
    # SDI extras (kept on consignment for now; review screen splits later)
    "declaration_choice", "incoterm",
    "freight_charge", "freight_charge_currency",
    "insurance", "insurance_currency",
    "vat_adjustment", "vat_adjust_currency",
    "postponed_vat", "exchange_rate",
    "supervising_customs_office",
    "procedure_code", "additional_procedure_code", "valuation_method",
    "preference", "nature_of_transaction",
}

_COMMODITY_NON_DIGIT_RE = re.compile(r'\D+')


def _normalise_commodity_for_ingest(raw):
    """Return (clean, original_clean, action) where clean is a 10-digit padded code.

    - Strips spaces, dashes, dots and any non-digit (preserves only digits).
    - Pads with trailing zeros to reach 10 digits.
    - Truncates to 10 digits if longer.
    - action is one of 'unchanged', 'padded', 'truncated', 'stripped', 'empty'.

    The padding mirrors `app.hmrc_api._normalise_commodity` so the live Trade
    Tariff API and our staging see the same shape. TSS still validates the
    actual code against its tariff database; this helper only guarantees the
    format constraint.
    """
    original_clean = str(raw or '').strip()
    if not original_clean:
        return '', '', 'empty'
    digits = _COMMODITY_NON_DIGIT_RE.sub('', original_clean)
    if not digits:
        return '', original_clean, 'empty'
    actions = []
    if digits != original_clean:
        actions.append('stripped')
    if len(digits) < 10:
        digits = digits.ljust(10, '0')
        actions.append('padded')
    elif len(digits) > 10:
        digits = digits[:10]
        actions.append('truncated')
    if not actions:
        return digits, original_clean, 'unchanged'
    return digits, original_clean, '+'.join(actions)


SYNOVIAFLOW_GOODS_KEYS = {
    "commodity_code", "taric_code", "goods_description",
    "type_of_packages", "number_of_packages", "number_of_individual_pieces", "package_marks",
    "gross_mass_kg", "net_mass_kg", "country_of_origin",
    "amount", "item_invoice_amount", "item_invoice_currency", "procedure_code",
    "additional_procedure_code", "controlled_goods_type", "cus_code",
    "national_additional_code", "ni_additional_information_codes",
    "country_of_preferential_origin", "preference", "valuation_method",
    "valuation_indicator", "invoice_number", "nature_of_transaction",
    "statistical_value", "quota_order_number", "supplementary_units",
    "tax_type", "tax_base_unit", "tax_base_quantity",
    "payable_tax_amount", "payable_tax_currency",
}


def parse_synoviaflow_template_excel(path_or_bytes: str | bytes | Path | io.BytesIO) -> tuple[EmailMetadata, ParsedSalesOrders]:
    """Parse the SynoviaFlow Excel upload template.

    Two sheets:
      - 'ENS Header'        — vertical Field/Value form. Hidden column D holds machine keys.
      - 'Consignments+Goods'— row 2 holds machine keys, row 3 human labels, row 4+ data.

    Returns (EmailMetadata, ParsedSalesOrders) so the caller can pass the
    same shape into stage_sales_orders_batch as the Birkdale parser does.
    """
    import openpyxl

    if isinstance(path_or_bytes, (bytes, bytearray)):
        stream: Any = io.BytesIO(bytes(path_or_bytes))
    elif isinstance(path_or_bytes, io.IOBase):
        stream = path_or_bytes
    else:
        stream = str(path_or_bytes)

    wb = openpyxl.load_workbook(stream, data_only=True, read_only=True)

    # ── ENS Header sheet ──
    meta = EmailMetadata()
    if SYNOVIAFLOW_ENS_SHEET in wb.sheetnames:
        ens_ws = wb[SYNOVIAFLOW_ENS_SHEET]
        kv: dict[str, str] = {}
        for row in ens_ws.iter_rows(min_row=3, values_only=True):
            if not row:
                continue
            # Layout: A=label, B=value, C=help, D=machine_key (hidden)
            machine_key = row[3] if len(row) > 3 else None
            value = row[1] if len(row) > 1 else None
            if machine_key and isinstance(machine_key, str):
                kv[machine_key.strip()] = "" if value is None else value

        # Map straight to EmailMetadata
        raw_mov = str(kv.get("movement_type") or "").strip()
        meta.movement_type = (
            MOVEMENT_TYPE_MAP.get(raw_mov.lower())
            or (raw_mov if raw_mov else "")  # accept TSS code already
        )
        raw_passive = str(kv.get("type_of_passive_transport") or "").strip()
        meta.type_of_passive_transport = map_passive_transport(raw_passive) or raw_passive
        if raw_passive and not meta.type_of_passive_transport:
            meta.parse_warnings.append(f"Unmapped passive transport: {raw_passive!r}")
        meta.identity_no_of_transport = str(kv.get("identity_no_of_transport") or "").strip()

        raw_nat = str(kv.get("nationality_of_transport") or "").strip()
        meta.nationality_of_transport = (
            NATIONALITY_MAP.get(raw_nat.lower())
            or (raw_nat if len(raw_nat) == 2 else "")
        )

        meta.carrier_eori = str(kv.get("carrier_eori") or "").strip().upper()
        meta.carrier_name = str(kv.get("carrier_name") or "").strip()
        meta.carrier_street_number = str(kv.get("carrier_street_number") or "").strip()
        meta.carrier_city = str(kv.get("carrier_city") or "").strip()
        meta.carrier_postcode = str(kv.get("carrier_postcode") or "").strip()
        raw_carrier_country = str(kv.get("carrier_country") or "").strip()
        meta.carrier_country = (
            NATIONALITY_MAP.get(raw_carrier_country.lower())
            or (raw_carrier_country.upper() if len(raw_carrier_country) == 2 else "")
        )
        meta.haulier_eori = str(kv.get("haulier_eori") or "").strip().upper()
        meta.conveyance_ref = str(kv.get("conveyance_ref") or "").strip()

        raw_arrival = kv.get("arrival_date_time")
        if isinstance(raw_arrival, datetime):
            meta.arrival_date_time = raw_arrival
        elif raw_arrival:
            meta.arrival_date_time = parse_arrival_datetime(str(raw_arrival), datetime.now(timezone.utc))

        raw_port = str(kv.get("arrival_port") or "").strip()
        meta.arrival_port = (
            ARRIVAL_PORT_HINTS.get(raw_port.lower())
            or (raw_port if raw_port else "")
        )
        meta.place_of_loading = str(kv.get("place_of_loading") or "").strip()
        meta.place_of_acceptance_same_as_loading = _yes_no(kv.get("place_of_acceptance_same_as_loading"))
        meta.place_of_acceptance = str(kv.get("place_of_acceptance") or "").strip()
        meta.place_of_unloading = str(kv.get("place_of_unloading") or "").strip()
        meta.place_of_delivery_same_as_unloading = _yes_no(kv.get("place_of_delivery_same_as_unloading"))
        meta.place_of_delivery = str(kv.get("place_of_delivery") or "").strip()

        raw_charges = str(kv.get("transport_charges") or "").strip()
        meta.transport_charges = (
            TRANSPORT_CHARGES_MAP.get(raw_charges.lower())
            or (raw_charges if raw_charges else "")
        )
    else:
        meta.parse_warnings.append(f"Sheet '{SYNOVIAFLOW_ENS_SHEET}' not found")

    # ── Consignments+Goods sheet ──
    parsed = ParsedSalesOrders()
    if SYNOVIAFLOW_DATA_SHEET not in wb.sheetnames:
        parsed.parse_warnings.append(f"Sheet '{SYNOVIAFLOW_DATA_SHEET}' not found")
        return meta, parsed

    data_ws = wb[SYNOVIAFLOW_DATA_SHEET]
    keys_row = next(data_ws.iter_rows(min_row=2, max_row=2, values_only=True), ())
    keys: list[str] = [str(v).strip() if isinstance(v, str) else "" for v in keys_row]
    if "trader_reference" not in keys:
        parsed.parse_warnings.append("Data sheet header row 2 missing 'trader_reference'")
        return meta, parsed

    consignment_by_doc: dict[str, ParsedConsignment] = {}
    line_counter: dict[str, int] = {}

    for row in data_ws.iter_rows(min_row=4, values_only=True):
        if not row or all(v in (None, "") for v in row):
            continue
        record: dict[str, Any] = {}
        for col_idx, key in enumerate(keys):
            if not key:
                continue
            if col_idx >= len(row):
                continue
            record[key] = row[col_idx]

        trader_ref = (record.get("trader_reference") or "")
        if isinstance(trader_ref, str):
            trader_ref = trader_ref.strip()
        if not trader_ref:
            parsed.parse_warnings.append("Row skipped — empty trader_reference")
            continue
        trader_ref = str(trader_ref)

        cons = consignment_by_doc.get(trader_ref)
        if cons is None:
            cons = ParsedConsignment(document_no=trader_ref)
            cons.sell_to_customer_no = str(record.get("importer_eori") or record.get("consignee_eori") or "")
            cons.ship_to_name = str(record.get("consignee_name") or "")
            cons.ship_to_address = str(record.get("consignee_street_number") or "")
            cons.ship_to_city = str(record.get("consignee_city") or "")
            cons.ship_to_postcode = str(record.get("consignee_postcode") or "")
            cons.ship_to_country = str(record.get("consignee_country") or "")
            cons.ship_to_email = ""
            consignment_by_doc[trader_ref] = cons
            line_counter[trader_ref] = 0

        line_counter[trader_ref] += 1
        line_no = line_counter[trader_ref]

        normalised_code, original_code, action = _normalise_commodity_for_ingest(record.get("commodity_code"))
        if not normalised_code:
            parsed.parse_warnings.append(f"{trader_ref}: row skipped - missing commodity_code")
            continue
        if action != 'unchanged':
            record["commodity_code"] = normalised_code
            parsed.parse_warnings.append(
                f"{trader_ref}: commodity_code '{original_code}' -> '{normalised_code}' ({action}). "
                "Verify the 10-digit code in UK Trade Tariff before submission to TSS."
            )
        sku = normalised_code

        line = ParsedGoodsLine(
            document_no=trader_ref,
            line_no=line_no,
            sku=sku,
            quantity=_to_float(record.get("number_of_packages")),
            quantity_base=_to_float(record.get("number_of_packages")),
            amount=_to_float(record.get("amount")),
            line_amount_excl_vat=_to_float(record.get("item_invoice_amount")),
            unit_price_excl_vat=None,
            qty_per_uom=None,
            uom_code=str(record.get("type_of_packages") or "").strip(),
            raw=dict(record),
        )
        if _is_non_goods_sales_line(line.sku, line.uom_code):
            logger.info(
                "Skipping non-goods sales order line %s/%s during Excel parse",
                trader_ref,
                line.sku,
            )
            continue
        cons.goods.append(line)

    parsed.consignments = list(consignment_by_doc.values())
    if not parsed.consignments:
        parsed.parse_warnings.append("No consignments parsed from data sheet")
    return meta, parsed


def detect_excel_format(path_or_bytes: str | bytes | Path | io.BytesIO) -> str:
    """Return 'synoviaflow' if the workbook has the SynoviaFlow ENS Header
    + Consignments+Goods sheet pair; 'birkdale' otherwise (default fall-back).
    """
    import openpyxl
    if isinstance(path_or_bytes, (bytes, bytearray)):
        stream: Any = io.BytesIO(bytes(path_or_bytes))
    elif isinstance(path_or_bytes, io.IOBase):
        stream = path_or_bytes
    else:
        stream = str(path_or_bytes)
    try:
        wb = openpyxl.load_workbook(stream, data_only=True, read_only=True)
        names = set(wb.sheetnames)
        if SYNOVIAFLOW_ENS_SHEET in names and SYNOVIAFLOW_DATA_SHEET in names:
            return "synoviaflow"
    except Exception:
        pass
    return "birkdale"


def parse_sales_orders_excel(path_or_bytes: str | bytes | Path | io.BytesIO) -> ParsedSalesOrders:
    """Parse a sales-orders Excel workbook into grouped consignments + goods.

    Detects the header row by looking for the cell value
    "Sales Header - Sell-to Customer No." within the first 50 rows of any
    worksheet that contains it. Subsequent non-empty rows are data.
    """
    import openpyxl

    if isinstance(path_or_bytes, (bytes, bytearray)):
        stream = io.BytesIO(bytes(path_or_bytes))
    elif isinstance(path_or_bytes, io.IOBase):
        stream = path_or_bytes
    else:
        stream = str(path_or_bytes)

    wb = openpyxl.load_workbook(stream, data_only=True, read_only=True)
    parsed = ParsedSalesOrders()
    parsed.source_filename = getattr(path_or_bytes, "filename", "") if hasattr(path_or_bytes, "filename") else (
        Path(path_or_bytes).name if isinstance(path_or_bytes, (str, Path)) else ""
    )

    target_label = "Sales Header - Sell-to Customer No."
    consignment_by_doc: dict[str, ParsedConsignment] = {}

    for ws in wb.worksheets:
        header_row_idx = None
        col_map: dict[int, tuple[str, str]] = {}

        # Locate header row: scan first 50 rows; pick the row whose mappable
        # column count is highest. Excel reports include sidecar metadata rows
        # (e.g. "Headers:", "Fields:") that contain label echoes — we only want
        # the actual data header, which has many mapped labels in low cols.
        best_row = None
        best_count = 0
        best_map: dict[int, tuple[str, str]] = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
            cm: dict[int, tuple[str, str]] = {}
            for ci, header in enumerate(row):
                if isinstance(header, str) and header in EXCEL_COLUMN_MAP:
                    cm[ci] = EXCEL_COLUMN_MAP[header]
            # `>=` instead of `>` so ties resolve to the later row. Excel reports
            # often have a sidecar "Headers:" row that echoes labels before the
            # real data header — the real one always comes after.
            if len(cm) >= best_count and len(cm) > 0:
                best_count = len(cm)
                best_row = row_idx
                best_map = cm
        if best_row and best_count >= 5:
            header_row_idx = best_row
            col_map = best_map

        if not header_row_idx:
            continue

        # Iterate data rows
        current_doc_no: str | None = None
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if all(v in (None, "") for v in row):
                continue

            cons_fields: dict[str, Any] = {}
            goods_fields: dict[str, Any] = {}
            raw: dict[str, Any] = {}

            for col_idx, value in enumerate(row):
                if col_idx not in col_map:
                    continue
                target, attr = col_map[col_idx]
                raw[attr] = value
                if target == "consignment":
                    cons_fields[attr] = value
                else:
                    goods_fields[attr] = value

            doc_no = (cons_fields.get("document_no") or "").strip() if isinstance(cons_fields.get("document_no"), str) else cons_fields.get("document_no")
            if not doc_no:
                if _can_fill_down_document_no(goods_fields, current_doc_no):
                    doc_no = current_doc_no
                elif _missing_document_no_needs_warning(goods_fields):
                    parsed.parse_warnings.append("Row skipped — missing Document No.")
                    continue
                else:
                    continue
            else:
                doc_no = str(doc_no)
                current_doc_no = doc_no

            cons = consignment_by_doc.get(doc_no)
            if cons is None:
                cons = ParsedConsignment(document_no=str(doc_no))
                for k, v in cons_fields.items():
                    if k == "document_no":
                        continue
                    if v not in (None, ""):
                        setattr(cons, k, str(v) if isinstance(v, str) else v)
                consignment_by_doc[doc_no] = cons

            goods = ParsedGoodsLine(
                document_no=str(doc_no),
                line_no=_to_int(goods_fields.get("line_no")),
                sku=str(goods_fields.get("sku") or "").strip(),
                quantity=_to_float(goods_fields.get("quantity")),
                quantity_base=_to_float(goods_fields.get("quantity_base")),
                amount=_to_float(goods_fields.get("amount")),
                line_amount_excl_vat=_to_float(goods_fields.get("line_amount_excl_vat")),
                unit_price_excl_vat=_to_float(goods_fields.get("unit_price_excl_vat")),
                qty_per_uom=_to_float(goods_fields.get("qty_per_uom")),
                uom_code=str(goods_fields.get("uom_code") or "").strip(),
                raw=raw,
            )
            if not goods.sku:
                parsed.parse_warnings.append(f"{doc_no}: row skipped — missing SKU")
                continue
            if _is_non_goods_sales_line(goods.sku, goods.uom_code):
                logger.info(
                    "Skipping non-goods sales order line %s/%s during Excel parse",
                    doc_no,
                    goods.sku,
                )
                continue
            cons.goods.append(goods)

        # Stop after first sheet that yielded data
        if consignment_by_doc:
            break

    parsed.consignments = list(consignment_by_doc.values())
    if not parsed.consignments:
        parsed.parse_warnings.append("No consignments parsed — check header row and column names")
    return parsed


def _to_int(v: Any) -> int | None:
    if v in (None, ""):
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _to_float(v: Any) -> float | None:
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ── Product master lookup ─────────────────────────────────────────────


def resolve_product_master(cursor, master_schema: str, customer_code: str, sku: str, barcode: str | None = None) -> dict | None:
    """Resolve product master with priority:

    1. DocProductCatalog WHERE customer_code matches AND sku/barcode matches, with active=1 when present
    2. DocProductCatalog WHERE customer_code='ALL' AND sku/barcode matches, with active=1 when present
    3. Products WHERE stock_code/sku/product_code=sku, with is_active/active=1 when present
    4. None — caller treats as blocker (queue review)
    """
    sku = (sku or "").strip()
    barcode = (barcode or "").strip()

    # 1 + 2: DocProductCatalog
    catalog_row = _query_doc_product_catalog(cursor, master_schema, customer_code, sku, barcode)
    if catalog_row:
        return catalog_row

    # 3: Products fallback
    if sku:
        products_row = _query_products(cursor, master_schema, sku)
        if products_row:
            return products_row

    return None


def _query_doc_product_catalog(cursor, schema: str, customer_code: str, sku: str, barcode: str) -> dict | None:
    if not sku and not barcode:
        return None

    # Schema may not have DocProductCatalog yet — guard
    cursor.execute(
        """
        SELECT 1 FROM INFORMATION_SCHEMA.TABLES
        WHERE TABLE_SCHEMA = ? AND TABLE_NAME = 'DocProductCatalog'
        """,
        [schema],
    )
    if not cursor.fetchone():
        return None

    cursor.execute(
        """
        SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = ? AND TABLE_NAME = 'DocProductCatalog'
        """,
        [schema],
    )
    cols = {r[0].lower() for r in cursor.fetchall()}

    where_parts: list[str] = []
    params: list[Any] = []
    if "active" in cols:
        where_parts.append("active = 1")
    if customer_code and "customer_code" in cols:
        where_parts.append("(customer_code = ? OR customer_code = 'ALL')")
        params.append(customer_code)
    elif "customer_code" in cols:
        where_parts.append("customer_code = 'ALL'")

    matchers: list[str] = []
    for lookup_col in ("sku", "product_code", "stock_code"):
        if sku and lookup_col in cols:
            matchers.append(f"{lookup_col} = ?")
            params.append(sku)
    if barcode and "barcode" in cols:
        matchers.append("barcode = ?")
        params.append(barcode)
    if not matchers:
        return None
    where_parts.append("(" + " OR ".join(matchers) + ")")

    order_clause = ""
    order_params: list[Any] = []
    if "customer_code" in cols:
        order_clause = "ORDER BY CASE WHEN customer_code = ? THEN 0 ELSE 1 END"
        order_params.append(customer_code or "")
        if "id" in cols:
            order_clause += ", id"
    elif "id" in cols:
        order_clause = "ORDER BY id"

    cursor.execute(
        f"""
        SELECT TOP 1 *
        FROM [{schema}].DocProductCatalog
        WHERE {' AND '.join(where_parts)}
        {order_clause}
        """,
        params + order_params,
    )
    return _row_as_dict(cursor)


def _query_products(cursor, schema: str, sku: str) -> dict | None:
    cursor.execute(
        """
        SELECT 1 FROM INFORMATION_SCHEMA.TABLES
        WHERE TABLE_SCHEMA = ? AND TABLE_NAME = 'Products'
        """,
        [schema],
    )
    if not cursor.fetchone():
        return None

    cursor.execute(
        """
        SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = ? AND TABLE_NAME = 'Products'
        """,
        [schema],
    )
    cols = {r[0].lower() for r in cursor.fetchall()}
    active_filter = ""
    if "is_active" in cols:
        active_filter = "AND is_active = 1"
    elif "active" in cols:
        active_filter = "AND active = 1"
    lookup_col = "stock_code" if "stock_code" in cols else (
        "sku" if "sku" in cols else (
            "product_code" if "product_code" in cols else None
        )
    )
    if not lookup_col:
        return None

    cursor.execute(
        f"""
        SELECT TOP 1 *
        FROM [{schema}].Products
        WHERE {lookup_col} = ?
        {active_filter}
        """,
        [sku],
    )
    return _row_as_dict(cursor)


def _row_as_dict(cursor) -> dict | None:
    row = cursor.fetchone()
    if not row:
        return None
    cols = [c[0] for c in cursor.description]
    return dict(zip(cols, row))
