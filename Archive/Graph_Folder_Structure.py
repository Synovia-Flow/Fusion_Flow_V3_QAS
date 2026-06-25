#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Graph_Folder_Structure.py
=========================

Connects to a Microsoft 365 mailbox via the Microsoft Graph API (app-only /
client-credentials flow) and outputs the FULL folder structure - every folder
and sub-folder, recursively - together with the email count held in each.

It reuses the configuration, authentication and folder-enumeration code from
``Graph_Inbox_Analyzer.py`` so both tools read the same Parameters workbook and
behave consistently.

Inputs
------
* Parameters workbook : F:\\Synovia_Flow_Quality\\Documentation_Layer\\Project_Build_Files\\Ingestion_Setup.xlsx
    The "Parameters" tab must contain GRAPH_CLIENT_SECRET / GRAPH_CLIENT_ID /
    GRAPH_TENANT_ID (parameter name in the first column, value in the second).

Output
------
For the target mailbox (default ``nexus@synoviaflow.cloud``) the script:
  * prints an indented folder tree to the console, each line showing the
    folder name and its total / unread email counts, and
  * writes a timestamped report into --output-dir containing:
      - Folder_Structure_<mailbox>_<YYYYMMDD_HHMMSS>.xlsx  (one row per folder:
        indented name, full path, depth, total emails, unread, sub-folder
        count, hidden flag, plus a grand-total row), and
      - Folder_Structure_<mailbox>_<YYYYMMDD_HHMMSS>.txt   (the same tree as a
        plain-text file).

This tool is read-only - it only needs the Mail.Read / Mail.ReadBasic.All
application permissions already declared in Manifest_V2.json.

Usage
-----
    python Graph_Folder_Structure.py
    python Graph_Folder_Structure.py --mailbox someone@synoviaflow.cloud
    python Graph_Folder_Structure.py --output-dir "C:\\path\\out" --no-hidden

Dependencies (see requirements.txt):
    msal, requests, openpyxl
"""

from __future__ import annotations

import argparse
import datetime as _dt
import logging
import os
import sys
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - dependency guidance
    sys.exit("Missing dependency 'openpyxl'. Install with: pip install -r requirements.txt")

# Reuse the shared building blocks so both tools stay consistent.
try:
    from Graph_Inbox_Analyzer import (
        DEFAULT_MAILBOX,
        DEFAULT_MANIFEST,
        DEFAULT_PARAMETERS_SHEET,
        DEFAULT_SETUP_XLSX,
        REQUIRED_PARAMETERS,
        GraphClient,
        acquire_token,
        enumerate_folders,
        load_manifest,
        read_parameters_from_xlsx,
        validate_against_manifest,
    )
except ImportError as exc:  # pragma: no cover - dependency guidance
    sys.exit(
        "Could not import Graph_Inbox_Analyzer.py (it must sit alongside this "
        f"script): {exc}"
    )


DEFAULT_OUTPUT_DIR = r"F:\Synovia_Flow_Quality\Documentation_Layer\Graph"

LOG = logging.getLogger("graph_folder_structure")

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=14, bold=True, color="1F4E78")


# --------------------------------------------------------------------------- #
# Rendering                                                                   #
# --------------------------------------------------------------------------- #
def render_tree(folders: List[Dict[str, Any]], include_hidden: bool = True) -> str:
    """Build an indented, alphabetised text tree of the folder structure.

    Sorting by full path keeps each child directly under its parent and orders
    siblings alphabetically, which makes the tree stable run-to-run.
    """
    visible = [f for f in folders if include_hidden or not f["hidden"]]
    lines: List[str] = []
    for f in sorted(visible, key=lambda x: x["path"].lower()):
        indent = "    " * f["depth"]
        hidden = " [hidden]" if f["hidden"] else ""
        lines.append(
            f"{indent}{f['name']}  "
            f"({f['total']} emails, {f['unread']} unread){hidden}"
        )
    total_emails = sum(f["total"] for f in visible)
    total_unread = sum(f["unread"] for f in visible)
    lines.append("")
    lines.append(
        f"TOTAL: {len(visible)} folders, {total_emails} emails "
        f"({total_unread} unread)"
    )
    return "\n".join(lines)


def _style_header(ws, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, max_width: int = 70) -> None:
    for col_cells in ws.columns:
        length = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def write_structure_xlsx(
    output_path: str,
    mailbox: str,
    folders: List[Dict[str, Any]],
    include_hidden: bool = True,
) -> None:
    """Write a single-sheet workbook listing every folder and its email count."""
    visible = [f for f in folders if include_hidden or not f["hidden"]]

    wb = Workbook()
    ws = wb.active
    ws.title = "Folder Structure"
    ws["A1"] = f"Mailbox folder structure - {mailbox}"
    ws["A1"].font = _TITLE_FONT

    headers = ["Folder", "Full Path", "Depth", "Total Emails", "Unread",
               "Sub-folders", "Hidden"]
    ws.append([])  # spacer row (row 2)
    ws.append(headers)
    _style_header(ws, ws.max_row, len(headers))

    for f in sorted(visible, key=lambda x: x["path"].lower()):
        indent = "    " * f["depth"]
        ws.append([
            f"{indent}{f['name']}",
            f["path"],
            f["depth"],
            f["total"],
            f["unread"],
            f["child_count"],
            "Yes" if f["hidden"] else "No",
        ])

    total_emails = sum(f["total"] for f in visible)
    total_unread = sum(f["unread"] for f in visible)
    ws.append([])
    ws.append(["TOTAL", f"{len(visible)} folders", "", total_emails, total_unread, "", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.freeze_panes = "A4"
    _autosize(ws)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


# --------------------------------------------------------------------------- #
# Orchestration                                                               #
# --------------------------------------------------------------------------- #
def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--setup-xlsx", default=DEFAULT_SETUP_XLSX,
                        help="Path to Ingestion_Setup.xlsx")
    parser.add_argument("--parameters-sheet", default=DEFAULT_PARAMETERS_SHEET,
                        help="Name of the parameters tab")
    parser.add_argument("--manifest", default=DEFAULT_MANIFEST,
                        help="Path to the app manifest JSON")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR,
                        help="Directory for the structure report")
    parser.add_argument("--mailbox", default=DEFAULT_MAILBOX,
                        help="Target mailbox (UPN) to enumerate")
    parser.add_argument("--no-hidden", dest="include_hidden", action="store_false",
                        default=True, help="Exclude hidden folders from the output")
    parser.add_argument("--no-files", dest="write_files", action="store_false",
                        default=True, help="Print the tree only; do not write report files")
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging")
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)-7s %(message)s",
    )

    LOG.info("Reading parameters from %s", args.setup_xlsx)
    params = read_parameters_from_xlsx(args.setup_xlsx, args.parameters_sheet)

    missing = [p for p in REQUIRED_PARAMETERS if not params.get(p)]
    if missing:
        LOG.error("Missing required parameter(s) on the '%s' tab: %s",
                  args.parameters_sheet, ", ".join(missing))
        return 2

    client_id = params["GRAPH_CLIENT_ID"]
    tenant_id = params["GRAPH_TENANT_ID"]
    client_secret = params["GRAPH_CLIENT_SECRET"]

    manifest = load_manifest(args.manifest)
    validate_against_manifest(manifest, client_id)

    token = acquire_token(tenant_id, client_id, client_secret)
    client = GraphClient(token)

    LOG.info("Enumerating mail folders for %s", args.mailbox)
    folders = enumerate_folders(client, args.mailbox)
    LOG.info("Found %d folders.", len(folders))

    tree = render_tree(folders, include_hidden=args.include_hidden)
    print()
    print(f"Folder structure for {args.mailbox}")
    print("=" * (len(args.mailbox) + 21))
    print(tree)
    print()

    if not args.write_files:
        return 0

    timestamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_mailbox = args.mailbox.replace("@", "_at_").replace(".", "_")
    base = f"Folder_Structure_{safe_mailbox}_{timestamp}"
    xlsx_path = os.path.join(args.output_dir, f"{base}.xlsx")
    txt_path = os.path.join(args.output_dir, f"{base}.txt")

    write_structure_xlsx(xlsx_path, args.mailbox, folders, include_hidden=args.include_hidden)
    os.makedirs(os.path.dirname(txt_path), exist_ok=True)
    with open(txt_path, "w", encoding="utf-8") as fh:
        fh.write(f"Folder structure for {args.mailbox}\n")
        fh.write("=" * (len(args.mailbox) + 21) + "\n")
        fh.write(tree + "\n")

    LOG.info("Wrote %s", xlsx_path)
    LOG.info("Wrote %s", txt_path)
    print(xlsx_path)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:  # pragma: no cover
        sys.exit(130)
    except Exception as exc:  # noqa: BLE001 - top-level guard for operator clarity
        LOG.error("Fatal error: %s", exc)
        sys.exit(1)
