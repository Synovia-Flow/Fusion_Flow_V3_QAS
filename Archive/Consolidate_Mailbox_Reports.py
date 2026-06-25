#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Consolidate_Mailbox_Reports.py
==============================

Standalone tool that merges several ``Mailbox_Report_*.xlsx`` workbooks (the
output of Graph_Inbox_Analyzer.py) into a SINGLE consolidated workbook.

Each source report holds the same five sheets - Folder Summary, Inbox Analysis,
Attachments, Processed Messages and Run Info. This tool stacks the row-level
data from every report into combined sheets, tagging each row with the source
report it came from, and builds a one-row-per-report "Runs Summary".

Only openpyxl is required - there is no Microsoft Graph dependency, so it runs
anywhere the report files can be read.

Inputs
------
Either pass report files / folders as positional arguments, or let the tool
discover them with --input-dir + --pattern (default: every
``Mailbox_Report_*.xlsx`` in the Inbound_Stage folder).

Output
------
A timestamped workbook written to --output-dir (default
F:\\Synovia_Flow_Quality\\Documentation_Layer):

    Consolidated_Mailbox_Report_<YYYYMMDD_HHMMSS>.xlsx

with these sheets:
    * Runs Summary       - one row per source report (its Run Info pivoted to
                           columns, plus per-report row counts).
    * Attachments        - every DOWNLOADED attachment row from every report
                           (status == saved only; skipped/image/error "noise"
                           is excluded), prefixed with Source Report + Run
                           Generated.
    * Processed Messages - the processed-message rows that actually yielded a
                           saved file (Files Saved > 0), similarly prefixed.
    * Analysis Prompt    - a ready-to-paste prompt describing the data and
                           asking an assistant how to analyse and sort the
                           downloaded files.

Usage
-----
    python Consolidate_Mailbox_Reports.py
    python Consolidate_Mailbox_Reports.py report1.xlsx report2.xlsx
    python Consolidate_Mailbox_Reports.py --input-dir "C:\\reports" \\
                                          --output-dir "C:\\out"

Dependencies (see requirements.txt):
    openpyxl
"""

from __future__ import annotations

import argparse
import datetime as _dt
import glob
import logging
import os
import re
import sys
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - dependency guidance
    sys.exit("Missing dependency 'openpyxl'. Install with: pip install -r requirements.txt")


DEFAULT_INPUT_DIR = r"F:\Synovia_Flow_Quality\Integration_Layer\BKD\Inbound\Inbound_Stage"
DEFAULT_OUTPUT_DIR = r"F:\Synovia_Flow_Quality\Documentation_Layer"
DEFAULT_PATTERN = "Mailbox_Report_*.xlsx"

# Sheets whose row-level data is stacked across reports: (sheet name, header row).
DATA_SHEETS = ("Attachments", "Processed Messages", "Folder Summary")
HEADER_ROW = 3  # title in row 1, spacer in row 2, headers in row 3 in every report.
RUN_INFO_SHEET = "Run Info"

# Trailing totals / summary rows to drop when reading a data sheet.
_SUMMARY_RE = re.compile(r"^\s*(TOTAL\b|\d+\s+(file|message)\(s\))", re.IGNORECASE)
# Pull the run timestamp out of a report file name, if present.
_NAME_TS_RE = re.compile(r"(\d{8}_\d{6})")

# Addresses that forwarded mail INTO the target mailbox (often before it
# existed), so for their rows the Sender column is the forwarder rather than the
# true origin - the real sender/date sit in the forwarded body headers.
DEFAULT_FORWARDERS = [
    "nexus@synoviaintegration.com",
    "aidan.harrington@synoviadigital.com",
]


def build_analysis_prompt(forwarders: List[str]) -> str:
    """Build the ready-to-paste analysis prompt, noting the forwarder addresses.

    Shown on the "Analysis Prompt" tab so an assistant can profile, de-duplicate
    and propose a sorting scheme for the downloaded files.
    """
    if forwarders:
        forwarder_lines = "\n".join(f"      * {addr}" for addr in forwarders)
        forwarded_section = (
            "\nFORWARDED MAIL (IMPORTANT)\n"
            "- Some messages were bulk-forwarded into this mailbox (in some cases "
            "before the mailbox existed), so for those rows the Sender column is "
            "the FORWARDER, not the real origin. Treat these forwarder addresses "
            "as NOT the true sender:\n"
            f"{forwarder_lines}\n"
            "- For any row from those addresses, the real sender, original sent "
            "date and original subject live in the forwarded email's quoted "
            "headers (the From:/Sent:/To:/Subject: block inside the message "
            "body). Use those original headers - not the forwarder or the "
            "forward date - when profiling, de-duplicating and sorting.\n"
            "- Where the original headers are NOT present in this workbook, list "
            "the affected files separately and flag that we need to re-open those "
            "messages to recover the original headers before filing them.\n"
        )
    else:
        forwarded_section = ""

    return (
        "You are helping me analyse and sort the email file attachments that "
        "were downloaded from a Microsoft 365 mailbox (nexus@synoviaflow.cloud) "
        "and consolidated into this workbook.\n"
        "\n"
        "CONTEXT\n"
        "- The \"Downloaded attachments\" tab lists every file that was actually "
        "saved to the Inbound_Stage folder. Each row has: Source Report, Run "
        "Generated, Received (when the email arrived), Sender, Subject, "
        "Attachment (original file name), Size (bytes), Inline (Yes/No), Saved "
        "As (the on-disk file name), Status.\n"
        "- The \"Processed Messages\" tab lists the source emails that produced "
        "at least one downloaded file: Received, Sender, Subject, Source Folder, "
        "Files Saved, Move Status.\n"
        "- The \"Runs Summary\" tab gives one row per source report run with "
        "totals.\n"
        "- Files are business documents (PDFs, spreadsheets, CSVs, Word docs, "
        "etc.); images were deliberately excluded.\n"
        f"{forwarded_section}"
        "\n"
        "WHAT I WANT FROM YOU\n"
        "1. Profile the data: how many files, total size, date range, and the "
        "breakdown by file type (extension), by sender, and by sender domain "
        "(using the TRUE sender for forwarded mail).\n"
        "2. Identify likely duplicates or re-downloads (same sender + subject + "
        "attachment name + size appearing more than once, e.g. across different "
        "runs) and recommend which copy to keep.\n"
        "3. Propose a clear, consistent folder/taxonomy scheme to SORT these "
        "files into - for example by supplier/sender domain, then document type, "
        "then year-month of the Received date. Justify the scheme and call out "
        "edge cases.\n"
        "4. For each file (or each group), suggest the destination folder path "
        "under that scheme and a normalised, human-readable file name.\n"
        "5. Flag anything that looks like noise, mis-classified, sensitive, or "
        "that a human should review before filing.\n"
        "6. Give me the result as a table I can act on (original Saved As -> "
        "proposed folder -> proposed name -> reason), plus a short summary of the "
        "rules applied.\n"
        "\n"
        "Ask me for any clarification you need about the business meaning of "
        "specific senders or document types before finalising the sorting scheme."
    )


LOG = logging.getLogger("consolidate_mailbox_reports")

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=14, bold=True, color="1F4E78")


# --------------------------------------------------------------------------- #
# Reading source reports                                                      #
# --------------------------------------------------------------------------- #
def _is_blank(row: Tuple[Any, ...]) -> bool:
    return all(c is None or (isinstance(c, str) and not c.strip()) for c in row)


def _is_summary(row: Tuple[Any, ...]) -> bool:
    first = row[0] if row else None
    return isinstance(first, str) and bool(_SUMMARY_RE.match(first))


def extract_table(ws, header_row: int = HEADER_ROW) -> Tuple[List[str], List[List[Any]]]:
    """Return (headers, data_rows) for a report sheet.

    Blank spacer rows and the trailing totals/summary row are dropped, so only
    genuine data rows are returned.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < header_row:
        return [], []
    headers = [("" if h is None else str(h)) for h in rows[header_row - 1]]
    # Trim trailing all-empty header cells.
    while headers and headers[-1] == "":
        headers.pop()
    width = len(headers)

    data: List[List[Any]] = []
    for row in rows[header_row:]:
        if _is_blank(row) or _is_summary(row):
            continue
        data.append(list(row[:width]))
    return headers, data


def extract_run_info(ws) -> Dict[str, str]:
    """Return the Run Info sheet as an ordered key/value dict."""
    info: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=HEADER_ROW, values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        if not key or key.lower() == "run information":
            continue
        value = row[1] if len(row) > 1 else None
        info[key] = "" if value is None else str(value)
    return info


def report_label(path: str, run_info: Dict[str, str]) -> Tuple[str, str]:
    """Return (source_name, run_generated) used to tag every consolidated row."""
    name = os.path.basename(path)
    generated = run_info.get("Generated (local time)", "")
    if not generated:
        m = _NAME_TS_RE.search(name)
        if m:
            ts = m.group(1)
            generated = f"{ts[0:4]}-{ts[4:6]}-{ts[6:8]} {ts[9:11]}:{ts[11:13]}:{ts[13:15]}"
    return name, generated


def read_report(path: str) -> Dict[str, Any]:
    """Read one report workbook into a structured dict."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        # Guard against feeding a previously-consolidated workbook back in: those
        # carry a "Runs Summary" / "Analysis Prompt" sheet and would double-stack.
        if "Runs Summary" in wb.sheetnames or "Analysis Prompt" in wb.sheetnames:
            LOG.warning("Skipping already-consolidated workbook: %s", os.path.basename(path))
            return None
        run_info = extract_run_info(wb[RUN_INFO_SHEET]) if RUN_INFO_SHEET in wb.sheetnames else {}
        name, generated = report_label(path, run_info)
        tables: Dict[str, Tuple[List[str], List[List[Any]]]] = {}
        for sheet in DATA_SHEETS:
            if sheet in wb.sheetnames:
                tables[sheet] = extract_table(wb[sheet])
            else:
                tables[sheet] = ([], [])
        return {
            "path": path,
            "name": name,
            "generated": generated,
            "run_info": run_info,
            "tables": tables,
        }
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# Writing the consolidated workbook                                           #
# --------------------------------------------------------------------------- #
def _style_header(ws, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, max_width: int = 70) -> None:
    for col_cells in ws.columns:
        length = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def _keep_saved_attachment(headers: List[str], row: List[Any]) -> bool:
    """Keep only attachment rows that were actually downloaded (status == saved)."""
    if "Status" not in headers:
        return True
    idx = headers.index("Status")
    val = row[idx] if idx < len(row) else None
    return isinstance(val, str) and val.strip().lower() == "saved"


def _keep_downloaded_message(headers: List[str], row: List[Any]) -> bool:
    """Keep only processed-message rows that yielded at least one saved file."""
    if "Files Saved" not in headers:
        return True
    idx = headers.index("Files Saved")
    val = row[idx] if idx < len(row) else None
    try:
        return int(val) > 0
    except (TypeError, ValueError):
        return False


def _write_data_sheet(ws, title: str, reports: List[Dict[str, Any]], sheet_key: str,
                      row_filter=None, extra_headers=None, enrich=None) -> int:
    """Stack ``sheet_key`` rows from every report; return the total row count.

    ``row_filter(headers, row) -> bool`` optionally drops rows (e.g. attachments
    that were not downloaded). ``extra_headers`` / ``enrich(headers, row) ->
    list`` optionally append derived columns (e.g. the true sender recovered from
    forwarded headers).
    """
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT

    # Use the widest header found across the reports (they should all match).
    base_headers: List[str] = []
    for rpt in reports:
        headers, _ = rpt["tables"].get(sheet_key, ([], []))
        if len(headers) > len(base_headers):
            base_headers = headers

    extra = list(extra_headers or [])
    full_headers = ["Source Report", "Run Generated", *base_headers, *extra]
    ws.append([])  # spacer (row 2)
    ws.append(full_headers)
    _style_header(ws, ws.max_row, len(full_headers))

    total = 0
    for rpt in reports:
        _, data = rpt["tables"].get(sheet_key, ([], []))
        for row in data:
            if row_filter is not None and not row_filter(base_headers, row):
                continue
            padded = list(row) + [None] * (len(base_headers) - len(row))
            extra_vals = list(enrich(base_headers, row)) if enrich else []
            extra_vals += [None] * (len(extra) - len(extra_vals))
            ws.append([rpt["name"], rpt["generated"],
                       *padded[: len(base_headers)], *extra_vals[: len(extra)]])
            total += 1

    ws.append([])
    ws.append([f"{total} row(s) from {len(reports)} report(s)"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.freeze_panes = "A4"
    _autosize(ws)
    return total


def read_forwarded_headers(path: str) -> Tuple[List[str], List[List[Any]], Dict[str, Dict[str, Any]]]:
    """Read a Forwarded_Headers mapping workbook.

    Returns (headers, rows, by_token) where by_token maps each "Msg Token" to a
    dict of its recovered original-header values, for joining onto attachments.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = "Original Headers" if "Original Headers" in wb.sheetnames else wb.sheetnames[0]
        headers, rows = extract_table(wb[sheet])
    finally:
        wb.close()

    idx = {h: i for i, h in enumerate(headers)}

    def cell(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    by_token: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        token = cell(row, "Msg Token")
        if token is None or str(token).strip() == "":
            continue
        by_token[str(token).strip()] = {
            "sender": cell(row, "Original Sender") or cell(row, "Original From") or "",
            "sent": cell(row, "Original Sent") or "",
            "subject": cell(row, "Original Subject") or "",
        }
    return headers, rows, by_token


def _token_from_saved_as(saved_as: Any) -> str:
    """Extract the 8-char message token from a '<stamp>_<token>_<name>' file name."""
    parts = str(saved_as or "").split("_")
    return parts[2] if len(parts) >= 3 else ""


def _make_true_sender_enricher(by_token: Dict[str, Dict[str, Any]]):
    """Build an enrich() that backfills True Sender / Original Sent for forwarded mail."""
    def enrich(base_headers: List[str], row: List[Any]) -> List[Any]:
        if "Saved As" not in base_headers:
            return ["", ""]
        sa = row[base_headers.index("Saved As")] if base_headers.index("Saved As") < len(row) else ""
        rec = by_token.get(_token_from_saved_as(sa))
        if not rec:
            return ["", ""]
        return [rec.get("sender", ""), rec.get("sent", "")]
    return enrich


def _write_headers_tab(ws, headers: List[str], rows: List[List[Any]]) -> None:
    """Write the recovered original headers as their own consolidated tab."""
    ws["A1"] = "Original headers (recovered from forwarded mail)"
    ws["A1"].font = _TITLE_FONT
    ws.append([])  # spacer (row 2)
    ws.append(headers)
    _style_header(ws, ws.max_row, len(headers))
    for row in rows:
        ws.append(list(row))
    ws.append([])
    ws.append([f"{len(rows)} forwarded message(s)"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.freeze_panes = "A4"
    _autosize(ws)


def _write_runs_summary(ws, reports: List[Dict[str, Any]]) -> None:
    ws["A1"] = "Runs summary - consolidated mailbox reports"
    ws["A1"].font = _TITLE_FONT

    # Column order: Source Report, then the union of Run Info keys (first-seen
    # order), then per-report row counts.
    info_keys: List[str] = []
    for rpt in reports:
        for key in rpt["run_info"]:
            if key not in info_keys:
                info_keys.append(key)

    headers = ["Source Report", *info_keys,
               "Attachment rows", "Processed rows", "Folder rows"]
    ws.append([])  # spacer (row 2)
    ws.append(headers)
    _style_header(ws, ws.max_row, len(headers))

    for rpt in reports:
        info = rpt["run_info"]
        tables = rpt["tables"]
        ws.append([
            rpt["name"],
            *[info.get(k, "") for k in info_keys],
            len(tables.get("Attachments", ([], []))[1]),
            len(tables.get("Processed Messages", ([], []))[1]),
            len(tables.get("Folder Summary", ([], []))[1]),
        ])

    ws.append([])
    ws.append([f"{len(reports)} report(s) consolidated"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.freeze_panes = "A4"
    _autosize(ws)


def _write_prompt_sheet(ws, prompt_text: str) -> None:
    """Write a ready-to-paste prompt for analysing/sorting the downloaded files."""
    ws["A1"] = "Analysis prompt - paste into your assistant"
    ws["A1"].font = _TITLE_FONT
    ws.column_dimensions["A"].width = 118
    row = 3
    for line in prompt_text.strip("\n").split("\n"):
        cell = ws.cell(row=row, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1


def write_consolidated(output_path: str, reports: List[Dict[str, Any]],
                       forwarders: Optional[List[str]] = None,
                       forwarded_headers_path: Optional[str] = None) -> Dict[str, int]:
    """Build and save the consolidated workbook; return per-sheet row counts."""
    wb = Workbook()

    _write_runs_summary(wb.active, reports)
    wb.active.title = "Runs Summary"

    # Optional: recovered original headers from forwarded mail. When supplied,
    # backfill True Sender / Original Sent onto the Attachments rows and add an
    # Original Headers tab.
    hdr_headers: List[str] = []
    hdr_rows: List[List[Any]] = []
    enrich = None
    extra_headers = None
    if forwarded_headers_path:
        hdr_headers, hdr_rows, by_token = read_forwarded_headers(forwarded_headers_path)
        enrich = _make_true_sender_enricher(by_token)
        extra_headers = ["True Sender", "Original Sent"]
        LOG.info("Merged %d recovered header record(s) from %s",
                 len(hdr_rows), os.path.basename(forwarded_headers_path))

    counts = {
        "Attachments": _write_data_sheet(
            wb.create_sheet("Attachments"),
            "Downloaded attachments (consolidated)", reports, "Attachments",
            row_filter=_keep_saved_attachment,
            extra_headers=extra_headers, enrich=enrich),
        "Processed Messages": _write_data_sheet(
            wb.create_sheet("Processed Messages"),
            "Processed messages with downloads (consolidated)", reports, "Processed Messages",
            row_filter=_keep_downloaded_message),
    }

    if forwarded_headers_path:
        _write_headers_tab(wb.create_sheet("Original Headers"), hdr_headers, hdr_rows)

    prompt_text = build_analysis_prompt(
        forwarders if forwarders is not None else DEFAULT_FORWARDERS
    )
    _write_prompt_sheet(wb.create_sheet("Analysis Prompt"), prompt_text)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return counts


# --------------------------------------------------------------------------- #
# Input discovery                                                             #
# --------------------------------------------------------------------------- #
def discover_inputs(inputs: List[str], input_dir: str, pattern: str) -> List[str]:
    """Resolve the list of report files from positional args or --input-dir."""
    files: List[str] = []
    if inputs:
        for item in inputs:
            if os.path.isdir(item):
                files.extend(sorted(glob.glob(os.path.join(item, pattern))))
            elif os.path.isfile(item):
                files.append(item)
            else:
                LOG.warning("Input not found, skipping: %s", item)
    else:
        files.extend(sorted(glob.glob(os.path.join(input_dir, pattern))))

    # De-duplicate while preserving order.
    seen: set = set()
    unique: List[str] = []
    for f in files:
        key = os.path.abspath(f)
        if key not in seen:
            seen.add(key)
            unique.append(f)
    return unique


# --------------------------------------------------------------------------- #
# Orchestration                                                               #
# --------------------------------------------------------------------------- #
def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("inputs", nargs="*",
                        help="Report .xlsx files or folders to consolidate. "
                             "If omitted, --input-dir is scanned with --pattern.")
    parser.add_argument("--input-dir", default=DEFAULT_INPUT_DIR,
                        help="Folder scanned for reports when no inputs are given")
    parser.add_argument("--pattern", default=DEFAULT_PATTERN,
                        help="Glob pattern for report files (default: Mailbox_Report_*.xlsx)")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR,
                        help="Directory for the consolidated workbook")
    parser.add_argument("--output-name", default=None,
                        help="Output file name (default: Consolidated_Mailbox_Report_<timestamp>.xlsx)")
    parser.add_argument("--forwarders", default=",".join(DEFAULT_FORWARDERS),
                        help="Comma-separated forwarder addresses whose rows show the "
                             "forwarder, not the true sender. Noted on the Analysis Prompt "
                             "tab. Use '' to omit the forwarded-mail note.")
    parser.add_argument("--forwarded-headers", default=None,
                        help="Optional Forwarded_Headers_*.xlsx mapping (from "
                             "Extract_Forwarded_Headers.py). When given, adds an "
                             "Original Headers tab and backfills True Sender / "
                             "Original Sent onto the Attachments rows.")
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging")
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)-7s %(message)s",
    )

    files = discover_inputs(args.inputs, args.input_dir, args.pattern)
    if not files:
        LOG.error("No report files found. Pass files explicitly or check --input-dir/--pattern.")
        return 2

    LOG.info("Consolidating %d report(s):", len(files))
    reports: List[Dict[str, Any]] = []
    for path in files:
        try:
            rpt = read_report(path)
            if rpt is None:
                continue
            reports.append(rpt)
            tables = rpt["tables"]
            LOG.info("  %s  (att=%d, processed=%d, folders=%d)",
                     rpt["name"],
                     len(tables.get("Attachments", ([], []))[1]),
                     len(tables.get("Processed Messages", ([], []))[1]),
                     len(tables.get("Folder Summary", ([], []))[1]))
        except Exception as exc:  # noqa: BLE001 - keep going on a bad file
            LOG.warning("Could not read %s: %s", path, exc)

    if not reports:
        LOG.error("None of the input files could be read.")
        return 1

    timestamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    name = args.output_name or f"Consolidated_Mailbox_Report_{timestamp}.xlsx"
    output_path = os.path.join(args.output_dir, name)

    forwarders = [a.strip() for a in (args.forwarders or "").split(",") if a.strip()]
    if args.forwarded_headers and not os.path.isfile(args.forwarded_headers):
        LOG.error("Forwarded-headers file not found: %s", args.forwarded_headers)
        return 2
    counts = write_consolidated(output_path, reports, forwarders=forwarders,
                                forwarded_headers_path=args.forwarded_headers)
    LOG.info("Consolidated: %d downloaded-attachment rows, %d processed-message rows.",
             counts["Attachments"], counts["Processed Messages"])
    LOG.info("Saved %s", output_path)
    print(output_path)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:  # pragma: no cover
        sys.exit(130)
    except Exception as exc:  # noqa: BLE001 - top-level guard for operator clarity
        LOG.error("Fatal error: %s", exc)
        sys.exit(1)
