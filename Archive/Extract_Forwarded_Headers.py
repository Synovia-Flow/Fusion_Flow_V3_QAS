#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract_Forwarded_Headers.py
============================

Some mail was bulk-forwarded into the target mailbox (in some cases before it
existed), so for those messages the visible Sender is the FORWARDER, not the
real origin. The true sender / original sent date / original subject live in the
forwarded email's quoted header block ("From:/Sent:/To:/Subject:") inside the
message body.

This tool connects to the mailbox via Microsoft Graph (app-only /
client-credentials flow), looks in the Inbox/Fusion_Processed folder for
messages whose sender is one of the configured forwarder addresses, reads each
message body, parses the original headers out of it, and writes a mapping
workbook. That mapping can then be merged into the consolidated report by
Consolidate_Mailbox_Reports.py --forwarded-headers.

It reuses the configuration, authentication and Graph code from
``Graph_Inbox_Analyzer.py`` so all the tools stay consistent.

Inputs
------
* Parameters workbook : the GRAPH_* credentials on the Parameters tab (same as
  the other tools).

Output
------
A timestamped workbook in --output-dir (default
F:\\Synovia_Flow_Quality\\Documentation_Layer):

    Forwarded_Headers_<mailbox>_<YYYYMMDD_HHMMSS>.xlsx

with a single "Original Headers" sheet, one row per forwarded message:
    Msg Token, Received, Received Stamp, Forwarder, Forward Subject,
    Original From, Original Sender, Original Sent, Original To,
    Original Subject, Parse Status.

"Msg Token" is the same 8-character id token that Graph_Inbox_Analyzer embeds in
each downloaded file's "Saved As" name, so the mapping joins cleanly back to the
consolidated Attachments rows.

Usage
-----
    python Extract_Forwarded_Headers.py
    python Extract_Forwarded_Headers.py --mailbox someone@synoviaflow.cloud
    python Extract_Forwarded_Headers.py --forwarders "a@x.com,b@y.com"

Dependencies (see requirements.txt):
    msal, requests, openpyxl
"""

from __future__ import annotations

import argparse
import datetime as _dt
import html as _html
import logging
import os
import re
import sys
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - dependency guidance
    sys.exit("Missing dependency 'openpyxl'. Install with: pip install -r requirements.txt")

# Reuse the shared building blocks so every tool stays consistent.
try:
    from Graph_Inbox_Analyzer import (
        DEFAULT_MAILBOX,
        DEFAULT_MANIFEST,
        DEFAULT_PARAMETERS_SHEET,
        DEFAULT_PROCESSED_FOLDER,
        DEFAULT_SETUP_XLSX,
        REQUIRED_PARAMETERS,
        GraphClient,
        _parse_dt,
        _sender_address,
        acquire_token,
        load_manifest,
        read_parameters_from_xlsx,
        resolve_inbox_id,
        validate_against_manifest,
    )
except ImportError as exc:  # pragma: no cover - dependency guidance
    sys.exit(
        "Could not import Graph_Inbox_Analyzer.py (it must sit alongside this "
        f"script): {exc}"
    )

# Default forwarder addresses (kept in step with Consolidate_Mailbox_Reports.py).
DEFAULT_FORWARDERS = [
    "nexus@synoviaintegration.com",
    "aidan.harrington@synoviadigital.com",
]
DEFAULT_OUTPUT_DIR = r"F:\Synovia_Flow_Quality\Documentation_Layer"

LOG = logging.getLogger("extract_forwarded_headers")

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=14, bold=True, color="1F4E78")

MESSAGE_LIST_SELECT = "id,subject,from,sender,receivedDateTime"
_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
_TAG_RE = re.compile(r"<[^>]+>")


# --------------------------------------------------------------------------- #
# Body parsing                                                                #
# --------------------------------------------------------------------------- #
def html_to_text(content: str) -> str:
    """Reduce an HTML (or plain) message body to newline-separated text.

    Block-level tags become line breaks so the forwarded "From:/Sent:/..."
    labels land on their own lines, which is what the header parser expects.
    """
    if not content:
        return ""
    text = re.sub(r"(?i)<\s*br\s*/?\s*>", "\n", content)
    text = re.sub(r"(?i)</\s*(p|div|tr|li|h[1-6]|table)\s*>", "\n", text)
    text = _TAG_RE.sub("", text)
    text = _html.unescape(text)
    text = text.replace(" ", " ").replace("\r", "")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n[ \t]*", "\n", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text


def _first_label(label: str, text: str) -> str:
    """Return the first ``Label: value`` line found in ``text`` (case-insensitive)."""
    m = re.search(rf"(?im)^\s*{label}\s*:\s*(.+?)\s*$", text)
    return m.group(1).strip() if m else ""


def parse_forwarded_headers(body_text: str) -> Dict[str, str]:
    """Pull the first forwarded header block (From/Sent/To/Subject) from a body.

    Returns the raw From line plus the email address parsed out of it, the
    original Sent/Date, To and Subject, and a parse status. The FIRST occurrence
    of each label is used - that is the most-recent forwarded original, which is
    what we want when mail was simply forwarded once into the mailbox.
    """
    frm = _first_label("From", body_text)
    sent = _first_label("Sent", body_text) or _first_label("Date", body_text)
    to = _first_label("To", body_text)
    subject = _first_label("Subject", body_text)

    sender = ""
    if frm:
        m = _EMAIL_RE.search(frm)
        sender = m.group(0) if m else ""

    if frm and (sent or subject):
        status = "parsed"
    elif frm or sent or subject:
        status = "partial"
    else:
        status = "not found"

    return {
        "from": frm,
        "sender": sender,
        "sent": sent,
        "to": to,
        "subject": subject,
        "status": status,
    }


def msg_token(msg_id: str) -> str:
    """Reproduce the 8-char id token Graph_Inbox_Analyzer uses in saved names."""
    return re.sub(r"[^A-Za-z0-9]", "", msg_id or "")[-8:] or "msg"


# --------------------------------------------------------------------------- #
# Graph helpers                                                               #
# --------------------------------------------------------------------------- #
def find_child_folder(client: GraphClient, mailbox: str, parent_id: str,
                      name: str) -> Optional[str]:
    """Return the id of child folder ``name`` under ``parent_id``, or None."""
    listing = client.get_all(
        f"/users/{mailbox}/mailFolders/{parent_id}/childFolders",
        params={"$filter": f"displayName eq '{name}'", "$select": "id,displayName", "$top": 10},
    )
    return next((f["id"] for f in listing
                 if (f.get("displayName") or "").lower() == name.lower()), None)


def fetch_body_text(client: GraphClient, mailbox: str, msg_id: str) -> str:
    """Fetch and flatten a single message body to text."""
    data = client.get(
        f"/users/{mailbox}/messages/{msg_id}",
        params={"$select": "id,body"},
    )
    body = data.get("body") or {}
    return html_to_text(body.get("content") or "")


def collect_forwarded_headers(
    client: GraphClient, mailbox: str, folder_id: str, forwarders: List[str]
) -> List[Dict[str, Any]]:
    """Find forwarder messages in ``folder_id`` and recover their original headers."""
    wanted = {a.strip().lower() for a in forwarders if a.strip()}
    messages = client.get_all(
        f"/users/{mailbox}/mailFolders/{folder_id}/messages",
        params={"$select": MESSAGE_LIST_SELECT, "$top": 100},
    )
    LOG.info("Folder holds %d message(s); filtering to %d forwarder address(es).",
             len(messages), len(wanted))

    records: List[Dict[str, Any]] = []
    for msg in messages:
        address = (_sender_address(msg) or "").lower()
        if address not in wanted:
            continue
        received = _parse_dt(msg.get("receivedDateTime"))
        stamp = received.strftime("%Y%m%d_%H%M%S") if received else "00000000_000000"
        try:
            body_text = fetch_body_text(client, mailbox, msg["id"])
            parsed = parse_forwarded_headers(body_text)
        except RuntimeError as exc:
            LOG.warning("Could not read body for %s: %s", msg_token(msg["id"]), exc)
            parsed = {"from": "", "sender": "", "sent": "", "to": "",
                      "subject": "", "status": f"error: {exc}"}
        records.append({
            "token": msg_token(msg["id"]),
            "received": received.isoformat() if received else "",
            "stamp": stamp,
            "forwarder": address,
            "forward_subject": msg.get("subject") or "",
            "orig_from": parsed["from"],
            "orig_sender": parsed["sender"],
            "orig_sent": parsed["sent"],
            "orig_to": parsed["to"],
            "orig_subject": parsed["subject"],
            "status": parsed["status"],
        })

    LOG.info("Recovered headers for %d forwarded message(s).", len(records))
    return records


# --------------------------------------------------------------------------- #
# Output                                                                      #
# --------------------------------------------------------------------------- #
HEADERS = ["Msg Token", "Received", "Received Stamp", "Forwarder",
           "Forward Subject", "Original From", "Original Sender",
           "Original Sent", "Original To", "Original Subject", "Parse Status"]


def _style_header(ws, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, max_width: int = 70) -> None:
    for col_cells in ws.columns:
        length = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def write_headers_workbook(output_path: str, mailbox: str,
                           records: List[Dict[str, Any]]) -> None:
    """Write the recovered-headers mapping workbook (one 'Original Headers' sheet)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Original Headers"
    ws["A1"] = f"Forwarded original headers - {mailbox}"
    ws["A1"].font = _TITLE_FONT
    ws.append([])  # spacer (row 2)
    ws.append(HEADERS)
    _style_header(ws, ws.max_row, len(HEADERS))

    for rec in records:
        ws.append([
            rec["token"], rec["received"], rec["stamp"], rec["forwarder"],
            rec["forward_subject"], rec["orig_from"], rec["orig_sender"],
            rec["orig_sent"], rec["orig_to"], rec["orig_subject"], rec["status"],
        ])

    parsed = sum(1 for r in records if r["status"] == "parsed")
    ws.append([])
    ws.append([f"{len(records)} message(s) with recovered headers", "",
               "", "", "", "", "", "", "", "", f"{parsed} fully parsed"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.freeze_panes = "A4"
    _autosize(ws)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


# --------------------------------------------------------------------------- #
# Orchestration                                                               #
# --------------------------------------------------------------------------- #
def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--setup-xlsx", default=DEFAULT_SETUP_XLSX,
                        help="Path to Ingestion_Setup.xlsx")
    parser.add_argument("--parameters-sheet", default=DEFAULT_PARAMETERS_SHEET,
                        help="Name of the parameters tab")
    parser.add_argument("--manifest", default=DEFAULT_MANIFEST,
                        help="Path to the app manifest JSON")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR,
                        help="Directory for the headers mapping workbook")
    parser.add_argument("--mailbox", default=DEFAULT_MAILBOX,
                        help="Target mailbox (UPN)")
    parser.add_argument("--processed-folder", default=DEFAULT_PROCESSED_FOLDER,
                        help="Folder beneath the Inbox to search (default: Fusion_Processed)")
    parser.add_argument("--forwarders", default=",".join(DEFAULT_FORWARDERS),
                        help="Comma-separated forwarder addresses to match")
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging")
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)-7s %(message)s",
    )

    forwarders = [a.strip() for a in (args.forwarders or "").split(",") if a.strip()]
    if not forwarders:
        LOG.error("No forwarder addresses given (--forwarders).")
        return 2

    LOG.info("Reading parameters from %s", args.setup_xlsx)
    params = read_parameters_from_xlsx(args.setup_xlsx, args.parameters_sheet)
    missing = [p for p in REQUIRED_PARAMETERS if not params.get(p)]
    if missing:
        LOG.error("Missing required parameter(s) on the '%s' tab: %s",
                  args.parameters_sheet, ", ".join(missing))
        return 2

    manifest = load_manifest(args.manifest)
    validate_against_manifest(manifest, params["GRAPH_CLIENT_ID"])

    token = acquire_token(params["GRAPH_TENANT_ID"], params["GRAPH_CLIENT_ID"],
                          params["GRAPH_CLIENT_SECRET"])
    client = GraphClient(token)

    inbox_id = resolve_inbox_id(client, args.mailbox)
    folder_id = find_child_folder(client, args.mailbox, inbox_id, args.processed_folder)
    if not folder_id:
        LOG.error("Folder 'Inbox/%s' not found in %s.", args.processed_folder, args.mailbox)
        return 1

    LOG.info("Searching Inbox/%s for forwarder messages.", args.processed_folder)
    records = collect_forwarded_headers(client, args.mailbox, folder_id, forwarders)

    timestamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_mailbox = args.mailbox.replace("@", "_at_").replace(".", "_")
    output_path = os.path.join(
        args.output_dir, f"Forwarded_Headers_{safe_mailbox}_{timestamp}.xlsx")

    write_headers_workbook(output_path, args.mailbox, records)
    LOG.info("Saved %s", output_path)
    print(output_path)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:  # pragma: no cover
        sys.exit(130)
    except Exception as exc:  # noqa: BLE001 - top-level guard for operator clarity
        LOG.error("Fatal error: %s", exc)
        sys.exit(1)
