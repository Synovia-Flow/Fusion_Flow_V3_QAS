#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Graph_Inbox_Analyzer.py
=======================

Connects to a Microsoft 365 mailbox via the Microsoft Graph API (app-only /
client-credentials flow), enumerates every mail folder together with the number
of emails it holds, performs an analysis of the Inbox, and writes the results to
an Excel workbook.

Inputs
------
* Parameters workbook : F:\\Synovia_Flow_Quality\\Documentation_Layer\\Project_Build_Files\\Ingestion_Setup.xlsx
    The "Parameters" tab must contain the following key/value rows
    (parameter name in the first column, value in the second):
        GRAPH_CLIENT_SECRET
        GRAPH_CLIENT_ID
        GRAPH_TENANT_ID
* App manifest        : E:\\Git\\FLow_3_1\\Flow_3_1_Development\\Development_Tools\\Config\\Manifest_V2.json
    Used to validate that the client id in the workbook matches the registered
    application ("Fusion Flow Mail Reader") and to confirm the requested Graph
    permissions before any call is made.

Output
------
* Excel report        : F:\\Synovia_Flow_Quality\\Integration_Layer\\BKD\\Inbound\\Inbound_Stage
    A timestamped workbook (Mailbox_Report_<mailbox>_<YYYYMMDD_HHMMSS>.xlsx)
    written alongside the downloaded attachments, containing:
        - Folder Summary  : every folder (recursively), full path, total &
                            unread email counts.
        - Inbox Analysis  : headline inbox metrics (totals, read/unread, top
                            senders, oldest/newest, volume by day).
        - Attachments     : every attachment downloaded (message, sender,
                            file name, size, saved path).
        - Processed Messages : each scanned message, its source folder,
                            files saved, and move status.
        - Run Info        : metadata about the run (mailbox, app, timestamp).
* Attachments         : F:\\Synovia_Flow_Quality\\Integration_Layer\\BKD\\Inbound\\Inbound_Stage
    Downloaded attachments are saved here with a collision-safe name
    (<received>_<msgid>_<original name>). ALL attachments are kept EXCEPT images
    (png/jpg/gif/bmp/tiff/webp/svg/heic and anything with an image/* content
    type). Use --exclude-extensions to change the excluded set.

Folder workflow (default)
-------------------------
The tool applies ONE consistent rule across the whole mailbox: it scans every
folder - top-level folders outside the Inbox and Inbox sub-folders alike -
EXCEPT the Inbox root itself (its sub-folders are still scanned) and the
excluded system folders (Sent Items, Drafts, Deleted Items, Junk Email, Outbox,
Archive, ... and everything beneath them). From the messages it finds it
downloads every non-image attachment, then MOVES each scanned message into a
Fusion_Processed folder created directly beneath the Inbox (--processed-folder),
creating it if needed. Moving requires the Mail.ReadWrite application permission
(already declared in Manifest_V2.json). Use --no-move-processed to download
without moving, or --only-with-attachments to move only the messages that
yielded a saved file.

The target mailbox defaults to ``nexus@synoviaflow.cloud`` and can be overridden
with ``--mailbox``.

This tool is read-only: it requires the Mail.Read / Mail.ReadBasic.All
*application* permissions that are already declared in Manifest_V2.json.

Usage
-----
    python Graph_Inbox_Analyzer.py
    python Graph_Inbox_Analyzer.py --mailbox someone@synoviaflow.cloud
    python Graph_Inbox_Analyzer.py --setup-xlsx "C:\\path\\Ingestion_Setup.xlsx" \\
                                   --output-dir "C:\\path\\out" --inbox-sample 1000

Dependencies (see requirements.txt):
    msal, requests, openpyxl
"""

from __future__ import annotations

import argparse
import base64
import collections
import datetime as _dt
import json
import logging
import os
import re
import sys
import time
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    import requests
except ImportError:  # pragma: no cover - dependency guidance
    sys.exit("Missing dependency 'requests'. Install with: pip install -r requirements.txt")

try:
    import msal
except ImportError:  # pragma: no cover - dependency guidance
    sys.exit("Missing dependency 'msal'. Install with: pip install -r requirements.txt")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - dependency guidance
    sys.exit("Missing dependency 'openpyxl'. Install with: pip install -r requirements.txt")


# --------------------------------------------------------------------------- #
# Defaults - these mirror the deployment environment described in the request. #
# Every value can be overridden on the command line.                          #
# --------------------------------------------------------------------------- #
DEFAULT_SETUP_XLSX = r"F:\Synovia_Flow_Quality\Documentation_Layer\Project_Build_Files\Ingestion_Setup.xlsx"
DEFAULT_MANIFEST = r"E:\Git\FLow_3_1\Flow_3_1_Development\Development_Tools\Config\Manifest_V2.json"
DEFAULT_OUTPUT_DIR = r"F:\Synovia_Flow_Quality\Integration_Layer\BKD\Inbound\Inbound_Stage"
DEFAULT_ATTACHMENTS_DIR = r"F:\Synovia_Flow_Quality\Integration_Layer\BKD\Inbound\Inbound_Stage"
DEFAULT_MAILBOX = "nexus@synoviaflow.cloud"
DEFAULT_PARAMETERS_SHEET = "Parameters"
DEFAULT_INBOX_SAMPLE = 2000  # max inbox messages to pull for the analysis sheet

# Processing workflow: scan every folder in the mailbox (top-level folders and
# Inbox sub-folders alike) except the Inbox root and the excluded system
# folders, download all non-image attachments, then move each scanned message
# into a Fusion_Processed folder created directly beneath the Inbox.
DEFAULT_PROCESSED_FOLDER = "Fusion_Processed"
# Image attachments are excluded by extension (and by image/* content type);
# every other file type is downloaded.
DEFAULT_EXCLUDED_EXTENSIONS = "png,jpg,jpeg,gif,bmp,tif,tiff,webp,svg,svgz,heic,heif,ico"
# Well-known system folders never scanned or moved (matched case-insensitively
# against any segment of a folder's path, so their sub-folders are pruned too).
EXCLUDED_FOLDER_NAMES = {
    "sent items", "drafts", "deleted items", "junk email", "junk e-mail",
    "outbox", "archive", "conversation history", "sync issues", "rss feeds",
}

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
GRAPH_SCOPE = ["https://graph.microsoft.com/.default"]
AUTHORITY_TEMPLATE = "https://login.microsoftonline.com/{tenant}"

# Microsoft Graph well-known resource app id, used to read the manifest's
# requiredResourceAccess so we can report the granted permissions.
GRAPH_RESOURCE_APP_ID = "00000003-0000-0000-c000-000000000000"

REQUIRED_PARAMETERS = ("GRAPH_CLIENT_SECRET", "GRAPH_CLIENT_ID", "GRAPH_TENANT_ID")

# Friendly names for the Graph permission ids that may appear in the manifest.
GRAPH_PERMISSION_NAMES = {
    "e1fe6dd8-ba31-4d61-89e7-88639da4683d": "User.Read",
    "810c84a8-4a9e-49e6-bf7d-12d183f40d01": "Mail.Read",
    "e2a3a72e-5f79-4c64-b1b1-878b674786c9": "Mail.ReadWrite",
    "6be147d2-ea4f-4b5a-a3fa-3eab6f3c140a": "Mail.ReadBasic.All",
}

LOG = logging.getLogger("graph_inbox_analyzer")


# --------------------------------------------------------------------------- #
# Configuration loading                                                       #
# --------------------------------------------------------------------------- #
def read_parameters_from_xlsx(path: str, sheet_name: str) -> Dict[str, str]:
    """Read the key/value pairs from the Parameters tab of the setup workbook.

    The tab is expected to hold a parameter name in the first column and its
    value in the second. Matching of the sheet name is case-insensitive and
    tolerant of the common 'Paramters' misspelling.
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Parameters workbook not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        target = _resolve_sheet(wb.sheetnames, sheet_name)
        if target is None:
            raise ValueError(
                f"Could not find a '{sheet_name}' tab in {path}. "
                f"Available tabs: {', '.join(wb.sheetnames)}"
            )
        ws = wb[target]

        params: Dict[str, str] = {}
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            if not key:
                continue
            value = row[1] if len(row) > 1 else None
            params[key] = "" if value is None else str(value).strip()
        return params
    finally:
        wb.close()


def _resolve_sheet(sheet_names: Iterable[str], wanted: str) -> Optional[str]:
    """Find a worksheet name, ignoring case and the 'Paramters' typo."""
    wanted_norm = wanted.strip().lower()
    candidates = {wanted_norm, "parameters", "paramters", "params"}
    for name in sheet_names:
        if name.strip().lower() in candidates:
            return name
    # Fall back to any tab that starts with 'param'.
    for name in sheet_names:
        if name.strip().lower().startswith("param"):
            return name
    return None


def load_manifest(path: str) -> Optional[Dict[str, Any]]:
    """Load the app manifest if present; missing/invalid manifests are non-fatal."""
    if not os.path.isfile(path):
        LOG.warning("Manifest not found at %s - continuing without manifest validation.", path)
        return None
    try:
        with open(path, "r", encoding="utf-8-sig") as fh:
            return json.load(fh)
    except (OSError, json.JSONDecodeError) as exc:
        LOG.warning("Could not parse manifest %s (%s) - continuing.", path, exc)
        return None


def validate_against_manifest(manifest: Optional[Dict[str, Any]], client_id: str) -> List[str]:
    """Cross-check the workbook client id with the manifest and list permissions.

    Returns a human-readable list of the Graph permissions declared in the
    manifest. A client-id mismatch is logged as a warning (not fatal) so the
    operator is alerted but the run can still proceed if intended.
    """
    permissions: List[str] = []
    if not manifest:
        return permissions

    manifest_app_id = manifest.get("appId")
    if manifest_app_id and client_id and manifest_app_id.lower() != client_id.lower():
        LOG.warning(
            "Client id in workbook (%s) does not match manifest appId (%s).",
            client_id,
            manifest_app_id,
        )
    else:
        LOG.info("Client id matches manifest app '%s'.", manifest.get("displayName", "<unknown>"))

    for resource in manifest.get("requiredResourceAccess", []) or []:
        if resource.get("resourceAppId") != GRAPH_RESOURCE_APP_ID:
            continue
        for access in resource.get("resourceAccess", []) or []:
            name = GRAPH_PERMISSION_NAMES.get(access.get("id"), access.get("id"))
            permissions.append(f"{name} ({access.get('type')})")
    return permissions


# --------------------------------------------------------------------------- #
# Authentication                                                              #
# --------------------------------------------------------------------------- #
def acquire_token(tenant_id: str, client_id: str, client_secret: str) -> str:
    """Acquire an app-only access token via MSAL client-credentials flow."""
    authority = AUTHORITY_TEMPLATE.format(tenant=tenant_id)
    app = msal.ConfidentialClientApplication(
        client_id=client_id,
        client_credential=client_secret,
        authority=authority,
    )
    LOG.info("Requesting application token from %s", authority)
    result = app.acquire_token_for_client(scopes=GRAPH_SCOPE)
    if "access_token" not in result:
        raise RuntimeError(
            "Failed to acquire token: "
            f"{result.get('error')} - {result.get('error_description')}"
        )
    LOG.info("Token acquired successfully.")
    return result["access_token"]


# --------------------------------------------------------------------------- #
# Graph client                                                                #
# --------------------------------------------------------------------------- #
class GraphClient:
    """Thin Microsoft Graph REST wrapper with paging and throttling handling."""

    def __init__(self, token: str, timeout: int = 60) -> None:
        self._session = requests.Session()
        self._session.headers.update(
            {"Authorization": f"Bearer {token}", "Accept": "application/json"}
        )
        self._timeout = timeout

    def get(self, url: str, params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """GET a single Graph resource, honouring 429/503 Retry-After backoff."""
        full_url = url if url.startswith("http") else f"{GRAPH_BASE}{url}"
        for attempt in range(6):
            resp = self._session.get(full_url, params=params, timeout=self._timeout)
            if resp.status_code in (429, 503):
                wait = int(resp.headers.get("Retry-After", 2 ** attempt))
                LOG.warning("Throttled (%s). Waiting %ss before retry.", resp.status_code, wait)
                time.sleep(wait)
                continue
            if resp.status_code >= 400:
                raise RuntimeError(
                    f"Graph request failed ({resp.status_code}) for {full_url}: {resp.text}"
                )
            return resp.json()
        raise RuntimeError(f"Graph request kept being throttled: {full_url}")

    def get_all(self, url: str, params: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """GET a collection, following @odata.nextLink until exhausted."""
        items: List[Dict[str, Any]] = []
        page = self.get(url, params=params)
        items.extend(page.get("value", []))
        next_link = page.get("@odata.nextLink")
        while next_link:
            page = self.get(next_link)
            items.extend(page.get("value", []))
            next_link = page.get("@odata.nextLink")
        return items

    def get_bytes(self, url: str) -> bytes:
        """GET the raw bytes of a resource (e.g. an attachment's /$value)."""
        full_url = url if url.startswith("http") else f"{GRAPH_BASE}{url}"
        for attempt in range(6):
            resp = self._session.get(full_url, timeout=self._timeout)
            if resp.status_code in (429, 503):
                wait = int(resp.headers.get("Retry-After", 2 ** attempt))
                LOG.warning("Throttled (%s). Waiting %ss before retry.", resp.status_code, wait)
                time.sleep(wait)
                continue
            if resp.status_code >= 400:
                raise RuntimeError(
                    f"Graph byte request failed ({resp.status_code}) for {full_url}: {resp.text}"
                )
            return resp.content
        raise RuntimeError(f"Graph byte request kept being throttled: {full_url}")

    def post(self, url: str, json_body: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """POST to a Graph endpoint (e.g. move a message, create a folder)."""
        full_url = url if url.startswith("http") else f"{GRAPH_BASE}{url}"
        for attempt in range(6):
            resp = self._session.post(full_url, json=json_body, timeout=self._timeout)
            if resp.status_code in (429, 503):
                wait = int(resp.headers.get("Retry-After", 2 ** attempt))
                LOG.warning("Throttled (%s). Waiting %ss before retry.", resp.status_code, wait)
                time.sleep(wait)
                continue
            if resp.status_code >= 400:
                raise RuntimeError(
                    f"Graph POST failed ({resp.status_code}) for {full_url}: {resp.text}"
                )
            return resp.json() if resp.content else {}
        raise RuntimeError(f"Graph POST kept being throttled: {full_url}")


# --------------------------------------------------------------------------- #
# Mailbox enumeration                                                         #
# --------------------------------------------------------------------------- #
FOLDER_SELECT = "id,displayName,parentFolderId,childFolderCount,unreadItemCount,totalItemCount,isHidden"


def enumerate_folders(client: GraphClient, mailbox: str) -> List[Dict[str, Any]]:
    """Recursively walk the mailbox folder tree, returning one record per folder.

    Each record carries the full folder path and email counts. Hidden folders
    are included so the report reflects the true mailbox layout.
    """
    folders: List[Dict[str, Any]] = []

    def walk(folder_id: str, parent_path: str, depth: int) -> None:
        url = f"/users/{mailbox}/mailFolders/{folder_id}/childFolders"
        children = client.get_all(
            url, params={"includeHiddenFolders": "true", "$select": FOLDER_SELECT, "$top": 100}
        )
        for child in children:
            name = child.get("displayName", "<unnamed>")
            path = f"{parent_path}/{name}" if parent_path else name
            folders.append(_folder_record(child, path, depth))
            if (child.get("childFolderCount") or 0) > 0:
                walk(child["id"], path, depth + 1)

    top = client.get_all(
        f"/users/{mailbox}/mailFolders",
        params={"includeHiddenFolders": "true", "$select": FOLDER_SELECT, "$top": 100},
    )
    for folder in top:
        name = folder.get("displayName", "<unnamed>")
        folders.append(_folder_record(folder, name, 0))
        if (folder.get("childFolderCount") or 0) > 0:
            walk(folder["id"], name, 1)

    return folders


def _folder_record(folder: Dict[str, Any], path: str, depth: int) -> Dict[str, Any]:
    return {
        "id": folder.get("id"),
        "name": folder.get("displayName", "<unnamed>"),
        "path": path,
        "depth": depth,
        "total": int(folder.get("totalItemCount") or 0),
        "unread": int(folder.get("unreadItemCount") or 0),
        "child_count": int(folder.get("childFolderCount") or 0),
        "hidden": bool(folder.get("isHidden")),
    }


INBOX_MESSAGE_SELECT = "id,subject,from,receivedDateTime,isRead,hasAttachments,importance"


def fetch_inbox_messages(client: GraphClient, mailbox: str, sample: int) -> List[Dict[str, Any]]:
    """Pull up to ``sample`` most-recent Inbox messages for analysis."""
    messages: List[Dict[str, Any]] = []
    params = {
        "$select": INBOX_MESSAGE_SELECT,
        "$orderby": "receivedDateTime desc",
        "$top": 100,
    }
    url = f"/users/{mailbox}/mailFolders/inbox/messages"
    page = client.get(url, params=params)
    messages.extend(page.get("value", []))
    next_link = page.get("@odata.nextLink")
    while next_link and len(messages) < sample:
        page = client.get(next_link)
        messages.extend(page.get("value", []))
        next_link = page.get("@odata.nextLink")
    return messages[:sample]


def analyse_inbox(messages: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Compute headline metrics over a list of Inbox messages."""
    total = len(messages)
    unread = sum(1 for m in messages if not m.get("isRead", True))
    with_attachments = sum(1 for m in messages if m.get("hasAttachments"))
    high_importance = sum(1 for m in messages if m.get("importance") == "high")

    senders: "collections.Counter[str]" = collections.Counter()
    by_day: "collections.Counter[str]" = collections.Counter()
    received_times: List[_dt.datetime] = []

    for m in messages:
        sender = _sender_address(m)
        if sender:
            senders[sender] += 1
        received = _parse_dt(m.get("receivedDateTime"))
        if received:
            received_times.append(received)
            by_day[received.date().isoformat()] += 1

    oldest = min(received_times).isoformat() if received_times else None
    newest = max(received_times).isoformat() if received_times else None

    return {
        "total": total,
        "read": total - unread,
        "unread": unread,
        "with_attachments": with_attachments,
        "high_importance": high_importance,
        "unique_senders": len(senders),
        "top_senders": senders.most_common(25),
        "by_day": sorted(by_day.items()),
        "oldest": oldest,
        "newest": newest,
    }


def _sender_address(message: Dict[str, Any]) -> Optional[str]:
    frm = (message.get("from") or {}).get("emailAddress") or {}
    return frm.get("address") or frm.get("name")


def _parse_dt(value: Optional[str]) -> Optional[_dt.datetime]:
    if not value:
        return None
    try:
        return _dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


# --------------------------------------------------------------------------- #
# Attachment download                                                         #
# --------------------------------------------------------------------------- #
FILE_ATTACHMENT_TYPE = "#microsoft.graph.fileAttachment"
# Metadata only - contentBytes is excluded for performance and fetched via /$value.
ATTACHMENT_SELECT = "id,name,contentType,size,isInline,lastModifiedDateTime"
_INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def _sanitize_filename(name: str) -> str:
    """Make a string safe to use as a Windows file name."""
    cleaned = _INVALID_FILENAME_CHARS.sub("_", (name or "").strip())
    cleaned = cleaned.rstrip(". ")  # Windows disallows trailing dots/spaces
    return cleaned or "attachment"


def _unique_path(directory: str, filename: str) -> str:
    """Return a path in ``directory`` for ``filename`` that does not yet exist."""
    candidate = os.path.join(directory, filename)
    if not os.path.exists(candidate):
        return candidate
    stem, ext = os.path.splitext(filename)
    index = 1
    while True:
        candidate = os.path.join(directory, f"{stem} ({index}){ext}")
        if not os.path.exists(candidate):
            return candidate
        index += 1


def download_attachments(
    client: GraphClient,
    mailbox: str,
    messages: List[Dict[str, Any]],
    dest_dir: str,
    include_inline: bool = True,
    excluded_exts: Optional[Iterable[str]] = None,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Download every non-image file attachment on the given messages to ``dest_dir``.

    All file attachments are saved EXCEPT images: anything whose extension is in
    ``excluded_exts`` (png/jpg/gif/... by default) or whose content type starts
    with ``image/`` is recorded as skipped. Pass ``excluded_exts=None`` to keep
    every file type.

    Returns a tuple of (per-attachment records, summary). Item and reference
    attachments are recorded but not downloaded (they carry no file bytes).
    Each saved file is named ``<received>_<msgid8>_<original name>`` to stay
    unique across messages, and any remaining clash is de-duplicated with a
    numeric suffix.
    """
    os.makedirs(dest_dir, exist_ok=True)
    excluded = {e.strip().lower().lstrip(".") for e in excluded_exts} if excluded_exts else set()
    records: List[Dict[str, Any]] = []
    summary = {
        "saved": 0,
        "bytes": 0,
        "skipped_inline": 0,
        "skipped_nonfile": 0,
        "skipped_images": 0,
        "errors": 0,
        "messages_with_attachments": 0,
    }

    for message in messages:
        if not message.get("hasAttachments"):
            continue
        summary["messages_with_attachments"] += 1
        msg_id = message.get("id", "")
        received = _parse_dt(message.get("receivedDateTime"))
        received_stamp = received.strftime("%Y%m%d_%H%M%S") if received else "00000000_000000"
        msg_token = re.sub(r"[^A-Za-z0-9]", "", msg_id)[-8:] or "msg"
        sender = _sender_address(message) or "unknown"
        subject = message.get("subject") or "(no subject)"

        try:
            attachments = client.get_all(
                f"/users/{mailbox}/messages/{msg_id}/attachments",
                params={"$select": ATTACHMENT_SELECT, "$top": 50},
            )
        except RuntimeError as exc:
            LOG.warning("Could not list attachments for message %s: %s", msg_token, exc)
            summary["errors"] += 1
            continue

        for att in attachments:
            name = att.get("name") or "attachment"
            is_inline = bool(att.get("isInline"))
            odata_type = att.get("@odata.type", "")
            content_type = (att.get("contentType") or "").lower()
            ext = os.path.splitext(name)[1].lower().lstrip(".")

            def record(size, saved_as, status):
                return _attachment_record(msg_id, received, sender, subject, name,
                                          size, is_inline, saved_as, status)

            if odata_type != FILE_ATTACHMENT_TYPE:
                summary["skipped_nonfile"] += 1
                records.append(record(att.get("size"), "", f"skipped ({odata_type or 'non-file'})"))
                continue
            if ext in excluded or content_type.startswith("image/"):
                summary["skipped_images"] += 1
                records.append(record(att.get("size"), "", f"skipped (image: .{ext or '?'})"))
                continue
            if is_inline and not include_inline:
                summary["skipped_inline"] += 1
                records.append(record(att.get("size"), "", "skipped (inline)"))
                continue

            safe_name = _sanitize_filename(f"{received_stamp}_{msg_token}_{name}")
            target = _unique_path(dest_dir, safe_name)
            try:
                data = _attachment_bytes(client, mailbox, msg_id, att)
                with open(target, "wb") as fh:
                    fh.write(data)
                summary["saved"] += 1
                summary["bytes"] += len(data)
                records.append(record(len(data), os.path.basename(target), "saved"))
                LOG.info("Saved attachment %s (%d bytes)", os.path.basename(target), len(data))
            except (RuntimeError, OSError) as exc:
                summary["errors"] += 1
                LOG.warning("Failed to save attachment '%s' from message %s: %s",
                            name, msg_token, exc)
                records.append(record(att.get("size"), "", f"error: {exc}"))

    LOG.info("Attachments: %d saved (%s), %d images skipped, %d inline skipped, %d non-file skipped, %d errors.",
             summary["saved"], _human_bytes(summary["bytes"]), summary["skipped_images"],
             summary["skipped_inline"], summary["skipped_nonfile"], summary["errors"])
    return records, summary


def _attachment_bytes(client: GraphClient, mailbox: str, msg_id: str, att: Dict[str, Any]) -> bytes:
    """Get the file bytes for an attachment, preferring inline contentBytes."""
    content_b64 = att.get("contentBytes")
    if content_b64:
        return base64.b64decode(content_b64)
    return client.get_bytes(f"/users/{mailbox}/messages/{msg_id}/attachments/{att['id']}/$value")


def _attachment_record(msg_id, received, sender, subject, name, size, is_inline, saved_as, status):
    return {
        "msg_id": msg_id,
        "received": received.isoformat() if received else "",
        "sender": sender,
        "subject": subject,
        "name": name,
        "size": int(size) if isinstance(size, int) else (size or ""),
        "inline": "Yes" if is_inline else "No",
        "saved_as": saved_as,
        "status": status,
    }


def _human_bytes(num: int) -> str:
    value = float(num)
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024
    return f"{value:.1f} GB"


# --------------------------------------------------------------------------- #
# Folder processing: pick folders to scan, ensure target folder, collect, move #
# --------------------------------------------------------------------------- #
MESSAGE_SCAN_SELECT = "id,subject,from,sender,receivedDateTime,isRead,hasAttachments,parentFolderId"


def resolve_inbox_id(client: GraphClient, mailbox: str) -> str:
    """Return the id of the mailbox's well-known Inbox folder."""
    data = client.get(
        f"/users/{mailbox}/mailFolders/inbox", params={"$select": "id,displayName"}
    )
    return data["id"]


def select_scan_folders(
    folders: List[Dict[str, Any]],
    inbox_id: str,
    destination_id: Optional[str],
    excluded_names: Iterable[str],
) -> List[Dict[str, Any]]:
    """Pick which folders to scan, applying one consistent rule to the whole tree.

    Starting from every folder in the mailbox (``folders`` from
    :func:`enumerate_folders`), a folder is SKIPPED when:

    * it is the Inbox root itself (its sub-folders are still scanned), or
    * it is the move destination, or
    * any segment of its path matches an excluded folder name (case-insensitive)
      - this prunes the excluded system folders (Sent Items, Drafts, Deleted
      Items, Junk Email, Outbox, Archive, ...) together with everything beneath
      them, and likewise the processed folder.

    Every other folder - top-level folders outside the Inbox and Inbox
    sub-folders alike - is scanned. The result is sorted by path for stable,
    predictable processing order.
    """
    excluded = {n.strip().lower() for n in excluded_names if n and n.strip()}
    selected: List[Dict[str, Any]] = []
    for folder in folders:
        if folder["id"] == inbox_id:
            continue
        if destination_id and folder["id"] == destination_id:
            continue
        segments = [seg.strip().lower() for seg in folder["path"].split("/")]
        if any(seg in excluded for seg in segments):
            continue
        selected.append(folder)
    return sorted(selected, key=lambda f: f["path"].lower())


def ensure_child_folder(
    client: GraphClient, mailbox: str, parent_id: str, name: str
) -> str:
    """Find or create a child folder ``name`` under ``parent_id``; return its id."""
    listing = client.get_all(
        f"/users/{mailbox}/mailFolders/{parent_id}/childFolders",
        params={"$filter": f"displayName eq '{name}'", "$select": "id,displayName", "$top": 10},
    )
    match = next((f for f in listing if (f.get("displayName") or "").lower() == name.lower()), None)
    if match:
        LOG.info("Found folder '%s'.", name)
        return match["id"]
    created = client.post(
        f"/users/{mailbox}/mailFolders/{parent_id}/childFolders", {"displayName": name}
    )
    LOG.info("Created folder '%s'.", name)
    return created["id"]


def collect_messages(
    client: GraphClient,
    mailbox: str,
    folders: List[Dict[str, Any]],
    exclude_folder_ids: Optional[Iterable[str]] = None,
) -> List[Dict[str, Any]]:
    """Collect every message across ``folders`` (no sender filter).

    Folders whose id is in ``exclude_folder_ids`` (e.g. the move target) are
    skipped, and duplicate message ids are de-duplicated. Each message records
    its source folder path under ``_source_path`` for reporting.
    """
    excluded = {fid for fid in (exclude_folder_ids or []) if fid}
    matched: List[Dict[str, Any]] = []
    seen_ids: set = set()

    for folder in folders:
        if folder["id"] in excluded:
            LOG.debug("Skipping excluded folder %s", folder["path"])
            continue
        try:
            messages = client.get_all(
                f"/users/{mailbox}/mailFolders/{folder['id']}/messages",
                params={"$select": MESSAGE_SCAN_SELECT, "$top": 100},
            )
        except RuntimeError as exc:
            LOG.warning("Could not list messages in folder %s: %s", folder["path"], exc)
            continue

        for msg in messages:
            if msg["id"] in seen_ids:
                continue
            seen_ids.add(msg["id"])
            msg["_source_path"] = folder["path"]
            matched.append(msg)

    LOG.info("Collected %d message(s) across %d folder(s).",
             len(matched), len(folders))
    return matched


def move_messages(
    client: GraphClient,
    mailbox: str,
    messages: List[Dict[str, Any]],
    destination_id: str,
) -> Dict[str, Any]:
    """Move each message to ``destination_id``; returns a status summary.

    Messages already in the destination are skipped. Per-message move status is
    written back onto each message dict under ``_move_status`` for reporting.
    """
    summary = {"moved": 0, "skipped": 0, "errors": 0}
    for msg in messages:
        if msg.get("parentFolderId") == destination_id:
            msg["_move_status"] = "already in target"
            summary["skipped"] += 1
            continue
        try:
            client.post(f"/users/{mailbox}/messages/{msg['id']}/move",
                        {"destinationId": destination_id})
            msg["_move_status"] = "moved"
            summary["moved"] += 1
        except RuntimeError as exc:
            msg["_move_status"] = f"error: {exc}"
            summary["errors"] += 1
            LOG.warning("Failed to move message %s: %s", msg.get("id", "")[:12], exc)
    LOG.info("Move: %d moved, %d skipped, %d errors.",
             summary["moved"], summary["skipped"], summary["errors"])
    return summary


# --------------------------------------------------------------------------- #
# Excel report                                                                #
# --------------------------------------------------------------------------- #
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=14, bold=True, color="1F4E78")


def _style_header(ws, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, max_width: int = 70) -> None:
    for col_cells in ws.columns:
        length = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def write_report(
    output_path: str,
    mailbox: str,
    folders: List[Dict[str, Any]],
    inbox_stats: Dict[str, Any],
    run_meta: Dict[str, Any],
    attachments: Optional[List[Dict[str, Any]]] = None,
    processed: Optional[List[Dict[str, Any]]] = None,
) -> None:
    """Build the multi-sheet Excel workbook and save it to ``output_path``."""
    wb = Workbook()

    _write_folder_sheet(wb.active, mailbox, folders)
    _write_inbox_sheet(wb.create_sheet("Inbox Analysis"), mailbox, inbox_stats)
    _write_attachments_sheet(wb.create_sheet("Attachments"), mailbox, attachments or [])
    if processed is not None:
        _write_processed_sheet(wb.create_sheet("Processed Messages"), mailbox, processed)
    _write_run_info_sheet(wb.create_sheet("Run Info"), run_meta)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


def _write_processed_sheet(ws, mailbox: str, processed: List[Dict[str, Any]]) -> None:
    ws["A1"] = f"Messages processed (Inbox sub-folders) - {mailbox}"
    ws["A1"].font = _TITLE_FONT

    headers = ["Received", "Sender", "Subject", "Source Folder",
               "Files Saved", "Move Status"]
    ws.append([])
    ws.append(headers)
    _style_header(ws, ws.max_row, len(headers))

    for rec in processed:
        ws.append([
            rec.get("received", ""), rec.get("sender", ""), rec.get("subject", ""),
            rec.get("source_path", ""), rec.get("files_saved", 0), rec.get("move_status", ""),
        ])

    moved = sum(1 for r in processed if r.get("move_status") == "moved")
    ws.append([])
    ws.append([f"{len(processed)} message(s) matched", "", "", "",
               sum(r.get("files_saved", 0) for r in processed), f"{moved} moved"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.freeze_panes = "A4"
    _autosize(ws)


def _write_attachments_sheet(ws, mailbox: str, attachments: List[Dict[str, Any]]) -> None:
    ws["A1"] = f"Attachments downloaded - {mailbox}"
    ws["A1"].font = _TITLE_FONT

    headers = ["Received", "Sender", "Subject", "Attachment", "Size (bytes)",
               "Inline", "Saved As", "Status"]
    ws.append([])
    ws.append(headers)
    _style_header(ws, ws.max_row, len(headers))

    for rec in attachments:
        ws.append([
            rec["received"], rec["sender"], rec["subject"], rec["name"],
            rec["size"], rec["inline"], rec["saved_as"], rec["status"],
        ])

    saved = sum(1 for r in attachments if r["status"] == "saved")
    total_bytes = sum(r["size"] for r in attachments
                      if r["status"] == "saved" and isinstance(r["size"], int))
    ws.append([])
    ws.append([f"{saved} file(s) saved", "", "", "", total_bytes, "", "", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.freeze_panes = "A4"
    _autosize(ws)


def _write_folder_sheet(ws, mailbox: str, folders: List[Dict[str, Any]]) -> None:
    ws.title = "Folder Summary"
    ws["A1"] = f"Mailbox folder summary - {mailbox}"
    ws["A1"].font = _TITLE_FONT

    headers = ["Folder", "Full Path", "Depth", "Total Emails", "Unread", "Sub-folders", "Hidden"]
    ws.append([])  # spacer row (row 2)
    ws.append(headers)
    _style_header(ws, ws.max_row, len(headers))

    # Indent the display name by depth so the tree is readable at a glance.
    for f in sorted(folders, key=lambda x: x["path"].lower()):
        indent = "    " * f["depth"]
        ws.append(
            [
                f"{indent}{f['name']}",
                f["path"],
                f["depth"],
                f["total"],
                f["unread"],
                f["child_count"],
                "Yes" if f["hidden"] else "No",
            ]
        )

    # Totals row.
    total_emails = sum(f["total"] for f in folders)
    total_unread = sum(f["unread"] for f in folders)
    ws.append([])
    ws.append(["TOTAL", f"{len(folders)} folders", "", total_emails, total_unread, "", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.freeze_panes = "A4"
    _autosize(ws)


def _write_inbox_sheet(ws, mailbox: str, stats: Dict[str, Any]) -> None:
    ws["A1"] = f"Inbox analysis - {mailbox}"
    ws["A1"].font = _TITLE_FONT

    row = 3
    metrics: List[Tuple[str, Any]] = [
        ("Messages analysed (sample)", stats["total"]),
        ("Read", stats["read"]),
        ("Unread", stats["unread"]),
        ("With attachments", stats["with_attachments"]),
        ("High importance", stats["high_importance"]),
        ("Unique senders", stats["unique_senders"]),
        ("Oldest in sample", stats["oldest"] or "n/a"),
        ("Newest in sample", stats["newest"] or "n/a"),
    ]
    ws.cell(row=row, column=1, value="Metric")
    ws.cell(row=row, column=2, value="Value")
    _style_header(ws, row, 2)
    row += 1
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Top senders block.
    row += 1
    ws.cell(row=row, column=1, value="Top senders").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Sender")
    ws.cell(row=row, column=2, value="Email count")
    _style_header(ws, row, 2)
    row += 1
    for sender, count in stats["top_senders"]:
        ws.cell(row=row, column=1, value=sender)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Volume by day block (alongside, starting at column D).
    day_row = 3
    ws.cell(row=day_row, column=4, value="Date")
    ws.cell(row=day_row, column=5, value="Emails received")
    _style_header(ws, day_row, 0)  # no-op fill guard
    for col in (4, 5):
        c = ws.cell(row=day_row, column=col)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
    day_row += 1
    for day, count in stats["by_day"]:
        ws.cell(row=day_row, column=4, value=day)
        ws.cell(row=day_row, column=5, value=count)
        day_row += 1

    _autosize(ws)


def _write_run_info_sheet(ws, meta: Dict[str, Any]) -> None:
    ws["A1"] = "Run information"
    ws["A1"].font = _TITLE_FONT
    row = 3
    for key, value in meta.items():
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(value))
        row += 1
    _autosize(ws)


# --------------------------------------------------------------------------- #
# Orchestration                                                               #
# --------------------------------------------------------------------------- #
def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--setup-xlsx", default=DEFAULT_SETUP_XLSX, help="Path to Ingestion_Setup.xlsx")
    parser.add_argument("--parameters-sheet", default=DEFAULT_PARAMETERS_SHEET, help="Name of the parameters tab")
    parser.add_argument("--manifest", default=DEFAULT_MANIFEST, help="Path to the app manifest JSON")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="Directory for the Excel report")
    parser.add_argument("--attachments-dir", default=DEFAULT_ATTACHMENTS_DIR,
                        help="Directory to save downloaded attachments")
    parser.add_argument("--mailbox", default=DEFAULT_MAILBOX, help="Target mailbox (UPN) to analyse")
    parser.add_argument("--inbox-sample", type=int, default=DEFAULT_INBOX_SAMPLE,
                        help="Maximum number of Inbox messages to pull for analysis and attachments")
    parser.add_argument("--download-attachments", dest="download_attachments",
                        action="store_true", default=True,
                        help="Download attachments from the matched messages (default)")
    parser.add_argument("--no-download-attachments", dest="download_attachments",
                        action="store_false", help="Skip downloading attachments")
    parser.add_argument("--skip-inline", action="store_true",
                        help="Skip inline attachments (e.g. signature images)")
    parser.add_argument("--exclude-extensions", default=DEFAULT_EXCLUDED_EXTENSIONS,
                        help="Comma-separated file types to EXCLUDE from download "
                             "(default: image types). All other types are downloaded. "
                             "Use 'none' to download every type including images.")
    parser.add_argument("--processed-folder", default=DEFAULT_PROCESSED_FOLDER,
                        help="Folder created beneath the Inbox to move scanned messages "
                             "into (default: Fusion_Processed)")
    parser.add_argument("--only-with-attachments", action="store_true",
                        help="Move only the messages that yielded a saved (non-image) "
                             "attachment, instead of every scanned message")
    parser.add_argument("--move-processed", dest="move_processed",
                        action="store_true", default=True,
                        help="Move scanned messages to the processed folder after download (default)")
    parser.add_argument("--no-move-processed", dest="move_processed",
                        action="store_false", help="Download only; do not move messages")
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging")
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)-7s %(message)s",
    )

    LOG.info("Reading parameters from %s", args.setup_xlsx)
    params = read_parameters_from_xlsx(args.setup_xlsx, args.parameters_sheet)

    missing = [p for p in REQUIRED_PARAMETERS if not params.get(p)]
    if missing:
        LOG.error("Missing required parameter(s) on the '%s' tab: %s",
                  args.parameters_sheet, ", ".join(missing))
        return 2

    client_id = params["GRAPH_CLIENT_ID"]
    tenant_id = params["GRAPH_TENANT_ID"]
    client_secret = params["GRAPH_CLIENT_SECRET"]

    manifest = load_manifest(args.manifest)
    permissions = validate_against_manifest(manifest, client_id)
    if permissions:
        LOG.info("Manifest declares Graph permissions: %s", ", ".join(permissions))

    token = acquire_token(tenant_id, client_id, client_secret)
    client = GraphClient(token)

    LOG.info("Enumerating mail folders for %s", args.mailbox)
    folders = enumerate_folders(client, args.mailbox)
    LOG.info("Found %d folders.", len(folders))

    LOG.info("Fetching up to %d Inbox messages for analysis", args.inbox_sample)
    messages = fetch_inbox_messages(client, args.mailbox, args.inbox_sample)
    inbox_stats = analyse_inbox(messages)
    LOG.info("Inbox sample: %d messages (%d unread).", inbox_stats["total"], inbox_stats["unread"])

    # Resolve the attachment extension exclude-list (empty = download everything).
    ext_arg = (args.exclude_extensions or "").strip().lower()
    excluded_exts = [] if ext_arg in ("", "none", "*") else [e for e in ext_arg.split(",") if e.strip()]

    processed_records: List[Dict[str, Any]] = []
    move_summary: Dict[str, Any] = {}
    destination_id: Optional[str] = None
    target_path = f"Inbox/{args.processed_folder}"

    # Resolve the Inbox and ensure the move target exists directly beneath it.
    inbox_id = resolve_inbox_id(client, args.mailbox)
    if args.download_attachments or args.move_processed:
        LOG.info("Ensuring destination folder '%s' exists", target_path)
        destination_id = ensure_child_folder(
            client, args.mailbox, inbox_id, args.processed_folder
        )

    # Scan every folder in the mailbox under one consistent rule: skip the Inbox
    # root (its sub-folders are still scanned), the move target, and the excluded
    # system folders (and everything beneath them). Top-level folders outside the
    # Inbox and Inbox sub-folders are otherwise all processed the same way.
    excluded_names = set(EXCLUDED_FOLDER_NAMES) | {args.processed_folder.strip().lower()}
    scan_folders = select_scan_folders(folders, inbox_id, destination_id, excluded_names)
    LOG.info("Scanning %d folder(s) for messages.", len(scan_folders))
    source_messages = collect_messages(
        client, args.mailbox, scan_folders,
        exclude_folder_ids=[destination_id] if destination_id else None,
    )

    attachment_records: List[Dict[str, Any]] = []
    attachment_summary: Dict[str, Any] = {}
    if args.download_attachments:
        LOG.info("Downloading all non-image attachments to %s (excluding: %s)",
                 args.attachments_dir, ",".join(excluded_exts) if excluded_exts else "none")
        attachment_records, attachment_summary = download_attachments(
            client, args.mailbox, source_messages, args.attachments_dir,
            include_inline=not args.skip_inline, excluded_exts=excluded_exts,
        )
    else:
        LOG.info("Attachment download disabled (--no-download-attachments).")

    # Tally saved files per message (used for reporting and optional move filter).
    saved_per_msg: "collections.Counter[str]" = collections.Counter()
    for rec in attachment_records:
        if rec.get("status") == "saved":
            saved_per_msg[rec.get("msg_id")] += 1

    # Decide which scanned messages to move (default: all of them).
    if args.only_with_attachments:
        move_candidates = [m for m in source_messages if saved_per_msg.get(m.get("id"), 0) > 0]
    else:
        move_candidates = source_messages

    # Move the chosen messages into Inbox/<processed-folder> (after downloading).
    if args.move_processed and destination_id:
        LOG.info("Moving %d message(s) to %s", len(move_candidates), target_path)
        move_summary = move_messages(client, args.mailbox, move_candidates, destination_id)

    # Build the per-message "processed" view for the report.
    for msg in source_messages:
        received = _parse_dt(msg.get("receivedDateTime"))
        processed_records.append({
            "received": received.isoformat() if received else "",
            "sender": _sender_address(msg) or "",
            "subject": msg.get("subject") or "(no subject)",
            "source_path": msg.get("_source_path", ""),
            "files_saved": saved_per_msg.get(msg.get("id"), 0),
            "move_status": msg.get("_move_status", "not moved"),
        })

    timestamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_mailbox = args.mailbox.replace("@", "_at_").replace(".", "_")
    output_path = os.path.join(args.output_dir, f"Mailbox_Report_{safe_mailbox}_{timestamp}.xlsx")

    run_meta = {
        "Mailbox": args.mailbox,
        "Application (clientId)": client_id,
        "Tenant": tenant_id,
        "App display name": (manifest or {}).get("displayName", "n/a"),
        "Graph permissions": ", ".join(permissions) if permissions else "n/a",
        "Folders enumerated": len(folders),
        "Folders scanned": len(scan_folders),
        "Inbox messages sampled": inbox_stats["total"],
        "Messages scanned": len(source_messages),
        "Attachment types excluded": ",".join(excluded_exts) if excluded_exts else "none",
        "Attachments downloaded": attachment_summary.get("saved", 0) if args.download_attachments else "disabled",
        "Images skipped": attachment_summary.get("skipped_images", 0) if args.download_attachments else "n/a",
        "Attachments total size": _human_bytes(attachment_summary.get("bytes", 0)) if args.download_attachments else "n/a",
        "Attachments directory": args.attachments_dir if args.download_attachments else "n/a",
        "Move scope": ("messages with attachments" if args.only_with_attachments else "all scanned messages"),
        "Messages moved": move_summary.get("moved", "n/a") if args.move_processed else "n/a",
        "Move destination": target_path if args.move_processed else "n/a",
        "Generated (local time)": _dt.datetime.now().isoformat(timespec="seconds"),
        "Parameters source": args.setup_xlsx,
    }

    LOG.info("Writing report to %s", output_path)
    write_report(output_path, args.mailbox, folders, inbox_stats, run_meta,
                 attachment_records, processed_records)
    LOG.info("Done. Report saved: %s", output_path)
    print(output_path)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:  # pragma: no cover
        sys.exit(130)
    except Exception as exc:  # noqa: BLE001 - top-level guard for operator clarity
        LOG.error("Fatal error: %s", exc)
        sys.exit(1)
