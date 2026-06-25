"""
Simple Microsoft Graph downloader for customer inbound emails.

This file is intentionally written as a clear step-by-step script.

Current confirmed tenant:
    Birkdale / BKD

Current confirmed behavior:
    - Emails from @birkdalesales.com go to BKD.
    - If the tenant uses body-to-ENS, save the email body as TXT evidence.
    - If the email has file attachments, save the attachments.
    - Generate an operational API pack with ENS PACK and DEC PACK sheets.
    - Excel attachments are validated from the Graph attachment content.
    - A validation CSV is written with errors, warnings, and pending mapping items.

Body-to-ENS note:
    The email body remains original evidence in Inbound/ENS_Source.
    The generated Process workbook is the review/mapping pack, not the source.

Important security note:
    This shared script does not store the real client secret in source code.
    Put GRAPH_TENANT_ID, GRAPH_CLIENT_ID, and GRAPH_CLIENT_SECRET in the
    runtime environment. Later, those values should come from CFG tables.
"""

from __future__ import annotations

import argparse
import base64
import csv
import datetime as dt
import hashlib
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import html
import io
import json
import os
import re
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET


SCRIPT_FOLDER = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_FOLDER.parents[1]


def load_dotenv_file() -> None:
    """
    Load .env values for simple manual execution.

    Supported key styles:
        GRAPH_TENANT_ID=...
        GRAPH.TENANT_ID=...
    """
    aliases = {
        "GRAPH.TENANT_ID": "GRAPH_TENANT_ID",
        "GRAPH.CLIENT_ID": "GRAPH_CLIENT_ID",
        "GRAPH.CLIENT_SECRET": "GRAPH_CLIENT_SECRET",
        "GRAPH.MAILBOX": "GRAPH_MAILBOX",
        "GRAPH.FOLDER": "GRAPH_FOLDER",
        "GRAPH.HISTORIC_START_DATE": "GRAPH_HISTORIC_START_DATE",
        "GRAPH.DB_ENABLED": "GRAPH_DB_ENABLED",
        "GRAPH.DB_CONNECTION_STRING": "GRAPH_DB_CONNECTION_STRING",
        "GRAPH.ENV_CODE": "GRAPH_ENV_CODE",
    }
    candidates = [
        Path.cwd() / ".env",
        REPO_ROOT / ".env",
        SCRIPT_FOLDER / ".env",
        SCRIPT_FOLDER.parent / ".env",
    ]

    for env_path in candidates:
        if not env_path.exists():
            continue
        for raw_line in env_path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip().lstrip("\ufeff")
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value
            alias = aliases.get(key)
            if alias and alias not in os.environ:
                os.environ[alias] = value
        return


load_dotenv_file()


def first_env(*names: str, default: str = "") -> str:
    """Return the first non-empty environment value from the supplied names."""
    for name in names:
        value = os.getenv(name, "").strip()
        if value:
            return value
    return default


def env_flag(name: str, default: bool = False) -> bool:
    """Read a simple true/false environment switch."""
    value = os.getenv(name, "").strip().lower()
    if not value:
        return default
    if value in {"1", "true", "yes", "y", "on"}:
        return True
    if value in {"0", "false", "no", "n", "off"}:
        return False
    return default


# =============================================================================
# STEP 1 - GRAPH CONFIGURATION
# =============================================================================
#
# These are the Microsoft Graph connection settings.
#
# Keep the real secret outside the script because this file is on a shared
# QAS workspace. For a manual test, set these in PowerShell before running:
#
#   $env:GRAPH_TENANT_ID = "<tenant-id>"
#   $env:GRAPH_CLIENT_ID = "<client-id>"
#   $env:GRAPH_CLIENT_SECRET = "<client-secret>"
#   $env:GRAPH_MAILBOX = "nexus@synoviaflow.cloud"
#
# Future database version:
#   These values should be loaded from CFG.GraphSetting or similar.

GRAPH_TENANT_ID = os.getenv("GRAPH_TENANT_ID", "").strip()
GRAPH_CLIENT_ID = os.getenv("GRAPH_CLIENT_ID", "").strip()
GRAPH_CLIENT_SECRET = os.getenv("GRAPH_CLIENT_SECRET", "").strip()
GRAPH_MAILBOX = os.getenv("GRAPH_MAILBOX", "nexus@synoviaflow.cloud").strip()
GRAPH_FOLDER = os.getenv("GRAPH_FOLDER", "inbox").strip() or "inbox"
GRAPH_HISTORIC_START_DATE = os.getenv("GRAPH_HISTORIC_START_DATE", "2026-05-07").strip()
GRAPH_ENV_CODE = os.getenv("GRAPH_ENV_CODE", "QAS").strip() or "QAS"
GRAPH_DB_CONNECTION_STRING = first_env(
    "DB_CONN_STR",
    "GRAPH_DB_CONNECTION_STRING",
    "FUSION_QAS_DB_CONNECTION_STRING",
    "FUSION_QAS_DB_CONNECTION",
)
GRAPH_DB_ENABLED = env_flag("GRAPH_DB_ENABLED", default=bool(GRAPH_DB_CONNECTION_STRING))

GRAPH_API_ROOT = "https://graph.microsoft.com/v1.0"
GRAPH_TOKEN_URL = (
    f"https://login.microsoftonline.com/"
    f"{urllib.parse.quote(GRAPH_TENANT_ID, safe='')}/oauth2/v2.0/token"
)


# =============================================================================
# STEP 2 - SHARED DRIVE CONFIGURATION
# =============================================================================

GRAPH_SCRIPT_FOLDER = SCRIPT_FOLDER
CUSTOMER_CONFIG_FOLDER = GRAPH_SCRIPT_FOLDER / "config" / "customers"

DEFAULT_INTEGRATION_LAYER_ROOT = REPO_ROOT / "Integration_Layer"
INTEGRATION_LAYER_ROOT = Path(
    os.getenv("FUSION_INTEGRATION_LAYER_ROOT", str(DEFAULT_INTEGRATION_LAYER_ROOT))
)

# Operational history for Graph step 01. Original customer files stay in the
# tenant inbound folder; CSV history/reports live with the FLOW_V3 run evidence.
RUN_LOG_FOLDER = INTEGRATION_LAYER_ROOT / "FLOW_V3" / "Run_History" / "Graph"

# Tracks destination names selected during the current run. This lets the script
# save two same-name files received on the same date as "_2", while keeping
# reruns idempotent when the files already exist.
USED_DESTINATION_PATHS: set[Path] = set()

# Tracks ENS source emails seen in the current run so the later DEC/Excel email
# can generate one combined ENS PACK + DEC PACK workbook.
RUN_ENS_SOURCE_BY_KEY: dict[tuple[str, str, str], dict[str, Any]] = {}


# =============================================================================
# STEP 3 - CUSTOMER YML CONFIGURATION
# =============================================================================
#
# Customer routing is stored as one .yml file per customer:
#
#   Integration_Layer/Graph/config/customers/BKD.yml
#   Integration_Layer/Graph/config/customers/CWH.yml
#
# The parser below is intentionally small. It supports key/value pairs and
# list values using "- item". This keeps the script dependency-free while the
# database CFG.Graph table is not ready.

FALLBACK_BIRKDALE_CONFIG = {
    "tenant_name": "Birkdale",
    "tenant_code": "BKD",
    "active": True,
    "sender_domains": ["birkdalesales.com"],
    "sender_addresses": [],
    "file_types": [".xlsx"],
    "save_mode": "email_body_and_attachment_to_api_pack",
    "destination_folder": (
        INTEGRATION_LAYER_ROOT / "BKD" / "Inbound" / "Sales_Order_files"
    ),
    "body_source_for_ens": "email_body",
    "body_text_folder": INTEGRATION_LAYER_ROOT / "BKD" / "Inbound" / "ENS_Source",
    "process_folder": INTEGRATION_LAYER_ROOT / "BKD" / "Processed",
    "fail_folder": INTEGRATION_LAYER_ROOT / "BKD" / "Fails",
    "body_text_file_pattern": "ENS_Source_{dd.MM.yyyy}.txt",
    "body_text_cut_after_marker": "customsadmin@primelineexpress.co.uk",
    "body_text_cut_marker_occurrence": "first",
    "process_pack_pattern": "BKD_API_PACK_{dd.MM.yyyy}.xlsx",
    "ens_pack_sheet_name": "ENS PACK",
    "dec_pack_sheet_name": "DEC PACK",
}

# BODY-TO-ENS BEHAVIOUR:
# BKD uses the email body as the ENS source. The script saves that body as
# text evidence and generates a Process workbook with ENS PACK and DEC PACK sheets.


# =============================================================================
# STEP 4 - FUTURE TENANTS TODO
# =============================================================================
#
# For each future tenant, investigate and confirm:
#
#   1. Sender domain(s), for example @customer-domain.com.
#   2. Sender address(es), if one exact mailbox is required.
#   3. Whether the ENS/consignment source data arrives in the email body, in attachments, or both.
#   4. Attachment file types and whether any file should be ignored.
#   5. Whether body-only emails should generate an ENS PACK even when no attachment exists.
#   6. Exact Integration Layer destination folder.
#   7. Whether the message can be marked as read after successful processing.
#
# When CFG tables are ready, this list should come from the database.

def parse_yml_value(value: str) -> Any:
    """Parse one simple YML scalar value used by customer config files."""
    value = value.strip()
    if value in {"[]", ""}:
        return []
    if (
        (value.startswith('"') and value.endswith('"'))
        or (value.startswith("'") and value.endswith("'"))
    ):
        value = value[1:-1]

    lower_value = value.lower()
    if lower_value in {"true", "yes", "on"}:
        return True
    if lower_value in {"false", "no", "off"}:
        return False
    return value


def read_simple_yml(path: Path) -> dict[str, Any]:
    """
    Read a small, dependency-free YML file.

    Supported syntax:
        key: value
        key:
          - value 1
          - value 2
    """
    data: dict[str, Any] = {}
    current_list_key = ""

    for line_number, raw_line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        line = raw_line.split("#", 1)[0].rstrip()
        if not line.strip():
            continue

        stripped = line.strip()
        if stripped.startswith("- "):
            if not current_list_key:
                raise ValueError(f"{path}:{line_number} list item without a key")
            data.setdefault(current_list_key, []).append(parse_yml_value(stripped[2:]))
            continue

        if ":" not in stripped:
            raise ValueError(f"{path}:{line_number} expected 'key: value'")

        key, value = stripped.split(":", 1)
        key = key.strip().lstrip("\ufeff")
        value = value.strip()
        if not key:
            raise ValueError(f"{path}:{line_number} empty key")

        if value:
            data[key] = parse_yml_value(value)
            current_list_key = ""
        else:
            data[key] = []
            current_list_key = key

    return data


def as_list(value: Any) -> list[str]:
    """Return config value as a clean list of strings."""
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [str(value).strip()]


def normalise_file_types(values: list[str]) -> list[str]:
    """Return file extensions in lowercase, always starting with a dot."""
    file_types = []
    for value in values:
        text = value.lower().strip()
        if not text:
            continue
        if not text.startswith("."):
            text = "." + text
        file_types.append(text)
    return file_types


def repo_path(value: Any) -> Path:
    """Return a config path as an absolute repo path."""
    path = Path(str(value or "").strip())
    if path.is_absolute():
        return path
    return REPO_ROOT / path


def default_tenant_folder(tenant_code: str, folder_name: str) -> Path:
    """Return the default Integration_Layer tenant folder."""
    return INTEGRATION_LAYER_ROOT / tenant_code / folder_name


def default_body_text_folder(destination_folder: Path) -> Path:
    """Keep original ENS source text beside inbound files, not in Process."""
    return destination_folder.parent / "ENS_Source"


def build_customer_config(path: Path, raw_config: dict[str, Any]) -> dict[str, Any] | None:
    """Validate one customer YML file and convert it to the script config shape."""
    active = bool(raw_config.get("active", True))
    if not active:
        return None

    tenant_code = str(raw_config.get("tenant_code") or path.stem).strip().upper()
    tenant_name = str(raw_config.get("tenant_name") or tenant_code).strip()
    destination_folder = str(raw_config.get("destination_folder") or "").strip()

    if not destination_folder:
        raise ValueError(f"{path}: destination_folder is required")

    destination_path = repo_path(destination_folder)
    process_path = repo_path(raw_config.get("process_folder") or default_tenant_folder(tenant_code, "Process"))
    fail_path = repo_path(raw_config.get("fail_folder") or default_tenant_folder(tenant_code, "Fails"))
    body_text_path = repo_path(raw_config.get("body_text_folder") or default_body_text_folder(destination_path))

    config = dict(raw_config)
    config.update(
        {
            "tenant_name": tenant_name,
            "tenant_code": tenant_code,
            "active": active,
            "sender_domains": [
                domain.lower().lstrip("@") for domain in as_list(raw_config.get("sender_domains"))
            ],
            "sender_addresses": [
                address.lower() for address in as_list(raw_config.get("sender_addresses"))
            ],
            "file_types": normalise_file_types(as_list(raw_config.get("file_types"))),
            "save_mode": str(raw_config.get("save_mode") or "attachments_only").strip().lower(),
            "historic_start_date": str(raw_config.get("historic_start_date") or "").strip(),
            "destination_folder": destination_path,
            "process_folder": process_path,
            "fail_folder": fail_path,
            "body_text_folder": body_text_path,
            "body_text_file_pattern": raw_config.get("body_text_file_pattern") or "ENS_Source_{dd.MM.yyyy}.txt",
            "body_text_cut_after_marker": raw_config.get("body_text_cut_after_marker")
            or raw_config.get("body_text_cut_after_last_marker"),
            "body_text_cut_marker_occurrence": raw_config.get("body_text_cut_marker_occurrence") or "first",
            "process_pack_pattern": raw_config.get("process_pack_pattern") or f"{tenant_code}_API_PACK_{{dd.MM.yyyy}}.xlsx",
            "config_file": str(path),
        }
    )
    return config

def load_tenant_configs() -> list[dict[str, Any]]:
    """Load active customer configs from Integration_Layer/Graph/config/customers/*.yml."""
    paths = sorted(CUSTOMER_CONFIG_FOLDER.glob("*.yml"))
    configs = []

    for path in paths:
        config = build_customer_config(path, read_simple_yml(path))
        if config:
            configs.append(config)

    if configs:
        return configs

    # Fallback keeps the script usable if the config folder has not been copied
    # yet. Normal operation should use the customer YML files.
    return [FALLBACK_BIRKDALE_CONFIG]


TENANT_CONFIGS = load_tenant_configs()


# =============================================================================
# STEP 4A - OPTIONAL DATABASE INGESTION TRACE
# =============================================================================
#
# If GRAPH.DB_CONNECTION_STRING is supplied, Step 01 can use the prepared MVP
# database as the first ingestion trace. Without it, the script remains file-only
# and keeps using customer YML files.

def split_delimited_values(value: Any) -> list[str]:
    """Split comma/semicolon/pipe/newline config values into a clean list."""
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return as_list(value)
    return [part.strip() for part in re.split(r"[\n,;|]+", str(value)) if part.strip()]


def parse_sender_rule(value: Any) -> tuple[list[str], list[str]]:
    """Convert CFG.Graph.SenderRule text into domain and address lists."""
    domains: list[str] = []
    addresses: list[str] = []

    for raw_item in split_delimited_values(value):
        item = raw_item.strip().lower()
        item = re.sub(r"^(sender_)?(domain|address)\s*[:=]\s*", "", item)
        if "@" in item and not item.startswith("@"):
            addresses.append(item)
        elif item:
            domains.append(item.lstrip("@"))

    return domains, addresses


def db_text(value: Any, max_length: int) -> str | None:
    """Trim optional text values to the database column size."""
    if value in (None, ""):
        return None
    return str(value)[:max_length]


def parse_graph_datetime(value: Any) -> dt.datetime | None:
    """Return a naive UTC datetime suitable for SQL Server datetime2."""
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = dt.datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=dt.timezone.utc)
    return parsed.astimezone(dt.timezone.utc).replace(tzinfo=None)


def database_destination_path(value: Any) -> Path:
    """Return a database destination as an absolute path."""
    path = Path(str(value or "").strip())
    if not path.is_absolute():
        path = REPO_ROOT / path
    return path


def build_database_customer_config(
    row: dict[str, Any],
    yml_configs_by_tenant: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """Convert one CFG.Graph row into the same config shape as the YML loader."""
    tenant_code = str(row.get("TenantCode") or "").strip().upper()
    base = dict(yml_configs_by_tenant.get(tenant_code, {}))
    sender_domains, sender_addresses = parse_sender_rule(row.get("SenderRule"))
    file_types = normalise_file_types(split_delimited_values(row.get("AllowedFileTypes")))
    destination_folder = row.get("DestinationFolder") or base.get("destination_folder")

    if not tenant_code:
        raise ValueError("CFG.Graph row is missing TenantCode")
    if not destination_folder:
        raise ValueError(f"CFG.Graph row for {tenant_code} is missing DestinationFolder")

    destination_path = database_destination_path(destination_folder)
    process_path = database_destination_path(row.get("ProcessFolder") or base.get("process_folder") or default_tenant_folder(tenant_code, "Process"))
    fail_path = database_destination_path(row.get("FailFolder") or base.get("fail_folder") or default_tenant_folder(tenant_code, "Fails"))
    body_text_path = base.get("body_text_folder") or default_body_text_folder(destination_path)

    config = dict(base)
    config.update(
        {
            "config_id": int(row["ConfigID"]) if row.get("ConfigID") is not None else None,
            "route_id": int(row["RouteID"]) if row.get("RouteID") is not None else None,
            "env_code": str(row.get("EnvCode") or GRAPH_ENV_CODE).strip() or GRAPH_ENV_CODE,
            "tenant_code": tenant_code,
            "tenant_name": str(row.get("TenantName") or base.get("tenant_name") or tenant_code),
            "active": True,
            "mailbox": str(row.get("Mailbox") or GRAPH_MAILBOX).strip() or GRAPH_MAILBOX,
            "sender_domains": sender_domains or base.get("sender_domains", []),
            "sender_addresses": sender_addresses or base.get("sender_addresses", []),
            "file_types": file_types or base.get("file_types", []),
            "save_mode": str(base.get("save_mode") or "attachments_only").strip().lower(),
            "destination_folder": destination_path,
            "process_folder": process_path,
            "fail_folder": fail_path,
            "body_text_folder": body_text_path,
            "body_text_file_pattern": base.get("body_text_file_pattern") or "ENS_Source_{dd.MM.yyyy}.txt",
            "body_text_cut_after_marker": base.get("body_text_cut_after_marker")
            or base.get("body_text_cut_after_last_marker"),
            "body_text_cut_marker_occurrence": base.get("body_text_cut_marker_occurrence") or "first",
            "process_pack_pattern": base.get("process_pack_pattern") or f"{tenant_code}_API_PACK_{{dd.MM.yyyy}}.xlsx",
            "output_file_pattern": row.get("OutputFilePattern") or base.get("output_file_pattern"),
            "ens_pack_sheet_name": row.get("EnsSheetName") or base.get("ens_pack_sheet_name"),
            "dec_pack_sheet_name": row.get("DecSheetName") or base.get("dec_pack_sheet_name"),
            "body_source_for_ens": row.get("BodySourceForEns") or base.get("body_source_for_ens"),
            "processing_environment": row.get("ProcessingEnvironment") or base.get("processing_environment"),
            "config_source": "CFG.Graph",
        }
    )
    return config


class DatabaseIngestionTrace:
    """Small SQL Server adapter for CFG.Graph, EXC.Graph and ING.Graph."""

    def __init__(self, connection: Any):
        self.connection = connection
        self.execution_id: int | None = None
        self._ing_graph_columns: set[str] | None = None

    @classmethod
    def open(cls, args: argparse.Namespace) -> "DatabaseIngestionTrace | None":
        """Open the optional ingestion database connection."""
        if getattr(args, "no_database", False):
            return None
        if not GRAPH_DB_ENABLED:
            return None
        if not GRAPH_DB_CONNECTION_STRING:
            raise RuntimeError(
                "GRAPH.DB_ENABLED is true but GRAPH.DB_CONNECTION_STRING is not set."
            )

        try:
            import pyodbc  # type: ignore[import-not-found]
        except ImportError as error:
            raise RuntimeError(
                "Database ingestion tracing requires pyodbc when GRAPH.DB_ENABLED is true."
            ) from error

        connection = pyodbc.connect(GRAPH_DB_CONNECTION_STRING, autocommit=False)
        return cls(connection)

    def close(self) -> None:
        """Close the SQL connection."""
        self.connection.close()

    def fetch_dicts(self, query: str, *params: Any) -> list[dict[str, Any]]:
        """Run a SELECT query and return rows as dictionaries."""
        cursor = self.connection.cursor()
        cursor.execute(query, *params)
        columns = [column[0] for column in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def table_columns(self, schema: str, table: str) -> set[str]:
        """Return SQL Server column names for optional backward-compatible inserts."""
        cursor = self.connection.cursor()
        cursor.execute(
            "SELECT name FROM sys.columns WHERE object_id = OBJECT_ID(?)",
            f"{schema}.{table}",
        )
        return {str(row[0]) for row in cursor.fetchall()}

    def ing_graph_columns(self) -> set[str]:
        """Cache ING.Graph columns for optional process/fail metadata."""
        if self._ing_graph_columns is None:
            self._ing_graph_columns = self.table_columns("ING", "Graph")
        return self._ing_graph_columns

    def load_tenant_configs(self) -> list[dict[str, Any]]:
        """Read active routing rows from CFG.Graph for this environment/mailbox."""
        rows = self.fetch_dicts(
            """
            SELECT
                ConfigID,
                RouteID,
                EnvCode,
                TenantCode,
                TenantName,
                Mailbox,
                SenderRule,
                AllowedFileTypes,
                DestinationFolder,
                ProcessFolder,
                FailFolder,
                OutputFilePattern,
                EnsSheetName,
                DecSheetName,
                BodySourceForEns,
                ProcessingEnvironment
            FROM CFG.Graph
            WHERE EnvCode = ?
              AND IsActive = 1
              AND LOWER(Mailbox) = LOWER(?)
            ORDER BY TenantCode
            """,
            GRAPH_ENV_CODE,
            GRAPH_MAILBOX,
        )
        yml_configs = {config["tenant_code"]: config for config in load_tenant_configs()}
        return [build_database_customer_config(row, yml_configs) for row in rows]

    def begin_execution(self, args: argparse.Namespace, started_at: dt.datetime) -> None:
        """Insert one EXC.Graph row for this script run."""
        if args.dry_run:
            return
        cursor = self.connection.cursor()
        cursor.execute(
            """
            INSERT INTO EXC.Graph (EnvCode, ProcessName, RunMode, StartedAt, Status)
            OUTPUT INSERTED.ExecutionID
            VALUES (?, ?, ?, ?, ?)
            """,
            GRAPH_ENV_CODE,
            "FLOW_V3_01_GRAPH_EMAIL_ING",
            args.run_mode,
            started_at.astimezone(dt.timezone.utc).replace(tzinfo=None),
            "RUNNING",
        )
        self.execution_id = int(cursor.fetchone()[0])
        self.connection.commit()

    def finish_execution(
        self,
        status: str,
        stats: dict[str, Any],
        error_message: str = "",
    ) -> None:
        """Update the EXC.Graph row with the final run outcome."""
        if not self.execution_id:
            return
        items_processed = max(
            0,
            int(stats.get("matched", 0)) + int(stats.get("unmatched", 0)) - int(stats.get("failed", 0)),
        )
        cursor = self.connection.cursor()
        cursor.execute(
            """
            UPDATE EXC.Graph
            SET EndedAt = SYSUTCDATETIME(),
                Status = ?,
                ItemsFound = ?,
                ItemsProcessed = ?,
                ErrorMessage = ?
            WHERE ExecutionID = ?
            """,
            status[:30],
            int(stats.get("scanned", 0)),
            items_processed,
            db_text(error_message, 2000),
            self.execution_id,
        )
        self.connection.commit()

    def insert_ing_trace(
        self,
        message: dict[str, Any],
        config: dict[str, Any] | None,
        status: str,
        original_file_name: str = "",
        saved_path: Path | str = "",
        content_type: str = "",
        size_bytes: int | None = None,
        content: bytes | None = None,
        has_attachments: bool | None = None,
        pack_code: str = "",
        source_part: str = "",
        generated_csv_path: Path | str = "",
        load_status: str = "",
        fail_reason: str = "",
    ) -> int | None:
        """Insert one ING.Graph source trace row and return its GraphID."""
        if not self.execution_id:
            return None

        sender_email = get_sender_email(message)
        sender_domain = sender_email.split("@")[-1] if "@" in sender_email else ""
        saved_path_text = str(saved_path) if saved_path else ""
        saved_name = Path(saved_path_text).name if saved_path_text else ""
        if content is not None:
            file_hash = hashlib.sha256(content).hexdigest()
            size_bytes = len(content)
        else:
            file_hash = None
            if size_bytes is not None:
                try:
                    size_bytes = int(size_bytes)
                except (TypeError, ValueError):
                    size_bytes = None
        if has_attachments is None:
            has_attachments = bool(message.get("hasAttachments"))

        row_values: dict[str, Any] = {
            "ExecutionID": self.execution_id,
            "ConfigID": config.get("config_id") if config else None,
            "RouteID": config.get("route_id") if config else None,
            "EnvCode": config.get("env_code", GRAPH_ENV_CODE) if config else GRAPH_ENV_CODE,
            "TenantCode": config.get("tenant_code") if config else None,
            "Mailbox": config.get("mailbox", GRAPH_MAILBOX) if config else GRAPH_MAILBOX,
            "GraphMessageID": db_text(message.get("id"), 450),
            "InternetMessageID": db_text(message.get("internetMessageId"), 1000),
            "SenderEmail": db_text(sender_email, 320),
            "SenderDomain": db_text(sender_domain, 320),
            "Subject": db_text(message.get("subject"), 998),
            "ReceivedAt": parse_graph_datetime(message.get("receivedDateTime")),
            "HasAttachments": 1 if has_attachments else 0,
            "OriginalFileName": db_text(original_file_name, 500),
            "SavedFileName": db_text(saved_name, 500),
            "SavedPath": db_text(saved_path_text, 1000),
            "ContentType": db_text(content_type, 255),
            "SizeBytes": size_bytes,
            "FileHash": file_hash,
            "PackCode": db_text(pack_code, 30),
            "SourcePart": db_text(source_part, 30),
            "ProcessFolder": db_text(config.get("process_folder") if config else "", 1000),
            "FailFolder": db_text(config.get("fail_folder") if config else "", 1000),
            "GeneratedCsvPath": db_text(str(generated_csv_path) if generated_csv_path else "", 1000),
            "LoadStatus": db_text(load_status, 40),
            "FailReason": db_text(fail_reason, 2000),
            "Status": status[:30],
        }
        preferred_columns = [
            "ExecutionID",
            "ConfigID",
            "RouteID",
            "EnvCode",
            "TenantCode",
            "Mailbox",
            "GraphMessageID",
            "InternetMessageID",
            "SenderEmail",
            "SenderDomain",
            "Subject",
            "ReceivedAt",
            "HasAttachments",
            "OriginalFileName",
            "SavedFileName",
            "SavedPath",
            "ContentType",
            "SizeBytes",
            "FileHash",
            "PackCode",
            "SourcePart",
            "ProcessFolder",
            "FailFolder",
            "GeneratedCsvPath",
            "LoadStatus",
            "FailReason",
            "Status",
        ]
        existing_columns = self.ing_graph_columns()
        columns = [column for column in preferred_columns if column in existing_columns]
        placeholders = ", ".join("?" for _ in columns)
        column_sql = ", ".join(columns)

        cursor = self.connection.cursor()
        cursor.execute(
            f"INSERT INTO ING.Graph ({column_sql}) OUTPUT INSERTED.GraphID VALUES ({placeholders})",
            *[row_values[column] for column in columns],
        )
        graph_id = int(cursor.fetchone()[0])
        self.connection.commit()
        return graph_id

    def insert_execution_log(
        self,
        step_name: str,
        message: str,
        level: str = "INFO",
        tenant_code: str = "",
        detail: dict[str, Any] | None = None,
    ) -> None:
        """Insert one detailed EXC.ExecutionLog row for operator visibility."""
        if not self.execution_id:
            return
        cursor = self.connection.cursor()
        cursor.execute(
            """
            INSERT INTO EXC.ExecutionLog
                (ExecutionID, EnvCode, TenantCode, ProcessName, StepName, LogLevel, Message, DetailJson)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            self.execution_id,
            GRAPH_ENV_CODE,
            db_text(tenant_code, 10),
            "FLOW_V3_01_GRAPH_EMAIL_ING",
            db_text(step_name, 100),
            db_text(level, 20) or "INFO",
            db_text(message, 2000) or "",
            json.dumps(detail or {}, ensure_ascii=False) if detail else None,
        )
        self.connection.commit()

    def insert_process_pack_load(
        self,
        graph_id: int | None,
        config: dict[str, Any],
        pack_path: Path | str,
        file_hash: str | None = None,
    ) -> dict[str, int]:
        """Load generated API pack rows into ING.ProcessFile and ING.LoadRow."""
        if not self.execution_id or not graph_id:
            return {"process_files": 0, "load_rows": 0}

        path = Path(pack_path)
        if not path.exists():
            return {"process_files": 0, "load_rows": 0}

        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_name = str(config.get("dec_pack_sheet_name") or "DEC PACK")
        if sheet_name not in workbook.sheetnames:
            sheet_name = workbook.sheetnames[-1]
        sheet = workbook[sheet_name]
        raw_headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers = [header or f"Column {index + 1}" for index, header in enumerate(raw_headers)]

        cursor = self.connection.cursor()
        cursor.execute(
            """
            INSERT INTO ING.ProcessFile
                (GraphID, ExecutionID, TenantID, RouteID, EnvCode, TenantCode, PackCode, SourcePart,
                 SourceFolder, ProcessFolder, FailFolder, OriginalFileName, SavedFileName, SavedPath,
                 GeneratedCsvPath, SheetName, FileHash, Status)
            OUTPUT INSERTED.ProcessFileID
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            graph_id,
            self.execution_id,
            config.get("tenant_id"),
            config.get("route_id"),
            config.get("env_code", GRAPH_ENV_CODE),
            config.get("tenant_code"),
            "API_PACK",
            "GENERATED_PACK",
            db_text(str(path.parent), 1000),
            db_text(config.get("process_folder"), 1000),
            db_text(config.get("fail_folder"), 1000),
            db_text(path.name, 500),
            db_text(path.name, 500),
            db_text(str(path), 1000),
            db_text(str(path), 1000),
            db_text(sheet.title, 128),
            file_hash,
            "LOADED",
        )
        process_file_id = int(cursor.fetchone()[0])

        load_rows = 0
        for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
            payload = {
                headers[index]: value
                for index, value in enumerate(cells)
                if index < len(headers) and value not in (None, "")
            }
            if not payload:
                continue
            cursor.execute(
                """
                INSERT INTO ING.LoadRow (ProcessFileID, RowNumber, PayloadJson, Status)
                VALUES (?, ?, ?, ?)
                """,
                process_file_id,
                row_number,
                json.dumps(payload, ensure_ascii=False, default=str),
                "LOADED",
            )
            load_rows += 1

        self.connection.commit()
        self.insert_execution_log(
            "LOAD_PROCESS_PACK",
            f"Loaded generated API pack into ING rows: {path.name}",
            tenant_code=str(config.get("tenant_code") or ""),
            detail={
                "graph_id": graph_id,
                "process_file_id": process_file_id,
                "load_rows": load_rows,
                "sheet_name": sheet.title,
            },
        )
        return {"process_files": 1, "load_rows": load_rows}


# =============================================================================
# STEP 5 - EXCEL MAPPING AND API VALIDATION CONFIGURATION
# =============================================================================
#
# Files are saved locally for audit/support, but the preferred processing path is
# direct: read the Graph attachment bytes, parse the Excel in memory, map columns,
# and report API readiness issues immediately.
#
# Tenant config can extend aliases with keys such as:
#
#   map_transport_document_number:
#     - Customer TDN
#     - Document Number
#
# The code below keeps every source column accounted for: mapped to an API field,
# retained for audit, or reported as an unmapped source column.

EXCEL_SUFFIXES = {".xlsx"}

FIELD_RULES: dict[str, dict[str, Any]] = {
    "transport_document_number": {
        "label": "Transport document number",
        "required": True,
        "max_length": 35,
        "aliases": ["transport_document_number", "transport doc number", "tdn", "transport document", "document number", "document no"],
    },
    "goods_description": {
        "label": "Goods description",
        "required": True,
        "max_length": 255,
        "aliases": ["goods_description", "description", "item description", "product description", "commodity description", "goods desc"],
    },
    "controlled_goods": {
        "label": "Controlled goods",
        "required": True,
        "value_type": "yes_no",
        "aliases": ["controlled_goods", "controlled goods", "is controlled goods", "controlled"],
    },
    "consignor_eori": {
        "label": "Consignor EORI",
        "required": True,
        "max_length": 17,
        "value_type": "eori",
        "aliases": ["consignor_eori", "consignor eori", "shipper eori", "sender eori"],
    },
    "consignee_eori": {
        "label": "Consignee EORI",
        "required": True,
        "max_length": 17,
        "value_type": "eori",
        "aliases": ["consignee_eori", "consignee eori", "receiver eori"],
    },
    "importer_eori": {
        "label": "Importer EORI",
        "required": True,
        "max_length": 17,
        "value_type": "eori",
        "aliases": ["importer_eori", "importer eori", "importer eori number"],
    },
    "exporter_eori": {
        "label": "Exporter EORI",
        "required": True,
        "max_length": 17,
        "value_type": "eori",
        "aliases": ["exporter_eori", "exporter eori"],
    },
    "type_of_packages": {
        "label": "Type of packages",
        "required": True,
        "max_length": 40,
        "aliases": ["type_of_packages", "type of packages", "package type", "kind of packages", "package kind"],
    },
    "number_of_packages": {
        "label": "Number of packages",
        "required": True,
        "value_type": "integer",
        "min": 1,
        "max": 99999,
        "aliases": ["number_of_packages", "number of packages", "packages", "package count", "no of packages", "number packages"],
    },
    "package_marks": {
        "label": "Package marks",
        "required": True,
        "max_length": 140,
        "aliases": ["package_marks", "package marks", "marks", "shipping marks"],
    },
    "gross_mass_kg": {
        "label": "Gross mass kg",
        "required": True,
        "value_type": "decimal",
        "max_digits": 13,
        "max_decimals": 2,
        "must_be_positive": True,
        "aliases": ["gross_mass_kg", "gross mass kg", "gross mass", "gross weight", "gross weight kg", "gross kg"],
    },
    "net_mass_kg": {
        "label": "Net mass kg",
        "required": False,
        "value_type": "decimal",
        "max_digits": 13,
        "max_decimals": 2,
        "aliases": ["net_mass_kg", "net mass kg", "net mass", "net weight", "net weight kg", "net kg"],
    },
    "commodity_code": {
        "label": "Commodity code",
        "required": False,
        "value_type": "commodity_code",
        "aliases": ["commodity_code", "commodity code", "hs code", "tariff code", "cn code", "goods code"],
    },
    "country_of_origin": {
        "label": "Country of origin",
        "required": False,
        "value_type": "country_code",
        "aliases": ["country_of_origin", "country of origin", "origin country", "coo", "origin"],
    },
    "item_invoice_amount": {
        "label": "Invoice amount",
        "required": False,
        "value_type": "decimal",
        "max_digits": 13,
        "max_decimals": 2,
        "aliases": ["item_invoice_amount", "invoice amount", "item value", "value", "customs value"],
    },
    "item_invoice_currency": {
        "label": "Invoice currency",
        "required": False,
        "value_type": "currency_code",
        "aliases": ["item_invoice_currency", "invoice currency", "currency", "value currency"],
    },
}

AUDIT_COLUMN_ALIASES = {
    "goods_item_number": ["goods_item_number", "goods item number", "item number", "item no", "line number", "line no"],
    "trader_reference": ["trader_reference", "trader reference", "customer reference", "manifest reference"],
    "destination_country": ["destination_country", "destination country", "country of destination"],
}


# =============================================================================
# STEP 6 - MICROSOFT GRAPH FUNCTIONS
# =============================================================================

def check_graph_configuration() -> None:
    """Stop early if the Graph credentials were not configured."""
    missing = []
    if not GRAPH_TENANT_ID:
        missing.append("GRAPH_TENANT_ID")
    if not GRAPH_CLIENT_ID:
        missing.append("GRAPH_CLIENT_ID")
    if not GRAPH_CLIENT_SECRET:
        missing.append("GRAPH_CLIENT_SECRET")

    if missing:
        raise RuntimeError(
            "Missing Graph configuration: "
            + ", ".join(missing)
            + ". Set these values in the runtime environment."
        )


def get_graph_token() -> str:
    """Get an app-only Microsoft Graph token using client credentials."""
    form_data = urllib.parse.urlencode(
        {
            "client_id": GRAPH_CLIENT_ID,
            "client_secret": GRAPH_CLIENT_SECRET,
            "scope": "https://graph.microsoft.com/.default",
            "grant_type": "client_credentials",
        }
    ).encode("utf-8")

    request = urllib.request.Request(
        GRAPH_TOKEN_URL,
        data=form_data,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )

    response = open_json_request(request)
    return response["access_token"]


def graph_get(token: str, url: str) -> dict[str, Any]:
    """Run one GET request against Microsoft Graph."""
    request = urllib.request.Request(
        url,
        headers={"Authorization": f"Bearer {token}"},
        method="GET",
    )
    return open_json_request(request)


def open_json_request(request: urllib.request.Request) -> dict[str, Any]:
    """Open an HTTP request and return JSON, with readable Graph errors."""
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as error:
        details = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Graph HTTP {error.code}: {details}") from error
    except urllib.error.URLError as error:
        raise RuntimeError(f"Graph connection error: {error}") from error


def build_messages_url(args: argparse.Namespace) -> str:
    """Build the first Graph URL used to read mailbox messages."""
    mailbox = urllib.parse.quote(GRAPH_MAILBOX, safe="")
    folder = urllib.parse.quote(GRAPH_FOLDER, safe="")

    query = {
        "$select": ",".join(
            [
                "id",
                "subject",
                "from",
                "receivedDateTime",
                "hasAttachments",
                "internetMessageId",
                "body",
            ]
        ),
        "$orderby": "receivedDateTime asc",
        "$top": str(args.page_size),
    }

    filters = []
    if args.received_from:
        filters.append(f"receivedDateTime ge {normalise_graph_date(args.received_from)}")
    if args.received_to:
        filters.append(
            f"receivedDateTime le {normalise_graph_date(args.received_to, end_of_day=True)}"
        )
    if filters:
        query["$filter"] = " and ".join(filters)

    return (
        f"{GRAPH_API_ROOT}/users/{mailbox}/mailFolders/{folder}/messages?"
        f"{urllib.parse.urlencode(query, safe=', :')}"
    )


def read_messages(token: str, args: argparse.Namespace) -> list[dict[str, Any]]:
    """Read mailbox messages, following Graph pagination."""
    url = build_messages_url(args)
    messages: list[dict[str, Any]] = []

    while url:
        data = graph_get(token, url)
        for message in data.get("value", []):
            messages.append(message)
            if args.max_messages and len(messages) >= args.max_messages:
                return messages
        url = data.get("@odata.nextLink")

    return messages


def read_attachments(token: str, message_id: str) -> list[dict[str, Any]]:
    """Read all attachments for one email message."""
    mailbox = urllib.parse.quote(GRAPH_MAILBOX, safe="")
    message = urllib.parse.quote(message_id, safe="")
    url = f"{GRAPH_API_ROOT}/users/{mailbox}/messages/{message}/attachments"
    attachments: list[dict[str, Any]] = []

    while url:
        data = graph_get(token, url)
        attachments.extend(data.get("value", []))
        url = data.get("@odata.nextLink")

    return attachments


def read_one_attachment(
    token: str,
    message_id: str,
    attachment_id: str,
) -> dict[str, Any]:
    """Read one attachment directly if Graph did not include file content."""
    mailbox = urllib.parse.quote(GRAPH_MAILBOX, safe="")
    message = urllib.parse.quote(message_id, safe="")
    attachment = urllib.parse.quote(attachment_id, safe="")
    url = f"{GRAPH_API_ROOT}/users/{mailbox}/messages/{message}/attachments/{attachment}"
    return graph_get(token, url)


# =============================================================================
# STEP 7 - TENANT MATCHING
# =============================================================================

def get_sender_email(message: dict[str, Any]) -> str:
    """Return the sender email address in lowercase."""
    return (
        message.get("from", {})
        .get("emailAddress", {})
        .get("address", "")
        .strip()
        .lower()
    )


def find_tenant_config(message: dict[str, Any]) -> dict[str, Any] | None:
    """Match one email to one tenant config using sender email/domain."""
    sender_email = get_sender_email(message)
    sender_domain = sender_email.split("@")[-1] if "@" in sender_email else ""

    for config in TENANT_CONFIGS:
        sender_domains = [domain.lower() for domain in config["sender_domains"]]
        sender_addresses = [address.lower() for address in config["sender_addresses"]]

        if sender_email in sender_addresses:
            return config
        if sender_domain in sender_domains:
            return config

    return None


# =============================================================================
# STEP 8 - FILE HELPERS
# =============================================================================

def file_attachments_only(attachments: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Keep real file attachments and ignore inline images/signatures."""
    files = []
    for attachment in attachments:
        is_file = attachment.get("@odata.type") == "#microsoft.graph.fileAttachment"
        is_inline = bool(attachment.get("isInline"))
        has_name = bool(attachment.get("name"))

        if is_file and not is_inline and has_name:
            files.append(attachment)

    return files


def attachment_allowed_for_config(attachment: dict[str, Any], config: dict[str, Any]) -> bool:
    """Return True when the attachment extension is allowed for this customer."""
    allowed_file_types = [value.lower() for value in config.get("file_types", [])]
    if not allowed_file_types:
        return True

    suffix = Path(str(attachment.get("name") or "")).suffix.lower()
    return suffix in allowed_file_types


def safe_filename(value: str, fallback: str) -> str:
    """Make a Windows-safe filename."""
    value = (value or fallback).strip()
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value)
    value = re.sub(r"\s+", " ", value)
    value = value.rstrip(". ")
    return value[:140] or fallback


def received_date_for_filename(message: dict[str, Any]) -> str:
    """Return received date as dd.mm.yyyy for filenames."""
    value = message["receivedDateTime"].replace("Z", "+00:00")
    received = dt.datetime.fromisoformat(value)
    received = received.astimezone(dt.timezone.utc)
    return received.strftime("%d.%m.%Y")


def save_bytes(path: Path, content: bytes, args: argparse.Namespace) -> str:
    """
    Save one file.

    Returns:
        SAVED, DRY_RUN, or SKIPPED_EXISTS
    """
    if path.exists() and not args.overwrite:
        return "SKIPPED_EXISTS"

    if args.dry_run:
        return "DRY_RUN"

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)
    return "SAVED"


def choose_run_unique_path(base_path: Path, args: argparse.Namespace) -> Path:
    """Choose base path, or _2/_3 when this run already used that name."""
    if args.overwrite:
        return base_path

    path = base_path
    counter = 2
    while path in USED_DESTINATION_PATHS:
        path = base_path.with_name(f"{base_path.stem}_{counter}{base_path.suffix}")
        counter += 1

    USED_DESTINATION_PATHS.add(path)
    return path


def save_attachment_file(
    message: dict[str, Any],
    config: dict[str, Any],
    attachment: dict[str, Any],
    args: argparse.Namespace,
) -> tuple[str, Path, bytes]:
    """Save one attachment for audit and return the same bytes for validation."""
    original_name = safe_filename(attachment.get("name", ""), "attachment")
    original_path = Path(original_name)
    filename = f"{original_path.stem}_{received_date_for_filename(message)}{original_path.suffix}"
    path = choose_run_unique_path(config["destination_folder"] / filename, args)

    content = base64.b64decode(attachment["contentBytes"])
    result = save_bytes(path, content, args)
    return result, path, content


# =============================================================================
# STEP 8A - ENS BODY TEXT AND API PACK HELPERS
# =============================================================================

ENS_BODY_LABELS: dict[str, tuple[str, str]] = {
    "type of movement": ("movement_type", "Accepted movement_type code required."),
    "type of passive transport": ("passive_transport_type", "Operational context; confirm if needed in the final API payload."),
    "identity number of transport": ("identity_no_of_transport", "Direct ENS header field."),
    "nationality of means of transport": ("nationality_of_transport", "Accepted country/nationality code required."),
    "carrier eori": ("carrier_eori", "Direct ENS header field."),
    "transport document number icr number": ("transport_document_number", "Used for grouping/reference, not an ENS header field."),
    "arrival date time": ("arrival_date_time", "Convert to dd/mm/yyyy hh:mm:ss UTC before API submit."),
    "port of arrival": ("arrival_port", "Accepted port code required."),
    "place s of loading": ("place_of_loading", "Direct ENS header field."),
    "is are the place s of acceptance same as place s of loading": ("place_of_acceptance_same_as_loading", "Confirm accepted yes/no value for API submit."),
    "place s of unloading": ("place_of_unloading", "Direct ENS header field."),
    "is are the place s of delivery same as place s of unloading": ("place_of_delivery_same_as_unloading", "Confirm accepted yes/no value for API submit."),
    "transport charges": ("transport_charges", "Accepted transport_charges code required."),
}

ENS_PACK_FIELDS: list[dict[str, str]] = [
    {"api_field": "movement_type", "source_label": "Type of Movement", "required": "yes", "note": "Map text such as RoRo Accompanied ICS2 to the accepted choice value."},
    {"api_field": "identity_no_of_transport", "source_label": "Identity number of transport", "required": "yes", "note": "Expected format includes ferry IMO plus vehicle registration."},
    {"api_field": "nationality_of_transport", "source_label": "Nationality Of Means of Transport", "required": "yes", "note": "Map country text to the accepted code."},
    {"api_field": "arrival_date_time", "source_label": "Arrival Date/Time", "required": "yes", "note": "Convert relative wording such as Tomorrow's Date to an absolute UTC date/time."},
    {"api_field": "arrival_port", "source_label": "Port of Arrival", "required": "yes", "note": "Map port text to the accepted port code."},
    {"api_field": "place_of_loading", "source_label": "Place(s) of Loading", "required": "yes", "note": "Direct text from the email body."},
    {"api_field": "place_of_acceptance_same_as_loading", "source_label": "Is/are the Place(s) of Acceptance same as Place(s) of Loading?", "required": "conditional", "note": "If yes, place_of_acceptance can normally stay blank."},
    {"api_field": "place_of_acceptance", "source_label": "", "required": "conditional", "note": "Needed only when acceptance differs from loading."},
    {"api_field": "place_of_unloading", "source_label": "Place(s) of Unloading", "required": "yes", "note": "Direct text from the email body."},
    {"api_field": "place_of_delivery_same_as_unloading", "source_label": "Is/are the Place(s) of Delivery same as Place(s) of Unloading?", "required": "conditional", "note": "If yes, place_of_delivery can normally stay blank."},
    {"api_field": "place_of_delivery", "source_label": "", "required": "conditional", "note": "Needed only when delivery differs from unloading."},
    {"api_field": "transport_charges", "source_label": "Transport Charges", "required": "yes", "note": "Map body text to the accepted transport charge value."},
    {"api_field": "carrier_eori", "source_label": "Carrier EORI", "required": "yes", "note": "Direct ENS header field."},
    {"api_field": "carrier_name", "source_label": "", "required": "yes", "note": "Resolve from tenant/carrier master data."},
    {"api_field": "carrier_street_number", "source_label": "", "required": "yes", "note": "Resolve from tenant/carrier master data."},
    {"api_field": "carrier_city", "source_label": "", "required": "yes", "note": "Resolve from tenant/carrier master data."},
    {"api_field": "carrier_postcode", "source_label": "", "required": "yes", "note": "Resolve from tenant/carrier master data."},
    {"api_field": "carrier_country", "source_label": "", "required": "yes", "note": "Resolve from tenant/carrier master data."},
    {"api_field": "transport_document_number", "source_label": "Transport document number (ICR Number)", "required": "business", "note": "Keep for consignment grouping/reference even if not part of ENS header create."},
    {"api_field": "passive_transport_type", "source_label": "Type of passive transport", "required": "business", "note": "Keep as audit/operational context."},
]

CODE_OR_DATE_FIELDS = {
    "movement_type",
    "nationality_of_transport",
    "arrival_date_time",
    "arrival_port",
    "transport_charges",
}


def message_short_id(message: dict[str, Any]) -> str:
    """Return a stable short id for filenames without exposing the full Graph id."""
    seed = str(message.get("internetMessageId") or message.get("id") or message.get("subject") or "message")
    return hashlib.sha1(seed.encode("utf-8", errors="replace")).hexdigest()[:8].upper()


def render_message_filename(pattern: str, message: dict[str, Any], config: dict[str, Any]) -> str:
    """Render the small filename tokens used by tenant config."""
    value = pattern or "{tenant_code}_{dd.MM.yyyy}"
    replacements = {
        "{tenant_code}": str(config.get("tenant_code", "TENANT")),
        "{dd.MM.yyyy}": received_date_for_filename(message),
        "{message_id_short}": message_short_id(message),
    }
    for token, replacement in replacements.items():
        value = value.replace(token, replacement)
    return safe_filename(value, "generated_file")


def body_text_enabled(config: dict[str, Any]) -> bool:
    """Return True when the tenant uses email body data for ENS pack creation."""
    return str(config.get("body_source_for_ens") or "").strip().lower() == "email_body"


def email_body_as_text(message: dict[str, Any]) -> str:
    """Convert the Graph email body to readable text for evidence and parsing."""
    body = message.get("body") if isinstance(message.get("body"), dict) else {}
    content = str(body.get("content") or message.get("bodyPreview") or "")
    content_type = str(body.get("contentType") or "").lower()

    if content_type == "html" or "<" in content and ">" in content:
        content = re.sub(r"(?i)<br\s*/?>", "\n", content)
        content = re.sub(r"(?i)</(p|div|li|tr|h[1-6])>", "\n", content)
        content = re.sub(r"<[^>]+>", " ", content)

    content = html.unescape(content)
    content = content.replace("\ufeff", "")
    content = re.sub(r"[\u200b\u200c\u200d]", "", content)
    lines = [re.sub(r"\s+", " ", line).strip() for line in content.splitlines()]
    return "\n".join(line for line in lines if line)


def trim_ens_source_body_text(body_text: str, config: dict[str, Any]) -> str:
    """Trim forwarded-chain noise after the configured body marker."""
    marker = str(
        config.get("body_text_cut_after_marker") or config.get("body_text_cut_after_last_marker") or ""
    ).strip()
    if not marker:
        return body_text

    body_lower = body_text.lower()
    marker_lower = marker.lower()
    occurrence = str(config.get("body_text_cut_marker_occurrence") or "first").strip().lower()
    index = body_lower.rfind(marker_lower) if occurrence == "last" else body_lower.find(marker_lower)
    if index < 0:
        return body_text

    return body_text[: index + len(marker)].rstrip()

def save_ens_body_text(
    message: dict[str, Any],
    config: dict[str, Any],
    body_text: str,
    args: argparse.Namespace,
) -> tuple[str, Path, bytes]:
    """Save the email body as the original ENS text evidence."""
    filename = render_message_filename(
        str(config.get("body_text_file_pattern") or "ENS_Source_{dd.MM.yyyy}.txt"),
        message,
        config,
    )
    path = choose_run_unique_path(Path(config["body_text_folder"]) / filename, args)
    content_text = "\n".join(
        [
            f"Tenant: {config.get('tenant_code', '')}",
            f"Received: {message.get('receivedDateTime', '')}",
            f"From: {get_sender_email(message)}",
            f"Subject: {message.get('subject', '')}",
            "",
            body_text,
            "",
        ]
    )
    content = content_text.encode("utf-8")
    return save_bytes(path, content, args), path, content


def message_utc_date(message: dict[str, Any]) -> str:
    """Return the UTC received date used for ENS/DEC email pairing."""
    value = str(message.get("receivedDateTime") or "").strip()
    if not value:
        return ""
    try:
        received = dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return ""
    return received.astimezone(dt.timezone.utc).date().isoformat()


def normalise_pairing_subject(value: str) -> str:
    """Normalise subjects enough to pair ENS body details with later replies."""
    text = html.unescape(str(value or "")).strip().lower()
    while True:
        updated = re.sub(r"^(re|fw|fwd)\s*:\s*", "", text).strip()
        if updated == text:
            break
        text = updated
    return re.sub(r"\s+", " ", text)


def ens_source_keys(message: dict[str, Any], config: dict[str, Any]) -> list[tuple[str, str, str]]:
    """Return specific and date-level keys for matching ENS and DEC emails."""
    tenant_code = str(config.get("tenant_code") or "").strip().upper()
    received_date = message_utc_date(message)
    subject_key = normalise_pairing_subject(str(message.get("subject") or ""))
    keys: list[tuple[str, str, str]] = []
    if tenant_code and received_date and subject_key:
        keys.append((tenant_code, received_date, subject_key))
    if tenant_code and received_date:
        keys.append((tenant_code, received_date, ""))
    return keys


def remember_run_ens_source(
    message: dict[str, Any],
    config: dict[str, Any],
    body_text_path: Path | str,
    body_text: str,
) -> None:
    """Remember the ENS source email for a later Excel/DEC email in this run."""
    source = {
        "message": message,
        "path": body_text_path,
        "body_text": body_text,
    }
    for key in ens_source_keys(message, config):
        RUN_ENS_SOURCE_BY_KEY[key] = source


def saved_ens_source_body_text(content: str) -> str:
    """Strip the small saved-file header and return the original body text."""
    lines = content.splitlines()
    for index, line in enumerate(lines[:8]):
        if not line.strip():
            return "\n".join(lines[index + 1 :]).strip()
    return content.strip()


def find_existing_ens_source(message: dict[str, Any], config: dict[str, Any]) -> dict[str, Any] | None:
    """Find an already saved ENS source file for this tenant/date."""
    folder = Path(config.get("body_text_folder") or "")
    if not folder.exists():
        return None
    date_tag = received_date_for_filename(message)
    candidates = sorted(
        folder.glob(f"ENS_Source_{date_tag}_*.txt"),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    for candidate in candidates:
        try:
            content = candidate.read_text(encoding="utf-8-sig", errors="replace")
        except OSError:
            continue
        return {
            "message": message,
            "path": candidate,
            "body_text": saved_ens_source_body_text(content),
        }
    return None


def find_ens_source_for_dec_email(message: dict[str, Any], config: dict[str, Any]) -> dict[str, Any] | None:
    """Find the ENS source captured earlier in this run, or already on disk."""
    for key in ens_source_keys(message, config):
        source = RUN_ENS_SOURCE_BY_KEY.get(key)
        if source:
            return source
    return find_existing_ens_source(message, config)
def normalise_body_label(value: str) -> str:
    """Normalise body labels enough to match the known BKD email format."""
    text = html.unescape(value or "").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_ens_body_values(body_text: str) -> dict[str, dict[str, str]]:
    """Extract known ENS values from the email body label/value pattern."""
    lines = [line.strip() for line in body_text.splitlines() if line.strip()]
    values: dict[str, dict[str, str]] = {}

    for index, line in enumerate(lines):
        key = normalise_body_label(line)
        if key not in ENS_BODY_LABELS:
            continue

        api_field, note = ENS_BODY_LABELS[key]
        value = ""
        for candidate in lines[index + 1 : index + 6]:
            if normalise_body_label(candidate) in ENS_BODY_LABELS:
                break
            if candidate:
                value = candidate
                break

        values[api_field] = {
            "source_label": line,
            "source_value": value,
            "note": note,
        }

    return values


def ens_pack_status(api_field: str, source_value: str, required: str) -> tuple[str, str]:
    """Return a practical readiness status for an ENS PACK row."""
    if source_value and api_field not in CODE_OR_DATE_FIELDS:
        return "READY", source_value
    if source_value and api_field in CODE_OR_DATE_FIELDS:
        return "NEEDS_MAPPING", ""
    if required in {"yes", "conditional"}:
        return "NEEDS_SOURCE", ""
    return "INFO", ""


def build_ens_pack_rows(
    message: dict[str, Any],
    config: dict[str, Any],
    body_text_path: Path | str,
    body_text: str,
) -> list[dict[str, str]]:
    """Build the ENS PACK rows from email metadata and parsed body fields."""
    parsed = parse_ens_body_values(body_text)
    rows = [
        {
            "section": "EMAIL",
            "source_label": "receivedDateTime",
            "api_field": "",
            "source_value": str(message.get("receivedDateTime", "")),
            "api_value": "",
            "status": "INFO",
            "notes": "Email received timestamp from Graph.",
        },
        {
            "section": "EMAIL",
            "source_label": "sender",
            "api_field": "",
            "source_value": get_sender_email(message),
            "api_value": "",
            "status": "INFO",
            "notes": "Matched sender used for tenant routing.",
        },
        {
            "section": "EMAIL",
            "source_label": "subject",
            "api_field": "",
            "source_value": str(message.get("subject", "")),
            "api_value": "",
            "status": "INFO",
            "notes": "Original email subject.",
        },
        {
            "section": "EMAIL",
            "source_label": "body_text_path",
            "api_field": "",
            "source_value": str(body_text_path),
            "api_value": "",
            "status": "INFO",
            "notes": "Saved original email body evidence.",
        },
    ]

    for field in ENS_PACK_FIELDS:
        api_field = field["api_field"]
        required = field.get("required", "")
        source = parsed.get(api_field, {})
        source_value = source.get("source_value", "")
        status, api_value = ens_pack_status(api_field, source_value, required)
        rows.append(
            {
                "section": "ENS_HEADER",
                "source_label": source.get("source_label") or field.get("source_label", ""),
                "api_field": api_field,
                "source_value": source_value,
                "api_value": api_value,
                "status": status,
                "notes": field.get("note", "") or source.get("note", ""),
            }
        )

    return rows


def safe_sheet_name(value: str, fallback: str) -> str:
    """Return an Excel-safe sheet name."""
    text = re.sub(r"[\\/*?:\[\]]", "_", value or fallback).strip() or fallback
    return text[:31]


def style_pack_sheet(sheet: Any) -> None:
    """Apply small readable formatting to a generated pack sheet."""
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column[:100])
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)


def include_process_pack_source_header(header: str) -> bool:
    """Hide generated blank/technical source columns from the operational DEC PACK."""
    normalized = normalize_column_name(header)
    return not normalized.startswith("column ")


def append_pack_row(sheet: Any, values: list[Any]) -> None:
    """Append one row while keeping formula-looking source text as text."""
    sheet.append(values)
    row_number = sheet.max_row
    for column_number, value in enumerate(values, 1):
        if isinstance(value, str) and value.startswith("="):
            sheet.cell(row=row_number, column=column_number).data_type = "s"


def write_process_pack(
    message: dict[str, Any],
    config: dict[str, Any],
    body_text_path: Path | str,
    body_text: str,
    excel_sources: list[dict[str, Any]],
    args: argparse.Namespace,
    ens_message: dict[str, Any] | None = None,
) -> tuple[str, Path]:
    """Create the operational ENS/DEC Excel pack in the tenant Process folder."""
    from openpyxl import Workbook

    ens_message = ens_message or message

    filename = render_message_filename(
        str(config.get("process_pack_pattern") or "{tenant_code}_API_PACK_{dd.MM.yyyy}.xlsx"),
        message,
        config,
    )
    path = choose_run_unique_path(Path(config["process_folder"]) / filename, args)

    if path.exists() and not args.overwrite:
        return "SKIPPED_EXISTS", path
    if args.dry_run:
        return "DRY_RUN", path

    workbook = Workbook()
    ens_sheet = workbook.active
    ens_sheet.title = safe_sheet_name(str(config.get("ens_pack_sheet_name") or "ENS PACK"), "ENS PACK")
    ens_columns = ["section", "source_label", "api_field", "source_value", "api_value", "status", "notes"]
    append_pack_row(ens_sheet, ens_columns)
    for row in build_ens_pack_rows(ens_message, config, body_text_path, body_text):
        append_pack_row(ens_sheet, [row.get(column, "") for column in ens_columns])
    style_pack_sheet(ens_sheet)

    dec_sheet = workbook.create_sheet(safe_sheet_name(str(config.get("dec_pack_sheet_name") or "DEC PACK"), "DEC PACK"))
    source_rows: list[dict[str, str]] = []
    source_headers: list[str] = []
    for source in excel_sources:
        try:
            headers, rows = read_xlsx_rows(source["content"], config)
        except Exception as error:
            source_rows.append(
                {
                    "source_file": str(source.get("name", "")),
                    "saved_path": str(source.get("path", "")),
                    "source_row_number": "",
                    "pack_status": "ERROR",
                    "pack_notes": str(error),
                }
            )
            continue
        mapping = mapped_columns(headers, config)
        api_headers: list[str] = []
        for details in mapping.values():
            if details["kind"] != "api":
                continue
            api_header = f"api_{details['field_name']}"
            if api_header not in api_headers:
                api_headers.append(api_header)
        visible_headers = [header for header in headers if include_process_pack_source_header(header)]
        for header in [*api_headers, *visible_headers]:
            if header not in source_headers:
                source_headers.append(header)
        for index, row in enumerate(rows, 1):
            item = {
                "source_file": str(source.get("name", "")),
                "saved_path": str(source.get("path", "")),
                "source_row_number": str(index),
                "pack_status": "SOURCE_ROW",
                "pack_notes": "Original Excel row prepared for DEC/consignment mapping.",
            }
            item.update(api_ready_values_for_row(row, mapping))
            item.update(row)
            source_rows.append(item)

    dec_columns = ["source_file", "saved_path", "source_row_number", "pack_status", "pack_notes"] + source_headers
    append_pack_row(dec_sheet, dec_columns)
    if source_rows:
        for row in source_rows:
            append_pack_row(dec_sheet, [row.get(column, "") for column in dec_columns])
    else:
        append_pack_row(dec_sheet, ["", "", "", "NO_DEC_SOURCE", "No supported Excel attachment was available."] + [""] * len(source_headers))
    style_pack_sheet(dec_sheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return "SAVED", path

def normalise_graph_date(value: str, end_of_day: bool = False) -> str:
    """Accept YYYY-MM-DD or full ISO date and return a Graph UTC datetime."""
    value = value.strip()

    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        time_part = "23:59:59" if end_of_day else "00:00:00"
        return f"{value}T{time_part}Z"

    parsed = dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=dt.timezone.utc)
    parsed = parsed.astimezone(dt.timezone.utc).replace(microsecond=0)
    return parsed.isoformat().replace("+00:00", "Z")


def utc_today() -> dt.date:
    """Return today's date in UTC because Graph receivedDateTime is UTC."""
    return dt.datetime.now(dt.timezone.utc).date()


def date_only(value: str) -> str:
    """Return YYYY-MM-DD from a YYYY-MM-DD or ISO datetime value."""
    return normalise_graph_date(value)[:10]


def historic_start_date_from_configs() -> str:
    """Return the earliest configured historic start date across active customers."""
    candidates = []
    for config in TENANT_CONFIGS:
        value = str(config.get("historic_start_date") or "").strip()
        if value:
            candidates.append(date_only(value))

    if candidates:
        return min(candidates)
    return date_only(GRAPH_HISTORIC_START_DATE)


def apply_run_mode_dates(args: argparse.Namespace) -> None:
    """
    Resolve the date window before the mailbox query is built.

    Modes:
        daily    - default; reads only today's Graph emails.
        historic - reads from customer historic_start_date up to yesterday.
        custom   - uses only the dates passed by the operator.
    """
    today = utc_today()

    if args.run_mode == "daily":
        args.received_from = args.received_from or today.isoformat()
        args.received_to = args.received_to or today.isoformat()
        return

    if args.run_mode == "historic":
        yesterday = today - dt.timedelta(days=1)
        args.received_from = args.received_from or historic_start_date_from_configs()
        args.received_to = args.received_to or yesterday.isoformat()


# =============================================================================
# STEP 9 - EXCEL PARSING AND API READINESS VALIDATION
# =============================================================================

def normalize_column_name(value: str) -> str:
    """Return a stable comparison key for Excel headers and aliases."""
    text = str(value or "").strip().lower()
    text = re.sub(r"[_\-/]+", " ", text)
    text = re.sub(r"[^a-z0-9 ]+", "", text)
    return re.sub(r"\s+", " ", text).strip()


def config_aliases_for_field(config: dict[str, Any], field_name: str) -> list[str]:
    """Read optional tenant aliases from map_<field_name> config keys."""
    return [normalize_column_name(value) for value in as_list(config.get(f"map_{field_name}"))]


def build_alias_lookup(config: dict[str, Any]) -> dict[str, tuple[str, str]]:
    """Build normalized alias -> (field kind, field name)."""
    lookup: dict[str, tuple[str, str]] = {}

    for field_name, rule in FIELD_RULES.items():
        aliases = [*rule.get("aliases", []), *config_aliases_for_field(config, field_name)]
        for alias in aliases:
            normalized = normalize_column_name(alias)
            if normalized:
                lookup[normalized] = ("api", field_name)

    for field_name, aliases in AUDIT_COLUMN_ALIASES.items():
        tenant_aliases = config_aliases_for_field(config, field_name)
        for alias in [*aliases, *tenant_aliases]:
            normalized = normalize_column_name(alias)
            if normalized and normalized not in lookup:
                lookup[normalized] = ("audit", field_name)

    for alias in as_list(config.get("audit_columns")):
        normalized = normalize_column_name(alias)
        if normalized and normalized not in lookup:
            lookup[normalized] = ("audit", "tenant_audit")

    return lookup


def configured_field_source(config: dict[str, Any], field_name: str) -> str:
    """Return non-Excel source configured for an API field, if present."""
    for key in (f"source_{field_name}", f"default_{field_name}", f"derive_{field_name}"):
        value = str(config.get(key, "")).strip()
        if value:
            return value
    return ""


def read_shared_strings(workbook: zipfile.ZipFile) -> list[str]:
    """Read XLSX shared strings with only the standard library."""
    try:
        xml = workbook.read("xl/sharedStrings.xml")
    except KeyError:
        return []

    root = ET.fromstring(xml)
    ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    values = []
    for item in root.findall("s:si", ns):
        values.append("".join(text.text or "" for text in item.findall(".//s:t", ns)))
    return values


def column_index_from_reference(cell_reference: str) -> int:
    """Return zero-based column index from an Excel cell reference like B12."""
    letters = re.sub(r"[^A-Z]", "", cell_reference.upper())
    index = 0
    for letter in letters:
        index = index * 26 + (ord(letter) - ord("A") + 1)
    return max(index - 1, 0)


def workbook_target_path(target: str) -> str:
    """Return a zip path for a workbook relationship target."""
    text = str(target or "").strip()
    if text.startswith("/"):
        return text.lstrip("/")
    if text.startswith("xl/"):
        return text
    return "xl/" + text.lstrip("/")


def worksheet_paths(workbook: zipfile.ZipFile) -> list[str]:
    """Return worksheet paths in workbook order."""
    ns = {
        "s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    book_root = ET.fromstring(workbook.read("xl/workbook.xml"))
    rels_root = ET.fromstring(workbook.read("xl/_rels/workbook.xml.rels"))
    relationships = {relationship.get("Id"): relationship.get("Target", "") for relationship in rels_root}
    paths: list[str] = []

    for sheet in book_root.findall("s:sheets/s:sheet", ns):
        relationship_id = sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        target = relationships.get(relationship_id)
        if target:
            paths.append(workbook_target_path(target))

    if not paths:
        raise ValueError("Workbook has no worksheets")
    return paths


def worksheet_path(workbook: zipfile.ZipFile) -> str:
    """Return the first worksheet path from an XLSX file."""
    return worksheet_paths(workbook)[0]


def cell_text(cell: ET.Element, shared_strings: list[str]) -> str:
    """Return a readable value from one XLSX cell."""
    ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    cell_type = cell.get("t", "")

    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.findall(".//s:t", ns)).strip()

    value = cell.find("s:v", ns)
    if value is None or value.text is None:
        return ""

    raw = value.text.strip()
    if cell_type == "s" and raw.isdigit():
        index = int(raw)
        if 0 <= index < len(shared_strings):
            return shared_strings[index].strip()
    return raw


def header_candidate_score(row: list[str], alias_lookup: dict[str, tuple[str, str]]) -> int:
    """Score a row as a possible Excel table header."""
    business_words = {
        "address",
        "amount",
        "commodity",
        "consignee",
        "consignor",
        "country",
        "currency",
        "customer",
        "description",
        "document",
        "email",
        "eori",
        "exporter",
        "gross",
        "importer",
        "invoice",
        "item",
        "line",
        "marks",
        "measure",
        "number",
        "origin",
        "package",
        "packages",
        "postcode",
        "price",
        "quantity",
        "qty",
        "ship",
        "tariff",
        "unit",
        "weight",
    }
    technical_words = {
        "autohide",
        "autotable",
        "fields",
        "filters",
        "formula",
        "formulas",
        "headers",
        "hide",
        "linkfield",
        "links",
        "lookup",
        "option",
        "tables",
        "values",
    }

    normalized_values = [normalize_column_name(value) for value in row if str(value or "").strip()]
    if len(normalized_values) < 2:
        return -100

    score = min(len(normalized_values), 25)
    for normalized in normalized_values:
        words = set(normalized.split())
        if normalized in alias_lookup:
            score += 10
        score += len(words & business_words)
        if any(word in normalized for word in technical_words):
            score -= 8

    return score


def select_header_candidate(raw_rows: list[list[str]], config: dict[str, Any] | None = None) -> tuple[int, int]:
    """Choose the most likely business header row and return its score."""
    alias_lookup = build_alias_lookup(config or {})
    best_index = 0
    best_score = -999

    for index, row in enumerate(raw_rows[:60]):
        score = header_candidate_score(row, alias_lookup)
        if score > best_score:
            best_index = index
            best_score = score

    if best_score > 2:
        return best_index, best_score

    for index, row in enumerate(raw_rows[:10]):
        if sum(1 for value in row if value.strip()) >= 2:
            return index, header_candidate_score(row, alias_lookup)
    return 0, best_score


def select_header_index(raw_rows: list[list[str]], config: dict[str, Any] | None = None) -> int:
    """Choose the most likely business header row from an Excel worksheet."""
    return select_header_candidate(raw_rows, config)[0]


def unique_headers(raw_header_row: list[str]) -> list[str]:
    """Return readable unique headers while keeping blank columns obvious."""
    headers: list[str] = []
    seen: dict[str, int] = {}

    for index, value in enumerate(raw_header_row):
        header = value.strip() or f"Column {index + 1}"
        count = seen.get(header, 0) + 1
        seen[header] = count
        if count > 1:
            header = f"{header} ({count})"
        headers.append(header)

    return headers


def is_summary_excel_row(row_data: dict[str, str], header_count: int) -> bool:
    """Skip worksheet total rows that are not customer goods lines."""
    values = [normalize_column_name(value) for value in row_data.values() if str(value or "").strip()]
    if not values:
        return True
    has_total_marker = any(value in {"total", "grand total"} for value in values)
    return has_total_marker and len(values) < max(6, header_count // 2)

def raw_rows_from_sheet_xml(sheet_xml: bytes, shared_strings: list[str]) -> list[list[str]]:
    """Read one worksheet XML into sparse row values."""
    ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(sheet_xml)
    raw_rows: list[list[str]] = []

    for row in root.findall(".//s:sheetData/s:row", ns):
        values: list[str] = []
        for cell in row.findall("s:c", ns):
            column_index = column_index_from_reference(cell.get("r", "A1"))
            while len(values) <= column_index:
                values.append("")
            values[column_index] = cell_text(cell, shared_strings)
        if any(value.strip() for value in values):
            raw_rows.append(values)
    return raw_rows


def is_export_formula_text(value: str) -> bool:
    """Return True when an exported cell contains formula mechanics, not business data."""
    text = str(value or "").strip().lower()
    return text.startswith("=") or text.startswith("formula(=")


def replace_export_formula_values(row_data: dict[str, str]) -> dict[str, str]:
    """Prefer calculated export values over formula text in business columns."""
    fallback_columns = {
        "Amount": ["Sales Amount 1", "Amount 1", "Amount1"],
        "Line Amount Excl. VAT": ["Line Amount Excl. VAT 1", "Line Amount Excl. VAT1"],
        "Unit Price Excl. VAT": ["Unit Price Excl. VAT1", "Unit Price Excl. VAT 1"],
    }

    for header, candidates in fallback_columns.items():
        if not is_export_formula_text(row_data.get(header, "")):
            continue
        for candidate in candidates:
            value = str(row_data.get(candidate, "")).strip()
            if value:
                row_data[header] = value
                break
    return row_data


def rows_from_raw_rows(raw_rows: list[list[str]], header_index: int) -> tuple[list[str], list[dict[str, str]]]:
    """Build headers and dictionaries from a selected header row."""
    headers = unique_headers(raw_rows[header_index])
    rows: list[dict[str, str]] = []

    for raw_row in raw_rows[header_index + 1:]:
        row_data = {}
        for index, header in enumerate(headers):
            row_data[header] = raw_row[index].strip() if index < len(raw_row) else ""
        row_data = replace_export_formula_values(row_data)
        if any(value.strip() for value in row_data.values()) and not is_summary_excel_row(row_data, len(headers)):
            rows.append(row_data)

    return headers, rows


def read_xlsx_rows(content: bytes, config: dict[str, Any] | None = None) -> tuple[list[str], list[dict[str, str]]]:
    """Read the best business worksheet into headers and row dictionaries."""
    best: tuple[int, int, list[str], list[dict[str, str]]] | None = None
    with zipfile.ZipFile(io.BytesIO(content)) as workbook:
        shared_strings = read_shared_strings(workbook)
        for sheet_index, sheet_path in enumerate(worksheet_paths(workbook)):
            raw_rows = raw_rows_from_sheet_xml(workbook.read(sheet_path), shared_strings)
            if not raw_rows:
                continue
            header_index, score = select_header_candidate(raw_rows, config)
            headers, rows = rows_from_raw_rows(raw_rows, header_index)
            candidate = (score, len(rows), headers, rows)
            if best is None or candidate[:2] > best[:2]:
                best = candidate

    if best is None:
        return [], []
    return best[2], best[3]

def add_validation_issue(
    validation_rows: list[dict[str, str]],
    message: dict[str, Any],
    config: dict[str, Any],
    attachment_name: str,
    saved_path: Path | str,
    row_number: int,
    source_column: str,
    technical_field: str,
    severity: str,
    rule: str,
    details: str,
    normalized_value: str = "",
) -> None:
    """Append one validation issue to the run validation report."""
    validation_rows.append(
        {
            "receivedDateTime": message.get("receivedDateTime", ""),
            "tenantCode": config.get("tenant_code", ""),
            "tenantName": config.get("tenant_name", ""),
            "sender": get_sender_email(message),
            "subject": message.get("subject", ""),
            "attachmentName": attachment_name,
            "savedPath": str(saved_path),
            "rowNumber": str(row_number),
            "sourceColumn": source_column,
            "technicalField": technical_field,
            "severity": severity,
            "rule": rule,
            "details": details,
            "normalizedValue": normalized_value,
        }
    )


def mapped_columns(headers: list[str], config: dict[str, Any]) -> dict[str, dict[str, str]]:
    """Return mapping details keyed by source header."""
    lookup = build_alias_lookup(config)
    mapping = {}
    for header in headers:
        normalized = normalize_column_name(header)
        if normalized.startswith("column ") or normalized in {"autotable", "auto table"}:
            mapping[header] = {"kind": "ignored", "field_name": ""}
            continue
        kind, field_name = lookup.get(normalized, ("unmapped", ""))
        mapping[header] = {"kind": kind, "field_name": field_name}
    return mapping


def first_source_column_for_field(mapping: dict[str, dict[str, str]], field_name: str) -> str:
    """Return the source column mapped to a technical field, if any."""
    for source_column, details in mapping.items():
        if details["field_name"] == field_name:
            return source_column
    return ""


def validate_required_mappings(
    headers: list[str],
    mapping: dict[str, dict[str, str]],
    validation_rows: list[dict[str, str]],
    message: dict[str, Any],
    config: dict[str, Any],
    attachment_name: str,
    saved_path: Path | str,
) -> None:
    """Report every source column and missing mandatory API mappings."""
    for source_column, details in mapping.items():
        if details["kind"] == "unmapped":
            add_validation_issue(
                validation_rows,
                message,
                config,
                attachment_name,
                saved_path,
                0,
                source_column,
                "",
                "PENDING",
                "Unmapped source column",
                "Column is present in the Excel but not currently mapped to API, derived, audit, or ignored data.",
            )
        elif details["kind"] == "audit":
            add_validation_issue(
                validation_rows,
                message,
                config,
                attachment_name,
                saved_path,
                0,
                source_column,
                details["field_name"],
                "INFO",
                "Audit column accounted for",
                "Column is retained for traceability/support and is not sent directly to the API.",
            )

    for field_name, rule in FIELD_RULES.items():
        if not rule.get("required"):
            continue
        if not first_source_column_for_field(mapping, field_name):
            configured_source = configured_field_source(config, field_name)
            if configured_source:
                add_validation_issue(
                    validation_rows,
                    message,
                    config,
                    attachment_name,
                    saved_path,
                    0,
                    configured_source,
                    field_name,
                    "INFO",
                    "Mandatory API field source accounted for",
                    "Field is not expected as an Excel column; it must be resolved from this configured source before API submission.",
                )
                continue
            add_validation_issue(
                validation_rows,
                message,
                config,
                attachment_name,
                saved_path,
                0,
                "",
                field_name,
                "ERROR",
                "Missing API mandatory mapping",
                "No Excel column is mapped to this mandatory API field. Provide a column or approved tenant/master-data source.",
            )


def is_blank(value: str) -> bool:
    """Return True for empty Excel values."""
    return str(value or "").strip() == ""


def decimal_parts(value: str) -> tuple[str, str]:
    """Return integer and decimal parts for simple numeric validation."""
    text = value.strip().replace(" ", "")
    if text.startswith("-"):
        text = text[1:]
    if "." in text:
        left, right = text.split(".", 1)
        return left, right
    return text, ""


def decimal_quant(max_decimals: int) -> Decimal:
    """Return a Decimal quantizer for the configured scale."""
    return Decimal(1).scaleb(-max_decimals)


def format_decimal_for_api(value: Decimal) -> str:
    """Format a Decimal without forcing trailing zeros."""
    text = format(value, "f").rstrip("0").rstrip(".")
    if text in {"", "-0"}:
        return "0"
    return text


def normalise_decimal_for_api(field_name: str, text: str, rule: dict[str, Any]) -> tuple[str, list[tuple[str, str, str]]]:
    """Validate and round a decimal value for API-ready output."""
    issues: list[tuple[str, str, str]] = []
    compact = text.replace(" ", "")
    if "," in compact:
        if not re.fullmatch(r"-?\d{1,3}(,\d{3})+(\.\d+)?", compact):
            return "", [("ERROR", "Decimal format", "Commas are only accepted as thousands separators.")]
        cleaned = compact.replace(",", "")
    else:
        cleaned = compact

    if not re.fullmatch(r"-?\d+(\.\d+)?", cleaned):
        return "", [("ERROR", "Decimal value", "Expected a valid number.")]

    try:
        number = Decimal(cleaned)
    except InvalidOperation:
        return "", [("ERROR", "Decimal value", "Expected a valid number.")]

    max_decimals = int(rule.get("max_decimals", 2))
    rounded = number.quantize(decimal_quant(max_decimals), rounding=ROUND_HALF_UP)
    normalized = format_decimal_for_api(rounded)
    _, right = decimal_parts(cleaned)

    if len(right) > max_decimals:
        issues.append(("WARNING", "Decimal rounded", f"Value has {len(right)} decimals; rounded to {normalized}."))

    left, rounded_right = decimal_parts(normalized)
    total_digits = len(left.lstrip("0")) + len(rounded_right)
    if total_digits > int(rule.get("max_digits", total_digits)):
        issues.append(("ERROR", "Decimal max digits", f"Maximum total digits is {rule.get('max_digits')}."))

    if rule.get("must_be_positive") and number <= 0:
        issues.append(("ERROR", "Decimal range", f"{field_name} must be greater than zero."))

    return normalized, issues


def normalise_field_value_for_api(field_name: str, value: str) -> tuple[str, list[tuple[str, str, str]]]:
    """Return API-ready value plus validation issues."""
    rule = FIELD_RULES[field_name]
    issues: list[tuple[str, str, str]] = []
    text = str(value or "").strip()

    if rule.get("required") and is_blank(text):
        return "", [("ERROR", "Missing API mandatory value", "Value is required by the future API contract.")]
    if is_blank(text):
        return "", []
    if text.startswith("="):
        return "", [("PENDING", "Formula value", "Excel formula must be resolved to a calculated value before API validation.")]

    normalized_value = text
    max_length = rule.get("max_length")
    if max_length and len(text) > int(max_length):
        issues.append(("ERROR", "API maximum length", f"Value has {len(text)} characters; maximum is {max_length}."))

    value_type = rule.get("value_type")
    if value_type == "yes_no" and text.lower() not in {"yes", "no", "y", "n", "true", "false", "1", "0"}:
        issues.append(("ERROR", "Yes/no value", "Expected yes/no, true/false, or 1/0."))

    if value_type == "integer":
        if not re.fullmatch(r"\d+", text):
            issues.append(("ERROR", "Integer value", "Expected a whole number."))
        else:
            number = int(text)
            if number < int(rule.get("min", 0)) or number > int(rule.get("max", number)):
                issues.append(("ERROR", "Integer range", f"Expected value between {rule.get('min')} and {rule.get('max')}."))

    if value_type == "decimal":
        normalized_value, decimal_issues = normalise_decimal_for_api(field_name, text, rule)
        issues.extend(decimal_issues)

    if value_type == "eori":
        if not re.fullmatch(r"[A-Za-z]{2}[A-Za-z0-9]{1,15}", text):
            issues.append(("ERROR", "EORI format", "Expected country prefix plus alphanumeric EORI, maximum 17 characters."))
        if field_name in {"consignor_eori", "consignee_eori"} and text.upper().startswith("GB"):
            issues.append(("ERROR", "EORI role rule", "API guidance says GB EORI is not accepted for consignor_eori or consignee_eori."))

    if value_type == "commodity_code":
        if not re.fullmatch(r"\d+", text):
            issues.append(("ERROR", "Commodity code format", "Commodity code must contain digits only."))
        elif len(text) < 6 or len(text) == 7 or len(text) > 10:
            issues.append(("ERROR", "Commodity code length", "Expected 8-10 digits, or 6 digits only when APC is 1SG."))
        elif len(text) == 6:
            issues.append(("WARNING", "Conditional commodity code", "6 digits are only valid when APC is 1SG; confirm before API submission."))

    if value_type == "country_code" and not re.fullmatch(r"[A-Za-z]{2}", text):
        issues.append(("ERROR", "Country code", "Expected a 2-letter country code."))

    if value_type == "currency_code" and not re.fullmatch(r"[A-Za-z]{3}", text):
        issues.append(("WARNING", "Currency code", "Expected a 3-letter currency code when supplied."))

    return normalized_value, issues


def validate_field_value(field_name: str, value: str) -> list[tuple[str, str, str]]:
    """Return validation issues as severity/rule/details tuples."""
    return normalise_field_value_for_api(field_name, value)[1]


def api_ready_values_for_row(row: dict[str, str], mapping: dict[str, dict[str, str]]) -> dict[str, str]:
    """Return api_<field> values from mapped source columns without altering source data."""
    values: dict[str, str] = {}
    for source_column, details in mapping.items():
        if details["kind"] != "api":
            continue
        field_name = details["field_name"]
        api_column = f"api_{field_name}"
        if api_column in values and values[api_column]:
            continue
        normalized_value, _ = normalise_field_value_for_api(field_name, row.get(source_column, ""))
        values[api_column] = normalized_value
    return values


def decimal_for_compare(field_name: str, row: dict[str, str], mapping: dict[str, dict[str, str]]) -> Decimal | None:
    """Return a normalized Decimal for cross-field checks when valid and present."""
    source_column = first_source_column_for_field(mapping, field_name)
    if not source_column:
        return None
    normalized_value, issues = normalise_field_value_for_api(field_name, row.get(source_column, ""))
    if not normalized_value or any(severity == "ERROR" for severity, _, _ in issues):
        return None
    try:
        return Decimal(normalized_value)
    except InvalidOperation:
        return None


def validate_row_relationships(
    row: dict[str, str],
    mapping: dict[str, dict[str, str]],
) -> list[tuple[str, str, str, str]]:
    """Return cross-field API blockers as severity/rule/details/normalized tuples."""
    issues: list[tuple[str, str, str, str]] = []
    gross = decimal_for_compare("gross_mass_kg", row, mapping)
    net = decimal_for_compare("net_mass_kg", row, mapping)
    if gross is not None and net is not None and net > gross:
        issues.append((
            "ERROR",
            "Net/gross mass",
            "net_mass_kg cannot be greater than gross_mass_kg.",
            f"net_mass_kg={format_decimal_for_api(net)}; gross_mass_kg={format_decimal_for_api(gross)}",
        ))
    return issues


def validate_xlsx_attachment(
    message: dict[str, Any],
    config: dict[str, Any],
    attachment_name: str,
    content: bytes,
    saved_path: Path | str,
    validation_rows: list[dict[str, str]],
    stats: dict[str, Any],
) -> None:
    """Parse and validate one Excel attachment directly from Graph bytes."""
    suffix = Path(attachment_name).suffix.lower()
    if suffix not in EXCEL_SUFFIXES:
        return

    stats["validated_excels"] += 1

    try:
        headers, data_rows = read_xlsx_rows(content, config)
    except Exception as error:
        stats["validation_errors"] += 1
        add_validation_issue(
            validation_rows,
            message,
            config,
            attachment_name,
            saved_path,
            0,
            "",
            "",
            "ERROR",
            "Unreadable Excel",
            str(error),
        )
        return

    mapping = mapped_columns(headers, config)
    issue_start = len(validation_rows)
    validate_required_mappings(headers, mapping, validation_rows, message, config, attachment_name, saved_path)

    for row_index, row in enumerate(data_rows, 2):
        for source_column, details in mapping.items():
            if details["kind"] != "api":
                continue
            field_name = details["field_name"]
            normalized_value, field_issues = normalise_field_value_for_api(field_name, row.get(source_column, ""))
            for severity, rule, details_text in field_issues:
                if severity == "ERROR":
                    stats["validation_errors"] += 1
                elif severity == "WARNING":
                    stats["validation_warnings"] += 1
                else:
                    stats["validation_pending"] += 1
                add_validation_issue(
                    validation_rows,
                    message,
                    config,
                    attachment_name,
                    saved_path,
                    row_index,
                    source_column,
                    field_name,
                    severity,
                    rule,
                    details_text,
                    normalized_value,
                )

        for severity, rule, details_text, normalized_value in validate_row_relationships(row, mapping):
            if severity == "ERROR":
                stats["validation_errors"] += 1
            elif severity == "WARNING":
                stats["validation_warnings"] += 1
            else:
                stats["validation_pending"] += 1
            add_validation_issue(
                validation_rows,
                message,
                config,
                attachment_name,
                saved_path,
                row_index,
                "",
                "net_mass_kg",
                severity,
                rule,
                details_text,
                normalized_value,
            )

    # Missing mandatory mappings are added after row counters so the summary also
    # reflects file-level errors.
    for issue in validation_rows[issue_start:]:
        if issue["rule"] == "Missing API mandatory mapping":
            stats["validation_errors"] += 1

# =============================================================================
# STEP 10 - QUALITY REPORT HELPERS
# =============================================================================

def new_stats() -> dict[str, Any]:
    """Create simple counters for console output and support checks."""
    return {
        "scanned": 0,
        "matched": 0,
        "unmatched": 0,
        "saved_attachments": 0,
        "saved_body_texts": 0,
        "generated_process_packs": 0,
        "missing_ens_sources": 0,
        "skipped_existing": 0,
        "no_file_attachments": 0,
        "failed": 0,
        "validated_excels": 0,
        "validation_errors": 0,
        "validation_warnings": 0,
        "validation_pending": 0,
        "ing_rows": 0,
        "database_enabled": False,
        "tenant_config_source": "YML",
        "file_types": {},
    }


def count_file_type(stats: dict[str, Any], suffix: str) -> None:
    """Count saved file types for historic review."""
    suffix = suffix.lower() or ".no_extension"
    file_types = stats["file_types"]
    file_types[suffix] = file_types.get(suffix, 0) + 1


def report_row(
    message: dict[str, Any],
    config: dict[str, Any] | None,
    action: str,
    saved_path: Path | str = "",
    file_type: str = "",
    note: str = "",
) -> dict[str, str]:
    """Build one technical log row."""
    return {
        "receivedDateTime": message.get("receivedDateTime", ""),
        "tenantCode": config["tenant_code"] if config else "",
        "tenantName": config["tenant_name"] if config else "",
        "sender": get_sender_email(message),
        "subject": message.get("subject", ""),
        "action": action,
        "fileType": file_type,
        "savedPath": str(saved_path),
        "note": note,
        "messageId": message.get("internetMessageId") or message.get("id", ""),
    }


def write_run_log(
    rows: list[dict[str, str]],
    stats: dict[str, Any],
    started_at: dt.datetime,
    args: argparse.Namespace,
) -> Path:
    """Write a small technical run log; this is not the business objective."""
    RUN_LOG_FOLDER.mkdir(parents=True, exist_ok=True)
    run_type = "dry_run" if args.dry_run else "run"
    path = RUN_LOG_FOLDER / f"graph_download_run_log_{started_at:%Y%m%d_%H%M%S}_{run_type}.csv"

    columns = [
        "receivedDateTime",
        "tenantCode",
        "tenantName",
        "sender",
        "subject",
        "action",
        "fileType",
        "savedPath",
        "note",
        "messageId",
    ]

    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)
        writer.writerow({})
        writer.writerow({"action": "SUMMARY", "note": f"Scanned: {stats['scanned']}"})
        writer.writerow({"action": "SUMMARY", "note": f"Matched: {stats['matched']}"})
        writer.writerow({"action": "SUMMARY", "note": f"Unmatched: {stats['unmatched']}"})
        writer.writerow({"action": "SUMMARY", "note": f"Saved attachments: {stats['saved_attachments']}"})
        writer.writerow({"action": "SUMMARY", "note": f"Saved ENS source texts: {stats['saved_body_texts']}"})
        writer.writerow({"action": "SUMMARY", "note": f"Generated process packs: {stats['generated_process_packs']}"})
        writer.writerow({"action": "SUMMARY", "note": f"Missing ENS sources: {stats['missing_ens_sources']}"})
        writer.writerow({"action": "SUMMARY", "note": f"Skipped existing: {stats['skipped_existing']}"})
        writer.writerow({
            "action": "SUMMARY",
            "note": (
                "No file attachments: "
                f"{stats['no_file_attachments']}"
            ),
        })
        writer.writerow({"action": "SUMMARY", "note": f"Failed: {stats['failed']}"})
        writer.writerow({"action": "SUMMARY", "note": f"Validation errors: {stats['validation_errors']}"})
        writer.writerow({"action": "SUMMARY", "note": f"Validation warnings: {stats['validation_warnings']}"})
        writer.writerow({"action": "SUMMARY", "note": f"Validation pending: {stats['validation_pending']}"})
        if stats["validation_errors"]:
            writer.writerow({
                "action": "VALIDATION_BLOCKERS",
                "note": "Local download completed; later API validation may reject validation blockers.",
            })

        for suffix, count in sorted(stats["file_types"].items()):
            writer.writerow({"action": "FILE_TYPE", "fileType": suffix, "note": str(count)})

    return path



def write_validation_report(
    rows: list[dict[str, str]],
    started_at: dt.datetime,
    args: argparse.Namespace,
) -> Path | None:
    """Write Excel mapping/API validation output for support and QA."""
    if not rows:
        return None

    RUN_LOG_FOLDER.mkdir(parents=True, exist_ok=True)
    run_type = "dry_run" if args.dry_run else "run"
    path = RUN_LOG_FOLDER / f"graph_validation_report_{started_at:%Y%m%d_%H%M%S}_{run_type}.csv"

    columns = [
        "receivedDateTime",
        "tenantCode",
        "tenantName",
        "sender",
        "subject",
        "attachmentName",
        "savedPath",
        "rowNumber",
        "sourceColumn",
        "technicalField",
        "severity",
        "rule",
        "details",
        "normalizedValue",
    ]

    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)

    return path


def record_ing_trace(
    db_trace: DatabaseIngestionTrace | None,
    stats: dict[str, Any],
    message: dict[str, Any],
    config: dict[str, Any] | None,
    status: str,
    **kwargs: Any,
) -> int | None:
    """Write one optional ING.Graph row, update the counter, and return GraphID."""
    if not db_trace:
        return None
    graph_id = db_trace.insert_ing_trace(message, config, status, **kwargs)
    if graph_id:
        stats["ing_rows"] += 1
    return graph_id


# =============================================================================
# STEP 11 - PROCESS ONE EMAIL
# =============================================================================

def process_one_message(
    token: str,
    message: dict[str, Any],
    stats: dict[str, Any],
    rows: list[dict[str, str]],
    validation_rows: list[dict[str, str]],
    args: argparse.Namespace,
    db_trace: DatabaseIngestionTrace | None = None,
) -> None:
    """Process one email, supporting BKD's separate ENS body and DEC Excel emails."""
    stats["scanned"] += 1
    config = find_tenant_config(message)

    if not config:
        stats["unmatched"] += 1
        status = "SKIPPED_UNMATCHED_SENDER"
        rows.append(report_row(message, None, status))
        record_ing_trace(
            db_trace,
            stats,
            message,
            None,
            status,
            has_attachments=bool(message.get("hasAttachments")),
        )
        return

    stats["matched"] += 1

    try:
        body_text = email_body_as_text(message)
        if body_text_enabled(config) and body_text:
            body_text = trim_ens_source_body_text(body_text, config)
        attachments = []
        if message.get("hasAttachments"):
            attachments = read_attachments(token, message["id"])

        all_file_attachments = file_attachments_only(attachments)
        file_attachments = [
            attachment
            for attachment in all_file_attachments
            if attachment_allowed_for_config(attachment, config)
        ]

        if all_file_attachments and not file_attachments:
            status = "SKIPPED_UNSUPPORTED_FILE_TYPE"
            allowed = ", ".join(config.get("file_types", [])) or "any"
            rows.append(
                report_row(
                    message,
                    config,
                    status,
                    note=f"No attachment matched allowed file types: {allowed}",
                )
            )
            record_ing_trace(
                db_trace,
                stats,
                message,
                config,
                status,
                has_attachments=True,
                source_part="ATTACHMENT",
                fail_reason=f"No attachment matched allowed file types: {allowed}",
            )
            return

        has_supported_files = bool(file_attachments)
        body_path: Path | str = ""
        body_content: bytes | None = None

        if body_text_enabled(config) and body_text and not has_supported_files:
            body_result, body_path, body_content = save_ens_body_text(message, config, body_text, args)
            if body_result in {"SAVED", "DRY_RUN"}:
                stats["saved_body_texts"] += 1
            if body_result in {"SAVED", "DRY_RUN", "SKIPPED_EXISTS"}:
                remember_run_ens_source(message, config, body_path, body_text)
            rows.append(
                report_row(
                    message,
                    config,
                    f"{body_result}_ENS_SOURCE_TEXT",
                    saved_path=body_path,
                    file_type=".txt",
                    note="Original email body saved as ENS source evidence.",
                )
            )
            record_ing_trace(
                db_trace,
                stats,
                message,
                config,
                f"{body_result}_ENS_SOURCE_TEXT",
                original_file_name=Path(str(body_path)).name if body_path else "ENS_Source.txt",
                saved_path=body_path,
                content_type="text/plain",
                size_bytes=len(body_content or b""),
                content=body_content,
                has_attachments=bool(message.get("hasAttachments")),
                pack_code="ENS_PACK",
                source_part="EMAIL_BODY",
                load_status=body_result,
            )

        excel_sources: list[dict[str, Any]] = []

        if file_attachments:
            for attachment in file_attachments:
                if not attachment.get("contentBytes") and attachment.get("id"):
                    attachment = read_one_attachment(token, message["id"], attachment["id"])

                result, path, content = save_attachment_file(message, config, attachment, args)
                suffix = path.suffix.lower() or ".no_extension"
                count_file_type(stats, suffix)

                if result in {"SAVED", "DRY_RUN"}:
                    stats["saved_attachments"] += 1
                elif result == "SKIPPED_EXISTS":
                    stats["skipped_existing"] += 1

                if suffix in EXCEL_SUFFIXES:
                    excel_sources.append(
                        {
                            "name": attachment.get("name", ""),
                            "path": path,
                            "content": content,
                        }
                    )

                validate_xlsx_attachment(
                    message,
                    config,
                    attachment.get("name", ""),
                    content,
                    path,
                    validation_rows,
                    stats,
                )

                status = f"{result}_ATTACHMENT"
                rows.append(
                    report_row(
                        message,
                        config,
                        status,
                        saved_path=path,
                        file_type=suffix,
                        note=attachment.get("name", ""),
                    )
                )
                record_ing_trace(
                    db_trace,
                    stats,
                    message,
                    config,
                    status,
                    original_file_name=attachment.get("name", ""),
                    saved_path=path,
                    content_type=attachment.get("contentType", ""),
                    size_bytes=attachment.get("size"),
                    content=content,
                    has_attachments=True,
                    pack_code="DEC_PACK",
                    source_part="ATTACHMENT",
                    load_status=status,
                )
        else:
            stats["no_file_attachments"] += 1
            status = "NO_FILE_ATTACHMENTS"
            rows.append(
                report_row(
                    message,
                    config,
                    status,
                    note="No file attachments found; stored as ENS source when configured.",
                )
            )
            record_ing_trace(
                db_trace,
                stats,
                message,
                config,
                status,
                has_attachments=False,
                pack_code="ENS_PACK" if body_text_enabled(config) else "",
                source_part="EMAIL_BODY" if body_text_enabled(config) else "",
                load_status=status,
            )
            return

        if body_text_enabled(config) and excel_sources:
            ens_source = find_ens_source_for_dec_email(message, config)
            if not ens_source:
                stats["missing_ens_sources"] += 1
                status = "MISSING_ENS_SOURCE"
                rows.append(
                    report_row(
                        message,
                        config,
                        status,
                        file_type=".xlsx",
                        note="Excel saved, but no ENS source text was found for the same tenant/date/subject.",
                    )
                )
                record_ing_trace(
                    db_trace,
                    stats,
                    message,
                    config,
                    status,
                    has_attachments=True,
                    pack_code="API_PACK",
                    source_part="GENERATED_PACK",
                    load_status=status,
                    fail_reason="Missing ENS source text for DEC Excel email.",
                )
                return

            pack_result, pack_path = write_process_pack(
                message,
                config,
                ens_source.get("path", ""),
                str(ens_source.get("body_text") or ""),
                excel_sources,
                args,
                ens_message=ens_source.get("message") or message,
            )
            if pack_result in {"SAVED", "DRY_RUN"}:
                stats["generated_process_packs"] += 1
            rows.append(
                report_row(
                    message,
                    config,
                    f"{pack_result}_PROCESS_PACK",
                    saved_path=pack_path,
                    file_type=".xlsx",
                    note="Generated operational ENS PACK + DEC PACK workbook from paired ENS and DEC emails.",
                )
            )
            pack_content: bytes | None = None
            if not args.dry_run and Path(pack_path).exists():
                try:
                    pack_content = Path(pack_path).read_bytes()
                except OSError:
                    pack_content = None
            pack_graph_id = record_ing_trace(
                db_trace,
                stats,
                message,
                config,
                f"{pack_result}_PROCESS_PACK",
                original_file_name=pack_path.name,
                saved_path=pack_path,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                size_bytes=len(pack_content) if pack_content is not None else None,
                content=pack_content,
                has_attachments=True,
                pack_code="API_PACK",
                source_part="GENERATED_PACK",
                generated_csv_path=pack_path,
                load_status=pack_result,
            )
            if pack_result == "SAVED" and db_trace and pack_graph_id:
                file_hash = hashlib.sha256(pack_content).hexdigest() if pack_content is not None else None
                db_trace.insert_process_pack_load(pack_graph_id, config, pack_path, file_hash=file_hash)

    except Exception as error:
        stats["failed"] += 1
        rows.append(report_row(message, config, "FAILED_MESSAGE", note=str(error)))
        try:
            record_ing_trace(db_trace, stats, message, config, "FAILED_MESSAGE", fail_reason=str(error))
        except Exception as trace_error:
            print(f"ERROR writing failed ING trace for message {message.get('id')}: {trace_error}")
        print(f"ERROR processing message {message.get('id')}: {error}")
# =============================================================================
# STEP 12 - MAIN SCRIPT FLOW
# =============================================================================

def run(args: argparse.Namespace) -> int:
    r"""
    Main execution flow.

    Main objective:
        Save inbound files into the correct tenant folder under:
        \\PL-AZ-SDF-PLINT\Fusion_Production\Synovia_Flow_Quality\Integration_Layer

    Current confirmed example:
        BKD files are saved into:
        \\PL-AZ-SDF-PLINT\Fusion_Production\Synovia_Flow_Quality\Integration_Layer\BKD\Inbound\Sales_Order_files

    This is the order the operator should understand:
        1. Load optional database routing/trace support.
        2. Validate Graph configuration.
        3. Get Graph token.
        4. Read emails from the mailbox.
        5. For each email, find the tenant by sender.
        6. Save file attachments and optionally write ING.Graph trace rows.
        7. Print a summary and keep support run evidence.
    """
    global TENANT_CONFIGS

    started_at = dt.datetime.now(dt.timezone.utc)
    stats = new_stats()
    rows: list[dict[str, str]] = []
    validation_rows: list[dict[str, str]] = []
    db_trace: DatabaseIngestionTrace | None = None

    try:
        print("STEP 1 - Preparing optional database ingestion trace")
        db_trace = DatabaseIngestionTrace.open(args)
        if db_trace:
            stats["database_enabled"] = True
            database_configs = db_trace.load_tenant_configs()
            if database_configs:
                TENANT_CONFIGS = database_configs
                stats["tenant_config_source"] = "CFG.Graph"
                print(f"Loaded {len(database_configs)} active CFG.Graph route(s) for {GRAPH_MAILBOX}")
            else:
                print("No active CFG.Graph rows matched this mailbox; using customer YML routing")

            db_trace.begin_execution(args, started_at)
            db_trace.insert_execution_log("START", "Graph ingestion run started", detail={
                "run_mode": args.run_mode,
                "received_from": args.received_from,
                "received_to": args.received_to,
            })
            if args.dry_run:
                print("Dry-run mode: database routing can be read, but EXC/ING writes are skipped")
        else:
            print("Database trace disabled; using customer YML routing and CSV run evidence")

        print("STEP 2 - Checking Graph configuration")
        check_graph_configuration()

        print("STEP 3 - Getting Microsoft Graph token")
        token = get_graph_token()

        print(f"STEP 4 - Reading mailbox messages ({args.run_mode}: {args.received_from or 'no lower date'} to {args.received_to or 'no upper date'})")
        messages = read_messages(token, args)
        print(f"Found {len(messages)} message(s) to review")

        print("STEP 5 - Processing messages by tenant rules")
        for message in messages:
            process_one_message(token, message, stats, rows, validation_rows, args, db_trace)

        print("STEP 6 - Finalising support summary")
        run_log_path = write_run_log(rows, stats, started_at, args)
        validation_report_path = write_validation_report(validation_rows, started_at, args)

        final_status = "FAILED" if stats["failed"] else "COMPLETED"
        if db_trace:
            db_trace.finish_execution(final_status, stats)
            db_trace.insert_execution_log("FINISH", "Graph ingestion run finished", detail={
                "status": final_status,
                "scanned": stats.get("scanned"),
                "matched": stats.get("matched"),
                "generated_process_packs": stats.get("generated_process_packs"),
                "ing_rows": stats.get("ing_rows"),
            })

        print_summary(stats, run_log_path, validation_report_path)
        return 1 if stats["failed"] else 0

    except Exception as error:
        if db_trace:
            try:
                db_trace.finish_execution("FAILED", stats, str(error))
                db_trace.insert_execution_log("FAILED", "Graph ingestion run failed", level="ERROR", detail={"error": str(error)})
            except Exception as trace_error:
                print(f"ERROR finalising EXC.Graph failure state: {trace_error}")
        raise

    finally:
        if db_trace:
            db_trace.close()

def print_summary(
    stats: dict[str, Any],
    run_log_path: Path,
    validation_report_path: Path | None,
) -> None:
    """Print the final result in the terminal."""
    print("")
    print("Graph download finished")
    print(f"  Scanned messages: {stats['scanned']}")
    print(f"  Matched messages: {stats['matched']}")
    print(f"  Unmatched messages: {stats['unmatched']}")
    print(f"  Saved attachments: {stats['saved_attachments']}")
    print(f"  Saved ENS source texts: {stats['saved_body_texts']}")
    print(f"  Generated process packs: {stats['generated_process_packs']}")
    print(f"  Missing ENS sources: {stats['missing_ens_sources']}")
    print(f"  Skipped existing files: {stats['skipped_existing']}")
    print(
        "  No file attachments: "
        f"{stats['no_file_attachments']}"
    )
    print(f"  Failed messages: {stats['failed']}")
    print(f"  Validated Excel files: {stats['validated_excels']}")
    print(f"  Validation errors: {stats['validation_errors']}")
    print(f"  Validation warnings: {stats['validation_warnings']}")
    print(f"  Validation pending items: {stats['validation_pending']}")
    print(f"  Tenant config source: {stats['tenant_config_source']}")
    if stats["database_enabled"]:
        print(f"  ING.Graph rows inserted: {stats['ing_rows']}")
    print(f"  Technical run log: {run_log_path}")
    if validation_report_path:
        print(f"  Validation report: {validation_report_path}")
    if stats["validation_errors"]:
        print("  API readiness warning: local download completed; later API validation may reject validation blockers.")
    elif stats["validation_warnings"] or stats["validation_pending"]:
        print("  API readiness review: local download completed with validation warnings or pending mapping items.")

    if stats["file_types"]:
        print("  File types:")
        for suffix, count in sorted(stats["file_types"].items()):
            print(f"    {suffix}: {count}")


# =============================================================================
# STEP 13 - COMMAND LINE OPTIONS
# =============================================================================

def build_parser() -> argparse.ArgumentParser:
    """Define optional filters for daily and historic downloads."""
    parser = argparse.ArgumentParser(
        description="Simple Microsoft Graph downloader by tenant sender domain."
    )
    parser.add_argument(
        "--run-mode",
        choices=["daily", "historic", "custom"],
        default="daily",
        help=(
            "daily reads today's emails only, historic reads from customer "
            "historic_start_date to yesterday, custom uses the explicit date filters."
        ),
    )
    parser.add_argument(
        "--received-from",
        help="Optional lower received date. Example: 2026-01-01",
    )
    parser.add_argument(
        "--received-to",
        help="Optional upper received date. Example: 2026-06-17",
    )
    parser.add_argument(
        "--max-messages",
        type=int,
        default=0,
        help="Maximum messages to read. 0 means no script limit.",
    )
    parser.add_argument(
        "--page-size",
        type=int,
        default=50,
        help="Graph page size. Use 1 to 100.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Report what would happen without writing customer files.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite destination files if they already exist.",
    )
    parser.add_argument(
        "--no-database",
        action="store_true",
        help="Disable optional CFG/EXC/ING database reads and writes for this run.",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    apply_run_mode_dates(args)

    if args.max_messages < 0:
        parser.error("--max-messages cannot be negative")
    if args.page_size < 1 or args.page_size > 100:
        parser.error("--page-size must be between 1 and 100")

    return run(args)


if __name__ == "__main__":
    raise SystemExit(main())
