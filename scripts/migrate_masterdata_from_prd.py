#!/usr/bin/env python3
"""Bridge BKD master data from the PRD DATA schema into Fusion_TSS.

Azure SQL Database does not support normal three-part cross-database
`INSERT ... SELECT` between independent databases, so this script opens one
connection to the source database and one to the target database.

Default mode is a dry run. Pass `--execute` to write.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import UTC, datetime
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pyodbc
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from config.db_connection import build_connection_string


SOURCE_DB = "TSS_FLOW_V2_PRD"
SOURCE_SCHEMA = "DATA"
TARGET_DB = "Fusion_TSS"
RAW_TARGET_SCHEMA = "DATA"
APP_TARGET_SCHEMA = "BKD"

RAW_TABLE_KEYS = {
    "BKD_Customers": ("CustomerNo",),
    "BKD_Item_Commodity_Code": ("Item",),
    "BKD_Sales_Orders": ("BKD_SalesOrderID",),
    "BKD_Sales_Orders_Daily": ("BKD_DailyID",),
}

CSV_TABLE_FILES = {
    "BKD_Customers": "customers.csv",
    "BKD_Item_Commodity_Code": "products.csv",
    "BKD_Sales_Orders": "salesOrder.csv",
    "BKD_Sales_Orders_Daily": "salesOrdersDaily.csv",
}

WEIGHT_PRODUCTS_COLUMNS = [
    "stg_item_id",
    "ClientCode",
    "stg_consignment_id",
    "sub_status",
    "error_id",
    "validated_at",
    "submitted_at",
    "completed_at",
    "last_sub_status_change",
    "ing_item_id",
    "goods_stage",
    "tss_hex_id",
    "tss_consignment_ref",
    "item_seq",
    "goods_description",
    "commodity_code",
    "gross_mass_kg",
    "net_mass_kg",
    "number_of_packages",
    "number_of_individual_pieces",
    "type_of_packages",
    "package_marks",
    "equipment_number",
    "procedure_code",
    "additional_procedure_code",
    "controlled_goods",
    "country_of_origin",
    "item_invoice_amount",
    "item_invoice_currency",
    "customs_value",
    "valuation_method",
    "statistical_value",
    "nature_of_transaction",
    "preference",
]

WEIGHT_PRODUCTS_REQUIRED_HEADER_COLUMNS = [
    "goods_description",
    "commodity_code",
    "gross_mass_kg",
    "net_mass_kg",
]

EORI_VALIDATION_COLUMNS = [
    "EORI",
    "Route",
    "Valid",
    "Status",
    "Trader Name",
    "Street",
    "City",
    "Postcode",
    "Country",
    "Processing Date",
    "Error / Notes",
]

RAW_TABLE_COLUMNS = {
    "BKD_Customers": [
        "CustomerNo",
        "Name",
        "Address",
        "Address2",
        "City",
        "County",
        "PostCode",
        "CountryRegionCode",
        "EORINumber",
        "SourceFile",
        "SourceFileDate",
        "LoadedAtUTC",
        "UpdatedAtUTC",
    ],
    "BKD_Item_Commodity_Code": [
        "Item",
        "Description",
        "ItemCategoryCode",
        "CommodityCode",
        "CountryOfOriginCode",
        "SourceFile",
        "LoadedAtUTC",
        "UpdatedAtUTC",
    ],
    "BKD_Sales_Orders": [
        "BKD_SalesOrderID",
        "DocumentNo",
        "No",
        "LineNo",
        "ItemNo",
        "EORINumber",
        "ShipToName",
        "ShipToAddress",
        "ShipToAddress2",
        "ShipToCity",
        "ShipToCounty",
        "ShipToPostCode",
        "ShipToCountryRegionCode",
        "Description",
        "Amount",
        "QtyBase",
        "NetWeight",
        "GrossWeight",
        "Qty",
        "SourceFile",
        "LoadedAtUTC",
        "UpdatedAtUTC",
    ],
    "BKD_Sales_Orders_Daily": [
        "BKD_DailyID",
        "DocumentNo",
        "No",
        "LineNo",
        "SellToCustomerNo",
        "ShipToName",
        "ShipToAddress",
        "ShipToAddress2",
        "ShipToCity",
        "ShipToCounty",
        "ShipToPhoneNo",
        "Email",
        "Quantity",
        "QuantityBase",
        "Amount",
        "LineAmountExclVAT",
        "UnitPriceExclVAT",
        "QtyPerUnitOfMeasure",
        "UnitOfMeasureCode",
        "SourceFile",
        "SourceFileDate",
        "LoadedAtUTC",
        "UpdatedAtUTC",
    ],
}


@dataclass(frozen=True)
class Column:
    name: str
    type_name: str
    max_length: int
    precision: int
    scale: int
    nullable: bool
    identity: bool


RAW_TABLE_COLUMN_SPECS = {
    "BKD_Customers": [
        Column("CustomerNo", "nvarchar", 80, 0, 0, False, False),
        Column("Name", "nvarchar", 400, 0, 0, True, False),
        Column("Address", "nvarchar", 400, 0, 0, True, False),
        Column("Address2", "nvarchar", 400, 0, 0, True, False),
        Column("City", "nvarchar", 200, 0, 0, True, False),
        Column("County", "nvarchar", 200, 0, 0, True, False),
        Column("PostCode", "nvarchar", 40, 0, 0, True, False),
        Column("CountryRegionCode", "nvarchar", 20, 0, 0, True, False),
        Column("EORINumber", "nvarchar", 40, 0, 0, True, False),
        Column("SourceFile", "nvarchar", 520, 0, 0, True, False),
        Column("SourceFileDate", "date", 3, 10, 0, True, False),
        Column("LoadedAtUTC", "datetime2", 7, 23, 3, False, False),
        Column("UpdatedAtUTC", "datetime2", 7, 23, 3, True, False),
    ],
    "BKD_Item_Commodity_Code": [
        Column("Item", "nvarchar", 100, 0, 0, False, False),
        Column("Description", "nvarchar", 500, 0, 0, True, False),
        Column("ItemCategoryCode", "nvarchar", 100, 0, 0, True, False),
        Column("CommodityCode", "nvarchar", 40, 0, 0, True, False),
        Column("CountryOfOriginCode", "nvarchar", 20, 0, 0, True, False),
        Column("SourceFile", "nvarchar", 520, 0, 0, True, False),
        Column("LoadedAtUTC", "datetime2", 7, 23, 3, False, False),
        Column("UpdatedAtUTC", "datetime2", 7, 23, 3, True, False),
    ],
    "BKD_Sales_Orders": [
        Column("BKD_SalesOrderID", "int", 4, 10, 0, False, True),
        Column("DocumentNo", "nvarchar", 80, 0, 0, False, False),
        Column("No", "nvarchar", 80, 0, 0, False, False),
        Column("LineNo", "int", 4, 10, 0, False, False),
        Column("ItemNo", "nvarchar", 100, 0, 0, True, False),
        Column("EORINumber", "nvarchar", 40, 0, 0, True, False),
        Column("ShipToName", "nvarchar", 400, 0, 0, True, False),
        Column("ShipToAddress", "nvarchar", 400, 0, 0, True, False),
        Column("ShipToAddress2", "nvarchar", 400, 0, 0, True, False),
        Column("ShipToCity", "nvarchar", 200, 0, 0, True, False),
        Column("ShipToCounty", "nvarchar", 200, 0, 0, True, False),
        Column("ShipToPostCode", "nvarchar", 40, 0, 0, True, False),
        Column("ShipToCountryRegionCode", "nvarchar", 20, 0, 0, True, False),
        Column("Description", "nvarchar", 1000, 0, 0, True, False),
        Column("Amount", "decimal", 9, 18, 4, True, False),
        Column("QtyBase", "decimal", 9, 18, 4, True, False),
        Column("NetWeight", "decimal", 9, 18, 4, True, False),
        Column("GrossWeight", "decimal", 9, 18, 4, True, False),
        Column("Qty", "decimal", 9, 18, 4, True, False),
        Column("SourceFile", "nvarchar", 520, 0, 0, True, False),
        Column("LoadedAtUTC", "datetime2", 7, 23, 3, False, False),
        Column("UpdatedAtUTC", "datetime2", 7, 23, 3, True, False),
    ],
    "BKD_Sales_Orders_Daily": [
        Column("BKD_DailyID", "int", 4, 10, 0, False, True),
        Column("DocumentNo", "nvarchar", 80, 0, 0, False, False),
        Column("No", "nvarchar", 100, 0, 0, False, False),
        Column("LineNo", "int", 4, 10, 0, False, False),
        Column("SellToCustomerNo", "nvarchar", 80, 0, 0, True, False),
        Column("ShipToName", "nvarchar", 400, 0, 0, True, False),
        Column("ShipToAddress", "nvarchar", 400, 0, 0, True, False),
        Column("ShipToAddress2", "nvarchar", 400, 0, 0, True, False),
        Column("ShipToCity", "nvarchar", 200, 0, 0, True, False),
        Column("ShipToCounty", "nvarchar", 200, 0, 0, True, False),
        Column("ShipToPhoneNo", "nvarchar", 80, 0, 0, True, False),
        Column("Email", "nvarchar", 400, 0, 0, True, False),
        Column("Quantity", "decimal", 9, 18, 4, True, False),
        Column("QuantityBase", "decimal", 9, 18, 4, True, False),
        Column("Amount", "decimal", 9, 18, 4, True, False),
        Column("LineAmountExclVAT", "decimal", 9, 18, 4, True, False),
        Column("UnitPriceExclVAT", "decimal", 9, 18, 4, True, False),
        Column("QtyPerUnitOfMeasure", "decimal", 9, 18, 4, True, False),
        Column("UnitOfMeasureCode", "nvarchar", 40, 0, 0, True, False),
        Column("SourceFile", "nvarchar", 520, 0, 0, True, False),
        Column("SourceFileDate", "date", 3, 10, 0, True, False),
        Column("LoadedAtUTC", "datetime2", 7, 23, 3, False, False),
        Column("UpdatedAtUTC", "datetime2", 7, 23, 3, True, False),
    ],
}

NULLISH = {"", "null", "none", "undefined", "nan"}
EORI_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{2,17}$")


def clean(value: Any) -> str:
    text = str(value or "").strip()
    return "" if text.lower() in NULLISH else text


def safe_eori(value: Any) -> str:
    text = re.sub(r"[\s-]+", "", clean(value).upper())
    return text if EORI_RE.match(text) else ""


def q(name: str) -> str:
    return "[" + name.replace("]", "]]") + "]"


def connection_string_for_database(database: str) -> str:
    conn_str = build_connection_string()
    if re.search(r"(?i)(^|;)\s*(database|initial catalog)\s*=", conn_str):
        return re.sub(
            r"(?i)(^|;)\s*(database|initial catalog)\s*=[^;]*",
            lambda m: f"{m.group(1)}DATABASE={database}",
            conn_str,
            count=1,
        )
    return conn_str.rstrip(";") + f";DATABASE={database};"


def connect(database: str) -> pyodbc.Connection:
    return pyodbc.connect(connection_string_for_database(database))


def table_exists(conn: pyodbc.Connection, schema: str, table: str) -> bool:
    cur = conn.cursor()
    cur.execute(
        """
        SELECT 1
        FROM INFORMATION_SCHEMA.TABLES
        WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
        """,
        [schema, table],
    )
    return cur.fetchone() is not None


def schema_exists(conn: pyodbc.Connection, schema: str) -> bool:
    cur = conn.cursor()
    cur.execute(
        """
        SELECT 1
        FROM INFORMATION_SCHEMA.SCHEMATA
        WHERE SCHEMA_NAME = ?
        """,
        [schema],
    )
    return cur.fetchone() is not None


def ensure_schema(conn: pyodbc.Connection, schema: str, *, execute: bool) -> str:
    if schema_exists(conn, schema):
        return "exists"
    if execute:
        conn.cursor().execute(f"CREATE SCHEMA {q(schema)}")
    return "created"


def columns(conn: pyodbc.Connection, schema: str, table: str) -> list[Column]:
    cur = conn.cursor()
    cur.execute(
        """
        SELECT c.name,
               t.name AS type_name,
               c.max_length,
               c.precision,
               c.scale,
               c.is_nullable,
               c.is_identity
        FROM sys.columns c
        JOIN sys.types t ON c.user_type_id = t.user_type_id
        WHERE c.object_id = OBJECT_ID(?)
        ORDER BY c.column_id
        """,
        [f"{schema}.{table}"],
    )
    return [
        Column(
            name=row.name,
            type_name=row.type_name,
            max_length=int(row.max_length),
            precision=int(row.precision),
            scale=int(row.scale),
            nullable=bool(row.is_nullable),
            identity=bool(row.is_identity),
        )
        for row in cur.fetchall()
    ]


def identity_columns(conn: pyodbc.Connection, schema: str, table: str) -> set[str]:
    cur = conn.cursor()
    cur.execute(
        """
        SELECT c.name
        FROM sys.columns c
        WHERE c.object_id = OBJECT_ID(?)
          AND c.is_identity = 1
        """,
        [f"{schema}.{table}"],
    )
    return {row.name for row in cur.fetchall()}


def key_constraints(conn: pyodbc.Connection, schema: str, table: str) -> list[tuple[str, str, list[str]]]:
    cur = conn.cursor()
    cur.execute(
        """
        SELECT kc.name AS constraint_name,
               kc.type_desc,
               c.name AS column_name,
               ic.key_ordinal
        FROM sys.key_constraints kc
        JOIN sys.index_columns ic
          ON kc.parent_object_id = ic.object_id
         AND kc.unique_index_id = ic.index_id
        JOIN sys.columns c
          ON c.object_id = ic.object_id
         AND c.column_id = ic.column_id
        WHERE kc.parent_object_id = OBJECT_ID(?)
        ORDER BY kc.name, ic.key_ordinal
        """,
        [f"{schema}.{table}"],
    )
    grouped: dict[tuple[str, str], list[str]] = {}
    for row in cur.fetchall():
        grouped.setdefault((row.constraint_name, row.type_desc), []).append(row.column_name)
    return [(name, type_desc, cols) for (name, type_desc), cols in grouped.items()]


def column_names(conn: pyodbc.Connection, schema: str, table: str) -> set[str]:
    return {col.name for col in columns(conn, schema, table)}


def render_type(col: Column) -> str:
    typ = col.type_name.lower()
    identity_suffix = " IDENTITY(1,1)" if col.identity and typ in {"int", "bigint", "smallint", "tinyint"} else ""
    if typ in {"nvarchar", "nchar"}:
        length = "MAX" if col.max_length == -1 else str(max(1, col.max_length // 2))
        return f"{typ.upper()}({length})"
    if typ in {"varchar", "char", "varbinary", "binary"}:
        length = "MAX" if col.max_length == -1 else str(max(1, col.max_length))
        return f"{typ.upper()}({length})"
    if typ in {"decimal", "numeric"}:
        return f"{typ.upper()}({col.precision},{col.scale})"
    if typ in {"datetime2", "datetimeoffset", "time"}:
        return f"{typ.upper()}({col.scale})"
    return typ.upper() + identity_suffix


def ensure_raw_table(
    source: pyodbc.Connection,
    target: pyodbc.Connection,
    table: str,
    *,
    execute: bool,
) -> str:
    if table_exists(target, RAW_TARGET_SCHEMA, table):
        return "exists"

    source_cols = columns(source, SOURCE_SCHEMA, table)
    if not source_cols:
        return "source_missing"

    column_defs = [
        f"{q(col.name)} {render_type(col)} {'NULL' if col.nullable else 'NOT NULL'}"
        for col in source_cols
    ]
    for constraint_name, type_desc, constraint_cols in key_constraints(source, SOURCE_SCHEMA, table):
        constraint_type = "PRIMARY KEY" if type_desc == "PRIMARY_KEY_CONSTRAINT" else "UNIQUE"
        column_defs.append(
            f"CONSTRAINT {q(constraint_name)} {constraint_type} ({', '.join(q(col) for col in constraint_cols)})"
        )
    joined_defs = ",\n        ".join(column_defs)
    sql = f"CREATE TABLE {q(RAW_TARGET_SCHEMA)}.{q(table)} (\n        {joined_defs}\n    )"
    if execute:
        target.cursor().execute(sql)
    return "created"


def fetch_rows(conn: pyodbc.Connection, schema: str, table: str, selected_columns: list[str]) -> list[dict[str, Any]]:
    cur = conn.cursor()
    cur.execute(
        f"SELECT {', '.join(q(c) for c in selected_columns)} FROM {q(schema)}.{q(table)}"
    )
    return [dict(zip(selected_columns, row)) for row in cur.fetchall()]


def parse_csv_value(value: Any) -> Any:
    text = str(value or "").strip()
    return None if text.lower() in NULLISH else text


def parse_decimal(value: Any) -> float | None:
    text = clean(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalized_text(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean(value).lower()).strip()


def load_csv_rows(path: Path, table: str, *, has_header: bool = False) -> list[dict[str, Any]]:
    expected_columns = RAW_TABLE_COLUMNS[table]
    rows: list[dict[str, Any]] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        if has_header:
            header = next(reader, [])
            missing = [col for col in expected_columns if col not in header]
            if missing:
                raise ValueError(f"{path.name} is missing columns: {', '.join(missing)}")
            positions = [header.index(col) for col in expected_columns]
            for row_num, row in enumerate(reader, start=2):
                rows.append({col: parse_csv_value(row[pos]) if pos < len(row) else None for col, pos in zip(expected_columns, positions)})
        else:
            for row_num, row in enumerate(reader, start=1):
                if len(row) != len(expected_columns):
                    raise ValueError(
                        f"{path.name} row {row_num}: expected {len(expected_columns)} columns, got {len(row)}"
                    )
                rows.append({col: parse_csv_value(value) for col, value in zip(expected_columns, row)})
    return rows


def load_weight_product_rows(path: Path, *, has_header: bool = False) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        if has_header:
            header = next(reader, [])
            missing = [col for col in WEIGHT_PRODUCTS_REQUIRED_HEADER_COLUMNS if col not in header]
            if missing:
                raise ValueError(f"{path.name} is missing columns: {', '.join(missing)}")
            for row in reader:
                rows.append({col: parse_csv_value(row[pos]) if pos < len(row) else None for pos, col in enumerate(header)})
        else:
            for row_num, row in enumerate(reader, start=1):
                if len(row) != len(WEIGHT_PRODUCTS_COLUMNS):
                    raise ValueError(
                        f"{path.name} row {row_num}: expected {len(WEIGHT_PRODUCTS_COLUMNS)} columns, got {len(row)}"
                )
                rows.append({col: parse_csv_value(value) for col, value in zip(WEIGHT_PRODUCTS_COLUMNS, row)})
    return rows


def load_eori_validation_xlsx(path: Path) -> list[dict[str, Any]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Validation Results"] if "Validation Results" in workbook.sheetnames else workbook.active
        rows = sheet.iter_rows(values_only=True)
        header = [clean(value) for value in next(rows, [])]
        missing = [col for col in EORI_VALIDATION_COLUMNS if col not in header]
        if missing:
            raise ValueError(f"{path.name} is missing columns: {', '.join(missing)}")
        positions = [header.index(col) for col in EORI_VALIDATION_COLUMNS]
        records: list[dict[str, Any]] = []
        for row in rows:
            record = {col: row[pos] if pos < len(row) else None for col, pos in zip(EORI_VALIDATION_COLUMNS, positions)}
            if clean(record.get("EORI")):
                records.append(record)
        return records
    finally:
        workbook.close()


def ensure_raw_table_from_specs(target: pyodbc.Connection, table: str, *, execute: bool) -> str:
    if table_exists(target, RAW_TARGET_SCHEMA, table):
        return "exists"
    source_cols = RAW_TABLE_COLUMN_SPECS[table]
    column_defs = [
        f"{q(col.name)} {render_type(col)} {'NULL' if col.nullable else 'NOT NULL'}"
        for col in source_cols
    ]
    for constraint_name, constraint_type, constraint_cols in raw_table_constraints(table):
        column_defs.append(
            f"CONSTRAINT {q(constraint_name)} {constraint_type} ({', '.join(q(col) for col in constraint_cols)})"
        )
    joined_defs = ",\n        ".join(column_defs)
    sql = f"CREATE TABLE {q(RAW_TARGET_SCHEMA)}.{q(table)} (\n        {joined_defs}\n    )"
    if execute:
        target.cursor().execute(sql)
    return "created"


def raw_table_constraints(table: str) -> list[tuple[str, str, list[str]]]:
    if table == "BKD_Customers":
        return [("PK_BKD_Customers", "PRIMARY KEY", ["CustomerNo"])]
    if table == "BKD_Item_Commodity_Code":
        return [("PK_BKD_Item_Commodity_Code", "PRIMARY KEY", ["Item"])]
    if table == "BKD_Sales_Orders":
        return [
            ("PK_BKD_Sales_Orders", "PRIMARY KEY", ["BKD_SalesOrderID"]),
            ("UQ_BKD_Sales_Orders_DocLine", "UNIQUE", ["DocumentNo", "No", "LineNo"]),
        ]
    if table == "BKD_Sales_Orders_Daily":
        return [
            ("PK_BKD_Sales_Orders_Daily", "PRIMARY KEY", ["BKD_DailyID"]),
            ("UQ_BKD_Sales_Orders_Daily_DocLine", "UNIQUE", ["DocumentNo", "No", "LineNo"]),
        ]
    return []


def existing_keys(
    conn: pyodbc.Connection,
    schema: str,
    table: str,
    key_columns: tuple[str, ...],
) -> set[tuple[Any, ...]]:
    cur = conn.cursor()
    cur.execute(f"SELECT {', '.join(q(c) for c in key_columns)} FROM {q(schema)}.{q(table)}")
    return {tuple(normalize_key_value(value) for value in row) for row in cur.fetchall()}


def normalize_key_value(value: Any) -> str:
    return "" if value is None else str(value)


def row_key(row: dict[str, Any], key_columns: tuple[str, ...]) -> tuple[str, ...]:
    return tuple(normalize_key_value(row[key]) for key in key_columns)


def copy_raw_rows(
    target: pyodbc.Connection,
    table: str,
    rows: list[dict[str, Any]],
    *,
    execute: bool,
    update_existing: bool,
) -> dict[str, int | str]:
    ensure_status = ensure_raw_table_from_specs(target, table, execute=execute)
    if not execute and ensure_status == "created":
        target_cols = RAW_TABLE_COLUMNS[table]
    else:
        target_cols = list(column_names(target, RAW_TARGET_SCHEMA, table))

    source_cols = RAW_TABLE_COLUMNS[table]
    common_cols = [col for col in source_cols if col in target_cols]
    key_cols = RAW_TABLE_KEYS[table]
    missing_keys = [key for key in key_cols if key not in common_cols]
    if missing_keys:
        raise RuntimeError(f"{table} is missing key columns: {', '.join(missing_keys)}")

    seen = existing_keys(target, RAW_TARGET_SCHEMA, table, key_cols) if table_exists(target, RAW_TARGET_SCHEMA, table) else set()
    inserts = [row for row in rows if row_key(row, key_cols) not in seen]
    updates = [row for row in rows if row_key(row, key_cols) in seen]

    if execute and inserts:
        placeholders = ", ".join("?" for _ in common_cols)
        sql = (
            f"INSERT INTO {q(RAW_TARGET_SCHEMA)}.{q(table)} ({', '.join(q(c) for c in common_cols)}) "
            f"VALUES ({placeholders})"
        )
        cur = target.cursor()
        cur.fast_executemany = True
        needs_identity_insert = bool(identity_columns(target, RAW_TARGET_SCHEMA, table).intersection(common_cols))
        if needs_identity_insert:
            cur.execute(f"SET IDENTITY_INSERT {q(RAW_TARGET_SCHEMA)}.{q(table)} ON")
        try:
            cur.executemany(sql, [[row[col] for col in common_cols] for row in inserts])
        finally:
            if needs_identity_insert:
                cur.execute(f"SET IDENTITY_INSERT {q(RAW_TARGET_SCHEMA)}.{q(table)} OFF")

    updated_count = 0
    update_cols = [col for col in common_cols if col not in key_cols]
    if execute and update_existing and updates and update_cols:
        set_sql = ", ".join(f"{q(col)} = ?" for col in update_cols)
        where_sql = " AND ".join(f"{q(col)} = ?" for col in key_cols)
        sql = f"UPDATE {q(RAW_TARGET_SCHEMA)}.{q(table)} SET {set_sql} WHERE {where_sql}"
        cur = target.cursor()
        cur.fast_executemany = True
        cur.executemany(
            sql,
            [[row[col] for col in update_cols] + [row[col] for col in key_cols] for row in updates],
        )
        updated_count = len(updates)

    return {
        "table": table,
        "status": ensure_status,
        "source": len(rows),
        "inserted": len(inserts),
        "updated": updated_count if execute else (len(updates) if update_existing else 0),
    }


def copy_raw_table(
    source: pyodbc.Connection,
    target: pyodbc.Connection,
    table: str,
    *,
    execute: bool,
    update_existing: bool,
) -> dict[str, int | str]:
    if not table_exists(source, SOURCE_SCHEMA, table):
        return {"table": table, "status": "source_missing", "source": 0, "inserted": 0, "updated": 0}

    ensure_status = ensure_raw_table(source, target, table, execute=execute)
    if ensure_status == "source_missing":
        return {"table": table, "status": ensure_status, "source": 0, "inserted": 0, "updated": 0}
    if not execute and ensure_status == "created":
        target_cols = [col.name for col in columns(source, SOURCE_SCHEMA, table)]
    else:
        target_cols = list(column_names(target, RAW_TARGET_SCHEMA, table))

    source_cols = [col.name for col in columns(source, SOURCE_SCHEMA, table)]
    common_cols = [col for col in source_cols if col in target_cols]
    key_cols = RAW_TABLE_KEYS[table]
    missing_keys = [key for key in key_cols if key not in common_cols]
    if missing_keys:
        raise RuntimeError(f"{table} is missing key columns: {', '.join(missing_keys)}")

    rows = fetch_rows(source, SOURCE_SCHEMA, table, common_cols)
    seen = existing_keys(target, RAW_TARGET_SCHEMA, table, key_cols) if table_exists(target, RAW_TARGET_SCHEMA, table) else set()
    inserts = [row for row in rows if row_key(row, key_cols) not in seen]
    updates = [row for row in rows if row_key(row, key_cols) in seen]

    if execute and inserts:
        placeholders = ", ".join("?" for _ in common_cols)
        sql = (
            f"INSERT INTO {q(RAW_TARGET_SCHEMA)}.{q(table)} ({', '.join(q(c) for c in common_cols)}) "
            f"VALUES ({placeholders})"
        )
        cur = target.cursor()
        cur.fast_executemany = True
        needs_identity_insert = bool(identity_columns(target, RAW_TARGET_SCHEMA, table).intersection(common_cols))
        if needs_identity_insert:
            cur.execute(f"SET IDENTITY_INSERT {q(RAW_TARGET_SCHEMA)}.{q(table)} ON")
        try:
            cur.executemany(sql, [[row[col] for col in common_cols] for row in inserts])
        finally:
            if needs_identity_insert:
                cur.execute(f"SET IDENTITY_INSERT {q(RAW_TARGET_SCHEMA)}.{q(table)} OFF")

    updated_count = 0
    update_cols = [col for col in common_cols if col not in key_cols]
    if execute and update_existing and updates and update_cols:
        set_sql = ", ".join(f"{q(col)} = ?" for col in update_cols)
        where_sql = " AND ".join(f"{q(col)} = ?" for col in key_cols)
        sql = f"UPDATE {q(RAW_TARGET_SCHEMA)}.{q(table)} SET {set_sql} WHERE {where_sql}"
        cur = target.cursor()
        cur.fast_executemany = True
        cur.executemany(
            sql,
            [[row[col] for col in update_cols] + [row[col] for col in key_cols] for row in updates],
        )
        updated_count = len(updates)

    return {
        "table": table,
        "status": ensure_status,
        "source": len(rows),
        "inserted": len(inserts),
        "updated": updated_count if execute else (len(updates) if update_existing else 0),
    }


def row_value(row: Any, key: str) -> Any:
    if isinstance(row, dict):
        return row.get(key)
    return getattr(row, key)


def ensure_doc_product_catalog_weight_columns(target: pyodbc.Connection, *, execute: bool) -> str:
    if not table_exists(target, APP_TARGET_SCHEMA, "DocProductCatalog"):
        return "missing"

    specs = {
        "gross_weight_kg": "DECIMAL(10,3) NULL",
        "net_weight_kg": "DECIMAL(10,3) NULL",
        "weight_source": "NVARCHAR(160) NULL",
        "weight_sample_count": "INT NULL",
    }
    existing = column_names(target, APP_TARGET_SCHEMA, "DocProductCatalog")
    missing = [name for name in specs if name not in existing]
    if execute:
        for name in missing:
            target.cursor().execute(
                f"ALTER TABLE {q(APP_TARGET_SCHEMA)}.DocProductCatalog ADD {q(name)} {specs[name]}"
            )
    return "created" if missing else "exists"


def product_lookup_from_rows(rows: list[Any]) -> dict[tuple[str, str], Any]:
    lookup: dict[tuple[str, str], Any] = {}
    ambiguous: set[tuple[str, str]] = set()
    for row in rows:
        item = clean(row_value(row, "Item")).upper()
        description = normalized_text(row_value(row, "Description"))
        commodity = clean(row_value(row, "CommodityCode")).replace(" ", "")
        if not item or not description:
            continue
        key = (description, commodity)
        existing = lookup.get(key)
        if existing and clean(row_value(existing, "Item")).upper() != item:
            ambiguous.add(key)
            continue
        lookup[key] = row
    for key in ambiguous:
        lookup.pop(key, None)
    return lookup


def build_weight_payloads(weight_rows: list[Any], product_rows: list[Any]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    product_lookup = product_lookup_from_rows(product_rows)
    aggregates: dict[str, dict[str, Any]] = {}
    stats = {
        "source": len(weight_rows),
        "matched": 0,
        "unmatched": 0,
        "invalid_weight": 0,
    }

    for row in weight_rows:
        description = normalized_text(row_value(row, "goods_description"))
        commodity = clean(row_value(row, "commodity_code")).replace(" ", "")
        product = product_lookup.get((description, commodity))
        if not product:
            stats["unmatched"] += 1
            continue

        gross = parse_decimal(row_value(row, "gross_mass_kg"))
        net = parse_decimal(row_value(row, "net_mass_kg"))
        divisor = (
            parse_decimal(row_value(row, "number_of_individual_pieces"))
            or parse_decimal(row_value(row, "number_of_packages"))
            or 1.0
        )
        if gross is None or net is None or divisor <= 0:
            stats["invalid_weight"] += 1
            continue

        sku = clean(row_value(product, "Item")).upper()
        bucket = aggregates.setdefault(
            sku,
            {
                "sku": sku,
                "product_code": sku,
                "description": clean(row_value(product, "Description")) or clean(row_value(row, "goods_description")),
                "commodity_code": clean(row_value(product, "CommodityCode")).replace(" ", "") or commodity,
                "country_of_origin": clean(row_value(product, "CountryOfOriginCode")).upper()[:2] or None,
                "gross_sum": 0.0,
                "net_sum": 0.0,
                "count": 0,
            },
        )
        bucket["gross_sum"] += gross / divisor
        bucket["net_sum"] += net / divisor
        bucket["count"] += 1
        stats["matched"] += 1

    payloads = []
    for bucket in aggregates.values():
        count = int(bucket["count"])
        payloads.append({
            "customer_code": "ALL",
            "sku": bucket["sku"],
            "product_code": bucket["product_code"],
            "description": bucket["description"],
            "commodity_code": bucket["commodity_code"],
            "country_of_origin": bucket["country_of_origin"],
            "gross_weight_kg": round(float(bucket["gross_sum"]) / count, 3),
            "net_weight_kg": round(float(bucket["net_sum"]) / count, 3),
            "weight_source": "STG.BKD_GoodsItems",
            "weight_sample_count": count,
            "controlled_goods": 0,
            "active": 1,
        })
    return payloads, stats


def upsert_doc_product_catalog_weights(
    weight_rows: list[Any],
    product_rows: list[Any],
    target: pyodbc.Connection,
    *,
    execute: bool,
) -> dict[str, int | str]:
    ensure_status = ensure_doc_product_catalog_weight_columns(target, execute=execute)
    if ensure_status == "missing":
        return {"table": "DocProductCatalog weights", "status": "missing", "source": 0, "inserted": 0, "updated": 0}

    payloads, stats = build_weight_payloads(weight_rows, product_rows)
    existing = existing_keys(target, APP_TARGET_SCHEMA, "DocProductCatalog", ("customer_code", "sku"))
    inserts = [row for row in payloads if (row["customer_code"], row["sku"]) not in existing]
    updates = [row for row in payloads if (row["customer_code"], row["sku"]) in existing]

    target_cols = column_names(target, APP_TARGET_SCHEMA, "DocProductCatalog") if execute else set(payloads[0].keys() if payloads else [])
    if execute and inserts:
        insert_cols = [col for col in (
            "customer_code",
            "sku",
            "product_code",
            "description",
            "commodity_code",
            "country_of_origin",
            "gross_weight_kg",
            "net_weight_kg",
            "weight_source",
            "weight_sample_count",
            "controlled_goods",
            "active",
        ) if col in target_cols]
        sql = (
            f"INSERT INTO {q(APP_TARGET_SCHEMA)}.DocProductCatalog ({', '.join(q(c) for c in insert_cols)}) "
            f"VALUES ({', '.join('?' for _ in insert_cols)})"
        )
        target.cursor().executemany(sql, [[row.get(col) for col in insert_cols] for row in inserts])

    if execute and updates:
        update_cols = [col for col in (
            "product_code",
            "description",
            "commodity_code",
            "country_of_origin",
            "gross_weight_kg",
            "net_weight_kg",
            "weight_source",
            "weight_sample_count",
            "active",
        ) if col in target_cols]
        assignments = [f"{q(col)} = ?" for col in update_cols]
        if "updated_at" in target_cols:
            assignments.append("[updated_at] = SYSUTCDATETIME()")
        sql = (
            f"UPDATE {q(APP_TARGET_SCHEMA)}.DocProductCatalog "
            f"SET {', '.join(assignments)} "
            "WHERE customer_code = ? AND sku = ?"
        )
        params = [[row.get(col) for col in update_cols] + [row["customer_code"], row["sku"]] for row in updates]
        target.cursor().executemany(sql, params)

    return {
        "table": "DocProductCatalog weights",
        "status": ensure_status,
        "source": stats["source"],
        "inserted": len(inserts),
        "updated": len(updates),
        "matched": stats["matched"],
        "unmatched": stats["unmatched"],
        "invalid_weight": stats["invalid_weight"],
    }


def eori_bool(value: Any, status: Any = None) -> bool:
    if isinstance(value, bool):
        return value
    text = clean(value).lower()
    if text in {"1", "true", "yes", "valid"}:
        return True
    if text in {"0", "false", "no", "invalid", "not valid"}:
        return False
    return clean(status).lower() == "valid"


def eori_checked_at(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    text = clean(value)
    if text:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                pass
    return datetime.now(UTC).replace(tzinfo=None)


def build_eori_cache_payloads(rows: list[Any]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    payloads: list[dict[str, Any]] = []
    seen: set[str] = set()
    stats = {
        "source": len(rows),
        "valid": 0,
        "invalid": 0,
        "skipped": 0,
    }

    for row in rows:
        eori = re.sub(r"[\s-]+", "", clean(row_value(row, "EORI")).upper())
        if not eori or eori in seen:
            stats["skipped"] += 1
            continue
        seen.add(eori)

        status = clean(row_value(row, "Status"))
        valid = eori_bool(row_value(row, "Valid"), status)
        note = clean(row_value(row, "Error / Notes"))
        error_detail = note or (None if valid else status or "Not valid")
        payloads.append({
            "eori_value": eori,
            "is_valid": 1 if valid else 0,
            "trader_name": clean(row_value(row, "Trader Name")) or None,
            "error_detail": error_detail,
            "checked_at": eori_checked_at(row_value(row, "Processing Date")),
        })
        stats["valid" if valid else "invalid"] += 1

    return payloads, stats


def upsert_eori_precheck_cache(
    rows: list[Any],
    target: pyodbc.Connection,
    *,
    execute: bool,
) -> dict[str, int | str]:
    if not table_exists(target, APP_TARGET_SCHEMA, "PrecheckEoriCache"):
        return {"table": "PrecheckEoriCache", "status": "missing", "source": 0, "inserted": 0, "updated": 0}

    payloads, stats = build_eori_cache_payloads(rows)
    if not payloads:
        return {
            "table": "PrecheckEoriCache",
            "status": "ok",
            "source": stats["source"],
            "inserted": 0,
            "updated": 0,
            "valid": stats["valid"],
            "invalid": stats["invalid"],
            "skipped": stats["skipped"],
        }

    target_cols = column_names(target, APP_TARGET_SCHEMA, "PrecheckEoriCache")
    key_cols = ("eori_value",)
    required_cols = {"eori_value", "is_valid"}
    if not required_cols.issubset(target_cols):
        return {
            "table": "PrecheckEoriCache",
            "status": "missing_columns",
            "source": stats["source"],
            "inserted": 0,
            "updated": 0,
            "valid": stats["valid"],
            "invalid": stats["invalid"],
            "skipped": stats["skipped"],
        }

    existing = {
        (str(values[0]).upper(),)
        for values in existing_keys(target, APP_TARGET_SCHEMA, "PrecheckEoriCache", key_cols)
    }
    inserts = [row for row in payloads if (row["eori_value"],) not in existing]
    updates = [row for row in payloads if (row["eori_value"],) in existing]

    insert_cols = [col for col in ("eori_value", "is_valid", "trader_name", "error_detail", "checked_at") if col in target_cols]
    update_cols = [col for col in ("is_valid", "trader_name", "error_detail", "checked_at") if col in target_cols]

    if execute and inserts:
        sql = (
            f"INSERT INTO {q(APP_TARGET_SCHEMA)}.PrecheckEoriCache ({', '.join(q(c) for c in insert_cols)}) "
            f"VALUES ({', '.join('?' for _ in insert_cols)})"
        )
        cur = target.cursor()
        cur.fast_executemany = True
        cur.executemany(sql, [[row.get(col) for col in insert_cols] for row in inserts])

    if execute and updates and update_cols:
        sql = (
            f"UPDATE {q(APP_TARGET_SCHEMA)}.PrecheckEoriCache "
            f"SET {', '.join(f'{q(col)} = ?' for col in update_cols)} "
            "WHERE eori_value = ?"
        )
        cur = target.cursor()
        cur.fast_executemany = True
        cur.executemany(sql, [[row.get(col) for col in update_cols] + [row["eori_value"]] for row in updates])

    return {
        "table": "PrecheckEoriCache",
        "status": "ok",
        "source": stats["source"],
        "inserted": len(inserts),
        "updated": len(updates),
        "valid": stats["valid"],
        "invalid": stats["invalid"],
        "skipped": stats["skipped"],
    }


def backfill_staging_goods_weights(target: pyodbc.Connection, *, execute: bool) -> dict[str, int | str]:
    if not table_exists(target, APP_TARGET_SCHEMA, "StagingGoodsItems"):
        return {"table": "StagingGoodsItems weights", "status": "missing", "source": 0, "inserted": 0, "updated": 0}
    if not table_exists(target, APP_TARGET_SCHEMA, "DocProductCatalog"):
        return {"table": "StagingGoodsItems weights", "status": "catalog_missing", "source": 0, "inserted": 0, "updated": 0}

    goods_cols = column_names(target, APP_TARGET_SCHEMA, "StagingGoodsItems")
    catalog_cols = column_names(target, APP_TARGET_SCHEMA, "DocProductCatalog")
    required_goods = {"gross_mass_kg", "net_mass_kg"}
    required_catalog = {"customer_code", "sku", "gross_weight_kg", "net_weight_kg"}
    has_sku_join = "sku" in goods_cols
    has_description_join = {"goods_description", "commodity_code"}.issubset(goods_cols) and {"description", "commodity_code"}.issubset(catalog_cols)
    if not required_goods.issubset(goods_cols) or not required_catalog.issubset(catalog_cols) or not (has_sku_join or has_description_join):
        return {"table": "StagingGoodsItems weights", "status": "missing_columns", "source": 0, "inserted": 0, "updated": 0}

    multiplier = "COALESCE(NULLIF(TRY_CONVERT(DECIMAL(18,3), g.[number_of_packages]), 0), 1)"
    if "number_of_packages" not in goods_cols:
        multiplier = "1"

    active_filter = "AND ISNULL(c.[active], 1) = 1" if "active" in catalog_cols else ""
    if has_sku_join:
        join_condition = "c.[customer_code] = N'ALL' AND c.[sku] = g.[sku]"
    else:
        join_condition = (
            "c.[customer_code] = N'ALL' "
            "AND UPPER(LTRIM(RTRIM(c.[description]))) = UPPER(LTRIM(RTRIM(g.[goods_description]))) "
            "AND ISNULL(c.[commodity_code], N'') = ISNULL(g.[commodity_code], N'')"
        )
    count_sql = f"""
        SELECT COUNT(*)
        FROM {q(APP_TARGET_SCHEMA)}.StagingGoodsItems g
        JOIN {q(APP_TARGET_SCHEMA)}.DocProductCatalog c
          ON {join_condition}
         {active_filter}
        WHERE (
              (g.[gross_mass_kg] IS NULL OR g.[gross_mass_kg] = 0)
              OR (g.[net_mass_kg] IS NULL OR g.[net_mass_kg] = 0)
        )
          AND c.[gross_weight_kg] IS NOT NULL
          AND c.[net_weight_kg] IS NOT NULL
    """
    cur = target.cursor()
    cur.execute(count_sql)
    eligible = int(cur.fetchone()[0])

    if execute and eligible:
        assignments = [
            f"g.[gross_mass_kg] = CASE WHEN g.[gross_mass_kg] IS NULL OR g.[gross_mass_kg] = 0 THEN ROUND(c.[gross_weight_kg] * {multiplier}, 3) ELSE g.[gross_mass_kg] END",
            f"g.[net_mass_kg] = CASE WHEN g.[net_mass_kg] IS NULL OR g.[net_mass_kg] = 0 THEN ROUND(c.[net_weight_kg] * {multiplier}, 3) ELSE g.[net_mass_kg] END",
        ]
        if "error_message" in goods_cols:
            cleaned = (
                "LTRIM(RTRIM(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE("
                "g.[error_message], N'missing gross/net weight', N''), "
                "N'missing gross weight', N''), "
                "N'missing net weight', N''), "
                "N'; ;', N';'), "
                "N'Pending: ;', N'Pending:')))"
            )
            assignments.append(
                "g.[error_message] = CASE WHEN g.[error_message] IS NULL THEN NULL "
                f"ELSE NULLIF(NULLIF(NULLIF({cleaned}, N'Pending:'), N';'), N'') END"
            )
        if "updated_at" in goods_cols:
            assignments.append("g.[updated_at] = SYSUTCDATETIME()")

        update_sql = f"""
            UPDATE g
               SET {', '.join(assignments)}
            FROM {q(APP_TARGET_SCHEMA)}.StagingGoodsItems g
            JOIN {q(APP_TARGET_SCHEMA)}.DocProductCatalog c
              ON {join_condition}
             {active_filter}
            WHERE (
                  (g.[gross_mass_kg] IS NULL OR g.[gross_mass_kg] = 0)
                  OR (g.[net_mass_kg] IS NULL OR g.[net_mass_kg] = 0)
            )
              AND c.[gross_weight_kg] IS NOT NULL
              AND c.[net_weight_kg] IS NOT NULL
        """
        target.cursor().execute(update_sql)

    return {
        "table": "StagingGoodsItems weights",
        "status": "ok",
        "source": eligible,
        "inserted": 0,
        "updated": eligible,
    }


def upsert_doc_product_catalog_rows(
    rows: list[Any],
    target: pyodbc.Connection,
    *,
    execute: bool,
) -> dict[str, int | str]:
    if not table_exists(target, APP_TARGET_SCHEMA, "DocProductCatalog"):
        return {"table": "DocProductCatalog", "status": "missing", "source": 0, "inserted": 0, "updated": 0}

    target_cols = column_names(target, APP_TARGET_SCHEMA, "DocProductCatalog")
    existing = existing_keys(target, APP_TARGET_SCHEMA, "DocProductCatalog", ("customer_code", "sku"))
    inserts: list[dict[str, Any]] = []
    updates: list[dict[str, Any]] = []

    for row in rows:
        item = clean(row_value(row, "Item")).upper()
        if not item:
            continue
        payload = {
            "customer_code": "ALL",
            "sku": item,
            "product_code": item,
            "description": clean(row_value(row, "Description")) or item,
            "commodity_code": clean(row_value(row, "CommodityCode")).replace(" ", ""),
            "country_of_origin": clean(row_value(row, "CountryOfOriginCode")).upper()[:2] or None,
            "controlled_goods": 0,
            "active": 1,
        }
        payload = {key: value for key, value in payload.items() if key in target_cols}
        if ("ALL", item) in existing:
            updates.append(payload)
        else:
            inserts.append(payload)

    if execute and inserts:
        insert_cols = [col for col in (
            "customer_code",
            "sku",
            "product_code",
            "description",
            "commodity_code",
            "country_of_origin",
            "controlled_goods",
            "active",
        ) if col in target_cols]
        value_expr = ["?" for _ in insert_cols]
        params = []
        for row in inserts:
            params.append([row.get(col) for col in insert_cols])
        sql = (
            f"INSERT INTO {q(APP_TARGET_SCHEMA)}.DocProductCatalog ({', '.join(q(c) for c in insert_cols)}) "
            f"VALUES ({', '.join(value_expr)})"
        )
        target.cursor().executemany(sql, params)

    if execute and updates:
        update_cols = [col for col in (
            "product_code",
            "description",
            "commodity_code",
            "country_of_origin",
            "controlled_goods",
            "active",
        ) if col in target_cols]
        assignments = [f"{q(col)} = ?" for col in update_cols]
        if "updated_at" in target_cols:
            assignments.append("[updated_at] = SYSUTCDATETIME()")
        sql = (
            f"UPDATE {q(APP_TARGET_SCHEMA)}.DocProductCatalog "
            f"SET {', '.join(assignments)} "
            "WHERE customer_code = ? AND sku = ?"
        )
        params = []
        for row in updates:
            params.append([row.get(col) for col in update_cols] + ["ALL", row["sku"]])
        target.cursor().executemany(sql, params)

    return {
        "table": "DocProductCatalog",
        "status": "ok",
        "source": len(rows),
        "inserted": len(inserts),
        "updated": len(updates),
    }


def upsert_doc_product_catalog(
    source: pyodbc.Connection,
    target: pyodbc.Connection,
    *,
    execute: bool,
) -> dict[str, int | str]:
    cur = source.cursor()
    cur.execute(
        """
        SELECT Item, Description, CommodityCode, CountryOfOriginCode
        FROM DATA.BKD_Item_Commodity_Code
        WHERE Item IS NOT NULL AND LTRIM(RTRIM(Item)) <> ''
        """
    )
    return upsert_doc_product_catalog_rows(cur.fetchall(), target, execute=execute)


def upsert_partners_from_customer_rows(
    rows: list[Any],
    target: pyodbc.Connection,
    *,
    execute: bool,
) -> dict[str, int | str]:
    if not table_exists(target, APP_TARGET_SCHEMA, "Partners"):
        return {"table": "Partners", "status": "missing", "source": 0, "inserted": 0, "updated": 0}
    target_cols = column_names(target, APP_TARGET_SCHEMA, "Partners")

    existing: set[tuple[Any, ...]] = set()
    if {"partner_type", "account_ref"}.issubset(target_cols):
        existing = existing_keys(target, APP_TARGET_SCHEMA, "Partners", ("partner_type", "account_ref"))

    inserts: list[dict[str, Any]] = []
    updates: list[dict[str, Any]] = []
    for row in rows:
        customer_no = clean(row_value(row, "CustomerNo")).upper()
        if not customer_no:
            continue
        payload = {
            "partner_type": "Consignee",
            "partner_name": clean(row_value(row, "Name")) or customer_no,
            "eori": safe_eori(row_value(row, "EORINumber")) or None,
            "address_line1": clean(row_value(row, "Address")) or None,
            "address_line2": clean(row_value(row, "Address2")) or None,
            "city": clean(row_value(row, "City")) or None,
            "county": clean(row_value(row, "County")) or None,
            "postcode": clean(row_value(row, "PostCode")).upper() or None,
            "country": clean(row_value(row, "CountryRegionCode")).upper()[:2] or None,
            "account_ref": customer_no,
            "active": 1,
            "source_system": "DATA.BKD_Customers",
            "source_loaded_at": clean(row_value(row, "UpdatedAtUTC") or row_value(row, "LoadedAtUTC")) or None,
            "notes": "Loaded from TSS_FLOW_V2_PRD.DATA.BKD_Customers",
        }
        payload = {key: value for key, value in payload.items() if key in target_cols}
        if ("Consignee", customer_no) in existing:
            updates.append(payload)
        else:
            inserts.append(payload)

    if execute and inserts:
        insert_cols = [col for col in (
            "partner_type",
            "partner_name",
            "eori",
            "address_line1",
            "address_line2",
            "city",
            "county",
            "postcode",
            "country",
            "account_ref",
            "active",
            "source_system",
            "source_loaded_at",
            "notes",
        ) if col in target_cols]
        sql = (
            f"INSERT INTO {q(APP_TARGET_SCHEMA)}.Partners ({', '.join(q(c) for c in insert_cols)}) "
            f"VALUES ({', '.join('?' for _ in insert_cols)})"
        )
        params = []
        for row in inserts:
            params.append([row.get(col) for col in insert_cols])
        target.cursor().executemany(sql, params)

    if execute and updates and {"partner_type", "account_ref"}.issubset(target_cols):
        update_cols = [col for col in (
            "partner_name",
            "eori",
            "address_line1",
            "address_line2",
            "city",
            "county",
            "postcode",
            "country",
            "active",
            "source_system",
            "source_loaded_at",
            "notes",
        ) if col in target_cols]
        assignments = [f"{q(col)} = ?" for col in update_cols]
        if "updated_at" in target_cols:
            assignments.append("[updated_at] = SYSUTCDATETIME()")
        sql = (
            f"UPDATE {q(APP_TARGET_SCHEMA)}.Partners "
            f"SET {', '.join(assignments)} "
            "WHERE partner_type = ? AND account_ref = ?"
        )
        params = []
        for row in updates:
            params.append([row.get(col) for col in update_cols] + ["Consignee", row["account_ref"]])
        target.cursor().executemany(sql, params)

    return {
        "table": "Partners",
        "status": "ok",
        "source": len(rows),
        "inserted": len(inserts),
        "updated": len(updates),
    }


def upsert_partners_from_customers(
    source: pyodbc.Connection,
    target: pyodbc.Connection,
    *,
    execute: bool,
) -> dict[str, int | str]:
    cur = source.cursor()
    cur.execute(
        """
        SELECT CustomerNo, Name, Address, Address2, City, County, PostCode,
               CountryRegionCode, EORINumber, LoadedAtUTC, UpdatedAtUTC
        FROM DATA.BKD_Customers
        WHERE CustomerNo IS NOT NULL AND LTRIM(RTRIM(CustomerNo)) <> ''
        """
    )
    return upsert_partners_from_customer_rows(cur.fetchall(), target, execute=execute)


def print_result(result: dict[str, int | str], dry_run: bool) -> None:
    prefix = "DRY-RUN " if dry_run else ""
    extras = [
        f"{key}={result[key]}"
        for key in ("matched", "unmatched", "invalid_weight", "valid", "invalid", "skipped")
        if key in result
    ]
    print(
        f"{prefix}{result['table']}: status={result['status']} "
        f"source={result['source']} inserted={result['inserted']} updated={result['updated']}"
        + (f" {' '.join(extras)}" if extras else "")
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-db", default=SOURCE_DB)
    parser.add_argument("--target-db", default=TARGET_DB)
    parser.add_argument("--csv-dir", help="Load no-header CSV files from this directory instead of source DB.")
    parser.add_argument("--csv-has-header", action="store_true", help="Treat CSV files as headered exports.")
    parser.add_argument("--execute", action="store_true", help="Write changes. Without this flag, only reports counts.")
    parser.add_argument("--no-raw", action="store_true", help="Skip raw BKD_* table copy.")
    parser.add_argument("--no-canonical", action="store_true", help="Skip DocProductCatalog/Partners integration.")
    parser.add_argument("--no-weights", action="store_true", help="Skip optional DocProductCatalog unit weight import.")
    parser.add_argument("--no-staging-weight-backfill", action="store_true", help="Do not backfill existing StagingGoodsItems weights after importing unit weights.")
    parser.add_argument("--weights-csv", help="Optional STG.BKD_GoodsItems CSV export used to calculate unit product weights.")
    parser.add_argument("--eori-xlsx", help="Optional EORI.xlsx validation workbook used to seed BKD.PrecheckEoriCache.")
    parser.add_argument("--no-update-existing", action="store_true", help="Only insert missing raw rows; do not update existing rows.")
    args = parser.parse_args()

    load_dotenv(ROOT / ".env")
    execute = bool(args.execute)
    csv_dir = Path(args.csv_dir) if args.csv_dir else None
    print(f"Source: {csv_dir if csv_dir else f'{args.source_db}.{SOURCE_SCHEMA}'}")
    print(f"Raw target: {args.target_db}.{RAW_TARGET_SCHEMA}")
    print(f"App target: {args.target_db}.{APP_TARGET_SCHEMA}")
    print("Mode: " + ("EXECUTE" if execute else "DRY-RUN"))

    needs_source = not csv_dir and (not args.no_raw or not args.no_canonical)
    source = connect(args.source_db) if needs_source else None
    target = connect(args.target_db)
    try:
        product_rows_for_weights: list[Any] = []
        if not args.no_raw:
            result = {"table": f"{RAW_TARGET_SCHEMA} schema", "status": ensure_schema(target, RAW_TARGET_SCHEMA, execute=execute), "source": 0, "inserted": 0, "updated": 0}
            print_result(result, not execute)
            for table in RAW_TABLE_KEYS:
                if csv_dir:
                    csv_path = csv_dir / CSV_TABLE_FILES[table]
                    rows = load_csv_rows(csv_path, table, has_header=args.csv_has_header)
                    result = copy_raw_rows(
                        target,
                        table,
                        rows,
                        execute=execute,
                        update_existing=not args.no_update_existing,
                    )
                else:
                    result = copy_raw_table(
                        source,
                        target,
                        table,
                        execute=execute,
                        update_existing=not args.no_update_existing,
                    )
                print_result(result, not execute)

        if not args.no_canonical:
            if csv_dir:
                product_rows = load_csv_rows(csv_dir / CSV_TABLE_FILES["BKD_Item_Commodity_Code"], "BKD_Item_Commodity_Code", has_header=args.csv_has_header)
                customer_rows = load_csv_rows(csv_dir / CSV_TABLE_FILES["BKD_Customers"], "BKD_Customers", has_header=args.csv_has_header)
                print_result(upsert_doc_product_catalog_rows(product_rows, target, execute=execute), not execute)
                print_result(upsert_partners_from_customer_rows(customer_rows, target, execute=execute), not execute)
                product_rows_for_weights = product_rows
            else:
                product_rows_for_weights = fetch_rows(source, SOURCE_SCHEMA, "BKD_Item_Commodity_Code", RAW_TABLE_COLUMNS["BKD_Item_Commodity_Code"])
                print_result(upsert_doc_product_catalog_rows(product_rows_for_weights, target, execute=execute), not execute)
                print_result(upsert_partners_from_customers(source, target, execute=execute), not execute)

        weight_path = Path(args.weights_csv) if args.weights_csv else None
        if not weight_path and csv_dir:
            candidate = csv_dir / "weight Products.csv"
            weight_path = candidate if candidate.exists() else None
        if weight_path and not args.no_weights:
            weight_rows = load_weight_product_rows(weight_path, has_header=args.csv_has_header)
            if not product_rows_for_weights:
                if table_exists(target, RAW_TARGET_SCHEMA, "BKD_Item_Commodity_Code"):
                    product_rows_for_weights = fetch_rows(
                        target,
                        RAW_TARGET_SCHEMA,
                        "BKD_Item_Commodity_Code",
                        RAW_TABLE_COLUMNS["BKD_Item_Commodity_Code"],
                    )
                elif source is not None and table_exists(source, SOURCE_SCHEMA, "BKD_Item_Commodity_Code"):
                    product_rows_for_weights = fetch_rows(
                        source,
                        SOURCE_SCHEMA,
                        "BKD_Item_Commodity_Code",
                        RAW_TABLE_COLUMNS["BKD_Item_Commodity_Code"],
                    )
            print_result(
                upsert_doc_product_catalog_weights(
                    weight_rows,
                    product_rows_for_weights,
                    target,
                    execute=execute,
                ),
                not execute,
            )
            if not args.no_staging_weight_backfill:
                print_result(backfill_staging_goods_weights(target, execute=execute), not execute)

        if args.eori_xlsx:
            eori_rows = load_eori_validation_xlsx(Path(args.eori_xlsx))
            print_result(upsert_eori_precheck_cache(eori_rows, target, execute=execute), not execute)

        if execute:
            target.commit()
            print("Committed.")
        else:
            target.rollback()
            print("Dry-run complete. Re-run with --execute to apply.")
        return 0
    except Exception:
        target.rollback()
        raise
    finally:
        if source is not None:
            source.close()
        target.close()


if __name__ == "__main__":
    raise SystemExit(main())
