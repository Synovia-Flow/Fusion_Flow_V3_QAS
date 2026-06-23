"""Build the customer-facing BKD SUPDEC action workbook.

The workbook is generated from the current official TSS API response plus the
PRD staging/source rows. It separates customer asks from internal Fusion fixes:

* customer values required: TSS says Missing Item Value and Fusion/source data
  has no real non-zero value.
* internal value updates: TSS says Missing Item Value but Fusion already has a
  non-zero value that can be resent without asking Birkdale.

Dry-run/read-only against TSS and SQL; writes only the output XLSX artifact.
"""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
import json
from pathlib import Path
import re
import sys
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    from dotenv import load_dotenv
except Exception:  # pragma: no cover - local convenience only
    load_dotenv = None

if load_dotenv:
    load_dotenv(ROOT / ".env")

from app import create_app  # noqa: E402
from app.db import get_standalone_connection  # noqa: E402
from app.ingestion.excel_sales_orders import parse_sales_orders_excel  # noqa: E402
from app.tss_api import build_cfg_client  # noqa: E402
from app.blueprints.supdec import routes as supdec_routes  # noqa: E402
from scripts.pull_inbound_email import _attachment_archive_dir  # noqa: E402


FINAL_STATUSES = {"CLOSED", "COMPLETED", "CANCELLED", "CANCELED", "CLEARED", "ACCEPTED"}
VALUE_ERROR_TEXT = "Missing Item Value"
ITEM_RE = re.compile(
    r"Goods Item Number:\s*(?P<start>\d+)(?:\s*(?:-|to|–)\s*(?P<end>\d+))?\s*-\s*(?P<issue>[^\r\n]+)",
    re.IGNORECASE,
)

DEFAULT_CANCELLATION_ROWS = [
    {
        "Transport Document": "S-ORD356007",
        "SUPDEC to keep": "SUP000000007513394",
        "Keep Arrival Date/Time": "2026-05-21 06:30:00",
        "Keep DEC/SFD": "DEC000000016867021",
        "SUPDEC to cancel": "SUP000000007519231",
        "Cancel Arrival Date/Time": "2026-05-22 06:30:00",
        "Cancel DEC/SFD": "DEC000000016878598",
        "Required Action": "Raise TSS ticket/request to cancel duplicate SUPDEC",
        "Reason": "Duplicate SUPDEC for the same June movement; TSS refused API cancellation",
    },
    {
        "Transport Document": "S-ORD356095",
        "SUPDEC to keep": "SUP000000007513579",
        "Keep Arrival Date/Time": "2026-05-21 06:30:00",
        "Keep DEC/SFD": "DEC000000016866128",
        "SUPDEC to cancel": "SUP000000007519213",
        "Cancel Arrival Date/Time": "2026-05-22 06:30:00",
        "Cancel DEC/SFD": "DEC000000016878599",
        "Required Action": "Raise TSS ticket/request to cancel duplicate SUPDEC",
        "Reason": "Duplicate SUPDEC for the same June movement; TSS refused API cancellation",
    },
    {
        "Transport Document": "S-ORD357240",
        "SUPDEC to keep": "SUP000000007549011",
        "Keep Arrival Date/Time": "2026-05-30 06:30:00",
        "Keep DEC/SFD": "DEC000000016948529",
        "SUPDEC to cancel": "SUP000000007557095",
        "Cancel Arrival Date/Time": "2026-06-02 06:30:00",
        "Cancel DEC/SFD": "DEC000000016964014",
        "Required Action": "Raise TSS ticket/request to cancel duplicate SUPDEC",
        "Reason": "Proven duplicate goods fingerprint; TSS refused API cancellation",
    },
]


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _norm_text(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", _clean(value).upper()).strip()


def _status(value: Any) -> str:
    return _clean(value).replace("_", " ").upper()


def _sup(row: dict[str, Any]) -> str:
    return _clean(row.get("sup_dec_number") or row.get("tss_sup_dec_number"))


def _date_text(value: Any) -> str:
    if not value:
        return ""
    return str(value)[:19]


def _decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def _nonzero(value: Any) -> Decimal | None:
    number = _decimal(value)
    if number is None or number == 0:
        return None
    return number


def _multiply(left: Any, right: Any) -> Decimal | None:
    left_number = _decimal(left)
    right_number = _decimal(right)
    if left_number is None or right_number is None:
        return None
    result = left_number * right_number
    return result if result != 0 else None


def _item_value(item: dict[str, Any], *names: str) -> Any:
    for name in names:
        value = (item or {}).get(name)
        if value in (None, ""):
            continue
        if isinstance(value, dict):
            for nested_key in ("value", "display_value", "displayValue", "label", "name"):
                nested = value.get(nested_key)
                if nested not in (None, ""):
                    return nested
            continue
        return value
    return None


def _tss_payload(result: Any) -> dict[str, Any]:
    return supdec_routes._tss_response_payload(result)


def _tss_value(payload: dict[str, Any], key: str) -> Any:
    return supdec_routes._tss_response_value(payload, key)


def _official_read(api: Any, sup_ref: str) -> dict[str, Any]:
    result = api.read_sdi(sup_ref, fields=list(supdec_routes.PRD_SDI_SUBMIT_READ_FIELDS))
    payload = _tss_payload(result)
    return {
        "ok": supdec_routes._sdi_tss_result_ok(result),
        "status": _status(_tss_value(payload, "status")),
        "error_message": _clean(_tss_value(payload, "error_message")),
        "raw": result,
    }


def _missing_item_numbers(error_message: str) -> list[tuple[int, str]]:
    found: list[tuple[int, str]] = []
    for match in ITEM_RE.finditer(error_message or ""):
        issue = _clean(match.group("issue"))
        if VALUE_ERROR_TEXT.lower() not in issue.lower():
            continue
        start = int(match.group("start"))
        end = int(match.group("end") or start)
        for item_no in range(start, end + 1):
            found.append((item_no, issue))
    return found


def _load_goods(cur: Any, stg_sdi_id: int) -> dict[int, dict[str, Any]]:
    cur.execute(
        """
        SELECT
            g.stg_sdi_item_id,
            g.item_seq,
            g.goods_description,
            g.commodity_code,
            g.country_of_origin,
            g.number_of_individual_pieces,
            g.item_invoice_amount,
            g.customs_value,
            g.line_amount_excl_vat,
            g.source_amount,
            g.unit_price_excl_vat,
            g.tss_goods_id,
            sg.sku AS source_sku,
            sg.ing_item_id AS source_ing_item_id,
            sg.item_invoice_amount AS source_item_invoice_amount,
            sg.customs_value AS source_customs_value,
            sg.line_amount_excl_vat AS source_line_amount_excl_vat,
            sg.source_amount AS source_source_amount,
            sg.unit_price_excl_vat AS source_unit_price_excl_vat,
            sg.number_of_individual_pieces AS source_pieces,
            sol.RecordId AS ing_record_id,
            sol.SourceRowNum AS ing_source_row,
            sol.ItemNo AS ing_item_no,
            sol.Amount AS ing_amount,
            sol.LineAmountExclVat AS ing_line_amount_excl_vat,
            sol.UnitPriceExclVat AS ing_unit_price_excl_vat,
            sol.Quantity AS ing_quantity,
            sol.DocumentNo AS ing_document_no,
            sfl.FileName AS source_file_name,
            sfl.ArchivePath AS source_archive_path
        FROM [STG].[BKD_SDI_GoodsItems] g
        LEFT JOIN [STG].[BKD_GoodsItems] sg
               ON sg.ClientCode = g.ClientCode
              AND sg.stg_item_id = g.source_stg_item_id
        LEFT JOIN [ING].[BKD_SalesOrderLine] sol
               ON sol.RecordId = sg.ing_item_id
        LEFT JOIN [ING].[BKD_SourceFileLog] sfl
               ON sfl.FileId = sol.FileId
        WHERE g.ClientCode = 'BKD'
          AND g.stg_sdi_id = ?
        ORDER BY COALESCE(g.item_seq, g.stg_sdi_item_id), g.stg_sdi_item_id
        """,
        [stg_sdi_id],
    )
    columns = [column[0] for column in cur.description or []]
    goods: dict[int, dict[str, Any]] = {}
    fallback_index = 1
    for row in cur.fetchall():
        item = dict(zip(columns, row))
        item_seq = int(item.get("item_seq") or fallback_index)
        goods[item_seq] = item
        fallback_index += 1
    return goods


def _table_columns(cur: Any, schema: str, table: str) -> set[str]:
    cur.execute(
        """
        SELECT LOWER(COLUMN_NAME)
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
        """,
        [schema, table],
    )
    return {str(row[0]).lower() for row in cur.fetchall()}


def _catalog_select(columns: set[str], name: str, fallback: str = "CAST(NULL AS NVARCHAR(4000))") -> str:
    return f"[{name}]" if name.lower() in columns else fallback


def _load_catalog_lookup(cur: Any) -> dict[tuple[str, str, str], dict[str, Any]]:
    try:
        columns = _table_columns(cur, "BKD", "DocProductCatalog")
    except Exception:
        return {}
    if not columns or not ({"sku", "product_code"} & columns) or "description" not in columns:
        return {}

    where_parts = ["NULLIF(LTRIM(RTRIM([description])), '') IS NOT NULL"]
    if "active" in columns:
        where_parts.append("COALESCE([active], 1) = 1")

    cur.execute(
        f"""
        SELECT
            {_catalog_select(columns, "sku")} AS sku,
            {_catalog_select(columns, "product_code")} AS product_code,
            {_catalog_select(columns, "description")} AS description,
            {_catalog_select(columns, "commodity_code")} AS commodity_code,
            {_catalog_select(columns, "country_of_origin")} AS country_of_origin
        FROM [BKD].[DocProductCatalog]
        WHERE {' AND '.join(where_parts)}
        """,
    )
    rows = []
    for row in cur.fetchall():
        item = {
            "sku": _clean(row[0] or row[1]),
            "product_code": _clean(row[1] or row[0]),
            "description": _clean(row[2]),
            "commodity_code": _clean(row[3]).replace(" ", ""),
            "country_of_origin": _clean(row[4]).upper(),
        }
        if item["sku"] and item["description"]:
            rows.append(item)

    exact: dict[tuple[str, str, str], dict[str, Any]] = {}
    buckets: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    for item in rows:
        keys = [
            (_norm_text(item["description"]), item["commodity_code"], item["country_of_origin"]),
            (_norm_text(item["description"]), item["commodity_code"], ""),
            (_norm_text(item["description"]), "", ""),
        ]
        for key in keys:
            buckets.setdefault(key, []).append(item)

    for key, matches in buckets.items():
        skus = {_clean(match.get("sku")) for match in matches if _clean(match.get("sku"))}
        if len(skus) == 1:
            exact[key] = matches[0]
    return exact


def _apply_catalog_source(
    goods: dict[str, Any],
    *,
    catalog_lookup: dict[tuple[str, str, str], dict[str, Any]],
    api_goods: dict[str, Any] | None,
) -> dict[str, Any]:
    if _clean(goods.get("ing_item_no") or goods.get("source_sku") or goods.get("archive_sku") or goods.get("catalog_sku")):
        return goods
    description = _clean(goods.get("goods_description")) or _clean(_item_value(api_goods or {}, "goods_description", "goodsDescription"))
    if not description:
        return goods
    commodity = _clean(goods.get("commodity_code") or _item_value(api_goods or {}, "commodity_code", "commodityCode")).replace(" ", "")
    origin = _clean(goods.get("country_of_origin") or _item_value(api_goods or {}, "country_of_origin", "countryOfOrigin")).upper()
    for key in (
        (_norm_text(description), commodity, origin),
        (_norm_text(description), commodity, ""),
        (_norm_text(description), "", ""),
    ):
        match = catalog_lookup.get(key)
        if match:
            merged = dict(goods)
            merged["catalog_sku"] = match.get("sku")
            merged["catalog_product_code"] = match.get("product_code")
            return merged
    return goods


def _source_values(goods: dict[str, Any]) -> list[tuple[str, Decimal]]:
    candidates = [
        ("STG SDI item_invoice_amount", _nonzero(goods.get("item_invoice_amount"))),
        ("STG SDI customs_value", _nonzero(goods.get("customs_value"))),
        ("STG SDI line_amount_excl_vat", _nonzero(goods.get("line_amount_excl_vat"))),
        ("STG SDI source_amount", _nonzero(goods.get("source_amount"))),
        (
            "STG SDI unit_price_excl_vat * pieces",
            _multiply(goods.get("unit_price_excl_vat"), goods.get("number_of_individual_pieces")),
        ),
        ("Source goods item_invoice_amount", _nonzero(goods.get("source_item_invoice_amount"))),
        ("Source goods customs_value", _nonzero(goods.get("source_customs_value"))),
        ("Source goods line_amount_excl_vat", _nonzero(goods.get("source_line_amount_excl_vat"))),
        ("Source goods source_amount", _nonzero(goods.get("source_source_amount"))),
        (
            "Source goods unit_price_excl_vat * pieces",
            _multiply(goods.get("source_unit_price_excl_vat"), goods.get("source_pieces")),
        ),
        ("ING Sales Order Amount", _nonzero(goods.get("ing_amount"))),
        ("ING Sales Order LineAmountExclVat", _nonzero(goods.get("ing_line_amount_excl_vat"))),
        (
            "ING Sales Order UnitPriceExclVat * Quantity",
            _multiply(goods.get("ing_unit_price_excl_vat"), goods.get("ing_quantity")),
        ),
        ("Archived Sales Order Amount", _nonzero(goods.get("archive_amount"))),
        ("Archived Sales Order LineAmountExclVat", _nonzero(goods.get("archive_line_amount_excl_vat"))),
        (
            "Archived Sales Order UnitPriceExclVat * Quantity",
            _multiply(goods.get("archive_unit_price_excl_vat"), goods.get("archive_quantity")),
        ),
    ]
    return [(name, value) for name, value in candidates if value is not None]


def _sku_value(goods: dict[str, Any]) -> tuple[str, str]:
    for label, key in (
        ("ING Sales Order", "ing_item_no"),
        ("Source goods", "source_sku"),
        ("Archived Sales Order", "archive_sku"),
        ("Product masterdata", "catalog_sku"),
        ("Product masterdata", "catalog_product_code"),
    ):
        value = _clean(goods.get(key))
        if value:
            return value, label
    return "(blank)", ""


def _format_decimal(value: Any) -> str:
    number = _decimal(value)
    if number is None:
        return ""
    return f"{number.normalize():f}"


def _value_evidence(goods: dict[str, Any]) -> str:
    parts = []
    for label, key in (
        ("STG SDI item", "item_invoice_amount"),
        ("STG SDI customs", "customs_value"),
        ("STG SDI line", "line_amount_excl_vat"),
        ("STG SDI source", "source_amount"),
        ("STG SDI unit", "unit_price_excl_vat"),
        ("Source item", "source_item_invoice_amount"),
        ("Source customs", "source_customs_value"),
        ("Source line", "source_line_amount_excl_vat"),
        ("Source amount", "source_source_amount"),
        ("Source unit", "source_unit_price_excl_vat"),
        ("ING amount", "ing_amount"),
        ("ING line", "ing_line_amount_excl_vat"),
        ("ING unit", "ing_unit_price_excl_vat"),
        ("Archive amount", "archive_amount"),
        ("Archive line", "archive_line_amount_excl_vat"),
        ("Archive unit", "archive_unit_price_excl_vat"),
    ):
        text = _format_decimal(goods.get(key))
        if text:
            parts.append(f"{label}={text}")
    return "; ".join(parts) or "No value found in STG/ING source columns"


def _api_goods_by_item(api: Any, sup_ref: str) -> dict[int, dict[str, Any]]:
    try:
        goods_items = api.lookup_sdi_goods(sup_ref) or []
    except Exception:
        return {}
    result: dict[int, dict[str, Any]] = {}
    fallback = 1
    for item in goods_items:
        if not isinstance(item, dict):
            continue
        raw_item_number = _item_value(item, "item_number", "itemNumber", "goods_item_number", "goodsItemNumber")
        try:
            item_number = int(raw_item_number or fallback)
        except Exception:
            item_number = fallback
        result[item_number] = item
        fallback += 1
    return result


def _load_archived_sales_order_lookup() -> dict[str, list[dict[str, Any]]]:
    lookup: dict[str, list[dict[str, Any]]] = {}
    archive_dir = _attachment_archive_dir()
    if not archive_dir.exists():
        return lookup
    for path in sorted(archive_dir.rglob("Sales Orders Synovia*.xlsx")):
        try:
            parsed = parse_sales_orders_excel(path)
        except Exception:
            continue
        for consignment in parsed.consignments:
            goods_rows = lookup.setdefault(_clean(consignment.document_no), [])
            for position, goods in enumerate(consignment.goods, start=1):
                goods_rows.append(
                    {
                        "archive_position": position,
                        "archive_file_name": path.name,
                        "archive_file_path": str(path),
                        "archive_sku": goods.sku,
                        "archive_line_no": goods.line_no,
                        "archive_amount": goods.amount,
                        "archive_line_amount_excl_vat": goods.line_amount_excl_vat,
                        "archive_unit_price_excl_vat": goods.unit_price_excl_vat,
                        "archive_quantity": goods.quantity,
                    }
                )
    return lookup


def _merge_archived_source(
    goods: dict[str, Any],
    *,
    transport_document_number: str,
    item_no: int,
    archived_lookup: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    archived_rows = archived_lookup.get(_clean(transport_document_number)) or []
    if not archived_rows:
        return goods
    selected = archived_rows[item_no - 1] if 0 <= item_no - 1 < len(archived_rows) else None
    if not selected:
        return goods
    merged = dict(goods)
    for key, value in selected.items():
        if key in {"archive_file_name", "archive_file_path"}:
            merged[key] = value
        elif not _clean(merged.get(key)):
            merged[key] = value
    return merged


def _build_value_row(
    *,
    header: dict[str, Any],
    official: dict[str, Any],
    item_no: int,
    issue: str,
    goods: dict[str, Any],
    api_goods: dict[str, Any] | None,
) -> dict[str, Any]:
    source_values = _source_values(goods)
    has_real_value = bool(source_values)
    description = (
        _clean(goods.get("goods_description"))
        or _clean(_item_value(api_goods or {}, "goods_description", "goodsDescription"))
    )
    sku, sku_source = _sku_value(goods)
    action_owner = "Synovia" if has_real_value else "Birkdale"
    return {
        "Action Owner": action_owner,
        "SUPDEC": _sup(header),
        "TSS Status": official.get("status"),
        "TSS Due Date": _date_text(header.get("tss_submission_due_date") or header.get("submission_due_date")),
        "Arrival Date/Time": _date_text(header.get("arrival_date_time")),
        "Transport Document": header.get("transport_document_number"),
        "SFD / DEC Reference": header.get("sfd_reference") or header.get("tss_sfd_consignment_ref"),
        "Goods Item": item_no,
        "TSS Issue": issue,
        "SKU / ItemNo": sku,
        "SKU Source": sku_source,
        "Product / Description": description,
        "Commodity Code": _clean(goods.get("commodity_code") or _item_value(api_goods or {}, "commodity_code", "commodityCode")),
        "Country of Origin": _clean(goods.get("country_of_origin") or _item_value(api_goods or {}, "country_of_origin", "countryOfOrigin")),
        "Quantity / Pieces": _format_decimal(
            goods.get("ing_quantity") or goods.get("number_of_individual_pieces") or goods.get("source_pieces")
        ),
        "Current STG Item Value": _format_decimal(goods.get("item_invoice_amount")),
        "Current STG Customs Value": _format_decimal(goods.get("customs_value")),
        "Source Row": goods.get("ing_source_row") or goods.get("archive_line_no") or goods.get("archive_position") or "",
        "Source File": goods.get("archive_file_name") or goods.get("source_file_name") or "",
        "Source Archive Path": goods.get("archive_file_path") or goods.get("source_archive_path") or "",
        "Value Evidence": _value_evidence(goods),
        "Required Action": (
            "Confirm correct commercial value for this goods line"
            if action_owner == "Birkdale"
            else "Internal Fusion update/resubmit using existing non-zero source value"
        ),
        "Official TSS Error": official.get("error_message"),
    }


def _build_goods_context_row(
    *,
    header: dict[str, Any],
    official: dict[str, Any],
    item_no: int,
    goods: dict[str, Any],
    api_goods: dict[str, Any] | None,
    missing_issue: str | None,
) -> dict[str, Any]:
    source_values = _source_values(goods)
    has_real_value = bool(source_values)
    is_missing_value_item = bool(missing_issue)
    if is_missing_value_item and not has_real_value:
        customer_action = "Yes"
        workbook_section = "Customer values required"
        fusion_position = "Missing/zero value in Fusion/source/archive; Birkdale needs to confirm commercial value"
    elif is_missing_value_item and has_real_value:
        customer_action = "No"
        workbook_section = "Internal value updates"
        fusion_position = "TSS reports missing value, but Fusion/source/archive has a non-zero value to resend"
    elif has_real_value:
        customer_action = "No"
        workbook_section = "Context only"
        fusion_position = "No customer action; Fusion/source/masterdata value evidence already present"
    else:
        customer_action = "No"
        workbook_section = "Context only"
        fusion_position = "No customer action in the current official TSS Missing Item Value response"

    description = (
        _clean(goods.get("goods_description"))
        or _clean(_item_value(api_goods or {}, "goods_description", "goodsDescription"))
    )
    sku, sku_source = _sku_value(goods)
    return {
        "SUPDEC": _sup(header),
        "Transport Document": header.get("transport_document_number"),
        "Arrival Date/Time": _date_text(header.get("arrival_date_time")),
        "TSS Due Date": _date_text(header.get("tss_submission_due_date") or header.get("submission_due_date")),
        "Official TSS Status": official.get("status"),
        "Goods Item": item_no,
        "Workbook Section": workbook_section,
        "Customer Action Required": customer_action,
        "TSS Missing Item Value?": "Yes" if is_missing_value_item else "No",
        "TSS Issue": missing_issue or "",
        "Fusion Data Position": fusion_position,
        "SKU / ItemNo": sku,
        "SKU Source": sku_source,
        "Product / Description": description,
        "Commodity Code": _clean(goods.get("commodity_code") or _item_value(api_goods or {}, "commodity_code", "commodityCode")),
        "Country of Origin": _clean(goods.get("country_of_origin") or _item_value(api_goods or {}, "country_of_origin", "countryOfOrigin")),
        "Quantity / Pieces": _format_decimal(
            goods.get("ing_quantity") or goods.get("number_of_individual_pieces") or goods.get("source_pieces")
        ),
        "Current STG Item Value": _format_decimal(goods.get("item_invoice_amount")),
        "Current STG Customs Value": _format_decimal(goods.get("customs_value")),
        "Source Row": goods.get("ing_source_row") or goods.get("archive_line_no") or goods.get("archive_position") or "",
        "Source File": goods.get("archive_file_name") or goods.get("source_file_name") or "",
        "Source Archive Path": goods.get("archive_file_path") or goods.get("source_archive_path") or "",
        "Value Evidence": _value_evidence(goods),
    }


def _load_existing_cancellation_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return list(DEFAULT_CANCELLATION_ROWS)
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if "Cancellation tickets" not in wb.sheetnames:
        return []
    ws = wb["Cancellation tickets"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return list(DEFAULT_CANCELLATION_ROWS)
    header_index = None
    for index, row in enumerate(rows):
        values = [_clean(value) for value in row]
        if "Transport Document" in values and "SUPDEC to cancel" in values:
            header_index = index
            break
    if header_index is None:
        return list(DEFAULT_CANCELLATION_ROWS)
    headers = [_clean(value) for value in rows[header_index]]
    parsed = [dict(zip(headers, row)) for row in rows[header_index + 1 :] if any(row)]
    return parsed or list(DEFAULT_CANCELLATION_ROWS)


def _write_sheet(ws: Any, rows: list[dict[str, Any]], *, title: str) -> None:
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    if not rows:
        ws.append(["No rows"])
        return
    headers = list(rows[0])
    ws.append(headers)
    header_row = ws.max_row
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
        row_number = ws.max_row
        for index, header in enumerate(headers, start=1):
            value = str(row.get(header, "") or "").strip().lower()
            if header in {"Customer Action Required", "TSS Missing Item Value?"}:
                if value == "yes":
                    ws.cell(row_number, index).fill = PatternFill("solid", fgColor="F8D7DA")
                    ws.cell(row_number, index).font = Font(bold=True, color="842029")
                elif value == "no":
                    ws.cell(row_number, index).fill = PatternFill("solid", fgColor="D1E7DD")
                    ws.cell(row_number, index).font = Font(bold=True, color="0F5132")
    ws.freeze_panes = f"A{header_row + 1}"
    for index, header in enumerate(headers, start=1):
        max_len = max(len(str(header)), *(len(str(row.get(header, "") or "")) for row in rows))
        ws.column_dimensions[get_column_letter(index)].width = min(max(12, max_len + 2), 60)


def build_workbook(
    *,
    output: Path,
    cancellation_source: Path,
    audit_json: Path,
) -> dict[str, Any]:
    app = create_app()
    with app.app_context():
        api = build_cfg_client()
        headers = supdec_routes._load_prd_sdi_headers("")
        open_headers = [
            row
            for row in headers
            if _sup(row) and _status(row.get("tss_status") or row.get("sub_status")) not in FINAL_STATUSES
        ]

        conn = get_standalone_connection()
        cur = conn.cursor()
        try:
            customer_rows: list[dict[str, Any]] = []
            internal_rows: list[dict[str, Any]] = []
            goods_context_rows: list[dict[str, Any]] = []
            audit_events: list[dict[str, Any]] = []
            archived_lookup = _load_archived_sales_order_lookup()
            catalog_lookup = _load_catalog_lookup(cur)
            for header in open_headers:
                sup_ref = _sup(header)
                official = _official_read(api, sup_ref)
                event = {
                    "sup_ref": sup_ref,
                    "transport_document_number": header.get("transport_document_number"),
                    "local_status": _status(header.get("tss_status") or header.get("sub_status")),
                    "official_status": official.get("status"),
                    "official_error_message": official.get("error_message"),
                }
                missing_items = _missing_item_numbers(official.get("error_message") or "")
                event["missing_item_value_items"] = [item for item, _issue in missing_items]
                if not missing_items:
                    audit_events.append(event)
                    continue

                local_goods = _load_goods(cur, int(header.get("staging_id") or header.get("stg_sdi_id") or 0))
                api_goods = _api_goods_by_item(api, sup_ref)
                missing_issue_by_item = {item_no: issue for item_no, issue in missing_items}
                context_item_numbers = sorted(set(local_goods) | set(api_goods) | set(missing_issue_by_item))
                for context_item_no in context_item_numbers:
                    goods = _merge_archived_source(
                        local_goods.get(context_item_no, {}),
                        transport_document_number=_clean(header.get("transport_document_number")),
                        item_no=context_item_no,
                        archived_lookup=archived_lookup,
                    )
                    goods = _apply_catalog_source(
                        goods,
                        catalog_lookup=catalog_lookup,
                        api_goods=api_goods.get(context_item_no),
                    )
                    goods_context_rows.append(
                        _build_goods_context_row(
                            header=header,
                            official=official,
                            item_no=context_item_no,
                            goods=goods,
                            api_goods=api_goods.get(context_item_no),
                            missing_issue=missing_issue_by_item.get(context_item_no),
                        )
                    )
                for item_no, issue in missing_items:
                    goods = _merge_archived_source(
                        local_goods.get(item_no, {}),
                        transport_document_number=_clean(header.get("transport_document_number")),
                        item_no=item_no,
                        archived_lookup=archived_lookup,
                    )
                    goods = _apply_catalog_source(
                        goods,
                        catalog_lookup=catalog_lookup,
                        api_goods=api_goods.get(item_no),
                    )
                    row = _build_value_row(
                        header=header,
                        official=official,
                        item_no=item_no,
                        issue=issue,
                        goods=goods,
                        api_goods=api_goods.get(item_no),
                    )
                    if row["Action Owner"] == "Birkdale":
                        customer_rows.append(row)
                    else:
                        internal_rows.append(row)
                audit_events.append(event)
        finally:
            cur.close()
            conn.close()

    cancellation_rows = _load_existing_cancellation_rows(cancellation_source)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    summary_rows = [
        ("Birkdale SUPDEC actions required - verified", ""),
        ("Generated", generated),
        ("Source", "Official TSS API read + PRD STG/ING source data"),
        ("Customer values required", len(customer_rows)),
        ("Internal value updates, not customer ask", len(internal_rows)),
        ("Affected SUPDEC goods context rows", len(goods_context_rows)),
        ("Duplicate cancellation tickets", len(cancellation_rows)),
        ("Open SUPDECs checked", len(open_headers)),
    ]
    for row in summary_rows:
        summary.append(row)
    summary["A1"].font = Font(bold=True, size=14)
    for col in ("A", "B"):
        summary.column_dimensions[col].width = 42

    cancel_ws = wb.create_sheet("Cancellation tickets")
    _write_sheet(cancel_ws, cancellation_rows, title="Duplicate SUPDECs to raise with TSS for cancellation")

    customer_ws = wb.create_sheet("Customer values required")
    _write_sheet(customer_ws, customer_rows, title="Commercial values Birkdale needs to confirm")

    internal_ws = wb.create_sheet("Internal value updates")
    _write_sheet(internal_ws, internal_rows, title="Missing Item Value cases with existing non-zero Fusion/source value")

    context_ws = wb.create_sheet("Affected SUPDEC goods")
    _write_sheet(
        context_ws,
        goods_context_rows,
        title="All goods for SUPDECs with Missing Item Value context",
    )

    audit_ws = wb.create_sheet("TSS audit")
    audit_rows = [
        {
            "SUPDEC": event["sup_ref"],
            "Transport Document": event.get("transport_document_number"),
            "Local Status": event.get("local_status"),
            "Official TSS Status": event.get("official_status"),
            "Missing Item Value Items": ", ".join(str(item) for item in event.get("missing_item_value_items") or []),
            "Official TSS Error": event.get("official_error_message"),
        }
        for event in audit_events
    ]
    _write_sheet(audit_ws, audit_rows, title="Official TSS read audit for open local SUPDECs")

    wb.save(output)

    audit_json.parent.mkdir(parents=True, exist_ok=True)
    audit_json.write_text(json.dumps(audit_events, indent=2, default=str, ensure_ascii=True), encoding="utf-8")

    return {
        "output": str(output),
        "audit_json": str(audit_json),
        "generated": generated,
        "open_checked": len(open_headers),
        "customer_values_required": len(customer_rows),
        "internal_value_updates": len(internal_rows),
        "goods_context_rows": len(goods_context_rows),
        "duplicate_cancellation_tickets": len(cancellation_rows),
        "status_counts": dict(Counter(event.get("official_status") or "UNKNOWN" for event in audit_events)),
        "customer_sups": sorted({row["SUPDEC"] for row in customer_rows}),
        "internal_sups": sorted({row["SUPDEC"] for row in internal_rows}),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output",
        default="artifacts/customer/Birkdale_SUPDEC_actions_required_20260612_verified.xlsx",
    )
    parser.add_argument(
        "--cancellation-source",
        default="artifacts/customer/Birkdale_SUPDEC_actions_required_20260612.xlsx",
    )
    parser.add_argument(
        "--audit-json",
        default="artifacts/customer/Birkdale_SUPDEC_actions_required_20260612_verified_audit.json",
    )
    args = parser.parse_args()
    summary = build_workbook(
        output=Path(args.output),
        cancellation_source=Path(args.cancellation_source),
        audit_json=Path(args.audit_json),
    )
    print(json.dumps(summary, indent=2, ensure_ascii=True, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
