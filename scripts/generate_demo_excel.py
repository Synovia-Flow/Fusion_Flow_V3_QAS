"""
NOT FOR PRD: reads BKD.StagingEnsHeaders / BKD.StagingConsignments removed by migration 078.
             Use STG.BKD_* or ING.BKD_* for new pipeline work.

Generate demo Excel upload files for the Fusion Flow demo.

Usage:
    python scripts/generate_demo_excel.py [--ens-id N] [--cons-id N] [--out-dir DIR]

    --ens-id   staging_id from BKD.StagingEnsHeaders (from SQL seed output)
    --cons-id  staging_id from BKD.StagingConsignments (from SQL seed output)
    --out-dir  output directory (default: current dir)

Generates:
    demo_consignment_upload.xlsx  — 1 consignment row for bulk upload
    demo_goods_upload.xlsx        — 2 goods item rows for bulk upload

Format matches the Fusion Flow bulk upload template (row 2 = DB column names,
row 5 = data). Upload via: Bulk Upload → Consignment / Goods Item.
"""

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

_NAVY  = '0B1D3A'
_BLUE  = '1A5CCC'
_AMBER = 'D97706'
_WHITE = 'FFFFFF'
_LT_GRN = 'F0FDF4'
_GRAY   = 'F3F4F6'

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _font(bold=False, color=_NAVY, size=9, italic=False):
    return Font(name='Calibri', bold=bold, color=color, size=size, italic=italic)

def _align(h='left', wrap=True):
    return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

def _thin_border():
    s = Side(style='thin', color='D1D5DB')
    return Border(left=s, right=s, top=s, bottom=s)

def _write_sheet(ws, fields, data_rows, title_label):
    cols = len(fields)

    # Row 1 — title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=f'  Fusion Flow  —  {title_label} Demo Upload  |  Birkdale TSS Portal')
    c.fill = _fill(_NAVY)
    c.font = Font(name='Calibri', bold=True, color=_WHITE, size=13)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 32

    # Row 2 — DB column keys (machine-readable header)
    for i, f in enumerate(fields, 1):
        cell = ws.cell(row=2, column=i, value=f['col'])
        cell.fill = _fill(_NAVY)
        cell.font = _font(bold=True, color=_WHITE, size=9)
        cell.alignment = _align('center', wrap=False)
        cell.border = _thin_border()
    ws.row_dimensions[2].height = 22

    # Row 3 — human labels
    for i, f in enumerate(fields, 1):
        label = f['label'] + (' ★' if f.get('required') else '')
        cell = ws.cell(row=3, column=i, value=label)
        cell.fill = _fill(_BLUE)
        cell.font = _font(bold=True, color=_WHITE, size=9)
        cell.alignment = _align('center', wrap=True)
        cell.border = _thin_border()
    ws.row_dimensions[3].height = 28

    # Row 4 — notes
    for i, f in enumerate(fields, 1):
        parts = []
        if f.get('required'):  parts.append('REQUIRED')
        if f.get('max'):       parts.append(f"max {f['max']} chars")
        if f.get('cv'):        parts.append(f"CV: {f['cv']}")
        if f.get('allowed'):   parts.append('/'.join(f['allowed']))
        if f.get('note'):      parts.append(f['note'])
        cell = ws.cell(row=4, column=i, value=' | '.join(parts) or '—')
        cell.fill = _fill(_AMBER)
        cell.font = _font(italic=True, color=_WHITE, size=8)
        cell.alignment = _align('left', wrap=True)
        cell.border = _thin_border()
    ws.row_dimensions[4].height = 42

    # Data rows starting at row 5
    for row_idx, row_data in enumerate(data_rows, 5):
        for col_idx, f in enumerate(fields, 1):
            val = row_data.get(f['col'], '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = _fill(_LT_GRN)
            cell.font = _font(color='14532D', size=9)
            cell.alignment = _align('left', wrap=False)
            cell.border = _thin_border()
        ws.row_dimensions[row_idx].height = 18

    # Column widths
    wide = {'goods_description': 32, 'label': 22, 'staging_ens_id': 16,
            'staging_cons_id': 16, 'transport_document_number': 24,
            'package_marks': 18, 'commodity_code': 14}
    for i, f in enumerate(fields, 1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = \
            wide.get(f['col'], max(14, min(len(f['label']) + 4, 24)))

    ws.freeze_panes = 'A5'


CONSIGNMENT_FIELDS = [
    {'col': 'staging_ens_id',            'label': 'ENS Staging ID',         'required': True,  'note': 'ID from SQL seed output'},
    {'col': 'label',                     'label': 'Label',                   'required': False},
    {'col': 'goods_description',         'label': 'Goods Description',       'required': True,  'max': 254},
    {'col': 'trader_reference',          'label': 'Trader Reference',        'required': False, 'max': 100},
    {'col': 'transport_document_number', 'label': 'Transport Doc Number',    'required': True,  'max': 35},
    {'col': 'controlled_goods',          'label': 'Controlled Goods',        'required': True,  'allowed': ['yes', 'no']},
    {'col': 'goods_domestic_status',     'label': 'Domestic Status',         'required': False},
    {'col': 'destination_country',       'label': 'Destination Country',     'required': False},
    {'col': 'consignor_eori',            'label': 'Consignor EORI',          'required': False},
    {'col': 'consignor_name',            'label': 'Consignor Name',          'required': False, 'max': 35},
    {'col': 'consignee_eori',            'label': 'Consignee EORI',          'required': False},
    {'col': 'consignee_name',            'label': 'Consignee Name',          'required': False, 'max': 35},
    {'col': 'importer_eori',             'label': 'Importer EORI',           'required': True},
    {'col': 'importer_name',             'label': 'Importer Name',           'required': False, 'max': 35},
    {'col': 'exporter_eori',             'label': 'Exporter EORI',           'required': False},
    {'col': 'exporter_name',             'label': 'Exporter Name',           'required': False, 'max': 35},
    {'col': 'buyer_same_as_importer',    'label': 'Buyer=Importer?',         'required': False, 'allowed': ['yes', 'no']},
    {'col': 'seller_same_as_exporter',   'label': 'Seller=Exporter?',        'required': False, 'allowed': ['yes', 'no']},
    {'col': 'container_indicator',       'label': 'Container?',              'required': False, 'allowed': ['0', '1']},
]

GOODS_FIELDS = [
    {'col': 'staging_cons_id',           'label': 'Consignment Staging ID',  'required': True,  'note': 'ID from SQL seed output'},
    {'col': 'item_number',               'label': 'Item Number',             'required': False},
    {'col': 'label',                     'label': 'Label',                   'required': False},
    {'col': 'goods_description',         'label': 'Goods Description',       'required': True,  'max': 255},
    {'col': 'type_of_packages',          'label': 'Package Type',            'required': True,  'cv': 'CV_type_of_package'},
    {'col': 'number_of_packages',        'label': 'Package Qty',             'required': True},
    {'col': 'package_marks',             'label': 'Package Marks',           'required': True,  'max': 140},
    {'col': 'gross_mass_kg',             'label': 'Gross Mass KG',           'required': True},
    {'col': 'net_mass_kg',               'label': 'Net Mass KG',             'required': False},
    {'col': 'controlled_goods',          'label': 'Controlled?',             'required': True,  'allowed': ['yes', 'no']},
    {'col': 'commodity_code',            'label': 'Commodity Code',          'required': False, 'max': 10},
    {'col': 'country_of_origin',         'label': 'Country of Origin',       'required': False},
    {'col': 'item_invoice_amount',       'label': 'Invoice Amount',          'required': False},
    {'col': 'item_invoice_currency',     'label': 'Invoice Currency',        'required': False},
]


def build_consignment_xlsx(ens_id, out_path):
    ens_val = str(ens_id) if ens_id else 'REPLACE_WITH_ENS_STAGING_ID'
    rows = [
        {
            'staging_ens_id':            ens_val,
            'label':                     'DEMO CONS 001 — Golf Equipment',
            'goods_description':         'Golf clubs, golf bags and accessories',
            'trader_reference':          'TRADER-REF-DEMO-001',
            'transport_document_number': 'BOL-DEMO-20260509-001',
            'controlled_goods':          'no',
            'goods_domestic_status':     '',
            'destination_country':       'GB',
            'consignor_eori':            'IE123456789',
            'consignor_name':            'Golf Supplies Ireland Ltd',
            'consignee_eori':            'XI123456789000',
            'consignee_name':            'Birkdale Sales Ltd',
            'importer_eori':             'XI123456789000',
            'importer_name':             'Birkdale Sales Ltd',
            'exporter_eori':             'IE123456789',
            'exporter_name':             'Golf Supplies Ireland Ltd',
            'buyer_same_as_importer':    'yes',
            'seller_same_as_exporter':   'yes',
            'container_indicator':       '0',
        }
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    _write_sheet(ws, CONSIGNMENT_FIELDS, rows, 'Consignment')
    wb.save(out_path)
    print(f'Created: {out_path}')


def build_goods_xlsx(cons_id, out_path):
    cons_val = str(cons_id) if cons_id else 'REPLACE_WITH_CONS_STAGING_ID'
    rows = [
        {
            'staging_cons_id':       cons_val,
            'item_number':           '1',
            'label':                 'DEMO GOODS 001 — Golf Clubs',
            'goods_description':     'Golf clubs — irons and woods',
            'type_of_packages':      'BX',
            'number_of_packages':    '24',
            'package_marks':         'DEMO-PKG-001',
            'gross_mass_kg':         '120.00',
            'net_mass_kg':           '95.00',
            'controlled_goods':      'no',
            'commodity_code':        '9506310000',
            'country_of_origin':     'IE',
            'item_invoice_amount':   '3600.00',
            'item_invoice_currency': 'GBP',
        },
        {
            'staging_cons_id':       cons_val,
            'item_number':           '2',
            'label':                 'DEMO GOODS 002 — Golf Bags',
            'goods_description':     'Golf bags and carry accessories',
            'type_of_packages':      'BX',
            'number_of_packages':    '12',
            'package_marks':         'DEMO-PKG-002',
            'gross_mass_kg':         '60.00',
            'net_mass_kg':           '48.00',
            'controlled_goods':      'no',
            'commodity_code':        '9506390000',
            'country_of_origin':     'IE',
            'item_invoice_amount':   '1800.00',
            'item_invoice_currency': 'GBP',
        },
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    _write_sheet(ws, GOODS_FIELDS, rows, 'Goods Items')
    wb.save(out_path)
    print(f'Created: {out_path}')


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument('--ens-id',  type=int, default=None, help='staging_id from StagingEnsHeaders')
    p.add_argument('--cons-id', type=int, default=None, help='staging_id from StagingConsignments')
    p.add_argument('--out-dir', default='.', help='output directory')
    args = p.parse_args()

    out = Path(args.out_dir)
    out.mkdir(parents=True, exist_ok=True)

    build_consignment_xlsx(args.ens_id,  out / 'demo_consignment_upload.xlsx')
    build_goods_xlsx(args.cons_id,       out / 'demo_goods_upload.xlsx')

    print()
    print('Demo Excel files ready.')
    if not args.ens_id:
        print('  ⚠  staging_ens_id not set — update column A in demo_consignment_upload.xlsx')
        print('     with the ID printed by manual_demo_data_seed.sql')
    if not args.cons_id:
        print('  ⚠  staging_cons_id not set — update column A in demo_goods_upload.xlsx')
        print('     with the ID printed by manual_demo_data_seed.sql')
    print()
    print('Upload via: Bulk Upload -> Consignment  (then Goods Item)')


if __name__ == '__main__':
    main()
