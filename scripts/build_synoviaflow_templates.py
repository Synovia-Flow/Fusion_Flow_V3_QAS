"""Generate SynoviaFlow Excel upload templates aligned to TSS API v2.9.5.

Produces two files in app/static/templates/:
  * SynoviaFlow_Consignment_v2.9.5.xlsx          (no SD)
  * SynoviaFlow_Consignment_with_SD_v2.9.5.xlsx  (with SD extras)

Each workbook has three sheets:
  README              — operator instructions
  ENS Header          — Field/Value vertical form for transport metadata
  Consignments+Goods  — flat: one row per goods item, grouped by trader_reference

Run from repo root:
    python scripts/build_synoviaflow_templates.py
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
REQUIRED_FONT = Font(name="Calibri", size=10, bold=True, color="C53030")
SAMPLE_FONT = Font(name="Calibri", size=10, italic=True, color="6C7A89")
SECTION_FILL = PatternFill(start_color="E9F2FF", end_color="E9F2FF", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="0D6EFD")
THIN_BORDER = Border(
    left=Side(style="thin", color="C7CFD9"),
    right=Side(style="thin", color="C7CFD9"),
    top=Side(style="thin", color="C7CFD9"),
    bottom=Side(style="thin", color="C7CFD9"),
)


# ── ENS Header form ────────────────────────────────────────────────────

ENS_HEADER_FIELDS = [
    # (field_key, label, required, default sample, help)
    ("movement_type", "Movement Type", True, "3a",
     "TSS choice. e.g. 1, 1a, 3, 3a, 3b. RoRo Accompanied = 3a."),
    ("type_of_passive_transport", "Type Of Passive Transport", False,
     "TRAILER",
     "TSS choice value. For RoRo trailer movements use TRAILER."),
    ("identity_no_of_transport", "Identity Number Of Transport", True,
     "IMO1234567#TEST123",
     "Format: IMO<ferry>#<vehicle reg>. Required by TSS."),
    ("nationality_of_transport", "Nationality Of Transport", True, "GB",
     "ISO2. GB | IE | FR | DE | NL | BE."),
    ("carrier_eori", "Carrier EORI", True, "XI000012340005",
     "XI/EU EORI of the carrier. Replace the demo value with the real carrier."),
    ("carrier_name", "Carrier Name", False, "Demo Carrier Ltd",
     "Carrier legal/trading name. Used when TSS cannot auto-populate from EORI."),
    ("carrier_street_number", "Carrier Street / Number", False, "10 Harbour Road",
     "Carrier address line required when TSS cannot auto-populate from EORI."),
    ("carrier_city", "Carrier City", False, "Belfast",
     "Carrier city required when TSS cannot auto-populate from EORI."),
    ("carrier_postcode", "Carrier Postcode", False, "BT1 1AA",
     "Carrier postcode required when TSS cannot auto-populate from EORI."),
    ("carrier_country", "Carrier Country", False, "GB",
     "ISO2 country code for carrier address."),
    ("conveyance_ref", "Conveyance Ref / ICR", True, "ICR-DEMO-001",
     "TSS conveyance reference. ICR number from carrier."),
    ("arrival_date_time", "Arrival Date/Time", True, "Tomorrow / 06:30",
     "DD/MM/YYYY HH:MM or Tomorrow / HH:MM. Cannot be in the past at submit time."),
    ("arrival_port", "Arrival Port", True, "GBAUBELBELBEL",
     "TSS port code. Belfast = GBAUBELBELBEL. Larne = GBAULARLARLAR."),
    ("place_of_loading", "Place Of Loading", True, "Birkenhead", "Free text origin."),
    ("place_of_acceptance_same_as_loading", "Accept = Loading", False, "yes",
     "TSS v2.9.5 RoRo field. Use yes when acceptance is the same as loading."),
    ("place_of_acceptance", "Place Of Acceptance", False, "",
     "Only required when Accept = Loading is no."),
    ("place_of_unloading", "Place Of Unloading", True, "Belfast", "Free text destination at port."),
    ("place_of_delivery_same_as_unloading", "Delivery = Unloading", False, "yes",
     "TSS v2.9.5 RoRo field. Use yes when delivery is the same as unloading."),
    ("place_of_delivery", "Place Of Delivery", False, "",
     "Only required when Delivery = Unloading is no."),
    ("transport_charges", "Transport Charges", True, "Y",
     "TSS code. Y = Account holder with carrier. A = Cash."),
    ("haulier_eori", "Haulier EORI", False, "",
     "GB EORI if different from carrier. Empty if same."),
]


# ── Consignments + Goods columns ──────────────────────────────────────

# (column_key, label, required, sample, help, group)
BASE_COLUMNS = [
    ("trader_reference", "Trader Reference", True, "DEMO-ORD-1001",
     "Grouping key. Same value = same consignment. New value = new consignment.", "consignment"),
    ("transport_document_number", "Transport Doc Number", True, "TDOC-DEMO-1001",
     "CMR / Bill of Lading. Falls back to trader_reference if no separate doc.", "consignment"),
    ("destination_country", "Destination Country", True, "GB", "ISO2.", "consignment"),
    ("goods_domestic_status", "Goods Domestic Status", True, "D",
     "D = Domestic / F = Foreign / N = Non-Union.", "consignment"),
    ("controlled_goods", "Controlled Goods", True, "no", "yes | no.", "consignment"),
    ("container_indicator", "Container Indicator", True, "0", "0 = uncontainerised, 1 = containerised.", "consignment"),
    ("no_sfd_reason", "No SFD Reason / ENS Only Reason", False, "",
     "Leave blank to use SFD. Fill only when this consignment must opt out of SFD creation.", "consignment"),
    ("ducr", "DUCR", False, "", "Declaration Unique Consignment Reference. Optional for ENS.", "consignment"),
    ("align_ukims", "Align UKIMS", False, "no", "yes | no. Leave no unless UKIMS alignment is required.", "consignment"),
    ("use_importer_sde", "Use Importer SDE", False, "no", "yes | no. Leave no unless instructed.", "consignment"),
    ("supervising_customs_office", "Supervising Customs Office", False, "",
     "Only required for customs warehouse / special flows.", "consignment"),
    ("customs_warehouse_identifier", "Customs Warehouse Identifier", False, "",
     "Only required for customs warehouse / special flows.", "consignment"),
    ("consignor_eori", "Consignor EORI", True, "XI000012340005",
     "XI/EU EORI. Replace demo value with the real shipper EORI.", "consignment"),
    ("consignor_name", "Consignor Name", True, "Demo Exporter Ltd",
     "Legal name. Required when TSS cannot auto-populate from EORI.", "consignment"),
    ("consignor_street_number", "Consignor Street And Number", True, "10 Export Road",
     "Legal address line.", "consignment"),
    ("consignor_city", "Consignor City", True, "London", "Legal address city.", "consignment"),
    ("consignor_postcode", "Consignor Postcode", True, "N1 8LN", "Legal address postcode.", "consignment"),
    ("consignor_country", "Consignor Country", True, "GB", "ISO2.", "consignment"),
    ("consignee_eori", "Consignee EORI", False, "XI000012340005",
     "Optional if consignee address is supplied and EORI is unknown.", "consignment"),
    ("consignee_name", "Consignee Name", True, "Demo Consignee Ltd", "Required if no EORI.", "consignment"),
    ("consignee_street_number", "Consignee Street And Number", True, "20 Delivery Street",
     "Required if no EORI or TSS asks for address.", "consignment"),
    ("consignee_city", "Consignee City", True, "Belfast", "Required if no EORI or TSS asks for address.", "consignment"),
    ("consignee_postcode", "Consignee Postcode", True, "BT1 1AA",
     "Required if no EORI or TSS asks for address.", "consignment"),
    ("consignee_country", "Consignee Country", True, "GB", "GB for NI movements.", "consignment"),
    ("importer_eori", "Importer EORI", True, "XI000012340005",
     "XI/EU/GB EORI. Required.", "consignment"),
    ("importer_name", "Importer Name", True, "Demo Importer Ltd", "Legal importer name.", "consignment"),
    ("importer_street_number", "Importer Street And Number", True, "30 Import Way",
     "Legal importer address line.", "consignment"),
    ("importer_city", "Importer City", True, "Belfast", "Legal importer city.", "consignment"),
    ("importer_postcode", "Importer Postcode", True, "BT2 8BG", "Legal importer postcode.", "consignment"),
    ("importer_country", "Importer Country", True, "GB", "ISO2.", "consignment"),
    ("buyer_same_as_importer", "Buyer Same As Importer", True, "yes",
     "yes = TSS derives buyer from importer. no = fill buyer fields.", "consignment"),
    ("buyer_eori", "Buyer EORI", False, "", "Only fill when buyer_same_as_importer is no.", "consignment"),
    ("buyer_name", "Buyer Name", False, "", "Only fill when buyer_same_as_importer is no.", "consignment"),
    ("buyer_street_and_number", "Buyer Street And Number", False, "",
     "Only fill when buyer_same_as_importer is no.", "consignment"),
    ("buyer_city", "Buyer City", False, "", "Only fill when buyer_same_as_importer is no.", "consignment"),
    ("buyer_postcode", "Buyer Postcode", False, "", "Only fill when buyer_same_as_importer is no.", "consignment"),
    ("buyer_country", "Buyer Country", False, "", "Only fill when buyer_same_as_importer is no.", "consignment"),
    ("exporter_eori", "Exporter EORI", True, "XI000012340005", "XI/EU EORI of exporter/seller.", "consignment"),
    ("exporter_name", "Exporter Name", True, "Demo Exporter Ltd", "Legal exporter name.", "consignment"),
    ("exporter_street_number", "Exporter Street And Number", True, "10 Export Road",
     "Legal exporter address line.", "consignment"),
    ("exporter_city", "Exporter City", True, "London", "Legal exporter city.", "consignment"),
    ("exporter_postcode", "Exporter Postcode", True, "N1 8LN", "Legal exporter postcode.", "consignment"),
    ("exporter_country", "Exporter Country", True, "GB", "ISO2.", "consignment"),
    ("seller_same_as_exporter", "Seller Same As Exporter", True, "yes",
     "yes = TSS derives seller from exporter. no = fill seller fields.", "consignment"),
    ("seller_eori", "Seller EORI", False, "", "Only fill when seller_same_as_exporter is no.", "consignment"),
    ("seller_name", "Seller Name", False, "", "Only fill when seller_same_as_exporter is no.", "consignment"),
    ("seller_street_and_number", "Seller Street And Number", False, "",
     "Only fill when seller_same_as_exporter is no.", "consignment"),
    ("seller_city", "Seller City", False, "", "Only fill when seller_same_as_exporter is no.", "consignment"),
    ("seller_postcode", "Seller Postcode", False, "", "Only fill when seller_same_as_exporter is no.", "consignment"),
    ("seller_country", "Seller Country", False, "", "Only fill when seller_same_as_exporter is no.", "consignment"),
    # Goods columns
    ("commodity_code", "Commodity Code", True, "8708299000",
     "10-digit HS code.", "goods"),
    ("taric_code", "TARIC Code", False, "87082990",
     "v2.9.5: compact 4-char segments concatenated, no spaces (e.g. 87082990).", "goods"),
    ("goods_description", "Goods Description", True, "Car parts for demo testing",
     "Max 255 chars. No EU-prohibited words.", "goods"),
    ("type_of_packages", "Type Of Packages", True, "BX",
     "TSS choice. BX = Box. PK = Package. PA = Pallet. CT = Carton.", "goods"),
    ("number_of_packages", "Number Of Packages", True, 1, "Integer >= 1.", "goods"),
    ("number_of_individual_pieces", "Number Of Individual Pieces", False, "",
     "Optional piece count inside packages.", "goods"),
    ("package_marks", "Package Marks", True, "ADDR", "ADDR if not known.", "goods"),
    ("gross_mass_kg", "Gross Mass KG", True, 100.0, "Decimal kilograms.", "goods"),
    ("net_mass_kg", "Net Mass KG", True, 90.0, "Decimal kilograms. <= gross.", "goods"),
    ("country_of_origin", "Country Of Origin", True, "GB", "ISO2.", "goods"),
    ("item_invoice_amount", "Item Invoice Amount", True, 500.00,
     "Decimal. Per goods line.", "goods"),
    ("item_invoice_currency", "Invoice Currency", True, "GBP", "ISO4217. GBP / EUR / USD.", "goods"),
    ("procedure_code", "Procedure Code", True, "4000",
     "TSS procedure_code. e.g. 4000 = home use of free-circulation goods.", "goods"),
    ("additional_procedure_code", "Additional Procedure", True, "000", "TSS code.", "goods"),
    ("controlled_goods_type", "Controlled Goods Type", False, "",
     "Required only when controlled_goods is yes.", "goods"),
    ("cus_code", "CUS Code", False, "", "Optional customs union and statistics code.", "goods"),
    ("national_additional_code", "National Additional Code", False, "", "Optional national code.", "goods"),
    ("ni_additional_information_codes", "NI Additional Information Codes", False, "NIDOM",
     "Optional NI code such as NIDOM / NIREM / NIIMP.", "goods"),
    ("country_of_preferential_origin", "Country Of Preferential Origin", False, "", "ISO2 if applicable.", "goods"),
    ("preference", "Preference", False, "100", "TSS preference code.", "goods"),
    ("valuation_method", "Valuation Method", True, "1", "TSS code 1-6.", "goods"),
    ("valuation_indicator", "Valuation Indicator", False, "", "Optional valuation indicator.", "goods"),
    ("invoice_number", "Invoice Number", False, "INV-DEMO-1001", "Invoice reference, max 35 chars.", "goods"),
    ("nature_of_transaction", "Nature Of Transaction", False, "1", "TSS code.", "goods"),
    ("statistical_value", "Statistical Value", False, "", "Optional statistical value.", "goods"),
    ("quota_order_number", "Quota Order Number", False, "", "Optional quota order number.", "goods"),
    ("supplementary_units", "Supplementary Units", False, "", "Only where tariff requires supplementary units.", "goods"),
    ("tax_type", "Tax Type", False, "", "Optional tax type.", "goods"),
    ("tax_base_unit", "Tax Base Unit", False, "", "Optional tax base unit.", "goods"),
    ("tax_base_quantity", "Tax Base Quantity", False, "", "Optional tax base quantity.", "goods"),
    ("payable_tax_amount", "Payable Tax Amount", False, "", "Optional payable tax amount.", "goods"),
    ("payable_tax_currency", "Payable Tax Currency", False, "", "Optional payable tax currency.", "goods"),
]

SD_EXTRA_COLUMNS = [
    ("generate_SD", "Generate SDI/SupDec", True, "no",
     "yes = SDI/SupDec known in advance. TSS/SFD discovery still decides what exists.", "sd"),
    ("declaration_choice", "Declaration Choice", True, "H1",
     "H1 / H2 / H3 / H4. SDI category.", "sd"),
    ("incoterm", "Incoterm", True, "CFR", "TSS incoterm code.", "sd"),
    ("freight_charge", "Freight Charge", False, 100.0, "Decimal.", "sd"),
    ("freight_charge_currency", "Freight Currency", False, "GBP", "ISO4217.", "sd"),
    ("insurance", "Insurance", False, 0.0, "Decimal.", "sd"),
    ("insurance_currency", "Insurance Currency", False, "GBP", "ISO4217.", "sd"),
    ("vat_adjustment", "VAT Adjustment", False, 0.0, "Decimal.", "sd"),
    ("vat_adjust_currency", "VAT Currency", False, "GBP", "ISO4217.", "sd"),
    ("postponed_vat", "Postponed VAT", False, "no", "yes | no.", "sd"),
    ("exchange_rate", "Exchange Rate", False, "", "Decimal. Optional override.", "sd"),
]


# ── Sample rows for the demo template ────────────────────────────────

DEMO_COMMON = {
    "destination_country": "GB",
    "goods_domestic_status": "D",
    "controlled_goods": "no",
    "container_indicator": "0",
    "no_sfd_reason": "",
    "ducr": "",
    "align_ukims": "no",
    "use_importer_sde": "no",
    "supervising_customs_office": "",
    "customs_warehouse_identifier": "",
    "consignor_eori": "XI000012340005",
    "consignor_name": "Demo Exporter Ltd",
    "consignor_street_number": "10 Export Road",
    "consignor_city": "London",
    "consignor_postcode": "N1 8LN",
    "consignor_country": "GB",
    "consignee_eori": "XI000012340005",
    "consignee_name": "Demo Consignee Ltd",
    "consignee_street_number": "20 Delivery Street",
    "consignee_city": "Belfast",
    "consignee_postcode": "BT1 1AA",
    "consignee_country": "GB",
    "importer_eori": "XI000012340005",
    "importer_name": "Demo Importer Ltd",
    "importer_street_number": "30 Import Way",
    "importer_city": "Belfast",
    "importer_postcode": "BT2 8BG",
    "importer_country": "GB",
    "buyer_same_as_importer": "yes",
    "buyer_eori": "",
    "buyer_name": "",
    "buyer_street_and_number": "",
    "buyer_city": "",
    "buyer_postcode": "",
    "buyer_country": "",
    "exporter_eori": "XI000012340005",
    "exporter_name": "Demo Exporter Ltd",
    "exporter_street_number": "10 Export Road",
    "exporter_city": "London",
    "exporter_postcode": "N1 8LN",
    "exporter_country": "GB",
    "seller_same_as_exporter": "yes",
    "seller_eori": "",
    "seller_name": "",
    "seller_street_and_number": "",
    "seller_city": "",
    "seller_postcode": "",
    "seller_country": "",
    "country_of_origin": "GB",
    "procedure_code": "4000",
    "additional_procedure_code": "000",
    "controlled_goods_type": "",
    "cus_code": "",
    "national_additional_code": "",
    "ni_additional_information_codes": "NIDOM",
    "country_of_preferential_origin": "",
    "preference": "100",
    "valuation_method": "1",
    "valuation_indicator": "",
    "nature_of_transaction": "1",
    "statistical_value": "",
    "quota_order_number": "",
    "supplementary_units": "",
    "tax_type": "",
    "tax_base_unit": "",
    "tax_base_quantity": "",
    "payable_tax_amount": "",
    "payable_tax_currency": "",
}

DEMO_ROWS = [
    # Consignment 1: two goods items.
    {
        **DEMO_COMMON,
        "trader_reference": "DEMO-ORD-1001",
        "transport_document_number": "TDOC-DEMO-1001",
        "commodity_code": "8708299000",
        "taric_code": "87082990",
        "goods_description": "Car parts for demo testing",
        "type_of_packages": "BX",
        "number_of_packages": 1,
        "number_of_individual_pieces": 12,
        "package_marks": "ADDR",
        "gross_mass_kg": 100.0,
        "net_mass_kg": 90.0,
        "item_invoice_amount": 500.00,
        "item_invoice_currency": "GBP",
        "invoice_number": "INV-DEMO-1001",
    },
    {
        **DEMO_COMMON,
        "trader_reference": "DEMO-ORD-1001",
        "transport_document_number": "TDOC-DEMO-1001",
        "commodity_code": "7318159000",
        "taric_code": "73181590",
        "goods_description": "Threaded fasteners for demo testing",
        "type_of_packages": "BX",
        "number_of_packages": 2,
        "number_of_individual_pieces": 40,
        "package_marks": "ADDR",
        "gross_mass_kg": 20.0,
        "net_mass_kg": 18.0,
        "item_invoice_amount": 150.00,
        "item_invoice_currency": "GBP",
        "invoice_number": "INV-DEMO-1001",
    },
    # Consignment 2: one goods item.
    {
        **DEMO_COMMON,
        "trader_reference": "DEMO-ORD-1002",
        "transport_document_number": "TDOC-DEMO-1002",
        "commodity_code": "3926909790",
        "taric_code": "39269097",
        "goods_description": "Plastic fittings for demo testing",
        "type_of_packages": "CT",
        "number_of_packages": 3,
        "number_of_individual_pieces": 24,
        "package_marks": "ADDR",
        "gross_mass_kg": 30.0,
        "net_mass_kg": 27.0,
        "item_invoice_amount": 225.00,
        "item_invoice_currency": "GBP",
        "invoice_number": "INV-DEMO-1002",
    },
]


def _set_col_widths(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _build_readme(ws, with_sd: bool):
    ws.title = "README"
    title = "Synovia Flow — Consignment Upload Template (v2.9.5)" + (
        " — with Supplementary Declaration" if with_sd else ""
    )

    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="0D6EFD")
    ws.merge_cells("A1:F1")

    rows = [
        "",
        ("Purpose", "Stage one ENS header plus N consignments and M goods items into Fusion Flow."),
        ("End-to-end flow", "Upload this xlsx via /ingest dropzone -> backend stages everything as PENDING_REVIEW -> review queue -> validate -> submit to TSS."),
        "",
        ("Sheet 'ENS Header'", "Vertical Field / Value form. Fill the carrier transport metadata (one ENS = one truck movement)."),
        ("Sheet 'Consignments+Goods'", "Flat. One row = one goods item. Group by trader_reference: rows sharing the same trader_reference become one consignment."),
        "",
        ("trader_reference", "Grouping key (e.g. DEMO-ORD-1001). Mandatory. Same value across rows = same consignment."),
        ("transport_document_number", "CMR / Bill of Lading. If you only have the order number, repeat trader_reference here."),
        ("commodity_code", "10-digit HS code per goods line. Required."),
        ("taric_code", "TSS v2.9.5 requires compact format: 4-char segments concatenated with no spaces."),
        ("gross_mass_kg / net_mass_kg", "Decimal kilograms. net <= gross."),
        ("item_invoice_amount", "Per-line invoice value. Currency in item_invoice_currency."),
        ("no_sfd_reason", "Leave blank for the normal TSS path where SFD is used. Fill a TSS no_sfd_reason choice only for ENS-only journeys."),
        "",
        ("Required vs optional", "Cells highlighted RED in row 3 are mandatory. Empty optional cells fall back to tenant defaults from Master Data > Operational Defaults."),
        ("Anti-duplicates", "Re-uploading the same workbook is idempotent. Natural keys: ENS by (conveyance_ref + arrival_date_time + identity_no_of_transport); consignment by (ens_id + trader_reference); goods by (cons_id + line index + commodity_code)."),
        "",
        ("Demo path", "This template ships with neutral sample rows only. Replace demo EORIs, parties, addresses, products and invoice values before live use."),
    ]
    if with_sd:
        rows.extend([
            "",
            ("SD extras", "Sheet 'Consignments+Goods' has additional columns at the right for SDI: generate_SD, declaration_choice, incoterm, freight_charge, etc. Fill them when SDI data is known at ingest time."),
            ("v2.9.5 mandatory", "header_additions_deductions[].addition_deduction_currency is now mandatory at submit. Operator fills it in the SDI review screen."),
        ])
    rows.extend([
        "",
        ("Endpoint", "POST /ingest/preview-batch (xlsx routes to sales-orders pipeline automatically)."),
    ])

    r = 2
    for row in rows:
        if row == "":
            r += 1
            continue
        if isinstance(row, tuple):
            ws.cell(row=r, column=1, value=row[0]).font = Font(bold=True)
            ws.cell(row=r, column=2, value=row[1]).alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 30
        else:
            ws.cell(row=r, column=1, value=row)
        r += 1

    _set_col_widths(ws, [28, 22, 22, 22, 22, 22])


def _build_ens_sheet(ws):
    ws.title = "ENS Header"
    ws["A1"] = "ENS Header — fill the Value column. Required = RED."
    ws["A1"].font = Font(bold=True, size=12, color="0D6EFD")
    ws.merge_cells("A1:C1")

    headers = ["Field", "Value", "Help"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left")

    for r_idx, (key, label, required, sample, helptext) in enumerate(ENS_HEADER_FIELDS, start=3):
        cell_field = ws.cell(row=r_idx, column=1, value=label)
        cell_field.font = REQUIRED_FONT if required else Font(name="Calibri", size=10)
        cell_field.alignment = Alignment(vertical="center")
        cell_field.border = THIN_BORDER

        # Hidden machine key in column D so parser can locate values robustly
        ws.cell(row=r_idx, column=4, value=key)

        cell_val = ws.cell(row=r_idx, column=2, value=sample)
        cell_val.alignment = Alignment(vertical="center")
        cell_val.border = THIN_BORDER

        cell_help = ws.cell(row=r_idx, column=3, value=helptext)
        cell_help.font = Font(name="Calibri", size=9, color="6C7A89")
        cell_help.alignment = Alignment(wrap_text=True, vertical="center")
        cell_help.border = THIN_BORDER

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].hidden = True
    ws.column_dimensions["D"].width = 28


def _build_data_sheet(ws, with_sd: bool):
    ws.title = "Consignments+Goods"
    columns = list(BASE_COLUMNS)
    if with_sd:
        columns.extend(SD_EXTRA_COLUMNS)

    # Row 1: section banner
    section_label = ("Consignment fields repeat across rows of same trader_reference. "
                     "Goods fields are unique per row. "
                     + ("SDI fields apply to all rows of a consignment. " if with_sd else "")
                     + "Required columns are RED in row 2.")
    ws.cell(row=1, column=1, value=section_label).font = SECTION_FONT
    ws.cell(row=1, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(columns), 18))

    # Row 2: machine keys (parser reads these)
    for c_idx, (key, _label, _req, _sample, _help, _grp) in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=c_idx, value=key)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # Row 3: human labels with required formatting
    for c_idx, (_key, label, required, _sample, helptext, _grp) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=c_idx, value=label)
        cell.font = REQUIRED_FONT if required else Font(name="Calibri", size=10, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER
        cell.fill = SECTION_FILL
        # Help comment in cell tooltip
        if helptext:
            from openpyxl.comments import Comment
            cell.comment = Comment(helptext, "Synovia Flow")

    # Row 4 onward: sample / demo data
    for r_offset, demo_row in enumerate(DEMO_ROWS, start=4):
        for c_idx, (key, _label, _req, _sample, _help, _grp) in enumerate(columns, start=1):
            value = demo_row.get(key)
            if value is None and key in dict([(k, s) for k, _l, _r, s, _h, _g in columns]):
                # If demo row doesn't override, fall back to column sample
                value = next((s for k, _l, _r, s, _h, _g in columns if k == key), "")
            ws.cell(row=r_offset, column=c_idx, value=value)

    # Column widths
    widths = []
    for key, _label, _req, _sample, _help, _grp in columns:
        if key in ("goods_description",):
            widths.append(36)
        elif key.endswith("_name") or key in ("type_of_passive_transport",):
            widths.append(28)
        elif key in ("trader_reference", "transport_document_number", "carrier_eori",
                     "consignor_eori", "consignee_eori", "importer_eori", "exporter_eori"):
            widths.append(20)
        else:
            widths.append(16)
    _set_col_widths(ws, widths)
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "A4"


def build(with_sd: bool, out_path: Path):
    wb = openpyxl.Workbook()
    _build_readme(wb.active, with_sd)
    _build_ens_sheet(wb.create_sheet())
    _build_data_sheet(wb.create_sheet(), with_sd)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")


def main():
    repo_root = Path(__file__).resolve().parent.parent
    out_dir = repo_root / "app" / "static" / "templates"
    build(with_sd=False, out_path=out_dir / "SynoviaFlow_Consignment_v2.9.5.xlsx")
    build(with_sd=True, out_path=out_dir / "SynoviaFlow_Consignment_with_SD_v2.9.5.xlsx")


if __name__ == "__main__":
    main()
