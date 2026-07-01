r"""
======================================================================
Synovia Fusion :: shared branding for scripts (Excel / decks / charts)
Module   : fusion_branding.py
Purpose  : One import so any solution's Python output matches the portal
           brand — same palette, fonts and helpers used by the PTI
           workbooks and the Process-Flow deck. Pure-stdlib; the
           openpyxl/pptx helpers import lazily so this has no hard deps.

Usage:
   from fusion_branding import COLORS, HEADER_FILL_HEX, FONT_NAME, xl_header
   # openpyxl:
   for c in ws[1]:
       xl_header(c)
   # python-pptx:
   from fusion_branding import rgb  # -> pptx RGBColor
======================================================================
"""
from __future__ import annotations

# ── palette (hex, no leading '#') — mirrors fusion_design_tokens.json ──
COLORS = {
    "navy":      "0B1D3A",
    "teal":      "00C4B4",
    "teal_deep": "00A99B",
    "accent":    "F97316",
    "sky":       "0EA5E9",
    "green":     "198754",
    "amber":     "F97316",
    "rose":      "DC3545",
    "violet":    "7C3AED",
    "bg":        "F0F4FF",
    "surface":   "FFFFFF",
    "text":      "1E2A3A",
    "muted":     "6B7A8D",
    "border":    "D8E0EE",
    "white":     "FFFFFF",
}

FONT_NAME = "Montserrat"          # web brand font
FONT_FALLBACK = "Segoe UI"        # for Excel/desktop where Montserrat is absent

HEADER_FILL_HEX = COLORS["navy"]  # table/KPI header band
HEADER_TEXT_HEX = COLORS["white"]
ACCENT_HEX = COLORS["sky"]

# Status -> colour, matching the portal badges.
STATUS_COLORS = {
    "SUCCESS": COLORS["green"], "OK": COLORS["green"], "CONFIRMED": COLORS["green"],
    "PARTIAL": COLORS["amber"], "WARN": COLORS["amber"], "PENDING": COLORS["amber"],
    "FAILED": COLORS["rose"], "ERROR": COLORS["rose"], "TIMEOUT": COLORS["rose"],
    "INFO": COLORS["violet"],
}


def hex_to_rgb(h: str) -> tuple[int, int, int]:
    h = h.lstrip("#")
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)


def rgb(name_or_hex: str):
    """python-pptx RGBColor for a palette name or hex (lazy import)."""
    from pptx.dml.color import RGBColor
    h = COLORS.get(name_or_hex, name_or_hex).lstrip("#")
    return RGBColor.from_string(h)


# ── openpyxl helpers (lazy import) ────────────────────────────────────
def xl_header(cell, fill_hex: str = HEADER_FILL_HEX, text_hex: str = HEADER_TEXT_HEX):
    """Style a header cell: brand fill, white bold Segoe UI."""
    from openpyxl.styles import Font, PatternFill
    cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.font = Font(color=text_hex, bold=True, name=FONT_FALLBACK)
    return cell


def xl_status_font(cell, status: str):
    """Colour a cell's text by status (green/amber/rose/...)."""
    from openpyxl.styles import Font
    col = STATUS_COLORS.get(str(status).upper())
    if col:
        cell.font = Font(color=col, bold=True, name=FONT_FALLBACK)
    return cell


def style_header_row(ws, ncols: int | None = None,
                     fill_hex: str = HEADER_FILL_HEX, freeze: bool = True):
    """Apply the brand header band to row 1 and freeze it + autofilter."""
    from openpyxl.utils import get_column_letter
    n = ncols or ws.max_column
    for c in ws[1][:n]:
        xl_header(c, fill_hex)
    if freeze:
        ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n)}{ws.max_row}"
    return ws


if __name__ == "__main__":
    print("Synovia Fusion palette:")
    for k, v in COLORS.items():
        print(f"  {k:10s} #{v}  rgb{hex_to_rgb(v)}")
    print(f"\nFont: {FONT_NAME} (web) / {FONT_FALLBACK} (Excel)")
